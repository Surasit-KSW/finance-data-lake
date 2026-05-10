"""
scripts/migrate_to_neon.py
===========================
Migrate Finance Data Lake (Silver Parquet + TB Excel) to Neon PostgreSQL.

Run once locally to seed the cloud database:
    python scripts/migrate_to_neon.py

Prerequisites:
    1. Set DATABASE_URL in .env:
       DATABASE_URL=postgresql://user:pass@ep-xxx.neon.tech/neondb?sslmode=require
    2. Schema must exist:
       psql "$DATABASE_URL" -f scripts/setup_neon_schema.sql
    3. pip install psycopg2-binary pandas pyarrow openpyxl python-dotenv
"""

import sys
import os
import glob
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd
import psycopg2
from psycopg2.extras import execute_values

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if not DATABASE_URL:
    print("ERROR: DATABASE_URL not set.")
    print("  Add to .env: DATABASE_URL=postgresql://user:pass@ep-xxx.neon.tech/neondb?sslmode=require")
    sys.exit(1)

SILVER = ROOT / "02_Silver_Cleaned"
GOLD   = ROOT / "03_Gold_DataMarts"
BRONZE = ROOT / "01_Bronze_Raw"


# ── Helpers ──────────────────────────────────────────────────────────────────

def get_conn():
    return psycopg2.connect(DATABASE_URL, connect_timeout=30)


def log_meta(conn, table, rows, source):
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO _lake_meta (table_name, row_count, loaded_at, source_file)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (table_name) DO UPDATE
              SET row_count    = EXCLUDED.row_count,
                  loaded_at    = EXCLUDED.loaded_at,
                  source_file  = EXCLUDED.source_file
            """,
            (table, rows, datetime.now(), source),
        )
    conn.commit()


def truncate_and_load(conn, table, df, source):
    """Truncate table and bulk-insert DataFrame rows."""
    df2 = df[[c for c in df.columns if c.lower() != "id"]]
    col_names = ", ".join(f'"{c}"' for c in df2.columns)
    rows = [tuple(r) for r in df2.itertuples(index=False, name=None)]
    with conn.cursor() as cur:
        cur.execute(f"TRUNCATE TABLE {table} RESTART IDENTITY")
        execute_values(
            cur,
            f"INSERT INTO {table} ({col_names}) VALUES %s",
            rows,
            page_size=500,
        )
    conn.commit()
    log_meta(conn, table, len(df2), source)
    print(f"  [{table}] {len(df2):>8,} rows  <- {source}")


# ── Loaders ──────────────────────────────────────────────────────────────────

def load_gl(conn):
    f = SILVER / "Master_GL_24_25.parquet"
    if not f.exists():
        print(f"  [v_gl] SKIPPED — {f.name} not found")
        return
    df = pd.read_parquet(f)
    for col in ("Year", "Month"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    schema_cols = [
        "Posting Date", "Document Date", "G/L Account",
        "G/L Account: Long Text", "Net_Amount", "Text",
        "Document Number", "Document Type",
        "Cost Center", "Cost Center: Short Text", "Profit Center",
        "Year", "Month",
    ]
    keep = [c for c in schema_cols if c in df.columns]
    truncate_and_load(conn, "v_gl", df[keep], f.name)


def load_sales(conn):
    files = sorted(glob.glob(str(SILVER / "master_sales_*.parquet")))
    if not files:
        print("  [v_sales] SKIPPED — no master_sales_*.parquet files found")
        return
    df = pd.concat([pd.read_parquet(p) for p in files], ignore_index=True)
    for col in ("Year", "Month"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    keep = [c for c in ["Year", "Month", "Product Group", "Product", "Customer",
                         "Net Sales", "Quantity", "Unit"] if c in df.columns]
    truncate_and_load(conn, "v_sales", df[keep], f"{len(files)} parquet files")


def load_production(conn):
    files = sorted(glob.glob(str(SILVER / "master_production_*.parquet")))
    if not files:
        print("  [v_production] SKIPPED — no master_production_*.parquet files found")
        return
    df = pd.concat([pd.read_parquet(p) for p in files], ignore_index=True)
    for col in ("Year", "Month"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    keep = [c for c in ["Year", "Month", "Plant", "Order No", "Material",
                         "Production Qty", "Total Cost"] if c in df.columns]
    truncate_and_load(conn, "v_production", df[keep], f"{len(files)} parquet files")


def load_tb(conn):
    """Extract TB from Excel files and insert into tb_data table."""
    import openpyxl
    loaded = 0

    # Mar 2026 — SAP TB Export
    sap_tb = BRONZE / "Inventory_RollStock" / "NRV" / "AMC_TB_03.2026_v9.XLSX"
    if sap_tb.exists():
        records = []
        wb = openpyxl.load_workbook(sap_tb, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[2] if len(row) > 2 else None
            if raw is None or str(raw).strip() in ("", "Account Number"):
                continue
            try:
                acct = str(int(float(str(raw)))).strip()
            except (ValueError, TypeError):
                continue
            name = str(row[1]).strip() if row[1] else ""
            if name.startswith(acct):
                name = name[len(acct):].strip()
            bal = float(row[3]) if isinstance(row[3], (int, float)) else 0.0
            records.append(("2026-03-31", acct, name, bal, str(row[0]) if row[0] else ""))
        wb.close()
        if records:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM tb_data WHERE period = %s", ("2026-03-31",))
                execute_values(
                    cur,
                    "INSERT INTO tb_data (period, account_code, account_name, balance, fs_item) VALUES %s",
                    records,
                    page_size=500,
                )
            conn.commit()
            print(f"  [tb_data] {len(records):>6,} rows  <- 2026-03-31 ({sap_tb.name})")
            loaded += 1

    # Dec 2025 — YE25 Leadsheet detail sheets (col G = Dec 31 balance)
    ye25 = BRONZE / "Templates" / "AMC Leadsheet YE25.xlsx"
    if ye25.exists():
        seen: dict = {}
        wb = openpyxl.load_workbook(ye25, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            if sh in {"F1", "F2", "F3"}:
                continue
            ws = wb[sh]
            for row in ws.iter_rows(min_row=6, values_only=True):
                raw = row[1] if len(row) > 1 else None
                if raw is None:
                    continue
                try:
                    acct = str(int(float(str(raw)))).strip()
                except (ValueError, TypeError):
                    continue
                bal = row[6] if len(row) > 6 else None
                if isinstance(bal, (int, float)) and acct not in seen:
                    name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    seen[acct] = ("2025-12-31", acct, name, float(bal), "")
        wb.close()
        if seen:
            rows = list(seen.values())
            with conn.cursor() as cur:
                cur.execute("DELETE FROM tb_data WHERE period = %s", ("2025-12-31",))
                execute_values(
                    cur,
                    "INSERT INTO tb_data (period, account_code, account_name, balance, fs_item) VALUES %s",
                    rows,
                    page_size=500,
                )
            conn.commit()
            print(f"  [tb_data] {len(rows):>6,} rows  <- 2025-12-31 ({ye25.name})")
            loaded += 1

    if loaded == 0:
        print("  [tb_data] SKIPPED — no TB source files found")
    log_meta(conn, "tb_data", loaded, "SAP TB + YE25 Leadsheet")


def load_ar(conn):
    f = SILVER / "master_ar.parquet"
    if not f.exists():
        print(f"  [v_ar] SKIPPED — {f.name} not found")
        return
    df = pd.read_parquet(f)
    keep = [c for c in ["Customer Code", "Customer Name", "Invoice No",
                         "Invoice Date", "Due Date", "Outstanding", "Aging Bucket"]
            if c in df.columns]
    truncate_and_load(conn, "v_ar", df[keep], f.name)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 60)
    print("  Finance Data Lake — Migrate to Neon PostgreSQL")
    print("=" * 60)
    host = DATABASE_URL.split("@")[-1] if "@" in DATABASE_URL else DATABASE_URL
    print(f"\n  Target : {host}")
    print(f"  Source : {ROOT}\n")

    conn = get_conn()
    print("  Connected to Neon PostgreSQL.\n")
    print("Loading tables:")

    load_gl(conn)
    load_sales(conn)
    load_production(conn)
    load_tb(conn)
    load_ar(conn)

    with conn.cursor() as cur:
        cur.execute(
            "SELECT table_name, row_count, loaded_at FROM _lake_meta ORDER BY table_name"
        )
        meta = cur.fetchall()

    print("\n" + "=" * 60)
    print("  Migration complete!")
    print("=" * 60)
    for table, count, ts in meta:
        print(f"  {table:<22} {count:>10,} rows  ({ts:%Y-%m-%d %H:%M})")

    conn.close()
    print("\n  Next steps:")
    print("  1. git push to GitHub")
    print("  2. Connect repo to Vercel")
    print("  3. Set DATABASE_URL in Vercel dashboard")
    print("  4. Deploy!\n")


if __name__ == "__main__":
    main()
