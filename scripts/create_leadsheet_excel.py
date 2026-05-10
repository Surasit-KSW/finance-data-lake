"""
create_leadsheet_excel.py
สร้าง Leadsheet_YE2025_STAT.xlsx (F1=Assets, F2=Liabilities+Equity, F3=P&L)
จากข้อมูลจริงใน finance_lake.duckdb
"""
import json, os
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DUCK_DB  = os.path.join(ROOT, "finance_lake.duckdb")
CFG      = os.path.join(ROOT, "config")
OUT_DIR  = os.path.join(ROOT, "03_Gold_DataMarts")
OUT_FILE = os.path.join(OUT_DIR, "Leadsheet_YE2025_STAT.xlsx")

# ── Styles ─────────────────────────────────────────────────────────────────────
thin = Side(border_style="thin", color="000000")
def box():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FILL = PatternFill("solid", fgColor="2F5597")
HDR_FNT  = Font(bold=True, color="FFFFFF")
SEC_FILL = PatternFill("solid", fgColor="D6E4F0")
SEC_FNT  = Font(bold=True, color="1F3864")
TOT_FILL = PatternFill("solid", fgColor="C6EFCE")
TOT_FNT  = Font(bold=True, color="375623")
NP_FILL  = PatternFill("solid", fgColor="FFF2CC")
NP_FNT   = Font(bold=True, size=11, color="7B3F00")
NUM_FMT  = '#,##0_);(#,##0);"-"'


def _set_num(cell, val):
    cell.value = round(float(val or 0), 0)
    cell.number_format = NUM_FMT
    cell.alignment = Alignment(horizontal="right")


def _write_headers(ws, py_label="Dec-2024", cy_label="Dec-2025"):
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 9
    for col in "CDEFGHI":
        ws.column_dimensions[col].width = 18

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Asia Metal Public Company Limited"
    c.font  = Font(bold=True, size=13)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = "Statutory Financial Statements Leadsheet (งบเฉพาะกิจการ)"
    c.font  = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:I3")
    c = ws["A3"]
    c.value = "For the year ended 31 December 2025 (ปีสิ้นสุด 31 ธันวาคม 2568)"
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:I4")
    c = ws["A4"]
    c.value = "Amounts in Thai Baht — หน่วย: บาท"
    c.alignment = Alignment(horizontal="center")

    headers = [
        ("A6", "Financial Statement Caption"),
        ("B6", "W/P Ref"),
        ("C6", f"Balance Per Book\n{py_label} (PY)"),
        ("D6", "Movement"),
        ("E6", f"Balance Per Book\n{cy_label} (TB)"),
        ("F6", "AJE / RJE\nDr."),
        ("G6", "AJE / RJE\nCr."),
        ("H6", f"Balance Per Audit\n{cy_label}"),
        ("I6", f"Balance Per Audit\n{py_label} (PY)"),
    ]
    for addr, txt in headers:
        c = ws[addr]
        c.value = txt
        c.font  = HDR_FNT
        c.fill  = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box()
    ws.row_dimensions[6].height = 36


def _write_section_header(ws, row, label):
    ws.merge_cells(f"A{row}:I{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font  = SEC_FNT
    c.fill  = SEC_FILL
    c.alignment = Alignment(horizontal="left", indent=1)
    c.border = box()
    ws.row_dimensions[row].height = 16


def _write_line(ws, row, caption, ref, py_val, cy_val, indent=2):
    mov = cy_val - py_val
    ws[f"A{row}"].value = caption
    ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=indent)
    ws[f"B{row}"].value = ref
    ws[f"B{row}"].alignment = Alignment(horizontal="center")
    _set_num(ws[f"C{row}"], py_val)
    _set_num(ws[f"D{row}"], mov)
    _set_num(ws[f"E{row}"], cy_val)
    ws[f"F{row}"].number_format = NUM_FMT
    ws[f"F{row}"].alignment = Alignment(horizontal="right")
    ws[f"G{row}"].number_format = NUM_FMT
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    ws[f"H{row}"].value = f"=E{row}+F{row}-G{row}"
    ws[f"H{row}"].number_format = NUM_FMT
    ws[f"H{row}"].alignment = Alignment(horizontal="right")
    _set_num(ws[f"I{row}"], py_val)
    for col in "ABCDEFGHI":
        ws[f"{col}{row}"].border = box()
    ws.row_dimensions[row].height = 15


def _write_subtotal(ws, row, label, ref, py_sum, cy_sum, fill=None, fnt=None):
    fill = fill or TOT_FILL
    fnt  = fnt  or TOT_FNT
    ws[f"A{row}"].value = label
    ws[f"A{row}"].font  = fnt
    ws[f"A{row}"].fill  = fill
    ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
    ws[f"B{row}"].value = ref
    ws[f"B{row}"].font  = Font(bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="center")
    _set_num(ws[f"C{row}"], py_sum); ws[f"C{row}"].font = fnt; ws[f"C{row}"].fill = fill
    _set_num(ws[f"D{row}"], cy_sum - py_sum); ws[f"D{row}"].font = fnt; ws[f"D{row}"].fill = fill
    _set_num(ws[f"E{row}"], cy_sum); ws[f"E{row}"].font = fnt; ws[f"E{row}"].fill = fill
    ws[f"F{row}"].fill = fill; ws[f"F{row}"].number_format = NUM_FMT
    ws[f"G{row}"].fill = fill; ws[f"G{row}"].number_format = NUM_FMT
    ws[f"H{row}"].value = f"=E{row}+F{row}-G{row}"
    ws[f"H{row}"].number_format = NUM_FMT; ws[f"H{row}"].font = fnt; ws[f"H{row}"].fill = fill
    ws[f"H{row}"].alignment = Alignment(horizontal="right")
    _set_num(ws[f"I{row}"], py_sum); ws[f"I{row}"].font = fnt; ws[f"I{row}"].fill = fill
    for col in "ABCDEFGHI":
        ws[f"{col}{row}"].border = box()
    ws.row_dimensions[row].height = 15


# ── Query + Aggregate ──────────────────────────────────────────────────────────
def _query_raw():
    con = duckdb.connect(DUCK_DB, read_only=True)

    def qd(sql, params):
        df = con.execute(sql, params).fetchdf()
        return dict(zip(df["account"].astype(str), df["amount"]))

    bs_py = qd(
        """SELECT CAST("G/L Account" AS VARCHAR) AS account, SUM(Net_Amount) AS amount
           FROM v_gl
           WHERE LEFT(CAST("G/L Account" AS VARCHAR), 1) IN ('1','2','3')
             AND (CAST(Year AS INTEGER) < ? OR (CAST(Year AS INTEGER) = ? AND CAST(Month AS INTEGER) <= ?))
           GROUP BY "G/L Account" """,
        [2024, 2024, 12],
    )
    bs_cy = qd(
        """SELECT CAST("G/L Account" AS VARCHAR) AS account, SUM(Net_Amount) AS amount
           FROM v_gl
           WHERE LEFT(CAST("G/L Account" AS VARCHAR), 1) IN ('1','2','3')
             AND (CAST(Year AS INTEGER) < ? OR (CAST(Year AS INTEGER) = ? AND CAST(Month AS INTEGER) <= ?))
           GROUP BY "G/L Account" """,
        [2025, 2025, 12],
    )
    pl_py = qd(
        """SELECT CAST("G/L Account" AS VARCHAR) AS account, SUM(Net_Amount) AS amount
           FROM v_gl
           WHERE LEFT(CAST("G/L Account" AS VARCHAR), 1) NOT IN ('1','2','3')
             AND CAST(Year AS INTEGER) = ? AND CAST(Month AS INTEGER) BETWEEN 1 AND 12
           GROUP BY "G/L Account" """,
        [2024],
    )
    pl_cy = qd(
        """SELECT CAST("G/L Account" AS VARCHAR) AS account, SUM(Net_Amount) AS amount
           FROM v_gl
           WHERE LEFT(CAST("G/L Account" AS VARCHAR), 1) NOT IN ('1','2','3')
             AND CAST(Year AS INTEGER) = ? AND CAST(Month AS INTEGER) BETWEEN 1 AND 12
           GROUP BY "G/L Account" """,
        [2025],
    )
    con.close()
    return bs_py, bs_cy, pl_py, pl_cy


def _aggregate(acct_dict, mapping, stmt_type):
    totals = {}
    acct_map = mapping.get("account_map", {})
    lines    = mapping["lines"]
    for acct, raw in acct_dict.items():
        lk = acct_map.get(str(acct))
        if not lk or lk not in lines:
            continue
        line_def = lines[lk]
        if stmt_type == "BS":
            sign = line_def.get("sign", 1)
            amt  = raw * sign
        else:
            flip = line_def.get("flip_sign", False)
            amt  = raw * (-1 if flip else 1)
        totals[lk] = totals.get(lk, 0.0) + amt
    return totals


# ── Sheet builders ─────────────────────────────────────────────────────────────
CA_KEYS = [
    ("O300", "Cash and Cash Equivalents / เงินสดและรายการเทียบเท่าเงินสด"),
    ("L300", "Trade Accounts Receivable / ลูกหนี้การค้า"),
    ("M300", "Other Receivables and Other Current Assets / ลูกหนี้อื่นและสินทรัพย์หมุนเวียนอื่น"),
    ("M400", "Other Current Financial Assets / สินทรัพย์ทางการเงินหมุนเวียนอื่น"),
    ("K300", "Inventories / สินค้าคงเหลือ"),
    ("I300", "Advance Payment for Purchasing of Goods / เงินจ่ายล่วงหน้าเพื่อซื้อสินค้า"),
    ("Q300", "Short-term Loans to Shareholders, Executives & Employees / เงินให้กู้ยืมระยะสั้น"),
]
NCA_KEYS = [
    ("G300", "Property, Plant and Equipment (net) / ที่ดิน อาคาร และอุปกรณ์ (สุทธิ)"),
    ("W300", "Right-of-Use Assets (net) / สินทรัพย์สิทธิการใช้ (สุทธิ)"),
    ("Y300", "Investment Property / อสังหาริมทรัพย์เพื่อการลงทุน"),
    ("H300", "Intangible Assets / สินทรัพย์ไม่มีตัวตน"),
    ("J300", "Investment in Securities / เงินลงทุนในหลักทรัพย์"),
    ("N300", "Other Non-Current Assets / สินทรัพย์ไม่หมุนเวียนอื่น"),
]
CL_KEYS = [
    ("Q400", "Bank Overdraft and Short-term Borrowings / เงินเบิกเกินบัญชีและเงินกู้ยืมระยะสั้น"),
    ("Q500", "Current Portion of Long-term Loans / ส่วนที่ถึงกำหนดชำระของเงินกู้ยืมระยะยาว"),
    ("R300", "Current Portion of Lease Liabilities / ส่วนที่ถึงกำหนดชำระของหนี้สินตามสัญญาเช่า"),
    ("V300", "Trade Accounts Payable / เจ้าหนี้การค้า"),
    ("X300", "Other Payables and Current Liabilities / เจ้าหนี้อื่นและหนี้สินหมุนเวียนอื่น"),
    ("T300", "Income Tax Payable / Deferred Tax / ภาษีเงินได้ค้างจ่าย"),
]
NCL_KEYS = [
    ("S300", "Employee Benefit Obligations / ภาระผูกพันผลประโยชน์พนักงาน"),
]
EQ_KEYS = [
    ("P300", "Shareholders' Equity / ส่วนของผู้ถือหุ้น"),
]
PL_SECTIONS = [
    ("Revenue / รายได้", [
        ("L400",  "Revenue from Sale of Goods / รายได้จากการขายสินค้า"),
        ("ZC300", "Other Income / รายได้อื่น"),
        ("ZA300", "Net FX Gain (Loss) / กำไร(ขาดทุน)สุทธิจากอัตราแลกเปลี่ยน"),
        ("ZE300", "Gains on Financial Instruments / กำไร(ขาดทุน)จากเครื่องมือทางการเงิน"),
    ]),
    ("Expenses / ค่าใช้จ่าย", [
        ("K400",  "Cost of Sale of Goods / ต้นทุนขาย"),
        ("ZD300", "Selling Expenses / ค่าใช้จ่ายในการขาย"),
        ("ZD400", "Administrative Expenses / ค่าใช้จ่ายในการบริหาร"),
        ("ZD500", "Management Remuneration / ผลตอบแทนผู้บริหาร"),
        ("ZF300", "Finance Costs / ต้นทุนทางการเงิน"),
        ("T400",  "Income Tax Expenses / ภาษีเงินได้"),
    ]),
]
REV_KEYS = ["L400", "ZC300", "ZA300", "ZE300"]
EXP_KEYS = ["K400", "ZD300", "ZD400", "ZD500", "ZF300", "T400"]


def _build_f1(ws, bs_py, bs_cy):
    _write_headers(ws)
    r = 8
    # Current Assets
    _write_section_header(ws, r, "Current Assets / สินทรัพย์หมุนเวียน"); r += 1
    ca_py, ca_cy = [], []
    for lk, cap in CA_KEYS:
        py_v, cy_v = bs_py.get(lk, 0), bs_cy.get(lk, 0)
        _write_line(ws, r, cap, lk, py_v, cy_v); r += 1
        ca_py.append(py_v); ca_cy.append(cy_v)
    _write_subtotal(ws, r, "Total Current Assets / รวมสินทรัพย์หมุนเวียน", "", sum(ca_py), sum(ca_cy)); r += 2
    # Non-Current Assets
    _write_section_header(ws, r, "Non-Current Assets / สินทรัพย์ไม่หมุนเวียน"); r += 1
    nca_py, nca_cy = [], []
    for lk, cap in NCA_KEYS:
        py_v, cy_v = bs_py.get(lk, 0), bs_cy.get(lk, 0)
        _write_line(ws, r, cap, lk, py_v, cy_v); r += 1
        nca_py.append(py_v); nca_cy.append(cy_v)
    _write_subtotal(ws, r, "Total Non-Current Assets / รวมสินทรัพย์ไม่หมุนเวียน", "", sum(nca_py), sum(nca_cy)); r += 1
    _write_subtotal(ws, r, "TOTAL ASSETS / รวมสินทรัพย์", "F1",
                    sum(ca_py) + sum(nca_py), sum(ca_cy) + sum(nca_cy),
                    fill=PatternFill("solid", fgColor="1F3864"),
                    fnt=Font(bold=True, color="FFFFFF"))


def _build_f2(ws, bs_py, bs_cy):
    _write_headers(ws)
    r = 8
    # Current Liabilities
    _write_section_header(ws, r, "Current Liabilities / หนี้สินหมุนเวียน"); r += 1
    cl_py, cl_cy = [], []
    for lk, cap in CL_KEYS:
        py_v, cy_v = bs_py.get(lk, 0), bs_cy.get(lk, 0)
        _write_line(ws, r, cap, lk, py_v, cy_v); r += 1
        cl_py.append(py_v); cl_cy.append(cy_v)
    _write_subtotal(ws, r, "Total Current Liabilities / รวมหนี้สินหมุนเวียน", "", sum(cl_py), sum(cl_cy)); r += 2
    # Non-Current Liabilities
    _write_section_header(ws, r, "Non-Current Liabilities / หนี้สินไม่หมุนเวียน"); r += 1
    ncl_py, ncl_cy = [], []
    for lk, cap in NCL_KEYS:
        py_v, cy_v = bs_py.get(lk, 0), bs_cy.get(lk, 0)
        _write_line(ws, r, cap, lk, py_v, cy_v); r += 1
        ncl_py.append(py_v); ncl_cy.append(cy_v)
    _write_subtotal(ws, r, "Total Non-Current Liabilities / รวมหนี้สินไม่หมุนเวียน", "", sum(ncl_py), sum(ncl_cy)); r += 2
    # Equity
    _write_section_header(ws, r, "Equity / ส่วนของผู้ถือหุ้น"); r += 1
    eq_py, eq_cy = [], []
    for lk, cap in EQ_KEYS:
        py_v, cy_v = bs_py.get(lk, 0), bs_cy.get(lk, 0)
        _write_line(ws, r, cap, lk, py_v, cy_v); r += 1
        eq_py.append(py_v); eq_cy.append(cy_v)
    _write_subtotal(ws, r, "Total Equity / รวมส่วนของผู้ถือหุ้น", "", sum(eq_py), sum(eq_cy)); r += 1
    _write_subtotal(ws, r, "TOTAL LIABILITIES AND EQUITY / รวมหนี้สินและส่วนของผู้ถือหุ้น", "F2",
                    sum(cl_py + ncl_py + eq_py), sum(cl_cy + ncl_cy + eq_cy),
                    fill=PatternFill("solid", fgColor="1F3864"),
                    fnt=Font(bold=True, color="FFFFFF"))


def _build_f3(ws, pl_py, pl_cy):
    _write_headers(ws, py_label="FY2024", cy_label="FY2025")
    r = 8
    for sec_label, lines in PL_SECTIONS:
        _write_section_header(ws, r, sec_label); r += 1
        sec_py, sec_cy = [], []
        for lk, cap in lines:
            py_v, cy_v = pl_py.get(lk, 0), pl_cy.get(lk, 0)
            _write_line(ws, r, cap, lk, py_v, cy_v); r += 1
            sec_py.append(py_v); sec_cy.append(cy_v)
        _write_subtotal(ws, r, f"Total {sec_label}", "", sum(sec_py), sum(sec_cy)); r += 2

    np_py = sum(pl_py.get(k, 0) for k in REV_KEYS) - sum(pl_py.get(k, 0) for k in EXP_KEYS)
    np_cy = sum(pl_cy.get(k, 0) for k in REV_KEYS) - sum(pl_cy.get(k, 0) for k in EXP_KEYS)
    _write_subtotal(ws, r, "NET PROFIT (LOSS) FOR THE YEAR / กำไร(ขาดทุน)สุทธิสำหรับปี", "F3",
                    np_py, np_cy,
                    fill=NP_FILL, fnt=NP_FNT)
    print(f"  Net Profit CY={np_cy:,.0f}  PY={np_py:,.0f}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("Loading data from DuckDB...")
    bs_py_raw, bs_cy_raw, pl_py_raw, pl_cy_raw = _query_raw()

    bs_map = json.load(open(os.path.join(CFG, "mapping_bs.json"), encoding="utf-8"))
    pl_map = json.load(open(os.path.join(CFG, "mapping_pl.json"), encoding="utf-8"))

    bs_py = _aggregate(bs_py_raw, bs_map, "BS")
    bs_cy = _aggregate(bs_cy_raw, bs_map, "BS")
    pl_py = _aggregate(pl_py_raw, pl_map, "PL")
    pl_cy = _aggregate(pl_cy_raw, pl_map, "PL")

    print(f"  Revenue 2025: {pl_cy.get('L400',0):,.0f}")
    print(f"  COGS   2025: {pl_cy.get('K400',0):,.0f}")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "F1 Assets"
    _build_f1(ws1, bs_py, bs_cy)
    print("F1 Assets - done")

    ws2 = wb.create_sheet("F2 Liabilities & Equity")
    _build_f2(ws2, bs_py, bs_cy)
    print("F2 Liabilities & Equity - done")

    ws3 = wb.create_sheet("F3 Profit & Loss")
    _build_f3(ws3, pl_py, pl_cy)
    print("F3 Profit & Loss - done")

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")


if __name__ == "__main__":
    main()
