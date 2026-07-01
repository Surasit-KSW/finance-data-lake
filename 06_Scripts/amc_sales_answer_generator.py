#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AMC Analytic Sales Q1'26 - Auto Answer Generator
=========================================================
อ่านข้อมูล Excel และสร้างคำตอบสำหรับแต่ละคำถามใน 5 sheets:
  1. BY Customer Q1FY26 and Q1FY25  -> ลูกค้า YoY
  2. BY Customer Q1FY26 and Q4FY25  -> ลูกค้า QoQ
  3. BY Product  Q1FY26 and Q1FY25  -> สินค้า YoY (labeled A1/A2/B1-B4/C1/D1)
  4. BY Product  Q1FY26 and Q4FY25  -> สินค้า QoQ (labeled A1/B1-B3/C1/D1)
  5. GP Analytic                     -> Gross Profit QoQ

Output:
  - Console report (UTF-8)
  - Excel file with answers written into Ans sections
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import os
from pathlib import Path as _Path

# โหลด .env จาก root โปรเจคอัตโนมัติ (ถ้ามี)
_env_file = _Path(__file__).parent.parent / ".env"
if _env_file.exists():
    for _line in _env_file.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

import anthropic
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FILE = (
    r"D:\_Work_Workspace\03_Data_Projects\_Finance_Data_Lake"
    r"\01_Bronze_Raw\templates\analytic_sales_q1_2026.xlsx"
)
OUTPUT_FILE = (
    r"D:\_Work_Workspace\03_Data_Projects\_Finance_Data_Lake"
    r"\04_Reports\AMC_Sales_Answers_Q1FY26.xlsx"
)

SALES_RAW_DIR = (
    r"D:\_Work_Workspace\03_Data_Projects\_Finance_Data_Lake"
    r"\01_Bronze_Raw\sales\amc"
)
SALES_FILES = {
    "Q1'26": [
        rf"{SALES_RAW_DIR}\2026\vf05_202601.xlsx",
        rf"{SALES_RAW_DIR}\2026\vf05_202602.xlsx",
        rf"{SALES_RAW_DIR}\2026\vf05_202603.xlsx",
    ],
    "Q1'25": [
        rf"{SALES_RAW_DIR}\2025\vf05_202501.xlsx",
        rf"{SALES_RAW_DIR}\2025\vf05_202502.xlsx",
        rf"{SALES_RAW_DIR}\2025\vf05_202503.xlsx",
    ],
    "Q4'25": [
        rf"{SALES_RAW_DIR}\2025\vf05_202510.xlsx",
        rf"{SALES_RAW_DIR}\2025\vf05_202511.xlsx",
        rf"{SALES_RAW_DIR}\2025\vf05_202512.xlsx",
    ],
}

# SAP billing column indices (0-based) — verified from file header
_COL_SOLD_TO  = 16
_COL_MAT_CODE = 60   # Material (SAP code e.g. PRGI, LCHR, GMC…)
_COL_MAT_DESC = 61   # Material Description
_COL_MAT_GRP  = 58   # Material Price Grp (HR, GI, GIM, CR…)
_COL_NET_THB  = 96   # Net Value (THB)
_COL_QTY      = 77   # Billed Qty (KG)
_COL_CANCEL   = 106  # Cancelled flag

# Claude API config — set ANTHROPIC_API_KEY in environment or .env
CLAUDE_MODEL   = "claude-opus-4-6"
CLAUDE_ENABLED = bool(os.getenv("ANTHROPIC_API_KEY"))

# ══════════════════════════════════════════════════════════════════════════════
# LOAD RAW SALES DATA: customer × product breakdown from SAP billing files
# ══════════════════════════════════════════════════════════════════════════════
import collections as _col

# Material Code prefix → Thai product family name
_MAT_FAMILY = {
    "LC":   "เหล็กตัวซี (C Channel)",
    "PR":   "ท่อแบน (Pipe Rect.)",
    "PS":   "ท่อเหลี่ยม (Pipe Square)",
    "PC":   "ท่อกลม (Pipe Round)",
    "GMC":  "Coil GIM กว้าง",
    "GML":  "Slit GIM",
    "GIC":  "Coil GI",
    "GIL":  "Slit GI",
    "HRC":  "Coil HR",
    "HRL":  "Slit HR",
    "CRC":  "Coil CR",
    "Slab": "เหล็กแท่งแบน (Slab HR)",
    "SLAB": "เหล็กแท่งแบน (Slab HR)",
    "Bloom":"เหล็กแท่ง (Bloom HR)",
}
_MAT_GRP_TH = {
    "HR":  "เหล็กรีดร้อน (HR)",
    "GI":  "เหล็กชุบสังกะสี (GI)",
    "GIM": "เหล็กชุบสังกะสีผสม (GIM)",
    "CR":  "เหล็กรีดเย็น (CR)",
    "EG":  "เหล็กชุบไฟฟ้า (EG)",
    "PO":  "เหล็กพิเศษ (PO)",
    "SL":  "เหล็กแผ่น (SL)",
    "AL":  "อะลูมิเนียม (AL)",
}

def _mat_family(code, desc, grp):
    """Map SAP material code/desc/grp to Thai product family."""
    code = str(code or "").strip()
    desc = str(desc or "").strip()
    # Slab / Bloom by description
    for kw in ("Slab", "SLAB", "Bloom"):
        if kw.lower() in desc.lower() or kw.lower() in code.lower():
            return _MAT_FAMILY.get(kw, kw), grp
    # By code prefix
    for prefix, name in _MAT_FAMILY.items():
        if code.upper().startswith(prefix.upper()):
            return name, grp
    # Fallback to material price group
    grp_th = _MAT_GRP_TH.get(str(grp or "").strip(), str(grp or "ไม่ระบุ"))
    return grp_th, grp


def load_customer_products():
    """
    Read raw SAP billing files and return:
      {sold_to: {period: {family_label: {'amt': float, 'qty': float}}}}
    Only for TARGET_CUSTOMERS, only non-cancelled lines.
    """
    TARGET = {1100005, 1100009, 1100019, 1100104, 1400001,
              1100001, 1100004, 1100842}
    result = _col.defaultdict(lambda: _col.defaultdict(
                              lambda: _col.defaultdict(lambda: {"amt": 0.0, "qty": 0.0})))

    for period, flist in SALES_FILES.items():
        print(f"  [RAW] loading {period}...", flush=True)
        for fpath in flist:
            try:
                wb_r = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                ws_r = wb_r.active
                for i, row in enumerate(ws_r.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    if row[_COL_CANCEL]:
                        continue
                    try:
                        sold_to = int(row[_COL_SOLD_TO])
                    except Exception:
                        continue
                    if sold_to not in TARGET:
                        continue
                    net = row[_COL_NET_THB]
                    qty = row[_COL_QTY]
                    if not isinstance(net, (int, float)):
                        continue
                    family, grp = _mat_family(row[_COL_MAT_CODE],
                                              row[_COL_MAT_DESC],
                                              row[_COL_MAT_GRP])
                    # Append material price group for disambiguation
                    key = f"{family} [{grp}]" if grp and grp not in family else family
                    result[sold_to][period][key]["amt"] += float(net)
                    result[sold_to][period][key]["qty"] += float(qty) if isinstance(qty, (int, float)) else 0.0
                wb_r.close()
            except Exception as e:
                print(f"  [WARN] {fpath}: {e}", flush=True)
    return result


def summarise_customer_products(cust_prods, sold_to, period, top_n=5):
    """
    Return list of (family_label, amt, pct_of_total) sorted by |amt| desc,
    limited to top_n entries.
    """
    period_data = cust_prods.get(sold_to, {}).get(period, {})
    if not period_data:
        return []
    total = sum(v["amt"] for v in period_data.values())
    if total == 0:
        return []
    rows = sorted(period_data.items(), key=lambda x: -abs(x[1]["amt"]))[:top_n]
    return [(lbl, v["amt"], v["amt"] / total * 100) for lbl, v in rows]


# ── Load all raw sales data (runs once at startup) ────────────────────────────
print("[RAW] กำลังโหลดข้อมูลยอดขาย raw จากไฟล์ SAP billing...", flush=True)
cust_prods = load_customer_products()
print("[RAW] โหลดเสร็จสิ้น", flush=True)


# ── Customer metadata: ชื่อบริษัท / ประเภทกิจการ (ยืนยันจาก SAP Billing) ──────
CUSTOMER_META = {
    1100001: {"name": "บริษัท สยามโกลบอลเฮ้าส์ จำกัด (มหาชน)",
              "biz_type": "ห้างค้าปลีกวัสดุก่อสร้าง (Home Improvement Retail)"},
    1100004: {"name": "บริษัท ศรีพูนทอง จำกัด",
              "biz_type": "ตัวแทนจำหน่ายเหล็กและวัสดุก่อสร้าง"},
    1100005: {"name": "บริษัท เจริญไชยสุริน จำกัด",
              "biz_type": "ผู้ค้าและผู้รับเหมาก่อสร้าง"},
    1100009: {"name": "บริษัท ศรีพูนทรัพย์ จำกัด",
              "biz_type": "ตัวแทนจำหน่ายเหล็กและวัสดุก่อสร้าง"},
    1100019: {"name": "บริษัท ดูโฮม จำกัด (มหาชน)",
              "biz_type": "ห้างค้าปลีกวัสดุก่อสร้าง (Home Improvement Retail)"},
    1100104: {"name": "บริษัท โฮมฮับ จำกัด",
              "biz_type": "ห้างค้าปลีกวัสดุก่อสร้าง (Home Improvement Retail)"},
    1100842: {"name": "บริษัท ฟินิกซ์ สตีล จำกัด",
              "biz_type": "ผู้ค้าและแปรรูปเหล็ก (Steel Trader/Processor)"},
    1400001: {"name": "บริษัท ไพร์ม สตีล มิลล์ จำกัด",
              "biz_type": "โรงงานแปรรูปเหล็ก (Steel Mill)"},
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def M(val, sign=True):
    """Format value in millions THB, e.g. +25.43M บาท"""
    if val is None:
        return "N/A"
    if sign:
        return f"{val / 1_000_000:+,.2f}M บาท"
    return f"{val / 1_000_000:,.2f}M บาท"

def K(val):
    """Format as KG with thousand separator."""
    if val is None:
        return "N/A"
    return f"{val:,.0f} KG"

def pct_str(val, base):
    if not base:
        return "N/A"
    return f"{val / base * 100:+.1f}%"

def direction(val):
    if val is None or val == 0:
        return "ไม่เปลี่ยนแปลง"
    return "เพิ่มขึ้น" if val > 0 else "ลดลง"

def safe_float(row, idx):
    if idx >= len(row):
        return 0.0
    v = row[idx]
    return float(v) if isinstance(v, (int, float)) and v is not None else 0.0

def avg_price(amt, qty):
    return amt / qty if qty else 0.0

def vol_effect(qty_new, qty_old, p_old):
    return (qty_new - qty_old) * p_old

def price_effect(p_new, p_old, qty_old):
    return (p_new - p_old) * qty_old


# ── Load workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

def sheet_rows(sheet_name):
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


# ══════════════════════════════════════════════════════════════════════════════
# EXTRACT: Customer data (Sheet 1 & 2)
# Structure: col0=Sold-to, col1=Name, col2=Q_curr, col3=Q_prev, col4=CHG, col5=Label
# ══════════════════════════════════════════════════════════════════════════════
def extract_customers(sheet_name):
    rows = sheet_rows(sheet_name)
    customers = []
    for row in rows:
        sold_to = row[0]
        if not isinstance(sold_to, (int, float)):
            continue
        sold_to = int(sold_to)
        name    = row[1] or ""
        q_curr  = safe_float(row, 2)
        q_prev  = safe_float(row, 3)
        chg     = q_curr - q_prev
        label   = row[5] if len(row) > 5 else None
        customers.append({
            "label": label, "sold_to": sold_to, "name": name,
            "q_curr": q_curr, "q_prev": q_prev, "chg": chg,
        })
    return customers

cust_yoy = extract_customers("BY Customer Q1FY26 and Q1FY25")
cust_qoq = extract_customers("BY Customer Q1FY26 and Q4FY25")


# ══════════════════════════════════════════════════════════════════════════════
# EXTRACT: Product data (Sheet 3 & 4) — indexed by LABEL
# Sheet 3: name@2, Q126_QTY@5, Q126_P@6, Q126_AMT@7,
#          Q125_QTY@8, Q125_P@9, Q125_AMT@10, CHG_AMT@13, LABEL@14
# Sheet 4: name@2, Q126_QTY@4, Q126_P@5, Q126_AMT@6,
#          Q425_QTY@7, Q425_P@8, Q425_AMT@9, CHG_AMT@12, LABEL@13
# ══════════════════════════════════════════════════════════════════════════════
def extract_products_by_label(sheet_name):
    """Returns dict {label: {name, q_curr_qty, q_curr_price, q_curr_amt,
                                    q_prev_qty, q_prev_price, q_prev_amt, chg_amt}}"""
    rows = sheet_rows(sheet_name)

    # Determine column offsets
    if "Q1FY25" in sheet_name:
        # Sheet 3
        nc, nq, np_, na = 2, 5, 6, 7
        pq, pp, pa      = 8, 9, 10
        ca, lc          = 13, 14
    else:
        # Sheet 4
        nc, nq, np_, na = 2, 4, 5, 6
        pq, pp, pa      = 7, 8, 9
        ca, lc          = 12, 13

    products = {}
    for row in rows:
        # Must have a label
        if len(row) <= lc:
            continue
        label = row[lc]
        if not isinstance(label, str) or len(label.strip()) < 2:
            continue
        # Label format: A1, A2, B1, B2, B3, B4, C1, D1
        if not (label[0].isalpha() and label[1].isdigit()):
            continue

        name = row[nc] if nc < len(row) else ""
        q_curr_qty   = safe_float(row, nq)
        q_curr_price = safe_float(row, np_)
        q_curr_amt   = safe_float(row, na)
        q_prev_qty   = safe_float(row, pq)
        q_prev_price = safe_float(row, pp)
        q_prev_amt   = safe_float(row, pa)
        # Always compute CHG from Q_curr - Q_prev (don't trust stored formula
        # which can have sign errors, e.g. Slab HR in Sheet 4)
        chg_amt      = q_curr_amt - q_prev_amt

        # Compute avg price if formula cached as 0
        if q_curr_qty and not q_curr_price:
            q_curr_price = avg_price(q_curr_amt, q_curr_qty)
        if q_prev_qty and not q_prev_price:
            q_prev_price = avg_price(q_prev_amt, q_prev_qty)

        products[label] = {
            "name": name,
            "q_curr_qty": q_curr_qty, "q_curr_price": q_curr_price, "q_curr_amt": q_curr_amt,
            "q_prev_qty": q_prev_qty, "q_prev_price": q_prev_price, "q_prev_amt": q_prev_amt,
            "chg_amt": chg_amt,
        }
    return products

prod_yoy = extract_products_by_label("BY Product Q1FY26 and Q1FY25")
prod_qoq = extract_products_by_label("BY Product Q1FY26 and Q4FY25")


# ══════════════════════════════════════════════════════════════════════════════
# EXTRACT: GP Data (Sheet 5)
# col1=label, col4=Q1FY26, col5=Q4FY25, col6=Change
# ══════════════════════════════════════════════════════════════════════════════
def extract_gp():
    rows = sheet_rows("GP Analytic")
    gp = {}
    for row in rows:
        if len(row) < 7:
            continue
        label = row[1]
        if not isinstance(label, str):
            continue
        v_curr, v_prev, v_chg = row[4], row[5], row[6]
        # Guard: only assign when the data cell is numeric (skip text rows)
        if not isinstance(v_curr, (int, float)):
            continue
        if "Revenue" in label:
            gp["rev_curr"], gp["rev_prev"], gp["rev_chg"] = v_curr, v_prev, v_chg
            gp["rev_pct"] = row[7] if len(row) > 7 else None
        elif "Cost of goods" in label:
            gp["cogs_curr"], gp["cogs_prev"], gp["cogs_chg"] = v_curr, v_prev, v_chg
        elif "Gross Profit Rate" in label:
            gp["gpr_curr"], gp["gpr_prev"], gp["gpr_chg"] = v_curr, v_prev, v_chg
        elif "Gross Profit" in label:
            gp["gp_curr"], gp["gp_prev"], gp["gp_chg"] = v_curr, v_prev, v_chg
    return gp

gp = extract_gp()


# ══════════════════════════════════════════════════════════════════════════════
# TOTALS
# ══════════════════════════════════════════════════════════════════════════════
def get_totals(sheet_name):
    rows = sheet_rows(sheet_name)
    for row in rows:
        if isinstance(row[1], str) and row[1].lower() == "total":
            q_curr = abs(safe_float(row, 2))
            q_prev = abs(safe_float(row, 3))
            return q_curr, q_prev, q_curr - q_prev
    return 0, 0, 0

tot_q126_yoy, tot_q125, _ = get_totals("BY Customer Q1FY26 and Q1FY25")
tot_q126_qoq, tot_q425, _ = get_totals("BY Customer Q1FY26 and Q4FY25")


# ══════════════════════════════════════════════════════════════════════════════
# ANSWER GENERATORS
# ══════════════════════════════════════════════════════════════════════════════
def _customer_header(sold_to, curr_period="Q1'26", prev_period=None):
    """Return formatted customer info block with real product breakdown from SAP billing."""
    meta     = CUSTOMER_META.get(sold_to, {})
    name     = meta.get("name", f"Sold-to {sold_to}")
    biz_type = meta.get("biz_type", "-")

    # Build product breakdown lines from raw data
    prod_lines = []
    for period in ([curr_period] + ([prev_period] if prev_period else [])):
        rows = summarise_customer_products(cust_prods, sold_to, period, top_n=5)
        if not rows:
            continue
        total_amt = sum(r[1] for r in rows)
        prod_lines.append(f"  สินค้าที่ซื้อ ({period})  :")
        for lbl, amt, pct in rows:
            prod_lines.append(f"    - {lbl:<45} {amt/1e6:>8.2f}M บาท ({pct:.1f}%)")

    header = (
        f"  ลูกค้า (Sold-to)  : {sold_to} — {name}\n"
        f"  ประเภทกิจการ      : {biz_type}\n"
    )
    if prod_lines:
        header += "\n".join(prod_lines) + "\n"
        header += "  (ข้อมูลจาก SAP Billing Document จริง)\n"
    else:
        header += "  สินค้าหลักที่ซื้อ  : ไม่พบข้อมูลใน SAP billing files\n"
    return header


def answer_customer(c, prev_period_label):
    """Generate audit-style explanation for one customer variance."""
    sold_to = c["sold_to"]
    q_curr  = c["q_curr"]
    q_prev  = c["q_prev"]
    chg     = c["chg"]
    pct     = pct_str(chg, q_prev) if q_prev else "N/A"
    dir_th  = "เพิ่มขึ้น" if chg >= 0 else "ลดลง"
    abs_chg = abs(chg)
    header  = _customer_header(sold_to, curr_period="Q1'26", prev_period=prev_period_label)

    # Customer 1100842: new/returning buyer (prev = 0)
    if sold_to == 1100842 and q_prev == 0:
        return (
            f"{header}"
            f"  ยอดขาย{dir_th} {M(abs_chg, sign=False)} "
            f"เนื่องจากในงวด {prev_period_label} ลูกค้ารายนี้ไม่มีรายการขาย "
            f"(ยอดขาย {prev_period_label} = 0 บาท) "
            f"ขณะที่งวด Q1'26 มียอดขาย {M(q_curr, sign=False)}"
        )

    # Customer 1400001: Slab HR — decompose volume/price effect
    if sold_to == 1400001:
        if "D1" in prod_yoy and "Q1'25" in prev_period_label:
            p    = prod_yoy["D1"]
            p_new = avg_price(p["q_curr_amt"], p["q_curr_qty"])
            p_old = p["q_prev_price"] or avg_price(p["q_prev_amt"], p["q_prev_qty"])
            ve    = vol_effect(p["q_curr_qty"], p["q_prev_qty"], p_old)
            pe    = price_effect(p_new, p_old, p["q_prev_qty"])
            return (
                f"{header}"
                f"  ยอดขาย{dir_th} {M(abs_chg, sign=False)} หรือร้อยละ {pct} เมื่อเทียบกับ {prev_period_label}\n\n"
                f"  สาเหตุของความแปรปรวน:\n"
                f"  1) ปริมาณขายลดลงจาก {K(p['q_prev_qty'])} เป็น {K(p['q_curr_qty'])} "
                f"({K(p['q_curr_qty'] - p['q_prev_qty'])}) "
                f"ส่งผลให้ยอดขายลดลง {M(ve)} (Volume Effect)\n"
                f"  2) ราคาขายเฉลี่ยลดลงจาก {p_old:.2f} บาท/กก. เป็น {p_new:.2f} บาท/กก. "
                f"({p_new - p_old:+.2f} บาท/กก.) "
                f"ส่งผลให้ยอดขายลดลง {M(pe)} (Price Effect)\n\n"
                f"  ยอดขาย Q1'26           : {M(q_curr, sign=False)}\n"
                f"  ยอดขาย {prev_period_label:8}     : {M(q_prev, sign=False)}\n"
                f"  ผลต่างสุทธิ             : {M(chg)}"
            )
        elif "D1" in prod_qoq and "Q4'25" in prev_period_label:
            p    = prod_qoq["D1"]
            p_new = avg_price(p["q_curr_amt"], p["q_curr_qty"])
            p_old = p["q_prev_price"] or avg_price(p["q_prev_amt"], p["q_prev_qty"])
            ve    = vol_effect(p["q_curr_qty"], p["q_prev_qty"], p_old)
            pe    = price_effect(p_new, p_old, p["q_prev_qty"])
            return (
                f"{header}"
                f"  ยอดขาย{dir_th} {M(abs_chg, sign=False)} หรือร้อยละ {pct} เมื่อเทียบกับ {prev_period_label}\n\n"
                f"  สาเหตุของความแปรปรวน:\n"
                f"  1) ปริมาณขายลดลงจาก {K(p['q_prev_qty'])} เป็น {K(p['q_curr_qty'])} "
                f"({K(p['q_curr_qty'] - p['q_prev_qty'])}) "
                f"ส่งผลให้ยอดขายลดลง {M(ve)} (Volume Effect)\n"
                f"  2) ราคาขายเฉลี่ยลดลงจาก {p_old:.2f} บาท/กก. เป็น {p_new:.2f} บาท/กก. "
                f"({p_new - p_old:+.2f} บาท/กก.) "
                f"ส่งผลให้ยอดขายลดลง {M(pe)} (Price Effect)\n\n"
                f"  ยอดขาย Q1'26           : {M(q_curr, sign=False)}\n"
                f"  ยอดขาย {prev_period_label:8}     : {M(q_prev, sign=False)}\n"
                f"  ผลต่างสุทธิ             : {M(chg)}"
            )

    # Standard customer
    return (
        f"{header}"
        f"  ยอดขาย{dir_th} {M(abs_chg, sign=False)} หรือร้อยละ {pct} เมื่อเทียบกับ {prev_period_label}\n\n"
        f"  ยอดขาย Q1'26           : {M(q_curr, sign=False)}\n"
        f"  ยอดขาย {prev_period_label:8}     : {M(q_prev, sign=False)}\n"
        f"  ผลต่างสุทธิ             : {M(chg)}"
    )


def answer_product(p, curr_period, prev_period):
    """Generate audit-style variance explanation for one labeled product."""
    q_curr_qty   = p["q_curr_qty"]
    q_curr_amt   = p["q_curr_amt"]
    q_prev_qty   = p["q_prev_qty"]
    q_prev_amt   = p["q_prev_amt"]
    chg_amt      = p["chg_amt"]
    name         = p.get("name", "")

    q_curr_price = p["q_curr_price"] or avg_price(q_curr_amt, q_curr_qty)
    q_prev_price = p["q_prev_price"] or avg_price(q_prev_amt, q_prev_qty)
    dir_th       = "เพิ่มขึ้น" if chg_amt >= 0 else "ลดลง"
    abs_chg      = abs(chg_amt)
    pct          = pct_str(chg_amt, q_prev_amt) if q_prev_amt else "N/A"

    # New product: no previous period data
    if q_prev_qty == 0 and q_curr_qty > 0:
        return (
            f"สินค้า {name} ไม่มีรายการขายในงวด {prev_period} "
            f"โดยเริ่มมียอดขายในงวด {curr_period} เป็นครั้งแรก\n\n"
            f"  ยอดขาย {curr_period}   : {M(q_curr_amt, sign=False)}\n"
            f"  ปริมาณ                : {K(q_curr_qty)}\n"
            f"  ราคาขายเฉลี่ย         : {q_curr_price:.2f} บาท/กก.\n"
            f"  ผลต่างสุทธิ           : {M(chg_amt)} (ไม่มีฐานเปรียบเทียบ)"
        )

    # Discontinued: no current period data
    if q_curr_qty == 0 and q_prev_qty > 0:
        return (
            f"สินค้า {name} ไม่มีรายการขายในงวด {curr_period} "
            f"(ปริมาณขาย = 0 กิโลกรัม)\n\n"
            f"  ยอดขาย {prev_period}   : {M(q_prev_amt, sign=False)}\n"
            f"  ปริมาณ                : {K(q_prev_qty)}\n"
            f"  ราคาขายเฉลี่ย         : {q_prev_price:.2f} บาท/กก.\n"
            f"  ผลต่างสุทธิ           : {M(chg_amt)}"
        )

    # Both periods have data — decompose
    ve = vol_effect(q_curr_qty, q_prev_qty, q_prev_price)
    pe = price_effect(q_curr_price, q_prev_price, q_prev_qty)
    ve_dir = "เพิ่มขึ้น" if ve >= 0 else "ลดลง"
    pe_dir = "เพิ่มขึ้น" if pe >= 0 else "ลดลง"

    return (
        f"ยอดขายสินค้า {name} {dir_th} {M(abs_chg, sign=False)} "
        f"หรือร้อยละ {pct} เมื่อเทียบกับงวด {prev_period}\n\n"
        f"  สาเหตุของความแปรปรวน:\n"
        f"  1) ปริมาณขาย{('เพิ่มขึ้น' if q_curr_qty >= q_prev_qty else 'ลดลง')}จาก "
        f"{K(q_prev_qty)} เป็น {K(q_curr_qty)} "
        f"({K(q_curr_qty - q_prev_qty)}, {pct_str(q_curr_qty - q_prev_qty, q_prev_qty)}) "
        f"ส่งผลให้ยอดขาย{ve_dir} {M(abs(ve), sign=False)} (Volume Effect)\n"
        f"  2) ราคาขายเฉลี่ย{('เพิ่มขึ้น' if q_curr_price >= q_prev_price else 'ลดลง')}จาก "
        f"{q_prev_price:.2f} บาท/กก. เป็น {q_curr_price:.2f} บาท/กก. "
        f"({q_curr_price - q_prev_price:+.2f} บาท/กก.) "
        f"ส่งผลให้ยอดขาย{pe_dir} {M(abs(pe), sign=False)} (Price Effect)\n\n"
        f"  ยอดขาย {curr_period}   : {M(q_curr_amt, sign=False)}\n"
        f"  ยอดขาย {prev_period}   : {M(q_prev_amt, sign=False)}\n"
        f"  ผลต่างสุทธิ           : {M(chg_amt)}"
    )


def answer_gp():
    rev_curr  = gp.get("rev_curr")  or 0
    rev_prev  = gp.get("rev_prev")  or 0
    rev_chg   = gp.get("rev_chg")   or 0
    cogs_curr = gp.get("cogs_curr") or 0
    cogs_prev = gp.get("cogs_prev") or 0
    cogs_chg  = gp.get("cogs_chg")  or 0
    gp_curr   = gp.get("gp_curr")   or 0
    gp_prev   = gp.get("gp_prev")   or 0
    gp_chg    = gp.get("gp_chg")    or 0
    gpr_curr  = (gp.get("gpr_curr") or 0) * 100
    gpr_prev  = (gp.get("gpr_prev") or 0) * 100

    rev_pct   = pct_str(rev_chg, rev_prev)
    cogs_pct  = pct_str(cogs_chg, cogs_prev)
    gp_dir    = direction(gp_chg)
    cogs_vs_rev = cogs_chg / rev_chg * 100 if rev_chg else 0

    gp_dir_th   = "เพิ่มขึ้น" if gp_chg >= 0 else "ลดลง"
    rev_dir_th  = "เพิ่มขึ้น" if rev_chg >= 0 else "ลดลง"
    cogs_dir_th = "เพิ่มขึ้น" if cogs_chg >= 0 else "ลดลง"

    return (
        f"Gross Profit งวด Q1'26 {gp_dir_th} {M(abs(gp_chg), sign=False)} "
        f"เมื่อเทียบกับงวด Q4'25 โดยมีรายละเอียดดังนี้\n\n"
        f"1. รายได้จากการขาย (Revenue)\n"
        f"   {rev_dir_th} {M(abs(rev_chg), sign=False)} หรือร้อยละ {rev_pct}\n"
        f"   - Q1'26 : {M(rev_curr, sign=False)}\n"
        f"   - Q4'25 : {M(rev_prev, sign=False)}\n"
        f"   สาเหตุหลักเกิดจากปริมาณขายสินค้ากลุ่มท่อ GI (Pipe Rect./Square GI), "
        f"ท่อ HR และ C Channel HR ที่เพิ่มขึ้น\n\n"
        f"2. ต้นทุนขาย (Cost of Goods Sold)\n"
        f"   {cogs_dir_th} {M(abs(cogs_chg), sign=False)} หรือร้อยละ {cogs_pct}\n"
        f"   - Q1'26 : {M(cogs_curr, sign=False)}\n"
        f"   - Q4'25 : {M(cogs_prev, sign=False)}\n"
        f"   ต้นทุนขายเพิ่มขึ้นในอัตราที่ต่ำกว่ารายได้ "
        f"(ต้นทุนเพิ่ม {cogs_vs_rev:.1f}% ของรายได้ที่เพิ่มขึ้น) "
        f"ส่งผลให้อัตรากำไรขั้นต้นปรับตัวดีขึ้น\n\n"
        f"3. Gross Profit\n"
        f"   - Q1'26 : {M(gp_curr, sign=False)} | GP Rate = {gpr_curr:.2f}%\n"
        f"   - Q4'25 : {M(gp_prev, sign=False)} | GP Rate = {gpr_prev:.2f}%\n"
        f"   - ผลต่าง : {M(gp_chg)}\n\n"
        f"หมายเหตุ: งวด Q4'25 มี Gross Profit ติดลบ ({M(gp_prev, sign=False)}) "
        f"เนื่องจากต้นทุนขายสูงกว่ารายได้จากการขายในงวดนั้น "
        f"งวด Q1'26 Revenue ฟื้นตัวทำให้ GP Rate ปรับเป็น {gpr_curr:.2f}%"
    )


# ══════════════════════════════════════════════════════════════════════════════
# QUALITATIVE ANALYSIS — Customer Behavior Engine
# ══════════════════════════════════════════════════════════════════════════════

def classify_customer(sold_to, chg_pct_val, q_curr, q_prev, period):
    """
    Return (tag, behavior_text, action_text) based on behavioral pattern.
    period = 'yoy' | 'qoq'
    """
    if q_prev == 0 and q_curr > 0:
        return (
            "REACTIVATED",
            "ลูกค้ากลับมาสั่งซื้อ / ลูกค้าใหม่ — ไม่มียอดใน period ก่อน",
            "ติดตามว่าจะสั่งซื้อต่อเนื่องหรือเป็น one-time order เพื่อประเมิน recurring potential",
        )
    if q_curr == 0 and q_prev > 0:
        return (
            "FULL_CHURN",
            "หยุดซื้อโดยสิ้นเชิง — ยอดเป็น 0 บาทใน period นี้",
            "สอบถามสาเหตุ: เปลี่ยน supplier / หยุดกิจการ / ปัญหาเครดิต",
        )
    if chg_pct_val is not None:
        pct_val = chg_pct_val
        if pct_val <= -85:
            return (
                "NEAR_CHURN",
                f"ยอดขายหดตัว {pct_val:+.1f}% — แทบหยุดซื้อ ถือเป็น Churn Risk สูงมาก",
                "ต้องสอบถามสาเหตุเร่งด่วน: เปลี่ยน supplier, credit hold, หรือ demand หาย",
            )
        if pct_val <= -40:
            return (
                "MAJOR_DECLINE",
                f"ยอดขายลดลงมาก {pct_val:+.1f}% — สูญเสีย market share อย่างมีนัยสำคัญ",
                "ตรวจสอบว่าลูกค้าหันไปซื้อจากคู่แข่ง หรือ demand ของลูกค้าลดจริง",
            )
        if pct_val <= -20:
            return (
                "MODERATE_DECLINE",
                f"ยอดขายลดลง {pct_val:+.1f}% — ต้องเฝ้าระวัง",
                "ติดตามแนวโน้มในไตรมาสถัดไป หากลดต่อเนื่องให้ escalate",
            )
        if pct_val <= -5:
            return (
                "MILD_DECLINE",
                f"ยอดขายลดลงเล็กน้อย {pct_val:+.1f}% — ยังอยู่ในระดับที่ยอมรับได้",
                "จับตาดูแนวโน้มไตรมาสหน้า",
            )
        if pct_val >= 80:
            tag = "STRONG_GROWTH"
            if period == "qoq":
                tag = "SEASONAL_SURGE"
                return (
                    tag,
                    f"ยอดขายเพิ่มขึ้น {pct_val:+.1f}% QoQ — อาจเป็น Seasonal Demand หรือ Project Order",
                    "ตรวจสอบว่าเป็น recurring demand หรือ one-time project เพื่อประมาณ Q2",
                )
            return (
                tag,
                f"ยอดขายเพิ่มขึ้น {pct_val:+.1f}% — การเติบโตแข็งแกร่ง",
                "Maintain relationship, explore upsell / cross-sell opportunities",
            )
        if pct_val >= 20:
            return (
                "GROWTH",
                f"ยอดขายเพิ่มขึ้น {pct_val:+.1f}% — แนวโน้มดี",
                "รักษาความสัมพันธ์และ service level",
            )
        return (
            "STABLE",
            f"ยอดขายใกล้เคียงกับ period ก่อน ({pct_val:+.1f}%)",
            "Maintain service level",
        )
    return ("UNKNOWN", "ไม่มีข้อมูลเพียงพอ", "N/A")


def behavioral_note(c, period):
    """Return audit observation note for one customer (no action recommendation)."""
    q_curr  = c["q_curr"]
    q_prev  = c["q_prev"]
    chg     = c["chg"]
    sold_to = c["sold_to"]

    pct_val = (chg / q_prev * 100) if q_prev else None
    tag, behavior, _ = classify_customer(sold_to, pct_val, q_curr, q_prev, period)

    # Reframe tag as audit observation language
    pv_str = f"{pct_val:+.1f}%" if pct_val is not None else "N/A"
    obs_map = {
        "REACTIVATED":     "ลูกค้ารายนี้ไม่มีรายการขายในงวดก่อนหน้า และกลับมามียอดขายในงวดปัจจุบัน",
        "FULL_CHURN":      "ลูกค้ารายนี้ไม่มีรายการขายในงวดปัจจุบัน ขณะที่งวดก่อนมียอดขาย",
        "NEAR_CHURN":      f"ยอดขายลูกค้ารายนี้ลดลงอย่างมีนัยสำคัญ ({pv_str}) เมื่อเทียบกับงวดก่อนหน้า",
        "MAJOR_DECLINE":   f"ยอดขายลดลง {pv_str} ซึ่งถือเป็นการเปลี่ยนแปลงที่มีนัยสำคัญ",
        "MODERATE_DECLINE":f"ยอดขายลดลง {pv_str} เมื่อเทียบกับงวดก่อนหน้า",
        "MILD_DECLINE":    f"ยอดขายลดลงเล็กน้อย {pv_str}",
        "SEASONAL_SURGE":  f"ยอดขายเพิ่มขึ้น {pv_str} ซึ่งอาจเป็นผลจาก Seasonal Demand หรือ Project Order",
        "STRONG_GROWTH":   f"ยอดขายเพิ่มขึ้น {pv_str} เมื่อเทียบกับงวดก่อนหน้า",
        "GROWTH":          f"ยอดขายเพิ่มขึ้น {pv_str}",
        "STABLE":          f"ยอดขายใกล้เคียงงวดก่อนหน้า ({pv_str})",
    }
    obs = obs_map.get(tag, behavior)

    # Cross-period factual observation
    cross_note = ""
    if sold_to == 1400001:
        cross_note = (
            "\n  หมายเหตุ: ยอดขายลูกค้ารายนี้ลดลงทั้งเมื่อเทียบ YoY และ QoQ "
            "สอดคล้องกับการลดลงของปริมาณขายสินค้า Slab HR ในทั้งสองงวด"
        )
    if sold_to == 1100005:
        cross_note = (
            "\n  หมายเหตุ: ยอดขายลูกค้ารายนี้เพิ่มขึ้นทั้งเมื่อเทียบ YoY (+147.5%) และ QoQ (+145.8%) "
            "แสดงถึงแนวโน้มการเติบโตที่ต่อเนื่องในทั้งสองมุมมอง"
        )

    return (
        f"\n  ข้อสังเกต: {obs}"
        f"{cross_note}"
    )


def qualitative_product_note(label, p, period):
    """Return qualitative behavioral note for a product."""
    q_curr_qty = p["q_curr_qty"]
    q_prev_qty = p["q_prev_qty"]
    chg_amt    = p["chg_amt"]

    if q_prev_qty == 0 and q_curr_qty > 0:
        return (
            "\n  [พฤติกรรมสินค้า | NEW_PRODUCT]\n"
            "  => สินค้าใหม่ที่เริ่มขายใน Q1'26 — ยังไม่มีฐานเปรียบเทียบ\n"
            "  => ติดตามว่าจะมี demand ต่อเนื่องหรือเป็น special order"
        )
    if q_curr_qty == 0 and q_prev_qty > 0:
        return (
            "\n  [พฤติกรรมสินค้า | DISCONTINUED]\n"
            "  => หยุดขายสินค้านี้ใน Q1'26 — ตรวจสอบว่าเปลี่ยน product code หรือ ปิด line\n"
            "  => หากเป็น reclassification ให้ map ยอดกับสินค้าที่ทดแทน"
        )

    pct_vol = (q_curr_qty - q_prev_qty) / q_prev_qty * 100 if q_prev_qty else 0
    pct_amt = chg_amt / p["q_prev_amt"] * 100 if p["q_prev_amt"] else 0
    p_new   = avg_price(p["q_curr_amt"], q_curr_qty)
    p_old   = p["q_prev_price"] or avg_price(p["q_prev_amt"], q_prev_qty)
    pct_p   = (p_new - p_old) / p_old * 100 if p_old else 0

    # Classify driver
    if abs(pct_vol) > abs(pct_p) * 2:
        driver = "Volume-driven"
        driver_th = "ปริมาณเป็นปัจจัยหลัก"
    elif abs(pct_p) > abs(pct_vol) * 2:
        driver = "Price-driven"
        driver_th = "ราคาเป็นปัจจัยหลัก"
    else:
        driver = "Mixed"
        driver_th = "ปริมาณ + ราคาเปลี่ยนแปลงพร้อมกัน"

    # Special pattern notes
    if label == "B3" and period == "yoy" and pct_amt < -30:
        pattern = "DEMAND_LOSS — ปริมาณลดมากที่สุดในกลุ่ม HR Pipe (-28.9%) อาจสูญเสียลูกค้ารายใหญ่"
    elif label == "B3" and period == "qoq" and pct_amt > 40:
        pattern = "PROJECT_SPIKE — เพิ่มขึ้น +43.9% QoQ สูงผิดปกติ น่าจะมาจาก project order"
    elif label == "D1" and pct_amt < -25:
        pattern = "ANCHOR_DECLINE — สินค้าหลักที่มียอดสูงสุดลดลงต่อเนื่อง ผูกกับลูกค้า 1400001"
    elif label in ("A1", "A2"):
        pattern = "PRODUCT_TRANSITION — Coil GIM แบบเดิม (A1) หายไป ถูกแทนด้วย กว้าง (A2) ผลสุทธิ -11.2M"
    elif label == "C1" and period == "qoq" and pct_amt > 35:
        pattern = "SEASONAL_RECOVERY — C Channel HR ฟื้นตัว Q1 หลังยอดต่ำใน Q4 (seasonal construction)"
    elif label in ("B1", "B2", "B3", "B4") and period == "yoy" and pct_amt < -15:
        pattern = f"INDUSTRY_PRESSURE — ท่อกลุ่ม HR/GI ทุกประเภทลดลง YoY ทั้ง volume และ price (market-wide)"
    elif label in ("B1", "B2") and period == "qoq" and pct_amt > 10:
        pattern = "SEASONAL_RECOVERY — ท่อ HR ฟื้นตัวใน Q1 เป็น high season ของ construction"
    else:
        pattern = f"{driver} — {driver_th}"

    return (
        f"\n  ข้อสังเกต: {pattern}\n"
        f"  (Volume {pct_vol:+.1f}% | Price {pct_p:+.1f}% | Amount {pct_amt:+.1f}%)"
    )


def qualitative_overall_summary():
    """Generate executive-level qualitative summary of all behavioral patterns."""

    # Compute customer stats
    yoy_map = {c["sold_to"]: c for c in cust_yoy if c["label"]}
    qoq_map = {c["sold_to"]: c for c in cust_qoq if c["label"]}

    def pct_val(c):
        return (c["chg"] / c["q_prev"] * 100) if c.get("q_prev") else None

    risk_yoy  = [c for c in cust_yoy if c["label"] and c.get("q_prev") and
                 (c["chg"] / c["q_prev"] * 100) <= -40]
    risk_qoq  = [c for c in cust_qoq if c["label"] and c.get("q_prev") and
                 (c["chg"] / c["q_prev"] * 100) <= -20]
    grow_yoy  = [c for c in cust_yoy if c["label"] and c.get("q_prev") and
                 (c["chg"] / c["q_prev"] * 100) >= 50]
    grow_qoq  = [c for c in cust_qoq if c["label"] and
                 (c["q_prev"] == 0 or (c.get("q_prev") and (c["chg"] / c["q_prev"] * 100) >= 50))]

    # Customer 1400001 cross-period
    e_yoy = yoy_map.get(1400001, {})
    e_qoq = qoq_map.get(1400001, {})
    e_yoy_pct = pct_val(e_yoy) if e_yoy else None
    e_qoq_pct = pct_val(e_qoq) if e_qoq else None

    # Slab HR
    slab_yoy = prod_yoy.get("D1", {})
    slab_qoq = prod_qoq.get("D1", {})

    lines = [
        "=" * 70,
        "  บันทึกอธิบายความแปรปรวน (Variance Explanation Memo) — Q1/2569",
        "  บริษัท AMC | จัดทำเพื่อประกอบการตรวจสอบ",
        "=" * 70,
        "",
        "─" * 70,
        "  ก. สรุปความแปรปรวนยอดขายโดยรวม",
        "─" * 70,
        "",
    ]

    # Total revenue context
    rev_c   = gp.get("rev_curr", 0) or 0
    rev_p   = gp.get("rev_prev", 0) or 0
    rev_chg = rev_c - rev_p
    gp_c    = gp.get("gp_curr", 0) or 0
    gp_p    = gp.get("gp_prev", 0) or 0
    gpr_c   = (gp.get("gpr_curr", 0) or 0) * 100
    gpr_p   = (gp.get("gpr_prev", 0) or 0) * 100

    lines += [
        f"  ยอดขายรวม Q1'26 (QoQ vs Q4'25): {M(rev_c, sign=False)} "
        f"เพิ่มขึ้น {M(rev_chg)} ({pct_str(rev_chg, rev_p)})",
        f"  Gross Profit Q1'26: {M(gp_c, sign=False)} | GP Rate: {gpr_c:.2f}%",
        f"  (เทียบ Q4'25: GP = {M(gp_p, sign=False)} | GP Rate: {gpr_p:.2f}%)",
        "",
        "  ความแปรปรวนเกิดจากปัจจัยหลัก 2 ประการ ดังนี้",
        "  (1) ปริมาณขายสินค้ากลุ่มท่อ GI และ C Channel HR ที่เพิ่มขึ้นตาม Seasonal Demand",
        "  (2) ยอดขาย Slab HR ที่ลดลงต่อเนื่องจากปริมาณสั่งซื้อของลูกค้ารายหลักที่ลดลง",
        "",
        "─" * 70,
        "  ข. ความแปรปรวนยอดขายตามลูกค้า",
        "─" * 70,
        "",
        "  ข.1 ลูกค้าที่มียอดขายลดลงอย่างมีนัยสำคัญ (YoY เทียบ Q1'25)",
    ]

    for c in risk_yoy:
        pv = pct_val(c)
        lines.append(
            f"       Sold-to {c['sold_to']} [{c['label']}]: "
            f"ยอดขายลดลง {M(c['chg'], sign=False)} ({pv:+.1f}%) "
            f"จาก {M(c['q_prev'], sign=False)} เหลือ {M(c['q_curr'], sign=False)}"
        )
    if not risk_yoy:
        lines.append("       - ไม่มีลูกค้าที่มียอดขายลดลงเกินร้อยละ 40")

    lines += ["", "  ข.2 ลูกค้าที่มียอดขายลดลงทั้งเมื่อเทียบ YoY และ QoQ"]
    double_declines = []
    for sold_to, cy in yoy_map.items():
        if sold_to in qoq_map:
            cq   = qoq_map[sold_to]
            pv_y = pct_val(cy) or 0
            pv_q = pct_val(cq) or 0
            if pv_y < 0 and pv_q < 0:
                double_declines.append((sold_to, cy["label"], cq["label"], pv_y, pv_q, cy["q_curr"]))
    if double_declines:
        for sold_to, ly, lq, pvy, pvq, curr in double_declines:
            lines.append(
                f"       Sold-to {sold_to}: YoY {pvy:+.1f}% | QoQ {pvq:+.1f}% "
                f"| ยอดขาย Q1'26 = {M(curr, sign=False)}"
            )
            lines.append(
                f"       (ยอดขายลดลงต่อเนื่องทั้งสองงวด — ควรระบุสาเหตุในบันทึกประกอบ)"
            )
    else:
        lines.append("       - ไม่มีลูกค้าที่มียอดขายลดลงทั้งสองงวด")

    lines += ["", "  ข.3 ลูกค้าที่มียอดขายเพิ่มขึ้นอย่างมีนัยสำคัญ"]
    for c in grow_qoq:
        sold_to = c["sold_to"]
        pv_q    = pct_val(c)
        cy      = yoy_map.get(sold_to)
        if cy:
            pv_y = pct_val(cy)
            if pv_y and pv_y > 0:
                lines.append(
                    f"       Sold-to {sold_to}: YoY {pv_y:+.1f}% | QoQ {pv_q:+.1f}% "
                    f"(เพิ่มขึ้นทั้งสองงวด)"
                )
                continue
        tag = "ไม่มียอดขายในงวดก่อน" if (c.get("q_prev") == 0) else f"QoQ {pv_q:+.1f}%"
        lines.append(f"       Sold-to {sold_to}: {tag}")

    lines += [
        "",
        "  ข.4 รูปแบบตามฤดูกาล (Seasonal Pattern)",
        "       จากข้อมูล QoQ พบว่าลูกค้าหลายรายมียอดขายสูงขึ้นอย่างมีนัยสำคัญในงวด Q1",
        "       ซึ่งสอดคล้องกับ Seasonal Demand ของอุตสาหกรรมก่อสร้างในไตรมาสที่ 1:",
    ]
    seasonal = [(c["sold_to"], pct_val(c)) for c in cust_qoq
                if c["label"] and c.get("q_prev") and pct_val(c) and pct_val(c) >= 50]
    for sold_to, pv in seasonal:
        lines.append(f"       Sold-to {sold_to}: QoQ +{pv:.1f}%")
    if not seasonal:
        lines.append("       - ไม่พบรูปแบบตามฤดูกาลที่ชัดเจน")

    lines += [
        "",
        "─" * 70,
        "  ค. ความแปรปรวนยอดขายตามสินค้า",
        "─" * 70,
        "",
        "  ค.1 สินค้า Slab HR (รหัส D1) — ยอดขายลดลงทั้ง YoY และ QoQ",
    ]
    if slab_yoy:
        p_new_y = avg_price(slab_yoy["q_curr_amt"], slab_yoy["q_curr_qty"])
        p_old_y = slab_yoy["q_prev_price"] or avg_price(slab_yoy["q_prev_amt"], slab_yoy["q_prev_qty"])
        lines += [
            f"       YoY: ปริมาณขายลดลง {slab_yoy['q_curr_qty'] - slab_yoy['q_prev_qty']:,.0f} กก. "
            f"({(slab_yoy['q_curr_qty'] - slab_yoy['q_prev_qty'])/slab_yoy['q_prev_qty']*100:+.1f}%) "
            f"ราคาขายเฉลี่ยลดลง {p_new_y - p_old_y:+.2f} บาท/กก.",
            f"       ยอดขายลดลงสุทธิ {M(slab_yoy['chg_amt'])}",
        ]
    if slab_qoq:
        p_new_q = avg_price(slab_qoq["q_curr_amt"], slab_qoq["q_curr_qty"])
        p_old_q = slab_qoq["q_prev_price"] or avg_price(slab_qoq["q_prev_amt"], slab_qoq["q_prev_qty"])
        lines += [
            f"       QoQ: ปริมาณขายลดลง {slab_qoq['q_curr_qty'] - slab_qoq['q_prev_qty']:,.0f} กก. "
            f"({(slab_qoq['q_curr_qty'] - slab_qoq['q_prev_qty'])/slab_qoq['q_prev_qty']*100:+.1f}%) "
            f"ราคาขายเฉลี่ยลดลง {p_new_q - p_old_q:+.2f} บาท/กก.",
            f"       ยอดขายลดลงสุทธิ {M(slab_qoq['chg_amt'])}",
        ]
    lines += [
        "       หมายเหตุ: ยอดขาย Slab HR ผูกกับลูกค้า Sold-to 1400001 เป็นหลัก",
        "       การลดลงของยอดขายสอดคล้องกับการลดลงของปริมาณสั่งซื้อของลูกค้ารายดังกล่าว",
        "",
        "  ค.2 กลุ่มสินค้าท่อ HR — ลดลง YoY จากทั้งปริมาณและราคา",
        "       ท่อกลุ่ม HR ทุกประเภท (Round / Rect. / Square) มียอดขายลดลงเมื่อเทียบ YoY",
        "       โดยสาเหตุหลักมาจากปริมาณขายที่ลดลง (Volume Effect) ประกอบกับราคาขายเฉลี่ยที่ปรับลดลง",
        "",
        "  ค.3 Coil GIM — การเปลี่ยนแปลงรหัสสินค้า (Product Reclassification)",
        "       Coil GIM (รหัสเดิม, A1) ไม่มีรายการขายใน Q1'26",
        "       ขณะที่ Coil GIM กว้าง (รหัสใหม่, A2) เริ่มมีรายการขายเป็นครั้งแรก",
        "       ผลต่างสุทธิรวมกัน: ยอดขายลดลงประมาณ 11.2 ล้านบาท",
        "       (ควรตรวจสอบว่าเป็นการเปลี่ยนรหัสสินค้าเดิมหรือสินค้าต่างประเภท)",
        "",
        "─" * 70,
        "  ง. ข้อสังเกตเพิ่มเติมสำหรับผู้สอบบัญชี",
        "─" * 70,
        "",
        "  1. ยอดขาย Slab HR ลดลงต่อเนื่องทั้ง YoY และ QoQ และมีการกระจุกตัวสูง",
        "     กับลูกค้าเพียงรายเดียว (Sold-to 1400001) — ควรพิจารณาความเสี่ยงด้านการกระจุกตัว",
        "  2. ความแตกต่างของ GP Rate ระหว่าง Q1'26 และ Q4'25 มีนัยสำคัญ",
        f"     ({gpr_c:.2f}% vs {gpr_p:.2f}%) — เกิดจากการฟื้นตัวของ Revenue ขณะที่ COGS เพิ่มขึ้นน้อยกว่า",
        "  3. Coil GIM รหัสเดิมหยุดขายและมีรหัสใหม่เข้ามาแทน",
        "     ควรตรวจสอบการรับรู้รายได้ว่าถูกต้องและครบถ้วน",
        "  4. ลูกค้า Sold-to 1100842 ไม่มียอดขายในงวด Q4'25 แต่มียอดขายใน Q1'26",
        "     ควรตรวจสอบเอกสารประกอบการรับรู้รายได้",
        "",
        "=" * 70,
    ]

    return "\n".join(lines)


# ── Build answers (quantitative + qualitative combined) ────────────────────
ans_cust_yoy = {
    c["label"]: answer_customer(c, "Q1'25") + behavioral_note(c, "yoy")
    for c in cust_yoy if c["label"]
}
ans_cust_qoq = {
    c["label"]: answer_customer(c, "Q4'25") + behavioral_note(c, "qoq")
    for c in cust_qoq if c["label"]
}
ans_prod_yoy = {
    label: answer_product(p, "Q1'26", "Q1'25")
           + qualitative_product_note(label, p, "yoy")
    for label, p in prod_yoy.items()
}
ans_prod_qoq = {
    label: answer_product(p, "Q1'26", "Q4'25")
           + qualitative_product_note(label, p, "qoq")
    for label, p in prod_qoq.items()
}
ans_gp = answer_gp()
ans_qual_summary = qualitative_overall_summary()


# ══════════════════════════════════════════════════════════════════════════════
# CLAUDE AI ANALYSIS PLUGIN
# ══════════════════════════════════════════════════════════════════════════════
def build_data_summary() -> str:
    """Build a structured text snapshot of all extracted data to pass to Claude."""
    lines = []

    # ── KPI ──
    lines.append("=== AMC SALES DATA SNAPSHOT — Q1 FY2026 ANALYSIS ===\n")

    rev_c = gp.get("rev_curr", 0) or 0
    rev_p = gp.get("rev_prev", 0) or 0
    gp_c  = gp.get("gp_curr",  0) or 0
    gp_p  = gp.get("gp_prev",  0) or 0
    gpr_c = (gp.get("gpr_curr", 0) or 0) * 100
    gpr_p = (gp.get("gpr_prev", 0) or 0) * 100

    lines.append("[REVENUE & GP — Q1'26 vs Q4'25 (QoQ)]")
    lines.append(f"  Revenue  Q1'26: {rev_c/1e6:,.2f}M THB")
    lines.append(f"  Revenue  Q4'25: {rev_p/1e6:,.2f}M THB  | Change: {(rev_c-rev_p)/1e6:+,.2f}M THB ({(rev_c-rev_p)/rev_p*100:+.1f}%)")
    lines.append(f"  GP       Q1'26: {gp_c/1e6:,.2f}M THB  | GP Rate: {gpr_c:.2f}%")
    lines.append(f"  GP       Q4'25: {gp_p/1e6:,.2f}M THB  | GP Rate: {gpr_p:.2f}%")
    lines.append(f"  GP Change:      {(gp_c-gp_p)/1e6:+,.2f}M THB\n")

    # ── Customer YoY ──
    lines.append("[CUSTOMER PERFORMANCE — Q1'26 vs Q1'25 (YoY)]")
    for c in sorted(cust_yoy, key=lambda x: x["label"] or "Z"):
        if not c["label"]:
            continue
        pct = c["chg"] / abs(c["q_prev"]) * 100 if c["q_prev"] else 0
        lines.append(f"  [{c['label']}] Sold-to {c['sold_to']:>8} | Q1'26: {c['q_curr']/1e6:,.2f}M"
                     f" | Q1'25: {c['q_prev']/1e6:,.2f}M | CHG: {c['chg']/1e6:+,.2f}M ({pct:+.1f}%)")
    lines.append("")

    # ── Customer QoQ ──
    lines.append("[CUSTOMER PERFORMANCE — Q1'26 vs Q4'25 (QoQ)]")
    for c in sorted(cust_qoq, key=lambda x: x["label"] or "Z"):
        if not c["label"]:
            continue
        pct = c["chg"] / abs(c["q_prev"]) * 100 if c["q_prev"] else 0
        lines.append(f"  [{c['label']}] Sold-to {c['sold_to']:>8} | Q1'26: {c['q_curr']/1e6:,.2f}M"
                     f" | Q4'25: {c['q_prev']/1e6:,.2f}M | CHG: {c['chg']/1e6:+,.2f}M ({pct:+.1f}%)")
    lines.append("")

    # ── Product YoY ──
    lines.append("[PRODUCT PERFORMANCE — Q1'26 vs Q1'25 (YoY)]")
    for lbl, p in sorted(prod_yoy.items()):
        pct = p["chg_amt"] / abs(p["q_prev_amt"]) * 100 if p["q_prev_amt"] else 0
        ve  = vol_effect(p["q_curr_qty"], p["q_prev_qty"], p["q_prev_price"])
        pe  = price_effect(p["q_curr_price"], p["q_prev_price"], p["q_prev_qty"])
        lines.append(f"  [{lbl}] {(p['name'] or '')[:30]:<30} | AMT: {p['q_curr_amt']/1e6:,.2f}M → {p['q_prev_amt']/1e6:,.2f}M"
                     f" | CHG: {p['chg_amt']/1e6:+,.2f}M ({pct:+.1f}%)"
                     f" | Vol.Effect: {ve/1e6:+,.2f}M | Price.Effect: {pe/1e6:+,.2f}M")
    lines.append("")

    # ── Product QoQ ──
    lines.append("[PRODUCT PERFORMANCE — Q1'26 vs Q4'25 (QoQ)]")
    for lbl, p in sorted(prod_qoq.items()):
        pct = p["chg_amt"] / abs(p["q_prev_amt"]) * 100 if p["q_prev_amt"] else 0
        ve  = vol_effect(p["q_curr_qty"], p["q_prev_qty"], p["q_prev_price"])
        pe  = price_effect(p["q_curr_price"], p["q_prev_price"], p["q_prev_qty"])
        lines.append(f"  [{lbl}] {(p['name'] or '')[:30]:<30} | AMT: {p['q_curr_amt']/1e6:,.2f}M → {p['q_prev_amt']/1e6:,.2f}M"
                     f" | CHG: {p['chg_amt']/1e6:+,.2f}M ({pct:+.1f}%)"
                     f" | Vol.Effect: {ve/1e6:+,.2f}M | Price.Effect: {pe/1e6:+,.2f}M")
    lines.append("")

    # ── Rule-based qualitative tags (context for Claude) ──
    lines.append("[RULE-BASED BEHAVIORAL TAGS (for context)]")
    for c in cust_yoy:
        if not c["label"]:
            continue
        tag, _, _ = classify_customer(c["sold_to"], c["chg"] / abs(c["q_prev"]) * 100 if c["q_prev"] else 0,
                                       c["q_curr"], c["q_prev"], "yoy")
        lines.append(f"  Cust YoY [{c['label']}] {c['sold_to']}: {tag}")
    for c in cust_qoq:
        if not c["label"]:
            continue
        tag, _, _ = classify_customer(c["sold_to"], c["chg"] / abs(c["q_prev"]) * 100 if c["q_prev"] else 0,
                                       c["q_curr"], c["q_prev"], "qoq")
        lines.append(f"  Cust QoQ [{c['label']}] {c['sold_to']}: {tag}")
    lines.append("")

    return "\n".join(lines)


def ask_claude(data_summary: str) -> str:
    """Call Claude API for deep strategic analysis. Returns analysis text."""
    if not CLAUDE_ENABLED:
        return (
            "[Claude AI Insights — DISABLED]\n"
            "ตั้งค่า environment variable ANTHROPIC_API_KEY เพื่อเปิดใช้งาน Claude AI analysis\n"
            "ตัวอย่าง: set ANTHROPIC_API_KEY=sk-ant-..."
        )

    print("\n[Claude AI] กำลังวิเคราะห์ข้อมูล... (ใช้เวลาประมาณ 20-40 วินาที)", flush=True)

    prompt = f"""\
คุณคือผู้สอบบัญชีอาวุโส (Senior Auditor) ที่เชี่ยวชาญด้านการตรวจสอบรายได้และต้นทุนของบริษัทในอุตสาหกรรมเหล็กและวัสดุก่อสร้างของไทย

ด้านล่างนี้คือข้อมูล Sales Variance Analysis ของบริษัท AMC สำหรับ Q1 FY2026 เปรียบเทียบกับ Q1 FY2025 (YoY) และ Q4 FY2025 (QoQ) ที่ทีมบัญชีจัดทำขึ้น:

{data_summary}

กรุณาเขียนบันทึกอธิบายความแปรปรวน (Variance Explanation Memo) ในรูปแบบที่เหมาะสำหรับการนำเสนอต่อผู้สอบบัญชีภายนอก โดยครอบคลุมหัวข้อต่อไปนี้:

1. **สรุปภาพรวม (Executive Summary)**
   - อธิบายภาพรวมความแปรปรวนของยอดขายและกำไรขั้นต้นใน Q1'26
   - ระบุปัจจัยหลักที่ทำให้เกิดความแปรปรวน (Volume / Price / Mix)

2. **อธิบายความแปรปรวนยอดขายตามลูกค้า**
   - ลูกค้าที่มียอดขายลดลงอย่างมีนัยสำคัญ — ระบุสาเหตุที่สมเหตุสมผล
   - ลูกค้าที่มียอดขายเพิ่มขึ้น — อธิบายว่าเกิดจากอะไร (Seasonal / New Order / Price)
   - ข้อสังเกตด้านการกระจุกตัวของรายได้ (Revenue Concentration Risk)

3. **อธิบายความแปรปรวนยอดขายตามสินค้า**
   - ระบุว่าความแปรปรวนเกิดจาก Volume Effect หรือ Price Effect เป็นหลัก
   - สินค้าที่มีการเปลี่ยนแปลงผิดปกติและควรมีเอกสารอธิบายประกอบ
   - Slab HR ที่ลดลงต่อเนื่อง และ Coil GIM ที่มีการเปลี่ยนรหัส

4. **อธิบายการเปลี่ยนแปลง Gross Profit Rate**
   - เหตุผลที่ GP Rate เปลี่ยนจาก Q4'25 เป็น Q1'26
   - ความสมเหตุสมผลของตัวเลข GP ที่ติดลบใน Q4'25

5. **ประเด็นที่ควรขอเอกสารประกอบเพิ่มเติม**
   - ระบุรายการที่ผู้สอบบัญชีควรขอเอกสารยืนยันเพิ่มเติม

เขียนด้วยภาษาทางการ สุภาพ ชัดเจน ไม่มีศัพท์ฟุ่มเฟือย
ใช้ภาษาไทยเป็นหลัก ยกเว้นคำศัพท์เทคนิคบัญชีที่นิยมใช้ภาษาอังกฤษ
ความยาวประมาณ 600-900 คำ
"""

    try:
        client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
        message = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        result = message.content[0].text
        print("[Claude AI] วิเคราะห์เสร็จสิ้น", flush=True)
        return result
    except Exception as e:
        return f"[Claude AI Error] {e}\n\nกรุณาตรวจสอบ ANTHROPIC_API_KEY และการเชื่อมต่ออินเทอร์เน็ต"


# Build Claude analysis (runs after all data is extracted)
data_summary   = build_data_summary()
ans_claude_ai  = ask_claude(data_summary)


# ══════════════════════════════════════════════════════════════════════════════
# CONSOLE REPORT
# ══════════════════════════════════════════════════════════════════════════════
SEP = "=" * 72

def print_report():
    print(SEP)
    print("  AMC ANALYTIC SALES Q1'26 — VARIANCE ANALYSIS ANSWERS")
    print(SEP)

    # KPI Summary
    print(f"\n[KPI Summary]")
    print(f"  ยอดขาย Q1'26 : {M(tot_q126_yoy, sign=False)}")
    print(f"  ยอดขาย Q1'25 : {M(tot_q125,     sign=False)}"
          f"  | YoY: {M(tot_q126_yoy - tot_q125)}"
          f" ({pct_str(tot_q126_yoy - tot_q125, tot_q125)})")
    print(f"  ยอดขาย Q4'25 : {M(tot_q425,     sign=False)}"
          f"  | QoQ: {M(tot_q126_qoq - tot_q425)}"
          f" ({pct_str(tot_q126_qoq - tot_q425, tot_q425)})")
    gp_c = gp.get("gp_curr") or 0
    gp_p = gp.get("gp_prev") or 0
    gpr_c = (gp.get("gpr_curr") or 0) * 100
    gpr_p = (gp.get("gpr_prev") or 0) * 100
    print(f"  GP Q1'26     : {M(gp_c, sign=False)} (GP Rate = {gpr_c:.2f}%)")
    print(f"  GP Q4'25     : {M(gp_p, sign=False)} (GP Rate = {gpr_p:.2f}%)")

    # --- Customer YoY ---
    print(f"\n{SEP}")
    print("  SHEET 1: ลูกค้า Q1'26 vs Q1'25 (YoY)")
    print(SEP)
    for label in sorted(ans_cust_yoy):
        c = next((x for x in cust_yoy if x["label"] == label), {})
        print(f"\n  [{label}] Sold-to: {c.get('sold_to','?')}")
        for line in ans_cust_yoy[label].split("\n"):
            print(f"      {line}")

    # --- Customer QoQ ---
    print(f"\n{SEP}")
    print("  SHEET 2: ลูกค้า Q1'26 vs Q4'25 (QoQ)")
    print(SEP)
    for label in sorted(ans_cust_qoq):
        c = next((x for x in cust_qoq if x["label"] == label), {})
        print(f"\n  [{label}] Sold-to: {c.get('sold_to','?')}")
        for line in ans_cust_qoq[label].split("\n"):
            print(f"      {line}")

    # --- Product YoY ---
    print(f"\n{SEP}")
    print("  SHEET 3: สินค้า Q1'26 vs Q1'25 (YoY)")
    print(SEP)
    for label in sorted(ans_prod_yoy):
        print(f"\n  [{label}]")
        for line in ans_prod_yoy[label].split("\n"):
            print(f"      {line}")

    # --- Product QoQ ---
    print(f"\n{SEP}")
    print("  SHEET 4: สินค้า Q1'26 vs Q4'25 (QoQ)")
    print(SEP)
    for label in sorted(ans_prod_qoq):
        print(f"\n  [{label}]")
        for line in ans_prod_qoq[label].split("\n"):
            print(f"      {line}")

    # --- GP ---
    print(f"\n{SEP}")
    print("  SHEET 5: Gross Profit Analytic Q1'26 vs Q4'25")
    print(SEP)
    for line in ans_gp.split("\n"):
        print(f"  {line}")

    # --- Qualitative Summary ---
    print(f"\n{ans_qual_summary}")
    print(f"\n{SEP}")

    # --- Claude AI Insights ---
    print(f"\n{SEP}")
    print("  CLAUDE AI STRATEGIC INSIGHTS")
    print(SEP)
    for line in ans_claude_ai.split("\n"):
        print(f"  {line}")
    print(f"\n{SEP}")


# ══════════════════════════════════════════════════════════════════════════════
# WRITE ANSWERS TO EXCEL
# Answer slot row numbers (1-indexed as in Excel)
# ══════════════════════════════════════════════════════════════════════════════
CUST_YOY_ROWS  = {"A": 17, "B": 20, "C": 23, "D": 26, "E": 29}
CUST_QOQ_ROWS  = {"A": 16, "B": 19, "C": 22, "D": 25, "E": 28}
PROD_YOY_ROWS  = {"A1": 100, "A2": 103,
                  "B1": 107, "B2": 110, "B3": 113, "B4": 116,
                  "C1": 120, "D1": 124}
PROD_QOQ_ROWS  = {"A1": 103,
                  "B1": 107, "B2": 110, "B3": 113,
                  "C1": 117, "D1": 121}
GP_ANS_ROW     = 19


def write_excel():
    wb_out = openpyxl.load_workbook(INPUT_FILE)
    ans_font      = Font(name="TH Sarabun New", size=13, color="1F4E79", bold=False)
    ans_alignment = Alignment(wrap_text=True, vertical="top")

    def write_cell(ws, row_i, col_i, text):
        cell = ws.cell(row=row_i, column=col_i)
        cell.value = text
        cell.font = ans_font
        cell.alignment = ans_alignment
        line_count = text.count("\n") + 1
        ws.row_dimensions[row_i].height = max(line_count * 16, 48)

    # Sheet 1
    ws1 = wb_out["BY Customer Q1FY26 and Q1FY25"]
    for lbl, row_i in CUST_YOY_ROWS.items():
        if lbl in ans_cust_yoy:
            write_cell(ws1, row_i, 3, ans_cust_yoy[lbl])

    # Sheet 2
    ws2 = wb_out["BY Customer Q1FY26 and Q4FY25"]
    for lbl, row_i in CUST_QOQ_ROWS.items():
        if lbl in ans_cust_qoq:
            write_cell(ws2, row_i, 3, ans_cust_qoq[lbl])

    # Sheet 3
    ws3 = wb_out["BY Product Q1FY26 and Q1FY25"]
    for lbl, row_i in PROD_YOY_ROWS.items():
        if lbl in ans_prod_yoy:
            write_cell(ws3, row_i, 3, ans_prod_yoy[lbl])

    # Sheet 4
    ws4 = wb_out["BY Product Q1FY26 and Q4FY25"]
    for lbl, row_i in PROD_QOQ_ROWS.items():
        if lbl in ans_prod_qoq:
            write_cell(ws4, row_i, 3, ans_prod_qoq[lbl])

    # Sheet 5
    ws5 = wb_out["GP Analytic"]
    write_cell(ws5, GP_ANS_ROW, 2, ans_gp)

    # Sheet 6 (new): Qualitative Analysis Summary
    if "Qualitative Analysis" in wb_out.sheetnames:
        del wb_out["Qualitative Analysis"]
    ws_qual = wb_out.create_sheet("Qualitative Analysis")

    # Title row
    title_cell = ws_qual.cell(row=1, column=1, value="Qualitative Analysis — AMC Sales Q1'26")
    title_cell.font = Font(name="TH Sarabun New", size=16, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left")
    ws_qual.row_dimensions[1].height = 28

    # Subtitle
    sub_cell = ws_qual.cell(row=2, column=1, value="สรุปพฤติกรรมลูกค้าและสินค้า — เชิงคุณภาพ")
    sub_cell.font = Font(name="TH Sarabun New", size=13, italic=True, color="595959")
    ws_qual.row_dimensions[2].height = 20

    ws_qual.cell(row=3, column=1, value="")

    # Write summary as individual rows (better for Excel readability)
    section_font   = Font(name="TH Sarabun New", size=13, bold=True, color="FFFFFF")
    section_fill   = PatternFill(
        start_color="2E74B5", end_color="2E74B5", fill_type="solid"
    )
    normal_font    = Font(name="TH Sarabun New", size=13, color="1F1F1F")
    risk_font      = Font(name="TH Sarabun New", size=13, color="C00000", bold=True)
    oppt_font      = Font(name="TH Sarabun New", size=13, color="375623", bold=True)
    normal_align   = Alignment(wrap_text=True, vertical="top")

    row_i = 4
    for line in ans_qual_summary.split("\n"):
        cell = ws_qual.cell(row=row_i, column=1, value=line)
        cell.alignment = normal_align
        stripped = line.strip()
        if line.startswith("=") or (line.startswith("─") and len(line) > 5):
            cell.value = ""  # replace separator lines with blank
            ws_qual.row_dimensions[row_i].height = 6
        elif stripped.startswith("[") and stripped.endswith("]"):
            # Section headers like [CUSTOMER OVERVIEW], [PRODUCT OVERVIEW]
            cell.font = section_font
            cell.fill = section_fill
            ws_qual.row_dimensions[row_i].height = 22
        elif stripped.startswith("URGENT") or "NEAR_CHURN" in line or "CHURN" in line:
            cell.font = risk_font
            ws_qual.row_dimensions[row_i].height = 20
        elif "Opportunity" in line or "GROWTH" in line or "Strategic customer" in line:
            cell.font = oppt_font
            ws_qual.row_dimensions[row_i].height = 20
        elif "─" in line and len(stripped) < 5:
            ws_qual.row_dimensions[row_i].height = 4
        else:
            cell.font = normal_font
            ws_qual.row_dimensions[row_i].height = max(
                18, 18 * (len(line) // 90 + 1)
            )
        row_i += 1

    ws_qual.column_dimensions["A"].width = 100

    # Sheet 7 (new): Claude AI Insights
    if "Claude AI Insights" in wb_out.sheetnames:
        del wb_out["Claude AI Insights"]
    ws_ai = wb_out.create_sheet("Claude AI Insights")

    ai_title = ws_ai.cell(row=1, column=1, value="Claude AI Strategic Insights — AMC Sales Q1'26")
    ai_title.font      = Font(name="TH Sarabun New", size=16, bold=True, color="1F4E79")
    ai_title.alignment = Alignment(horizontal="left")
    ws_ai.row_dimensions[1].height = 28

    ai_sub = ws_ai.cell(row=2, column=1, value=f"Generated by {CLAUDE_MODEL} | {__import__('datetime').date.today()}")
    ai_sub.font      = Font(name="TH Sarabun New", size=11, italic=True, color="595959")
    ws_ai.row_dimensions[2].height = 18

    ws_ai.cell(row=3, column=1, value="")

    ai_normal_font  = Font(name="TH Sarabun New", size=13, color="1F1F1F")
    ai_header_font  = Font(name="TH Sarabun New", size=13, bold=True, color="FFFFFF")
    ai_header_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ai_risk_font    = Font(name="TH Sarabun New", size=13, color="C00000", bold=True)
    ai_action_font  = Font(name="TH Sarabun New", size=13, color="375623", bold=True)
    ai_bold_font    = Font(name="TH Sarabun New", size=13, bold=True, color="1F1F1F")
    ai_wrap         = Alignment(wrap_text=True, vertical="top")

    row_i = 4
    for line in ans_claude_ai.split("\n"):
        cell = ws_ai.cell(row=row_i, column=1, value=line)
        cell.alignment = ai_wrap
        stripped = line.strip()

        if stripped.startswith("**") and stripped.endswith("**") and len(stripped) > 4:
            # Markdown bold header — strip ** and style as section header
            cell.value = stripped.strip("*")
            cell.font  = ai_header_font
            cell.fill  = ai_header_fill
            ws_ai.row_dimensions[row_i].height = 22
        elif stripped.startswith("#"):
            cell.value = stripped.lstrip("# ")
            cell.font  = ai_header_font
            cell.fill  = ai_header_fill
            ws_ai.row_dimensions[row_i].height = 22
        elif "เสี่ยง" in stripped or "Risk" in stripped or "ลดลง" in stripped:
            cell.font = ai_risk_font
            ws_ai.row_dimensions[row_i].height = max(18, 18 * (len(line) // 90 + 1))
        elif stripped.startswith(("1.", "2.", "3.", "4.", "5.")) or "Action" in stripped:
            cell.font = ai_action_font
            ws_ai.row_dimensions[row_i].height = max(18, 18 * (len(line) // 90 + 1))
        elif stripped == "" or stripped == "---":
            cell.value = ""
            ws_ai.row_dimensions[row_i].height = 6
        else:
            cell.font = ai_normal_font
            ws_ai.row_dimensions[row_i].height = max(18, 18 * (len(line) // 90 + 1))
        row_i += 1

    ws_ai.column_dimensions["A"].width = 110

    Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(OUTPUT_FILE)
    print(f"\n[OK] saved: {OUTPUT_FILE}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print_report()
    write_excel()
