"""
nrv_rm_calc_explained_2026_q1.py
==================================
อธิบายการคำนวณ NRV ของ Raw Material (RM) สำหรับ Q1 2026

สูตรหลัก (IAS 2 / TAS 2):
  NRV(RM) = NRV(FG ที่ผลิตได้) − Conversion Cost per KG

ที่มา Conversion Cost per KG = ไฟล์ต้นทุนผลิต AMC_PRD_Q1'2026.XLSX
ที่มา FG NRV price           = ไฟล์ NRV_SKU_Analysis_2026_Q1_v2.xlsx (sheet: SKU_Detail, sku_type=FG)

Output: 04_Reports/NRV_RM_Calc_Explained_2026_Q1.xlsx
  Sheet 1: RM_Calc_Summary     — สรุปสูตรคำนวณ NRV RM แยกตาม RM Type
  Sheet 2: RM_SKU_Detail       — ราย SKU: cost_price, nrv_price, write_down + แสดงที่มา
  Sheet 3: Conv_from_Production— Conversion/KG จากข้อมูลผลิต แยก Mat Group
  Sheet 4: FG_NRV_Reference    — FG NRV price ที่ใช้อ้างอิง แยก product line
  Sheet 5: Methodology         — อธิบายขั้นตอนการคำนวณ
"""

import os, sys
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 0. Config
# ─────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PATH_NRV  = os.path.join(BASE_DIR, "01_Bronze_Raw", "Inventory_RollStock", "NRV",
                         "NRV_SKU_Analysis_2026_Q1_v2.xlsx")
PATH_PRD  = os.path.join(BASE_DIR, "04_Reports", "AMC_PRD_Q1'2026.XLSX")
OUTPUT    = os.path.join(BASE_DIR, "04_Reports", "NRV_RM_Calc_Explained_2026_Q1.xlsx")
REPORT_DATE = date.today().strftime("%d %B %Y")

# RM Type → FG Mat_Group ที่ใช้อ้างอิง NRV
RM_TO_FG_MAP = {
    "RM→FG(HR)":  ["Slit HR", "Pipe Square HR", "Pipe Rect. HR", "Pipe Round HR",
                   "C Channel HR", "Coil HR แคบ"],
    "RM→FG(GI)":  ["Slit GI", "Pipe Square GI", "Pipe Rect. GI", "Pipe Round GI",
                   "Coil GI แคบ", "C Channel GI"],
    "RM→FG(GIM)": ["Slit GIM", "Coil GIM แคบ"],
    "RM→FG(CR)":  ["Coil CR กว้าง"],
}

# Production Mat_Group → RM type mapping
PRD_TO_RM_TYPE = {
    "Slit HR":         "RM→FG(HR)",
    "Pipe Square HR":  "RM→FG(HR)",
    "Pipe Rect. HR":   "RM→FG(HR)",
    "Pipe Round HR":   "RM→FG(HR)",
    "C Channel HR":    "RM→FG(HR)",
    "Coil HR แคบ":     "RM→FG(HR)",
    "Slit GI":         "RM→FG(GI)",
    "Pipe Square GI":  "RM→FG(GI)",
    "Pipe Rect. GI":   "RM→FG(GI)",
    "Pipe Round GI":   "RM→FG(GI)",
    "Coil GI แคบ":     "RM→FG(GI)",
    "C Channel GI":    "RM→FG(GI)",
    "Slit GIM":        "RM→FG(GIM)",
    "Coil GIM แคบ":    "RM→FG(GIM)",
    "Coil CR แคบ":     "RM→FG(CR)",
    "Coil CR กว้าง":   "RM→FG(CR)",
}

# ─────────────────────────────────────────────
# 1. Load Data
# ─────────────────────────────────────────────
print("[1/5] Loading data...")

# NRV file
nrv_all = pd.read_excel(PATH_NRV, sheet_name="SKU_Detail")
nrv_rm  = pd.read_excel(PATH_NRV, sheet_name="RM_Detail")
nrv_fg  = nrv_all[nrv_all["sku_type"] == "FG"].copy()

# Production file
prd = pd.read_excel(PATH_PRD, sheet_name="Sheet1")
cols = list(prd.columns)
prd.rename(columns={
    cols[6]:  "Mat_Group_Name",
    cols[10]: "Mat_Group_1_Name",
    cols[11]: "Mat_Group_2_Name",
    cols[12]: "Mat_Group_3_Name",
    cols[13]: "Mat_Group_4_Name",
}, inplace=True)
prd["Plant"] = prd["Plant"].astype(str).str.strip()

num_cols = [
    "Actual GI Amount", "Actual GI QTY",
    "Actual ByProduct Scrap Amount", "Actual ByProduct Grade B Amount",
    "Actual ByProduct Grade C Amount",
    "Actual GR QTY", "Actual ByProduct Grade B QTY", "Actual ByProduct Grade C QTY",
    "Actual D101 Amount", "Actual D102 Amount",
    "Actual I101 Amount", "Actual I102 Amount", "Actual I103 Amount",
    "Actual I104 Amount", "Actual I105 Amount", "Actual I107 Amount",
    "Actual I108 Amount", "Actual I109 Amount", "Actual I110 Amount",
    "Actual I111 Amount",
]
for c in num_cols:
    prd[c] = pd.to_numeric(prd[c], errors="coerce").fillna(0)

prd["DM_ByProd"]    = (prd["Actual ByProduct Scrap Amount"] +
                       prd["Actual ByProduct Grade B Amount"] +
                       prd["Actual ByProduct Grade C Amount"])
prd["DM_Amount"]    = prd["Actual GI Amount"] - prd["DM_ByProd"]
prd["Direct"]       = prd[["Actual D101 Amount", "Actual D102 Amount"]].sum(axis=1)
prd["Indirect"]     = prd[["Actual I101 Amount", "Actual I102 Amount", "Actual I103 Amount",
                             "Actual I104 Amount", "Actual I105 Amount", "Actual I107 Amount",
                             "Actual I108 Amount", "Actual I109 Amount", "Actual I110 Amount",
                             "Actual I111 Amount"]].sum(axis=1)
prd["Conversion"]   = prd["Direct"] + prd["Indirect"]
prd["Total_Cost"]   = prd["DM_Amount"] + prd["Conversion"]
prd["Total_GR_QTY"] = (prd["Actual GR QTY"] +
                        prd["Actual ByProduct Grade B QTY"] +
                        prd["Actual ByProduct Grade C QTY"])

# ─────────────────────────────────────────────
# 2. Conversion per KG — จากข้อมูลผลิต แยก Mat Group
# ─────────────────────────────────────────────
print("[2/5] Calculating Conversion/KG from production...")

conv_mg = (
    prd.groupby("Mat_Group_Name")[
        ["Actual GI Amount", "Actual GI QTY", "DM_Amount",
         "Direct", "Indirect", "Conversion", "Total_Cost", "Total_GR_QTY"]
    ]
    .sum()
    .reset_index()
)
conv_mg["DM_per_KG"]    = np.where(conv_mg["Total_GR_QTY"] > 0,
                                    conv_mg["DM_Amount"]  / conv_mg["Total_GR_QTY"], np.nan)
conv_mg["Conv_per_KG"]  = np.where(conv_mg["Total_GR_QTY"] > 0,
                                    conv_mg["Conversion"] / conv_mg["Total_GR_QTY"], np.nan)
conv_mg["Direct_per_KG"]   = np.where(conv_mg["Total_GR_QTY"] > 0,
                                        conv_mg["Direct"]   / conv_mg["Total_GR_QTY"], np.nan)
conv_mg["Indirect_per_KG"] = np.where(conv_mg["Total_GR_QTY"] > 0,
                                        conv_mg["Indirect"] / conv_mg["Total_GR_QTY"], np.nan)
conv_mg["RM_Type"]      = conv_mg["Mat_Group_Name"].map(PRD_TO_RM_TYPE)

# Weighted avg Conversion/KG per RM type
conv_rm_type = (
    conv_mg.dropna(subset=["RM_Type"])
    .groupby("RM_Type")
    .apply(lambda g: pd.Series({
        "Total_GR_QTY":   g["Total_GR_QTY"].sum(),
        "Total_Conv":     g["Conversion"].sum(),
        "Total_DM":       g["DM_Amount"].sum(),
        "Conv_per_KG_wtd": g["Conversion"].sum() / g["Total_GR_QTY"].sum()
                              if g["Total_GR_QTY"].sum() > 0 else np.nan,
        "DM_per_KG_wtd":  g["DM_Amount"].sum() / g["Total_GR_QTY"].sum()
                          if g["Total_GR_QTY"].sum() > 0 else np.nan,
    }), include_groups=False)
    .reset_index()
)

# ─────────────────────────────────────────────
# 3. FG NRV price — weighted avg แยก RM Type
# ─────────────────────────────────────────────
print("[3/5] Calculating FG NRV reference prices...")

fg_nrv_rows = []
for rm_type, fg_groups in RM_TO_FG_MAP.items():
    fg_sub = nrv_fg[nrv_fg["mat_group"].isin(fg_groups)].copy()
    fg_sub = fg_sub[fg_sub["nrv_price"].notna() & (fg_sub["stock_qty_kg"] > 0)]

    if len(fg_sub) == 0:
        fg_nrv_rows.append({
            "RM_Type": rm_type, "FG_Groups": ", ".join(fg_groups),
            "FG_SKU_Count": 0, "FG_Total_QTY": 0,
            "FG_NRV_WtdAvg": np.nan, "FG_NRV_Min": np.nan, "FG_NRV_Max": np.nan,
        })
        continue

    wtd = (fg_sub["nrv_price"] * fg_sub["stock_qty_kg"]).sum() / fg_sub["stock_qty_kg"].sum()
    fg_nrv_rows.append({
        "RM_Type":       rm_type,
        "FG_Groups":     ", ".join(fg_groups),
        "FG_SKU_Count":  len(fg_sub),
        "FG_Total_QTY":  fg_sub["stock_qty_kg"].sum(),
        "FG_NRV_WtdAvg": wtd,
        "FG_NRV_Min":    fg_sub["nrv_price"].min(),
        "FG_NRV_Max":    fg_sub["nrv_price"].max(),
    })

fg_nrv_df = pd.DataFrame(fg_nrv_rows)

# ─────────────────────────────────────────────
# 4. RM NRV Summary — merge and show calculation
# ─────────────────────────────────────────────
print("[4/5] Building RM NRV summary...")

rm_summary = fg_nrv_df.merge(conv_rm_type[["RM_Type", "Conv_per_KG_wtd", "DM_per_KG_wtd",
                                            "Total_GR_QTY", "Total_Conv"]],
                               on="RM_Type", how="left")

# Actual RM NRV from file (one value per RM type)
rm_actual_nrv = (
    nrv_rm[nrv_rm["nrv_source"].isin(RM_TO_FG_MAP.keys())]
    .groupby("nrv_source")["nrv_price"]
    .first()
    .rename("Actual_NRV_in_File")
    .reset_index()
    .rename(columns={"nrv_source": "RM_Type"})
)
rm_summary = rm_summary.merge(rm_actual_nrv, on="RM_Type", how="left")

# Derived NRV = FG_NRV_WtdAvg - Conv_per_KG_wtd
rm_summary["NRV_Derived"] = rm_summary["FG_NRV_WtdAvg"] - rm_summary["Conv_per_KG_wtd"]
rm_summary["Diff_vs_File"] = rm_summary["NRV_Derived"] - rm_summary["Actual_NRV_in_File"]

# RM SKU detail with source explanation
rm_sku = nrv_rm.copy()
rm_sku = rm_sku.merge(
    conv_rm_type[["RM_Type", "Conv_per_KG_wtd", "DM_per_KG_wtd"]].rename(
        columns={"RM_Type": "nrv_source"}),
    on="nrv_source", how="left"
)
rm_sku = rm_sku.merge(
    fg_nrv_df[["RM_Type", "FG_NRV_WtdAvg", "FG_Groups"]].rename(
        columns={"RM_Type": "nrv_source"}),
    on="nrv_source", how="left"
)
rm_sku["NRV_Formula"] = rm_sku.apply(
    lambda r: (f"FG NRV ({r['FG_NRV_WtdAvg']:.4f}) − Conv/KG ({r['Conv_per_KG_wtd']:.4f}) = {r['nrv_price']:.4f}"
               if pd.notna(r.get("FG_NRV_WtdAvg")) and pd.notna(r.get("Conv_per_KG_wtd"))
               else "No FG reference data"),
    axis=1
)

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
NAVY    = "1F4E79"
GREEN   = "375623"
RED     = "7B241C"
ORANGE  = "C55A11"
PURPLE  = "4A235A"
LGRAY   = "F2F2F2"
LBLUE   = "D9E1F2"
LGREEN  = "E2EFDA"
LRED    = "FCE4D6"
LYELLOW = "FFF2CC"

def _fill(h):  return PatternFill("solid", fgColor=h)
def _bdr(c="D9D9D9"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _font(color="000000", bold=False, sz=10):
    return Font(color=color, bold=bold, size=sz, name="Calibri")
def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def write_hdr(ws, row, vals, bg, fg="FFFFFF", col=1):
    for i, v in enumerate(vals, start=col):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = _fill(bg); c.font = _font(fg, bold=True)
        c.alignment = _ctr(); c.border = _bdr("999999")
    ws.row_dimensions[row].height = 22

def write_row(ws, row, vals, bg=None, fmts=None, bold=False, col=1):
    for i, v in enumerate(vals, start=col):
        c = ws.cell(row=row, column=i, value=v)
        if bg: c.fill = _fill(bg)
        c.font = _font(bold=bold)
        c.alignment = _lft() if isinstance(v, str) else _ctr()
        c.border = _bdr()
        if fmts and i in fmts: c.number_format = fmts[i]
    ws.row_dimensions[row].height = 20

def auto_col(ws, mn=12, mx=48):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[ltr].width = min(max(w + 2, mn), mx)

# ─────────────────────────────────────────────
# Build Workbook
# ─────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════
# Sheet 1 – RM_Calc_Summary
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "RM_Calc_Summary"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:K1")
t = ws1["A1"]
t.value = f"NRV RM Calculation Summary — Q1 2026  (จัดทำ: {REPORT_DATE})"
t.font = _font("FFFFFF", bold=True, sz=13)
t.fill = _fill(NAVY); t.alignment = _ctr(); t.border = _bdr()
ws1.row_dimensions[1].height = 30

# Formula box
ws1["A3"] = "สูตรคำนวณ (IAS 2 / TAS 2):"
ws1["A3"].font = _font(RED, bold=True, sz=11)
ws1["B3"] = "NRV(RM)  =  NRV ของ Finished Goods ที่ผลิตได้  −  Conversion Cost per KG"
ws1["B3"].font = _font("000000", bold=True, sz=11)
ws1["A4"] = "ที่มา Conversion/KG:"
ws1["A4"].font = _font(NAVY, bold=True)
ws1["B4"] = "ไฟล์ต้นทุนผลิต AMC_PRD_Q1'2026.XLSX — คำนวณ Weighted Avg แยกตาม Product Line"
ws1["B4"].font = _font()
ws1["A5"] = "ที่มา FG NRV:"
ws1["A5"].font = _font(NAVY, bold=True)
ws1["B5"] = "NRV_SKU_Analysis_2026_Q1_v2.xlsx (Sheet: SKU_Detail, sku_type=FG) — Weighted Avg โดย stock_qty_kg"
ws1["B5"].font = _font()

for r in [3, 4, 5]:
    ws1.row_dimensions[r].height = 20

# Main calculation table
r = 7
write_hdr(ws1, r, [
    "RM Type (nrv_source)", "FG Mat_Group อ้างอิง",
    "FG SKU (มี NRV)", "FG QTY (KG)",
    "FG NRV WtdAvg\n(THB/KG)",
    "Conversion/KG\n(จากผลิต Q1'26)",
    "= NRV(RM) คำนวณ\n(THB/KG)",
    "NRV(RM) ในไฟล์\n(THB/KG)",
    "ผลต่าง",
    "Production QTY\nอ้างอิง (KG)",
    "Total Conversion\nQ1'26 (THB)",
], NAVY)
r += 1

for _, row in rm_summary.iterrows():
    is_no_data = row["RM_Type"] == "RM→FG(CR)" or pd.isna(row.get("FG_NRV_WtdAvg"))
    bg = LGREEN if pd.notna(row.get("Actual_NRV_in_File")) else LRED

    diff = row.get("Diff_vs_File", np.nan)
    diff_val = diff if pd.notna(diff) else None

    write_row(ws1, r, [
        row["RM_Type"],
        row["FG_Groups"],
        int(row["FG_SKU_Count"]) if pd.notna(row["FG_SKU_Count"]) else "",
        row["FG_Total_QTY"] if pd.notna(row.get("FG_Total_QTY")) else "",
        row["FG_NRV_WtdAvg"] if pd.notna(row.get("FG_NRV_WtdAvg")) else "N/A",
        row["Conv_per_KG_wtd"] if pd.notna(row.get("Conv_per_KG_wtd")) else "",
        row["NRV_Derived"] if pd.notna(row.get("NRV_Derived")) else "N/A",
        row["Actual_NRV_in_File"] if pd.notna(row.get("Actual_NRV_in_File")) else "N/A",
        diff_val,
        row["Total_GR_QTY"] if pd.notna(row.get("Total_GR_QTY")) else "",
        row["Total_Conv"] if pd.notna(row.get("Total_Conv")) else "",
    ], bg=bg, fmts={
        4: "#,##0", 5: "#,##0.0000", 6: "#,##0.0000",
        7: "#,##0.0000", 8: "#,##0.0000", 9: "0.0000",
        10: "#,##0", 11: "#,##0",
    })
    r += 1

# Note row
ws1.row_dimensions[r].height = 14
r += 1
ws1.cell(row=r, column=1, value=(
    "หมายเหตุ: 'ผลต่าง' อาจเกิดจาก (1) FG NRV ใช้ราคาเฉพาะ SKU ไม่ใช่ WtdAvg  "
    "(2) Conv/KG ใช้ค่าเฉพาะ Product Sub-group  "
    "(3) RM No FG Data = ไม่มี FG ที่ match กับ RM type นี้ในไฟล์ NRV"
)).font = _font(RED, sz=9)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 18
ws1.column_dimensions["G"].width = 18
ws1.column_dimensions["H"].width = 18
ws1.column_dimensions["I"].width = 12
ws1.column_dimensions["J"].width = 18
ws1.column_dimensions["K"].width = 22
ws1.freeze_panes = "A8"

# ══════════════════════════════════════════════
# Sheet 2 – RM_SKU_Detail
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("RM_SKU_Detail")
ws2.sheet_view.showGridLines = False

write_hdr(ws2, 1, [
    "SKU", "Description", "Mat Group", "Grade",
    "Stock QTY (KG)", "Cost Price\n(THB/KG)", "NRV Source",
    "FG Mat_Groups อ้างอิง",
    "FG NRV WtdAvg\n(THB/KG)",
    "Conv/KG\nจากผลิต",
    "NRV Price\n(THB/KG)",
    "Write Down\n/KG",
    "Write Down\nAmount",
    "สูตรที่มา NRV",
], PURPLE)

rm_sku_out = rm_sku[[
    "sku", "description", "mat_group", "grade",
    "stock_qty_kg", "cost_price", "nrv_source",
    "FG_Groups", "FG_NRV_WtdAvg", "Conv_per_KG_wtd",
    "nrv_price", "write_down_per_kg", "write_down_amt",
    "NRV_Formula",
]].sort_values(["nrv_source", "write_down_amt"])

prev_src = None
for ri, row in enumerate(rm_sku_out.itertuples(index=False), start=2):
    is_wd = pd.notna(row.write_down_amt) and row.write_down_amt < 0
    new_src = (row.nrv_source != prev_src)
    bg = LRED if is_wd else (LGREEN if not is_wd and pd.notna(row.nrv_price) else LRED)
    if new_src: bg = LBLUE if row.nrv_source != "RM No FG Data" else LGRAY
    prev_src = row.nrv_source

    write_row(ws2, ri, [
        row.sku, row.description, row.mat_group, row.grade,
        row.stock_qty_kg, row.cost_price, row.nrv_source,
        row.FG_Groups if pd.notna(row.FG_Groups) else "—",
        row.FG_NRV_WtdAvg if pd.notna(row.FG_NRV_WtdAvg) else None,
        row.Conv_per_KG_wtd if pd.notna(row.Conv_per_KG_wtd) else None,
        row.nrv_price if pd.notna(row.nrv_price) else None,
        row.write_down_per_kg if pd.notna(row.write_down_per_kg) else None,
        row.write_down_amt if pd.notna(row.write_down_amt) else None,
        row.NRV_Formula,
    ], bg=bg, fmts={
        5: "#,##0", 6: "#,##0.0000",
        9: "#,##0.0000", 10: "#,##0.0000",
        11: "#,##0.0000", 12: "#,##0.0000", 13: "#,##0",
    })

# Color write_down_amt cells
wd_col = 13
for ri in range(2, ws2.max_row + 1):
    cell = ws2.cell(row=ri, column=wd_col)
    if isinstance(cell.value, (int, float)) and cell.value < 0:
        cell.fill = _fill(LRED)
        cell.font = _font(RED, bold=True)
    elif isinstance(cell.value, (int, float)) and cell.value == 0:
        cell.fill = _fill(LGREEN)

auto_col(ws2)
ws2.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 3 – Conv_from_Production
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Conv_from_Production")
ws3.sheet_view.showGridLines = False

write_hdr(ws3, 1, [
    "Mat Group (Production)", "RM Type Map",
    "GI QTY (KG)", "GR QTY (KG)", "Yield %",
    "GI Amount (THB)", "DM Amount (THB)",
    "Direct (THB)", "Indirect (THB)", "Conversion (THB)",
    "DM/KG", "Direct/KG", "Indirect/KG", "Conversion/KG",
], GREEN)

conv_out = conv_mg.sort_values("Conversion", ascending=False)
for ri, row in enumerate(conv_out.itertuples(index=False), start=2):
    bg = LGREEN if pd.notna(row.RM_Type) else LGRAY
    gi_qty = getattr(row, "Actual GI QTY", 0) or 0
    gi_amt = getattr(row, "Actual GI Amount", 0) or 0
    yield_pct = (row.Total_GR_QTY / gi_qty if gi_qty > 0 else np.nan)
    write_row(ws3, ri, [
        row.Mat_Group_Name,
        row.RM_Type if pd.notna(row.RM_Type) else "—",
        gi_qty,
        row.Total_GR_QTY,
        yield_pct,
        gi_amt,
        row.DM_Amount,
        row.Direct,
        row.Indirect,
        row.Conversion,
        row.DM_per_KG,
        row.Direct_per_KG,
        row.Indirect_per_KG,
        row.Conv_per_KG,
    ], bg=bg, fmts={
        3: "#,##0", 4: "#,##0", 5: "0.00%",
        6: "#,##0", 7: "#,##0", 8: "#,##0", 9: "#,##0", 10: "#,##0",
        11: "#,##0.0000", 12: "#,##0.0000", 13: "#,##0.0000", 14: "#,##0.0000",
    })

# Weighted avg summary per RM type
ws3.row_dimensions[ws3.max_row + 1].height = 8
r_wtd = ws3.max_row + 2
ws3.cell(row=r_wtd, column=1, value="Weighted Average Conversion/KG by RM Type").font = \
    _font(NAVY, bold=True, sz=11)
r_wtd += 1

write_hdr(ws3, r_wtd, [
    "RM Type", "Total GR QTY (KG)", "Total Conversion (THB)", "Conv/KG (Wtd Avg)"
], NAVY)
r_wtd += 1
for _, row in conv_rm_type.iterrows():
    write_row(ws3, r_wtd, [
        row["RM_Type"], row["Total_GR_QTY"], row["Total_Conv"], row["Conv_per_KG_wtd"]
    ], bg=LBLUE, fmts={2: "#,##0", 3: "#,##0", 4: "#,##0.0000"}, bold=True)
    r_wtd += 1

auto_col(ws3)
ws3.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 4 – FG_NRV_Reference
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("FG_NRV_Reference")
ws4.sheet_view.showGridLines = False

write_hdr(ws4, 1, [
    "RM Type", "FG Mat_Group", "FG SKU", "Description",
    "Stock QTY (KG)", "NRV Price (THB/KG)", "NRV Source (FG)",
    "avg_price_4m", "so_price",
], RED)

for rm_type, fg_groups in RM_TO_FG_MAP.items():
    fg_sub = nrv_fg[nrv_fg["mat_group"].isin(fg_groups)].copy()
    fg_sub = fg_sub[fg_sub["nrv_price"].notna()].sort_values("mat_group")
    for ri, row in enumerate(fg_sub.itertuples(index=False)):
        write_row(ws4, ws4.max_row + 1, [
            rm_type, row.mat_group, row.sku, row.description,
            row.stock_qty_kg, row.nrv_price,
            row.nrv_source if pd.notna(row.nrv_source) else "",
            row.avg_price_4m if pd.notna(row.avg_price_4m) else None,
            row.so_price if pd.notna(row.so_price) else None,
        ], fmts={5: "#,##0", 6: "#,##0.0000", 8: "#,##0.0000", 9: "#,##0.0000"})

auto_col(ws4)
ws4.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 5 – Methodology
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("Methodology")
ws5.sheet_view.showGridLines = False

steps = [
    ("หลักการ (IAS 2 / TAS 2)",
     "สินค้าคงเหลือต้องแสดงมูลค่าด้วย ราคาทุน หรือ NRV แล้วแต่อย่างใดต่ำกว่า\n"
     "สำหรับ RM: ถ้า FG ที่ผลิตจาก RM นั้นมี NRV < ต้นทุน → RM อาจต้อง write-down ด้วย"),

    ("ขั้นที่ 1 – จัดกลุ่ม RM ตาม Product Line",
     "แบ่ง RM ออกเป็น 4 กลุ่มตามประเภท FG ที่ผลิตได้:\n"
     "  • RM→FG(HR)  = Coil HR แคบ/กว้าง, Slab HR → ผลิต Slit HR, Pipe HR, C Channel HR\n"
     "  • RM→FG(GI)  = Coil GI กว้าง, Ingot GI → ผลิต Slit GI, Pipe GI, C Channel GI\n"
     "  • RM→FG(GIM) = Coil GIM กว้าง → ผลิต Slit GIM, Coil GIM\n"
     "  • RM→FG(CR)  = Coil CR กว้าง → ผลิต Coil CR"),

    ("ขั้นที่ 2 – หา FG NRV Price (อ้างอิง)",
     "จากไฟล์ NRV_SKU_Analysis_2026_Q1_v2.xlsx Sheet: SKU_Detail (sku_type=FG)\n"
     "คำนวณ Weighted Average NRV Price โดยน้ำหนัก = stock_qty_kg\n"
     "แยกตาม mat_group ที่อยู่ใน FG group ของแต่ละ RM type"),

    ("ขั้นที่ 3 – หา Conversion Cost/KG (อ้างอิงจากผลิตจริง)",
     "จากไฟล์ AMC_PRD_Q1'2026.XLSX\n"
     "Conversion = Direct (D101+D102) + Indirect (I101–I111)\n"
     "Conversion/KG = Total Conversion ÷ Total_GR_QTY (Grade A + B + C)\n"
     "คำนวณ Weighted Average แยกตาม RM type (รวม mat_group ที่ relate)"),

    ("ขั้นที่ 4 – คำนวณ NRV(RM)",
     "NRV(RM) = FG NRV WtdAvg − Conversion/KG\n\n"
     "ตัวอย่าง (HR):\n"
     f"  FG NRV WtdAvg (HR line)  ≈ ค่าที่คำนวณ (ดู Sheet RM_Calc_Summary)\n"
     f"  Conversion/KG (HR line)  ≈ ค่าที่คำนวณ (ดู Sheet Conv_from_Production)\n"
     f"  → NRV(RM HR) = ค่าที่บันทึกใน nrv_price ในไฟล์ NRV"),

    ("ขั้นที่ 5 – เปรียบเทียบ Cost vs NRV",
     "cost_price (= stock_price = ราคาทุน RM จาก Inventory)\n"
     "nrv_price = ค่าที่คำนวณได้จากขั้นที่ 4\n\n"
     "ถ้า cost_price > nrv_price:\n"
     "  write_down_per_kg = nrv_price − cost_price  (ค่าลบ)\n"
     "  write_down_amt    = write_down_per_kg × stock_qty_kg\n\n"
     "ถ้า cost_price ≤ nrv_price → ไม่ต้อง write-down (write_down = 0)"),

    ("RM No FG Data — กรณีพิเศษ",
     "RM ที่ nrv_source = 'RM No FG Data' ได้แก่:\n"
     "  Coil EG, Coil GA, Coil GL, Coil PCM, Coil PO, Coil SI\n"
     "ไม่มีข้อมูล FG NRV ในไฟล์ → ไม่สามารถคำนวณ NRV ได้จากสูตรนี้\n"
     "→ nrv_price = NaN, write_down = NaN\n"
     "→ แนะนำ: ใช้ราคาตลาดหรือ recent purchase price เปรียบเทียบแทน"),
]

write_hdr(ws5, 1, ["ขั้นตอน / หัวข้อ", "รายละเอียด"], NAVY)
ws5.column_dimensions["A"].width = 38
ws5.column_dimensions["B"].width = 80

for ri, (title, detail) in enumerate(steps, start=2):
    ws5.row_dimensions[ri].height = 80
    c1 = ws5.cell(row=ri, column=1, value=title)
    c1.fill = _fill(LBLUE if ri % 2 == 0 else LGREEN)
    c1.font = _font(NAVY, bold=True)
    c1.alignment = _lft(); c1.border = _bdr()

    c2 = ws5.cell(row=ri, column=2, value=detail)
    c2.fill = _fill(LGRAY if ri % 2 == 0 else "FFFFFF")
    c2.font = _font()
    c2.alignment = _lft(); c2.border = _bdr()

ws5.freeze_panes = "A2"

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
print("[5/5] Saving...")
wb.save(OUTPUT)

# Summary print
print(f"\n  Output : {OUTPUT}")
print(f"  Sheets : {[s.title for s in wb.worksheets]}")

total_wd   = nrv_rm["write_down_amt"].sum()
total_wd_n = nrv_rm[nrv_rm["write_down_amt"] < 0]["write_down_amt"].sum()
print(f"\n=== RM Write-Down Summary ===")
print(f"  RM SKUs total       : {len(nrv_rm):,}")
print(f"  SKUs with write-down: {(nrv_rm['write_down_amt'] < 0).sum():,}")
print(f"  Total write-down amt: {total_wd_n:>20,.2f} THB")
print(f"  RM No FG Data       : {(nrv_rm['nrv_source']=='RM No FG Data').sum():,} SKUs (nrv_price = NaN)")
