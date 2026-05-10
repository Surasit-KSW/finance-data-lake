"""
production_variance_allocation_2026_q1.py
==========================================
1. กระทบยอด Actual Production Cost vs GR Amount (ตัด 2 rows anomalous ออก)
2. คำนวณ Variance per KG แยก Mat Group
3. กระจาย Variance เข้า FG Inventory (จากไฟล์ NRV) ตามสัดส่วน stock_qty_kg
4. คำนวณ Adjusted Cost Price และ Revised NRV Comparison

Output: 04_Reports/Production_Variance_Allocation_2026_Q1.xlsx
  Sheet 1: Reconciliation      — กระทบยอด Actual vs GR Amount
  Sheet 2: Variance_by_MatGroup— Variance แยก Mat Group + Variance Rate/KG
  Sheet 3: FG_Allocation       — กระจาย Variance ลงราย SKU
  Sheet 4: Revised_NRV         — Adjusted Cost vs NRV หลังกระจาย Variance
  Sheet 5: WIP_Cost            — WIP orders ที่ยังไม่ได้รับ (ข้อมูลประกอบ)
"""

import os
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 0. Config
# ─────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PATH_PRD  = os.path.join(BASE_DIR, "04_Reports", "AMC_PRD_Q1'2026.XLSX")
PATH_NRV  = os.path.join(BASE_DIR, "01_Bronze_Raw", "Inventory_RollStock", "NRV",
                         "NRV_SKU_Analysis_2026_Q1_v2.xlsx")
OUTPUT    = os.path.join(BASE_DIR, "04_Reports", "Production_Variance_Allocation_2026_Q1.xlsx")
REPORT_DATE = date.today().strftime("%d %B %Y")

GR_AMT_THRESHOLD = 1_000_000_000   # ตัด rows ที่ GR Amount > 1B (anomalous)

# ─────────────────────────────────────────────
# 1. Load & Prepare Production Data
# ─────────────────────────────────────────────
print("[1/4] Loading production data...")

prd = pd.read_excel(PATH_PRD, sheet_name="Sheet1")
cols = list(prd.columns)
prd.rename(columns={
    cols[6]:  "Mat_Group_Name",
    cols[10]: "Mat_Group_1_Name",
    cols[11]: "Mat_Group_2_Name",
    cols[12]: "Mat_Group_3_Name",
    cols[13]: "Mat_Group_4_Name",
}, inplace=True)
prd["Plant"] = prd["Plant"].astype(str).str.strip()

num_cols = [
    "Actual GI Amount", "Actual GI QTY", "Actual GR Amount", "Actual GR QTY",
    "Actual ByProduct Scrap Amount", "Actual ByProduct Grade B Amount",
    "Actual ByProduct Grade C Amount",
    "Actual ByProduct Grade B QTY", "Actual ByProduct Grade C QTY",
    "Actual D101 Amount", "Actual D102 Amount",
    "Actual I101 Amount", "Actual I102 Amount", "Actual I103 Amount",
    "Actual I104 Amount", "Actual I105 Amount", "Actual I107 Amount",
    "Actual I108 Amount", "Actual I109 Amount", "Actual I110 Amount",
    "Actual I111 Amount",
]
for c in num_cols:
    prd[c] = pd.to_numeric(prd[c], errors="coerce").fillna(0)

prd["DM_ByProd"]  = (prd["Actual ByProduct Scrap Amount"] +
                     prd["Actual ByProduct Grade B Amount"] +
                     prd["Actual ByProduct Grade C Amount"])
prd["DM_Amount"]  = prd["Actual GI Amount"] - prd["DM_ByProd"]
prd["Conversion"] = prd[[
    "Actual D101 Amount", "Actual D102 Amount",
    "Actual I101 Amount", "Actual I102 Amount", "Actual I103 Amount",
    "Actual I104 Amount", "Actual I105 Amount", "Actual I107 Amount",
    "Actual I108 Amount", "Actual I109 Amount", "Actual I110 Amount",
    "Actual I111 Amount",
]].sum(axis=1)
prd["Total_Cost"]   = prd["DM_Amount"] + prd["Conversion"]
prd["Total_GR_QTY"] = (prd["Actual GR QTY"] +
                        prd["Actual ByProduct Grade B QTY"] +
                        prd["Actual ByProduct Grade C QTY"])

# ─────────────────────────────────────────────
# 2. Exclude anomalous rows + split WIP
# ─────────────────────────────────────────────
excl    = prd[prd["Actual GR Amount"] > GR_AMT_THRESHOLD].copy()
prd_all = prd[prd["Actual GR Amount"] <= GR_AMT_THRESHOLD].copy()

# WIP = GR Amount = 0 AND Total_Cost > 0  (ผลิตแล้ว แต่ยังไม่ GR)
wip       = prd_all[(prd_all["Actual GR Amount"] == 0) & (prd_all["Total_Cost"] > 0)].copy()
completed = prd_all[~((prd_all["Actual GR Amount"] == 0) & (prd_all["Total_Cost"] > 0))].copy()

print(f"  ตัดออก (anomalous)  : {len(excl):,} rows")
print(f"  WIP (GR=0, Cost>0)  : {len(wip):,} rows  | Cost = {wip['Total_Cost'].sum():,.2f}")
print(f"  Completed orders    : {len(completed):,} rows")

# ─────────────────────────────────────────────
# 3. Variance Calculation
# ─────────────────────────────────────────────
print("[2/4] Calculating variance...")

# Overall
total_actual = completed["Total_Cost"].sum()
total_gr     = completed["Actual GR Amount"].sum()
total_var    = total_actual - total_gr
total_qty    = completed["Total_GR_QTY"].sum()

# By Plant
var_plant = completed.groupby("Plant").agg(
    GI_Amount    = ("Actual GI Amount",  "sum"),
    DM_Amount    = ("DM_Amount",         "sum"),
    Conversion   = ("Conversion",        "sum"),
    Total_Cost   = ("Total_Cost",        "sum"),
    GR_Amount    = ("Actual GR Amount",  "sum"),
    Total_GR_QTY = ("Total_GR_QTY",     "sum"),
).reset_index()
var_plant["Variance"]    = var_plant["Total_Cost"] - var_plant["GR_Amount"]
var_plant["Var_per_KG"]  = var_plant["Variance"]   / var_plant["Total_GR_QTY"].replace(0, np.nan)
var_plant["Var_pct"]     = var_plant["Variance"]   / var_plant["GR_Amount"].replace(0, np.nan)
var_plant["Actual_CpKG"] = var_plant["Total_Cost"] / var_plant["Total_GR_QTY"].replace(0, np.nan)
var_plant["GR_CpKG"]     = var_plant["GR_Amount"]  / var_plant["Total_GR_QTY"].replace(0, np.nan)

# By Mat Group
var_mg = completed.groupby("Mat_Group_Name").agg(
    GI_Amount    = ("Actual GI Amount",  "sum"),
    DM_Amount    = ("DM_Amount",         "sum"),
    Conversion   = ("Conversion",        "sum"),
    Total_Cost   = ("Total_Cost",        "sum"),
    GR_Amount    = ("Actual GR Amount",  "sum"),
    Total_GR_QTY = ("Total_GR_QTY",     "sum"),
    Order_Count  = ("Material",          "count"),
).reset_index()
var_mg["Variance"]      = var_mg["Total_Cost"] - var_mg["GR_Amount"]
var_mg["Var_per_KG"]    = var_mg["Variance"]   / var_mg["Total_GR_QTY"].replace(0, np.nan)
var_mg["Var_pct"]       = var_mg["Variance"]   / var_mg["GR_Amount"].replace(0, np.nan)
var_mg["Actual_CpKG"]   = var_mg["Total_Cost"] / var_mg["Total_GR_QTY"].replace(0, np.nan)
var_mg["GR_CpKG"]       = var_mg["GR_Amount"]  / var_mg["Total_GR_QTY"].replace(0, np.nan)
var_mg = var_mg.sort_values("Variance", ascending=False)

# ─────────────────────────────────────────────
# 4. Allocate Variance to FG Inventory
# ─────────────────────────────────────────────
print("[3/4] Allocating variance to FG inventory...")

nrv = pd.read_excel(PATH_NRV, sheet_name="SKU_Detail")
nrv_fg = nrv[nrv["sku_type"] == "FG"].copy()

# Build variance rate map: mat_group → Var_per_KG
var_rate = var_mg.set_index("Mat_Group_Name")["Var_per_KG"].to_dict()

# Match NRV mat_group to production Mat_Group_Name (same naming)
nrv_fg["Var_rate_per_KG"]    = nrv_fg["mat_group"].map(var_rate).fillna(0)
nrv_fg["Variance_Allocated"] = nrv_fg["stock_qty_kg"] * nrv_fg["Var_rate_per_KG"]
nrv_fg["Adj_Cost_per_KG"]    = nrv_fg["cost_price"] + nrv_fg["Var_rate_per_KG"]
nrv_fg["Adj_Stock_Amt"]      = nrv_fg["Adj_Cost_per_KG"] * nrv_fg["stock_qty_kg"]

# Revised NRV comparison
nrv_fg["Adj_Price_Diff"]     = nrv_fg["nrv_price"] - nrv_fg["Adj_Cost_per_KG"]
nrv_fg["Adj_WriteDown_per_KG"] = np.where(
    nrv_fg["Adj_Price_Diff"] < 0, nrv_fg["Adj_Price_Diff"], 0)
nrv_fg["Adj_WriteDown_Amt"]  = nrv_fg["Adj_WriteDown_per_KG"] * nrv_fg["stock_qty_kg"]

# Summary of allocation
total_alloc    = nrv_fg["Variance_Allocated"].sum()
total_orig_wd  = nrv_fg["write_down_amt"].fillna(0).sum()
total_adj_wd   = nrv_fg["Adj_WriteDown_Amt"].sum()

print(f"  Total production variance allocated : {total_alloc:>15,.2f} THB")
print(f"  Original write-down (FG)            : {total_orig_wd:>15,.2f} THB")
print(f"  Revised  write-down (FG)            : {total_adj_wd:>15,.2f} THB")
print(f"  Impact on write-down                : {total_adj_wd - total_orig_wd:>15,.2f} THB")

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
NAVY   = "1F4E79"; GREEN  = "375623"; RED    = "7B241C"; ORANGE = "C55A11"
LGRAY  = "F2F2F2"; LBLUE  = "D9E1F2"; LGREEN = "E2EFDA"
LRED   = "FCE4D6"; LYELLOW= "FFF2CC"; PURPLE = "4A235A"

def _f(h):  return PatternFill("solid", fgColor=h)
def _b(c="D9D9D9"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _fn(color="000000", bold=False, sz=10):
    return Font(color=color, bold=bold, size=sz, name="Calibri")
def _c(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _l(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def _r(): return Alignment(horizontal="right",  vertical="center")

def whdr(ws, row, vals, bg, fg="FFFFFF", col=1):
    for i, v in enumerate(vals, start=col):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = _f(bg); c.font = _fn(fg, bold=True)
        c.alignment = _c(); c.border = _b("999999")
    ws.row_dimensions[row].height = 22

def wrow(ws, row, vals, bg=None, fmts=None, bold=False, col=1):
    for i, v in enumerate(vals, start=col):
        c = ws.cell(row=row, column=i, value=v)
        if bg: c.fill = _f(bg)
        c.font = _fn(bold=bold)
        c.alignment = _l() if isinstance(v, str) else _c()
        c.border = _b()
        if fmts and i in fmts: c.number_format = fmts[i]
    ws.row_dimensions[row].height = 18

def autowid(ws, mn=10, mx=46):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[ltr].width = min(max(w + 2, mn), mx)

def section(ws, row, txt, ncols, bg=LGRAY):
    c = ws.cell(row=row, column=1, value=txt)
    c.font = _fn(NAVY, bold=True, sz=11); c.fill = _f(bg)
    c.alignment = _l(); c.border = _b()
    for i in range(2, ncols + 1):
        cc = ws.cell(row=row, column=i)
        cc.fill = _f(bg); cc.border = _b()
    ws.row_dimensions[row].height = 22

# ─────────────────────────────────────────────
# Build Workbook
# ─────────────────────────────────────────────
print("[4/4] Writing Excel...")
wb = Workbook()

# ══════════════════════════════════════════════
# Sheet 1 – Reconciliation
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Reconciliation"
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:F1")
t = ws1["A1"]
t.value = f"Production Cost Reconciliation & Variance — Q1 2026  (จัดทำ: {REPORT_DATE})"
t.font = _fn("FFFFFF", bold=True, sz=13); t.fill = _f(NAVY)
t.alignment = _c(); t.border = _b(); ws1.row_dimensions[1].height = 30

# ─ Exclusion note
r = 3
section(ws1, r, "Orders ที่ตัดออกจากการคำนวณ (Anomalous GR Amount)", 6)
r += 1
whdr(ws1, r, ["Plant", "Material", "Description", "GR QTY (KG)", "GR Amount (THB)", "เหตุผลที่ตัดออก"], RED)
r += 1
for _, row in excl.iterrows():
    wrow(ws1, r, [
        row["Plant"], row["Material"], row["Description (EN)"],
        row["Actual GR QTY"], row["Actual GR Amount"],
        f"GR Amount สูงผิดปกติ ({row['Actual GR Amount']/1e9:.2f}B THB สำหรับ {row['Actual GR QTY']:,.0f} KG)",
    ], bg=LRED, fmts={4: "#,##0", 5: "#,##0"})
    r += 1

ws1.row_dimensions[r].height = 10; r += 1

# ─ Overall reconciliation
section(ws1, r, "กระทบยอดรวม (หลังตัด Anomalous Orders)", 6)
r += 1
whdr(ws1, r, ["รายการ", "จำนวน (THB)", "% ของ GR Amount", "", "", ""], NAVY)
r += 1

recon_rows = [
    ("Actual GI Amount (วัตถุดิบที่ใช้จริง)",          completed["Actual GI Amount"].sum(),  None, LGRAY),
    ("  หัก ByProduct Recovery",                         -completed["DM_ByProd"].sum(),         None, LGRAY),
    ("= Direct Material (DM) สุทธิ",                    completed["DM_Amount"].sum(),           None, LBLUE),
    ("+ Conversion Cost (D101+D102+I101–I111)",          completed["Conversion"].sum(),          None, LGRAY),
    ("= Total Actual Production Cost  [A]",              total_actual,                           1.0,  LGREEN),
    ("GR Amount (ยอดรับเข้าสต็อกที่ SAP บันทึก) [B]",   total_gr,    total_gr/total_actual,    LGRAY),
    ("VARIANCE  [A − B]",                                total_var,   total_var/total_gr,        LYELLOW),
    ("",                                                 None,                             None,  "FFFFFF"),
    ("WIP Cost (97 orders ยังไม่ GR)",                  wip["Total_Cost"].sum(),           None,  LRED),
    ("Total GR QTY (KG)",                                total_qty,                         None,  LGRAY),
    ("Variance per KG (เฉลี่ย)",                         total_var/total_qty if total_qty>0 else 0, None, LYELLOW),
]
for label, val, pct, bg in recon_rows:
    is_var = "VARIANCE" in label or "Variance" in label
    is_total = label.startswith("= Total")
    vals = [label, val if val is not None else "", pct if pct is not None else ""]
    wrow(ws1, r, vals + ["", "", ""], bg=bg,
         fmts={2: "#,##0.00", 3: "0.0000%"}, bold=(is_total or is_var))
    r += 1

ws1.row_dimensions[r].height = 10; r += 1

# ─ By Plant
section(ws1, r, "Variance แยกตาม Plant", 7)
r += 1
whdr(ws1, r, ["Plant", "Total Cost (THB)", "GR Amount (THB)", "Variance (THB)",
              "Var %", "Var / KG (THB)", "GR QTY (KG)"], NAVY)
r += 1
for _, row in var_plant.iterrows():
    bg = LGREEN if row["Variance"] >= 0 else LRED
    wrow(ws1, r, [
        row["Plant"], row["Total_Cost"], row["GR_Amount"], row["Variance"],
        row["Var_pct"], row["Var_per_KG"], row["Total_GR_QTY"],
    ], bg=bg, fmts={2:"#,##0",3:"#,##0",4:"#,##0.00",5:"0.00%",6:"#,##0.0000",7:"#,##0"}, bold=True)
    r += 1

# Grand total
wrow(ws1, r, ["TOTAL", total_actual, total_gr, total_var,
              total_var/total_gr, total_var/total_qty, total_qty],
     bg=LYELLOW, fmts={2:"#,##0",3:"#,##0",4:"#,##0.00",5:"0.00%",6:"#,##0.0000",7:"#,##0"}, bold=True)

ws1.column_dimensions["A"].width = 52
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 12
ws1.column_dimensions["F"].width = 16
ws1.column_dimensions["G"].width = 16
ws1.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 2 – Variance by Mat Group
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Variance_by_MatGroup")
ws2.sheet_view.showGridLines = False

whdr(ws2, 1, [
    "Mat Group", "Orders", "GR QTY (KG)",
    "DM Amount", "Conversion", "Total Actual Cost",
    "GR Amount (SAP)", "VARIANCE (THB)",
    "Var %", "Actual Cost/KG", "GR Cost/KG", "Variance/KG",
    "ใช้ Rate นี้กระจายต้นทุน",
], NAVY)

for ri, row in enumerate(var_mg.itertuples(index=False), start=2):
    has_rate = pd.notna(row.Var_per_KG)
    bg = (LGREEN if row.Variance >= 0 else LRED) if has_rate else LGRAY
    rate_note = f"{row.Var_per_KG:+.4f} THB/KG" if has_rate else "ไม่มีข้อมูล GR"
    wrow(ws2, ri, [
        row.Mat_Group_Name, row.Order_Count, row.Total_GR_QTY,
        row.DM_Amount, row.Conversion, row.Total_Cost,
        row.GR_Amount, row.Variance,
        row.Var_pct if pd.notna(row.Var_pct) else None,
        row.Actual_CpKG, row.GR_CpKG, row.Var_per_KG,
        rate_note,
    ], bg=bg, fmts={
        3:"#,##0",4:"#,##0",5:"#,##0",6:"#,##0",7:"#,##0",
        8:"#,##0.00",9:"0.00%",10:"#,##0.0000",11:"#,##0.0000",12:"#,##0.0000",
    })

# Grand total
ri += 1
wrow(ws2, ri, [
    "GRAND TOTAL", var_mg["Order_Count"].sum(), total_qty,
    completed["DM_Amount"].sum(), completed["Conversion"].sum(),
    total_actual, total_gr, total_var,
    total_var/total_gr, None, None, total_var/total_qty,
    "",
], bg=LYELLOW, fmts={
    3:"#,##0",4:"#,##0",5:"#,##0",6:"#,##0",7:"#,##0",
    8:"#,##0.00",9:"0.00%",10:"#,##0.0000",11:"#,##0.0000",12:"#,##0.0000",
}, bold=True)

autowid(ws2)
ws2.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 3 – FG Allocation (ราย SKU)
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("FG_Allocation")
ws3.sheet_view.showGridLines = False

whdr(ws3, 1, [
    "SKU", "Description", "Mat Group", "Grade",
    "Stock QTY (KG)", "Stock Amt (THB)",
    "Original Cost/KG",
    "Variance Rate/KG\n(จากผลิต Q1'26)",
    "Variance Allocated\n(THB)",
    "Adjusted Cost/KG",
    "Adjusted Stock Amt",
    "NRV Price",
    "Adj Price Diff\n(NRV−Cost)",
    "Adj WriteDown/KG",
    "Adj WriteDown Amt",
], GREEN)

alloc_out = nrv_fg[[
    "sku", "description", "mat_group", "grade",
    "stock_qty_kg", "stock_amt", "cost_price",
    "Var_rate_per_KG", "Variance_Allocated",
    "Adj_Cost_per_KG", "Adj_Stock_Amt",
    "nrv_price", "Adj_Price_Diff",
    "Adj_WriteDown_per_KG", "Adj_WriteDown_Amt",
]].sort_values(["mat_group", "Variance_Allocated"])

prev_mg = None
for ri, row in enumerate(alloc_out.itertuples(index=False), start=2):
    new_mg = (row.mat_group != prev_mg)
    prev_mg = row.mat_group

    has_wd = pd.notna(row.Adj_WriteDown_Amt) and row.Adj_WriteDown_Amt < 0
    no_rate = row.Var_rate_per_KG == 0
    bg = LRED if has_wd else (LGRAY if no_rate else ("FFFFFF" if ri % 2 == 0 else LGREEN))
    if new_mg: bg = LBLUE

    wrow(ws3, ri, [
        row.sku, row.description, row.mat_group, row.grade,
        row.stock_qty_kg, row.stock_amt, row.cost_price,
        row.Var_rate_per_KG,
        row.Variance_Allocated,
        row.Adj_Cost_per_KG,
        row.Adj_Stock_Amt,
        row.nrv_price if pd.notna(row.nrv_price) else None,
        row.Adj_Price_Diff if pd.notna(row.Adj_Price_Diff) else None,
        row.Adj_WriteDown_per_KG if pd.notna(row.Adj_WriteDown_per_KG) else None,
        row.Adj_WriteDown_Amt if pd.notna(row.Adj_WriteDown_Amt) else None,
    ], bg=bg, fmts={
        5:"#,##0",6:"#,##0",7:"#,##0.0000",8:"+#,##0.0000;-#,##0.0000",
        9:"#,##0.00",10:"#,##0.0000",11:"#,##0",
        12:"#,##0.0000",13:"#,##0.0000",14:"#,##0.0000",15:"#,##0",
    })

# Grand total
ri += 1
wrow(ws3, ri, [
    "GRAND TOTAL", "", "", "",
    nrv_fg["stock_qty_kg"].sum(),
    nrv_fg["stock_amt"].sum(),
    None,
    None,
    nrv_fg["Variance_Allocated"].sum(),
    None,
    nrv_fg["Adj_Stock_Amt"].sum(),
    None,
    None,
    None,
    nrv_fg["Adj_WriteDown_Amt"].sum(),
], bg=LYELLOW, fmts={
    5:"#,##0",6:"#,##0",9:"#,##0.00",11:"#,##0",15:"#,##0",
}, bold=True)

autowid(ws3)
ws3.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 4 – Revised NRV (Summary by Mat Group)
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("Revised_NRV")
ws4.sheet_view.showGridLines = False

# Summary by mat_group
rev_mg = nrv_fg.groupby("mat_group").agg(
    SKU_Count         = ("sku",                   "count"),
    Stock_QTY         = ("stock_qty_kg",           "sum"),
    Stock_Amt         = ("stock_amt",              "sum"),
    Orig_WriteDown    = ("write_down_amt",          lambda x: x.fillna(0).sum()),
    Variance_Alloc    = ("Variance_Allocated",      "sum"),
    Adj_Stock_Amt     = ("Adj_Stock_Amt",           "sum"),
    Adj_WriteDown     = ("Adj_WriteDown_Amt",       "sum"),
    SKU_WithWriteDown = ("Adj_WriteDown_Amt",       lambda x: (x < 0).sum()),
).reset_index()
rev_mg["Orig_CpKG"]     = rev_mg["Stock_Amt"]     / rev_mg["Stock_QTY"].replace(0, np.nan)
rev_mg["Adj_CpKG"]      = rev_mg["Adj_Stock_Amt"] / rev_mg["Stock_QTY"].replace(0, np.nan)
rev_mg["WD_Impact"]     = rev_mg["Adj_WriteDown"]  - rev_mg["Orig_WriteDown"]
rev_mg = rev_mg.sort_values("Adj_WriteDown")

whdr(ws4, 1, [
    "Mat Group", "SKU Count",
    "Stock QTY (KG)", "Stock Amt (THB)",
    "Orig Write-Down", "Variance Allocated",
    "Adj Stock Amt", "Adj Write-Down",
    "Impact on W/D\n(Adj − Orig)",
    "Orig Cost/KG", "Adj Cost/KG",
    "SKUs with W/D",
], ORANGE)

for ri, row in enumerate(rev_mg.itertuples(index=False), start=2):
    impact = row.WD_Impact
    bg = LRED if impact < -10000 else (LGREEN if impact > 10000 else LGRAY)
    wrow(ws4, ri, [
        row.mat_group, row.SKU_Count,
        row.Stock_QTY, row.Stock_Amt,
        row.Orig_WriteDown, row.Variance_Alloc,
        row.Adj_Stock_Amt, row.Adj_WriteDown,
        row.WD_Impact,
        row.Orig_CpKG, row.Adj_CpKG,
        int(row.SKU_WithWriteDown),
    ], bg=bg, fmts={
        3:"#,##0",4:"#,##0",5:"#,##0.00",6:"#,##0.00",
        7:"#,##0",8:"#,##0.00",9:"#,##0.00",10:"#,##0.0000",11:"#,##0.0000",
    })

# Grand total
ri += 1
wrow(ws4, ri, [
    "GRAND TOTAL", int(rev_mg["SKU_Count"].sum()),
    rev_mg["Stock_QTY"].sum(), rev_mg["Stock_Amt"].sum(),
    total_orig_wd, total_alloc,
    rev_mg["Adj_Stock_Amt"].sum(), total_adj_wd,
    total_adj_wd - total_orig_wd,
    None, None,
    int(rev_mg["SKU_WithWriteDown"].sum()),
], bg=LYELLOW, fmts={
    3:"#,##0",4:"#,##0",5:"#,##0.00",6:"#,##0.00",
    7:"#,##0",8:"#,##0.00",9:"#,##0.00",
}, bold=True)

# Summary box
ri += 2
for label, val, fmt, bg in [
    ("Original Write-Down (FG, ก่อนกระจาย Variance)", total_orig_wd, "#,##0.00", LGRAY),
    ("Variance ที่กระจายเข้า FG Inventory",            total_alloc,  "+#,##0.00;-#,##0.00", LBLUE),
    ("Revised Write-Down (FG, หลังกระจาย Variance)",   total_adj_wd, "#,##0.00", LRED if total_adj_wd < total_orig_wd else LGREEN),
    ("Impact (เพิ่ม/ลด Write-Down)",                   total_adj_wd - total_orig_wd, "+#,##0.00;-#,##0.00", LYELLOW),
]:
    c1 = ws4.cell(row=ri, column=1, value=label)
    c1.font = _fn(bold=True); c1.fill = _f(bg); c1.alignment = _l(); c1.border = _b()
    c2 = ws4.cell(row=ri, column=2, value=val)
    c2.font = _fn(bold=True); c2.fill = _f(bg); c2.number_format = fmt
    c2.alignment = _r(); c2.border = _b()
    ws4.row_dimensions[ri].height = 22; ri += 1

autowid(ws4)
ws4.freeze_panes = "A2"

# ══════════════════════════════════════════════
# Sheet 5 – WIP Cost
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("WIP_Cost")
ws5.sheet_view.showGridLines = False

whdr(ws5, 1, [
    "Plant", "Material", "Description", "Mat Group",
    "GI QTY (KG)", "DM Amount", "Conversion", "Total WIP Cost",
    "หมายเหตุ",
], PURPLE)

wip_out = wip[["Plant", "Material", "Description (EN)", "Mat_Group_Name",
               "Actual GI QTY", "DM_Amount", "Conversion", "Total_Cost"]].copy()
wip_out = wip_out.rename(columns={"Description (EN)": "Desc", "Actual GI QTY": "GI_QTY"})
wip_out["Note"] = "WIP — ยังไม่มี GR (ยังไม่รับเข้าสต็อก)"
wip_out = wip_out.sort_values(["Plant", "Total_Cost"], ascending=[True, False])

for ri, row in enumerate(wip_out.itertuples(index=False), start=2):
    wrow(ws5, ri, [
        row.Plant, row.Material, row.Desc,
        row.Mat_Group_Name, row.GI_QTY,
        row.DM_Amount, row.Conversion, row.Total_Cost,
        row.Note,
    ], bg=LGRAY if ri % 2 == 0 else "FFFFFF",
       fmts={5:"#,##0",6:"#,##0",7:"#,##0",8:"#,##0"})

# WIP summary by mat group
ri += 2
section(ws5, ri, "WIP สรุปตาม Mat Group", 9)
ri += 1
whdr(ws5, ri, ["Mat Group", "Orders", "Total WIP Cost (THB)", "% ของ Total WIP"], PURPLE)
ri += 1
wip_mg = wip.groupby("Mat_Group_Name")["Total_Cost"].agg(["count","sum"]).reset_index()
wip_mg.columns = ["Mat_Group_Name","Count","WIP_Cost"]
wip_total_cost = wip_mg["WIP_Cost"].sum()
for _, row in wip_mg.sort_values("WIP_Cost", ascending=False).iterrows():
    wrow(ws5, ri, [row["Mat_Group_Name"], row["Count"], row["WIP_Cost"],
                   row["WIP_Cost"]/wip_total_cost],
         fmts={3:"#,##0",4:"0.00%"})
    ri += 1
wrow(ws5, ri, ["TOTAL", wip_mg["Count"].sum(), wip_total_cost, 1.0],
     bg=LYELLOW, fmts={3:"#,##0",4:"0.00%"}, bold=True)

autowid(ws5)
ws5.freeze_panes = "A2"

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
wb.save(OUTPUT)
print(f"\n  Output: {OUTPUT}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
print()
print("=" * 60)
print("  PRODUCTION VARIANCE ALLOCATION SUMMARY — Q1 2026")
print("=" * 60)
print(f"  Total Actual Production Cost : {total_actual:>20,.2f} THB")
print(f"  Total GR Amount (SAP)        : {total_gr:>20,.2f} THB")
print(f"  Net Variance                 : {total_var:>20,.2f} THB")
print(f"  Variance / KG (avg)          : {total_var/total_qty:>20,.4f} THB/KG")
print(f"  WIP Cost (not yet GR)        : {wip['Total_Cost'].sum():>20,.2f} THB")
print()
print(f"  FG SKUs with allocation      : {(nrv_fg['Var_rate_per_KG']!=0).sum():,} / {len(nrv_fg):,}")
print(f"  Total Variance Allocated     : {total_alloc:>20,.2f} THB")
print(f"  Original FG Write-Down       : {total_orig_wd:>20,.2f} THB")
print(f"  Revised  FG Write-Down       : {total_adj_wd:>20,.2f} THB")
print(f"  Impact on Write-Down         : {total_adj_wd-total_orig_wd:>20,.2f} THB")
print("=" * 60)
