"""
production_cost_nrv_reference.py
=================================
สร้างเอกสาร Reference สำหรับการคำนวณ NRV ส่งออดิท
อธิบายที่มาที่ไปของการคำนวณ Production Cost 2025

Output: 04_Reports/Production_Cost_NRV_Reference_2025.xlsx
  Sheet 1: Document_Control   — ปกเอกสาร / วัตถุประสงค์
  Sheet 2: Calculation_Logic  — สูตรคำนวณ DM และ Conversion ทีละขั้น
  Sheet 3: Cost_Code_Ref      — นิยามและยอดรวม D101-I111
  Sheet 4: ByProduct_Treatment— การปฏิบัติ Grade B / C / Scrap
  Sheet 5: Unit_Cost_NRV      — Conversion per KG แยก Plant × MatGroup สำหรับ NRV
  Sheet 6: Data_Traceability  — Source files, row count, reconciliation ยอด
"""

import os
import sys
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 0. Config
# ─────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
INPUT_FILE  = os.path.join(BASE_DIR, "02_Silver_Cleaned", "master_production_2025.parquet")
OUTPUT_DIR  = os.path.join(BASE_DIR, "04_Reports")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Production_Cost_NRV_Reference_2025.xlsx")
os.makedirs(OUTPUT_DIR, exist_ok=True)

REPORT_DATE  = date.today().strftime("%d %B %Y")
FISCAL_YEAR  = "2025"
PREPARED_BY  = "Finance – Cost Accounting"

# ─────────────────────────────────────────────
# 1. Load & Prepare Data
# ─────────────────────────────────────────────
print("[1/6] Loading data...")
df = pd.read_parquet(INPUT_FILE, engine="pyarrow")
cols = list(df.columns)

rename_map = {
    cols[6]:  "Mat_Group_Name",
    cols[7]:  "Mat_Type_Code",
    cols[10]: "Mat_Group_1_Name",
    cols[11]: "Mat_Group_2_Name",
    cols[12]: "Mat_Group_3_Name",
    cols[13]: "Mat_Group_4_Name",
}
df.rename(columns=rename_map, inplace=True)
df["Month"] = df["Source_File"].str.extract(r"_(\d{2})\.XLSX$").astype(float).astype("Int64")
df["Plant"] = df["Plant"].astype(str).str.strip()

# Numeric cast
num_cols = [
    "Actual GI Amount",
    "Actual ByProduct Scrap Amount",
    "Actual ByProduct Grade B Amount",
    "Actual ByProduct Grade C Amount",
    "Actual GR QTY", "Actual GI QTY",
    "Actual ByProduct Grade B QTY",
    "Actual ByProduct Grade C QTY",
    "Actual D101 Amount", "Actual D102 Amount",
    "Actual I101 Amount", "Actual I102 Amount", "Actual I103 Amount",
    "Actual I104 Amount", "Actual I105 Amount", "Actual I107 Amount",
    "Actual I108 Amount", "Actual I109 Amount", "Actual I110 Amount",
    "Actual I111 Amount",
]
for c in num_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# Cost components
df["DM_GI_Amount"]     = df["Actual GI Amount"]
df["DM_ByProd_Scrap"]  = df["Actual ByProduct Scrap Amount"]
df["DM_ByProd_GradeB"] = df["Actual ByProduct Grade B Amount"]
df["DM_ByProd_GradeC"] = df["Actual ByProduct Grade C Amount"]
df["DM_ByProd_Total"]  = df["DM_ByProd_Scrap"] + df["DM_ByProd_GradeB"] + df["DM_ByProd_GradeC"]
df["DM_Amount"]        = df["DM_GI_Amount"] - df["DM_ByProd_Total"]

direct_codes   = ["D101", "D102"]
indirect_codes = ["I101","I102","I103","I104","I105","I107","I108","I109","I110","I111"]
all_codes      = direct_codes + indirect_codes

df["Direct_Amount"]     = df[[f"Actual {c} Amount" for c in direct_codes]].sum(axis=1)
df["Indirect_Amount"]   = df[[f"Actual {c} Amount" for c in indirect_codes]].sum(axis=1)
df["Conversion_Amount"] = df["Direct_Amount"] + df["Indirect_Amount"]
df["Total_Cost"]        = df["DM_Amount"] + df["Conversion_Amount"]

df["GR_QTY"]       = df["Actual GR QTY"]
df["GR_QTY_GradeB"]= df["Actual ByProduct Grade B QTY"]
df["GR_QTY_GradeC"]= df["Actual ByProduct Grade C QTY"]
df["GI_QTY"]       = df["Actual GI QTY"]
df["Total_GR_QTY"] = df["GR_QTY"] + df["GR_QTY_GradeB"] + df["GR_QTY_GradeC"]

# Summary scalars
total_gi      = df["DM_GI_Amount"].sum()
total_scrap   = df["DM_ByProd_Scrap"].sum()
total_gradeB  = df["DM_ByProd_GradeB"].sum()
total_gradeC  = df["DM_ByProd_GradeC"].sum()
total_byprod  = df["DM_ByProd_Total"].sum()
total_dm      = df["DM_Amount"].sum()
total_direct  = df["Direct_Amount"].sum()
total_indirect= df["Indirect_Amount"].sum()
total_conv    = df["Conversion_Amount"].sum()
total_cost    = df["Total_Cost"].sum()
total_qty     = df["Total_GR_QTY"].sum()
avg_conv_kg   = total_conv / total_qty if total_qty > 0 else 0
avg_cost_kg   = total_cost / total_qty if total_qty > 0 else 0

code_totals = {c: df[f"Actual {c} Amount"].sum() for c in all_codes}

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
NAVY   = "1F4E79"
GREEN  = "375623"
RED    = "7B241C"
PURPLE = "4A235A"
ORANGE = "C55A11"
GRAY   = "404040"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREEN= "E2EFDA"
LIGHT_RED  = "FCE4D6"
LIGHT_GRAY = "F2F2F2"
YELLOW     = "FFF2CC"

def hdr_font(color="FFFFFF", sz=10, bold=True):
    return Font(color=color, bold=bold, size=sz, name="Calibri")

def body_font(sz=10, bold=False, color="000000"):
    return Font(color=color, bold=bold, size=sz, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(color="D9D9D9"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_header_row(ws, row, values, bg, fg="FFFFFF", sz=10, bold=True, col_start=1):
    for i, v in enumerate(values, start=col_start):
        c = ws.cell(row=row, column=i, value=v)
        c.fill      = fill(bg)
        c.font      = hdr_font(fg, sz, bold)
        c.alignment = center()
        c.border    = border("AAAAAA")

def write_data_row(ws, row, values, bg=None, fmt_map=None, col_start=1, bold=False):
    for i, v in enumerate(values, start=col_start):
        c = ws.cell(row=row, column=i, value=v)
        if bg:
            c.fill = fill(bg)
        c.font      = body_font(bold=bold)
        c.alignment = left() if isinstance(v, str) else center()
        c.border    = border()
        if fmt_map and i in fmt_map:
            c.number_format = fmt_map[i]

def auto_width(ws, min_w=12, max_w=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[col_letter].width = min(max(w + 2, min_w), max_w)

def section_label(ws, row, text, span_cols, bg=LIGHT_GRAY, col_start=1):
    c = ws.cell(row=row, column=col_start, value=text)
    c.font      = Font(bold=True, size=10, name="Calibri", color=GRAY)
    c.fill      = fill(bg)
    c.alignment = left()
    c.border    = border()
    for i in range(col_start + 1, col_start + span_cols):
        cc = ws.cell(row=row, column=i)
        cc.fill   = fill(bg)
        cc.border = border()

# ─────────────────────────────────────────────
# Sheet 1 – Document Control
# ─────────────────────────────────────────────
print("[2/6] Sheet 1: Document Control...")

wb = Workbook()
ws1 = wb.active
ws1.title = "Document_Control"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 55

# Title block
title_rows = [
    ("เอกสาร",         "Production Cost – NRV Calculation Reference"),
    ("ปีงบประมาณ",      f"FY {FISCAL_YEAR}"),
    ("วัตถุประสงค์",    "ใช้เป็น Reference ประกอบการคำนวณ NRV สำหรับการสอบบัญชี (Audit)"),
    ("ขอบเขต",         "การผลิต Plant 1100 / 1200 / 1300 ทุกเดือน (ม.ค. – ธ.ค. 2025)"),
    ("แหล่งข้อมูล",     "SAP CO Production Order Actual Cost — extract รายเดือน"),
    ("จัดทำโดย",       PREPARED_BY),
    ("วันที่จัดทำ",    REPORT_DATE),
    ("สถานะ",          "Draft for Audit Review"),
]

ws1.row_dimensions[1].height = 8
for ri, (label, value) in enumerate(title_rows, start=2):
    ws1.row_dimensions[ri].height = 24
    la = ws1.cell(row=ri, column=1, value=label)
    la.font      = hdr_font(color="FFFFFF", sz=11)
    la.fill      = fill(NAVY)
    la.alignment = left()
    la.border    = border("FFFFFF")

    va = ws1.cell(row=ri, column=2, value=value)
    va.font      = body_font(sz=11, bold=(ri == 2))
    va.fill      = fill(LIGHT_BLUE if ri % 2 == 0 else "FFFFFF")
    va.alignment = left()
    va.border    = border()

ws1.row_dimensions[len(title_rows) + 3].height = 8

# Summary totals table
r = len(title_rows) + 4
write_header_row(ws1, r, ["รายการ", "จำนวน (THB)", "% ของ Total Cost"], NAVY)
r += 1

summary_items = [
    ("GI Amount (วัตถุดิบที่เบิกเข้าสายการผลิต)",    total_gi,     total_gi / total_cost),
    ("  หัก ByProduct – Scrap",                        -total_scrap,  -total_scrap / total_cost),
    ("  หัก ByProduct – Grade B",                      -total_gradeB, -total_gradeB / total_cost),
    ("  หัก ByProduct – Grade C",                      -total_gradeC, -total_gradeC / total_cost),
    ("= Direct Material (DM)",                          total_dm,     total_dm / total_cost),
    ("Direct Conversion (D101 + D102)",                 total_direct,  total_direct / total_cost),
    ("Indirect Conversion (I101 – I111)",               total_indirect,total_indirect / total_cost),
    ("= Conversion Cost",                               total_conv,   total_conv / total_cost),
    ("Total Production Cost",                           total_cost,   1.0),
]

for label, amt, pct in summary_items:
    is_total = label.startswith("=") or label.startswith("Total")
    bg = LIGHT_GREEN if label.startswith("=") else (LIGHT_BLUE if label.startswith("Total") else "FFFFFF")
    row_vals = [label, amt, pct]
    write_data_row(ws1, r, row_vals, bg=bg,
                   fmt_map={2: '#,##0', 3: '0.00%'}, bold=is_total)
    r += 1

ws1.row_dimensions[r].height = 8
r += 1

# Unit cost summary
write_header_row(ws1, r, ["Unit Cost Reference (THB / KG)", "ค่า"], PURPLE)
r += 1
unit_items = [
    ("Total GR Output QTY (KG)",         total_qty),
    ("Average Conversion Cost / KG",      avg_conv_kg),
    ("Average Total Cost / KG",           avg_cost_kg),
]
for label, val in unit_items:
    c1 = ws1.cell(row=r, column=1, value=label)
    c1.font = body_font(); c1.alignment = left(); c1.border = border()
    c2 = ws1.cell(row=r, column=2, value=val)
    c2.font = body_font(bold=True); c2.alignment = center()
    c2.number_format = '#,##0.00'; c2.border = border()
    r += 1

auto_width(ws1, min_w=28)

# ─────────────────────────────────────────────
# Sheet 2 – Calculation Logic
# ─────────────────────────────────────────────
print("[3/6] Sheet 2: Calculation Logic...")
ws2 = wb.create_sheet("Calculation_Logic")
ws2.sheet_view.showGridLines = False

steps = [
    # (step_no, component, formula, description, note)
    (
        "ขั้นที่ 1",
        "Direct Material (DM)",
        "DM = GI_Amount − (ByProd_Scrap + ByProd_GradeB + ByProd_GradeC)",
        "วัตถุดิบสุทธิที่ถูกใช้ในการผลิต หักด้วยมูลค่าที่กู้คืนจาก ByProduct",
        "GI Amount = Actual Goods Issue Amount จาก SAP CO\n"
        "ByProduct ที่หักออก = Scrap + Grade B + Grade C (ไม่รวม Waste ที่ไม่มีมูลค่า)",
    ),
    (
        "ขั้นที่ 2",
        "Direct Conversion",
        "Direct = D101_Amount + D102_Amount",
        "ต้นทุนแปรสภาพทางตรง — ค่าแรงและค่าเครื่องจักรที่ trace ได้โดยตรงกับ Production Order",
        "D101 = Direct Labor\nD102 = Direct Machine / Power",
    ),
    (
        "ขั้นที่ 3",
        "Indirect Conversion",
        "Indirect = I101+I102+I103+I104+I105+I107+I108+I109+I110+I111",
        "ต้นทุนแปรสภาพทางอ้อม — overhead ที่ปันส่วนลงมาที่ Production Order",
        "I106 ไม่ถูกนำมาคำนวณ (ยอดเป็น 0 ตลอดปี)\n"
        "ดูรายละเอียดแต่ละ I-code ใน Sheet Cost_Code_Ref",
    ),
    (
        "ขั้นที่ 4",
        "Conversion Cost",
        "Conversion = Direct + Indirect",
        "ต้นทุนแปรสภาพรวม ใช้เป็นตัวหารในการคำนวณ Conversion per KG สำหรับ NRV",
        "",
    ),
    (
        "ขั้นที่ 5",
        "Total Production Cost",
        "Total_Cost = DM + Conversion",
        "ต้นทุนการผลิตรวมต่อ Production Order",
        "ใช้ประกอบการตรวจสอบมูลค่าสินค้าคงเหลือ (Inventory Valuation)",
    ),
    (
        "ขั้นที่ 6",
        "Output QTY (denominator)",
        "Total_GR_QTY = GR_QTY_GradeA + GR_QTY_GradeB + GR_QTY_GradeC",
        "ปริมาณ output ที่รับเข้าคลัง รวม Grade A + B + C (ไม่รวม Scrap)",
        "Scrap ไม่ถูกรวมใน denominator เนื่องจากถือเป็น waste ไม่มีมูลค่าเชิงพาณิชย์",
    ),
    (
        "ขั้นที่ 7",
        "Conversion per KG",
        "Conv_per_KG = Conversion / Total_GR_QTY",
        "ต้นทุนแปรสภาพต่อหน่วย ใช้เป็นอัตราอ้างอิงสำหรับการประมาณ NRV ของ WIP",
        "แยกคำนวณต่อ Plant และ Mat_Group — ดู Sheet Unit_Cost_NRV",
    ),
    (
        "ขั้นที่ 8",
        "NRV Reference (แนวทาง)",
        "NRV = Est. Selling Price − Est. Conv. Cost to Complete − Selling Cost",
        "เปรียบเทียบ Cost กับ NRV เพื่อพิจารณา write-down ตาม IAS 2 / TAS 2",
        "Conversion per KG (ขั้นที่ 7) ใช้ประมาณ 'Cost to complete' ของ WIP\n"
        "Selling Price อ้างอิงจาก Price List หรือ Recent Sale",
    ),
]

write_header_row(ws2, 1, ["ขั้นตอน", "รายการ", "สูตร / Formula", "คำอธิบาย", "หมายเหตุ"], NAVY)
ws2.row_dimensions[1].height = 22

col_widths = [12, 24, 52, 52, 52]
for i, w in enumerate(col_widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for ri, (step, comp, formula, desc, note) in enumerate(steps, start=2):
    bg = LIGHT_BLUE if ri % 2 == 0 else "FFFFFF"
    is_key = comp in ("Conversion per KG", "NRV Reference (แนวทาง)", "Total Production Cost")
    if is_key:
        bg = LIGHT_GREEN
    ws2.row_dimensions[ri].height = 48
    for ci, val in enumerate([step, comp, formula, desc, note], start=1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font      = body_font(bold=is_key)
        c.fill      = fill(bg)
        c.alignment = left()
        c.border    = border()

ws2.freeze_panes = "A2"

# ─────────────────────────────────────────────
# Sheet 3 – Cost Code Reference
# ─────────────────────────────────────────────
print("[4/6] Sheet 3: Cost Code Reference...")
ws3 = wb.create_sheet("Cost_Code_Ref")
ws3.sheet_view.showGridLines = False

CODE_META = {
    "D101": ("Direct",   "ค่าแรงงานทางตรง (Direct Labor)",           "เวลาเครื่องจักร × อัตราค่าแรงที่ตั้งไว้ใน Work Center"),
    "D102": ("Direct",   "ค่าเครื่องจักรทางตรง (Direct Machine)",     "Machine Hour × Machine Rate จาก Work Center"),
    "I101": ("Indirect", "ค่าแรงงานทางอ้อม (Indirect Labor)",         "ปันส่วนจาก Cost Center Overhead"),
    "I102": ("Indirect", "ค่าเสื่อมราคา (Depreciation)",               "ค่าเสื่อมราคาเครื่องจักร / โรงงาน ปันส่วนตาม Machine Hour"),
    "I103": ("Indirect", "ค่าซ่อมบำรุง (Repair & Maintenance)",        "ค่าใช้จ่ายซ่อมบำรุงเครื่องจักรและอาคาร"),
    "I104": ("Indirect", "สาธารณูปโภค (Utility)",                       "ไฟฟ้า / น้ำ / ก๊าซ"),
    "I105": ("Indirect", "วัสดุสิ้นเปลือง (Consumable)",               "น้ำมันหล่อลื่น, อุปกรณ์ขนาดเล็ก, สารเคมี"),
    "I106": ("Indirect", "ไม่ใช้งาน (Always Zero)",                    "I106 ไม่มียอด — ไม่ถูกนำมาคำนวณ"),
    "I107": ("Indirect", "ต้นทุนทางอ้อมอื่น (Other Indirect Cost)",    "รายการ Overhead ที่ไม่จัดอยู่ใน I101-I105"),
    "I108": ("Indirect", "Overhead Allocation 1",                       "ปันส่วน Overhead จาก Cost Center ระดับที่ 1"),
    "I109": ("Indirect", "Overhead Allocation 2",                       "ปันส่วน Overhead จาก Cost Center ระดับที่ 2"),
    "I110": ("Indirect", "Overhead Allocation 3",                       "ปันส่วน Overhead จาก Cost Center ระดับที่ 3"),
    "I111": ("Indirect", "Overhead Allocation 4",                       "ปันส่วน Overhead จาก Cost Center ระดับที่ 4"),
}

write_header_row(ws3, 1,
    ["Cost Code", "ประเภท", "ชื่อรายการ", "วิธีการคำนวณ / ที่มา",
     "ยอดรวม FY2025 (THB)", "% ของ Conversion", "นำมาคำนวณ"],
    NAVY)

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 38
ws3.column_dimensions["D"].width = 52
ws3.column_dimensions["E"].width = 24
ws3.column_dimensions["F"].width = 18
ws3.column_dimensions["G"].width = 16

all_code_list = ["D101","D102","I101","I102","I103","I104","I105","I106",
                 "I107","I108","I109","I110","I111"]

prev_type = None
for ri, code in enumerate(all_code_list, start=2):
    cat, name, method = CODE_META[code]
    used = "YES" if code != "I106" else "NO"
    amount = code_totals.get(code, 0)
    pct    = amount / total_conv if total_conv > 0 and code != "I106" else 0

    is_direct = (cat == "Direct")
    if cat != prev_type:
        bg = LIGHT_BLUE if is_direct else LIGHT_GREEN
    else:
        bg = "FFFFFF" if ri % 2 == 0 else (LIGHT_BLUE if is_direct else LIGHT_GREEN)
    prev_type = cat

    if code == "I106":
        bg = LIGHT_RED

    row_vals = [code, cat, name, method, amount if code != "I106" else 0, pct, used]
    write_data_row(ws3, ri, row_vals, bg=bg,
                   fmt_map={5: '#,##0', 6: '0.00%'})
    ws3.row_dimensions[ri].height = 32

# Grand total row
ri += 1
ws3.row_dimensions[ri].height = 22
write_data_row(ws3, ri,
    ["", "TOTAL", "Conversion Cost รวม", "", total_conv, 1.0, ""],
    bg=YELLOW, fmt_map={5: '#,##0', 6: '0.00%'}, bold=True)

ws3.freeze_panes = "A2"

# ─────────────────────────────────────────────
# Sheet 4 – ByProduct Treatment
# ─────────────────────────────────────────────
print("[4b/6] Sheet 4: ByProduct Treatment...")
ws4 = wb.create_sheet("ByProduct_Treatment")
ws4.sheet_view.showGridLines = False

write_header_row(ws4, 1,
    ["ประเภท ByProduct", "การปฏิบัติทางบัญชี", "ผลต่อ DM Cost",
     "ยอดรวม FY2025 (THB)", "ยอดรวม QTY (KG)", "% ของ GI Amount"],
    RED)

# ByProduct QTY
scrap_qty  = df["Actual ByProduct Scrap QTY"].apply(pd.to_numeric, errors="coerce").fillna(0).sum() if "Actual ByProduct Scrap QTY" in df.columns else 0
gradeB_qty = df["GR_QTY_GradeB"].sum()
gradeC_qty = df["GR_QTY_GradeC"].sum()

bp_rows = [
    ("Scrap (เศษวัสดุ)",
     "หักออกจาก GI Amount\nไม่รวมใน Output QTY denominator (ไม่มีมูลค่าเชิงพาณิชย์)",
     "ลด DM Cost (ลบ Scrap Amount ออก)",
     total_scrap, scrap_qty, total_scrap / total_gi),
    ("Grade B (สินค้า Grade B)",
     "หักออกจาก GI Amount\nรวมใน Output QTY denominator (GR_QTY_GradeB)\nมีราคาขายต่ำกว่า Grade A",
     "ลด DM Cost\nแต่ยังนับเป็น Output ทำให้ Cost/KG ถูกกระจายออกไป",
     total_gradeB, gradeB_qty, total_gradeB / total_gi),
    ("Grade C (สินค้า Grade C)",
     "หักออกจาก GI Amount\nรวมใน Output QTY denominator (GR_QTY_GradeC)\nมีราคาขายต่ำกว่า Grade B",
     "ลด DM Cost\nแต่ยังนับเป็น Output",
     total_gradeC, gradeC_qty, total_gradeC / total_gi),
    ("รวม ByProduct ทั้งหมด",
     "—",
     "—",
     total_byprod, scrap_qty + gradeB_qty + gradeC_qty, total_byprod / total_gi),
]

ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 52
ws4.column_dimensions["C"].width = 44
ws4.column_dimensions["D"].width = 24
ws4.column_dimensions["E"].width = 20
ws4.column_dimensions["F"].width = 18

for ri, (typ, treatment, impact, amt, qty, pct) in enumerate(bp_rows, start=2):
    is_total = typ.startswith("รวม")
    bg = LIGHT_RED if is_total else ("FFFFFF" if ri % 2 == 0 else LIGHT_GRAY)
    ws4.row_dimensions[ri].height = 52
    for ci, val in enumerate([typ, treatment, impact, amt, qty, pct], start=1):
        c = ws4.cell(row=ri, column=ci, value=val)
        c.font      = body_font(bold=is_total)
        c.fill      = fill(bg)
        c.alignment = left()
        c.border    = border()
        if ci == 4:
            c.number_format = '#,##0'
        elif ci == 5:
            c.number_format = '#,##0.00'
        elif ci == 6:
            c.number_format = '0.00%'

# NRV note block
ri = len(bp_rows) + 3
ws4.row_dimensions[ri].height = 18
c = ws4.cell(row=ri, column=1,
    value="แนวทางใช้ข้อมูล ByProduct ใน NRV Calculation")
c.font = Font(bold=True, size=11, name="Calibri", color=RED)
c.alignment = left()

nrv_notes = [
    "1. สินค้า Grade A: NRV = Estimated Selling Price − Selling Cost",
    "2. สินค้า Grade B / C: NRV ต้องคำนวณแยก เนื่องจาก Selling Price ต่ำกว่า Grade A",
    "   → ควรใช้ราคาขายจริงของ Grade B / C แต่ละ Material ไม่ใช่ราคา Grade A",
    "3. Scrap: ไม่นำมาคำนวณ NRV (มูลค่าเท่ากับ 0 หรือราคาขายเศษ)",
    "4. Cost ที่ใช้เปรียบเทียบ = Total_Cost / Total_GR_QTY (รวม Grade A+B+C)",
]
for i, note in enumerate(nrv_notes, start=1):
    ri += 1
    ws4.row_dimensions[ri].height = 20
    c = ws4.cell(row=ri, column=1, value=note)
    c.font      = body_font(sz=10)
    c.alignment = left()

ws4.freeze_panes = "A2"

# ─────────────────────────────────────────────
# Sheet 5 – Unit Cost NRV (Conversion/KG by Plant × MatGroup)
# ─────────────────────────────────────────────
print("[5/6] Sheet 5: Unit Cost NRV...")
ws5 = wb.create_sheet("Unit_Cost_NRV")
ws5.sheet_view.showGridLines = False

uc = df.groupby(["Plant", "Mat_Group_Name", "Mat_Group_1_Name"], dropna=False).agg(
    Total_GR_QTY    = ("Total_GR_QTY",      "sum"),
    DM_Amount       = ("DM_Amount",          "sum"),
    Conversion      = ("Conversion_Amount",  "sum"),
    Total_Cost      = ("Total_Cost",         "sum"),
    Direct          = ("Direct_Amount",      "sum"),
    Indirect        = ("Indirect_Amount",    "sum"),
).reset_index()

uc = uc[uc["Total_GR_QTY"] > 0].copy()
uc["DM_per_KG"]         = uc["DM_Amount"]   / uc["Total_GR_QTY"]
uc["Conv_per_KG"]       = uc["Conversion"]  / uc["Total_GR_QTY"]
uc["Direct_per_KG"]     = uc["Direct"]      / uc["Total_GR_QTY"]
uc["Indirect_per_KG"]   = uc["Indirect"]    / uc["Total_GR_QTY"]
uc["Total_Cost_per_KG"] = uc["Total_Cost"]  / uc["Total_GR_QTY"]
uc = uc.sort_values(["Plant", "Conv_per_KG"], ascending=[True, False])

write_header_row(ws5, 1, [
    "Plant", "Mat Group", "Mat Group 1",
    "Output QTY (KG)", "DM / KG (THB)", "Conv / KG (THB)",
    "Direct / KG", "Indirect / KG", "Total Cost / KG (THB)",
    "หมายเหตุ NRV"
], GREEN)

ws5.column_dimensions["A"].width = 8
ws5.column_dimensions["B"].width = 28
ws5.column_dimensions["C"].width = 24
ws5.column_dimensions["D"].width = 18
ws5.column_dimensions["E"].width = 16
ws5.column_dimensions["F"].width = 18
ws5.column_dimensions["G"].width = 16
ws5.column_dimensions["H"].width = 16
ws5.column_dimensions["I"].width = 20
ws5.column_dimensions["J"].width = 38

prev_plant = None
for ri, row in enumerate(uc.itertuples(index=False), start=2):
    new_plant = (row.Plant != prev_plant)
    bg = LIGHT_BLUE if new_plant else ("FFFFFF" if ri % 2 == 0 else LIGHT_GRAY)
    prev_plant = row.Plant
    ws5.row_dimensions[ri].height = 20

    nrv_note = (
        f"Conv/KG = {row.Conv_per_KG:,.2f} THB\n"
        f"ใช้เป็น 'Cost to Complete' กรณี WIP ยังไม่สำเร็จรูป"
    )
    vals = [
        row.Plant, row.Mat_Group_Name, row.Mat_Group_1_Name,
        row.Total_GR_QTY, row.DM_per_KG, row.Conv_per_KG,
        row.Direct_per_KG, row.Indirect_per_KG, row.Total_Cost_per_KG,
        nrv_note,
    ]
    fmt_map = {4: '#,##0', 5: '#,##0.00', 6: '#,##0.00',
               7: '#,##0.00', 8: '#,##0.00', 9: '#,##0.00'}
    write_data_row(ws5, ri, vals, bg=bg, fmt_map=fmt_map, bold=new_plant)

ws5.freeze_panes = "D2"

# ─────────────────────────────────────────────
# Sheet 6 – Data Traceability
# ─────────────────────────────────────────────
print("[6/6] Sheet 6: Data Traceability...")
ws6 = wb.create_sheet("Data_Traceability")
ws6.sheet_view.showGridLines = False

# --- Data flow ---
flow_rows = [
    ("Layer",        "รูปแบบ",        "Path",                                           "หมายเหตุ"),
    ("Bronze (Raw)", "Excel (.XLSX)", "01_Bronze_Raw/Production/2025/PPPP_2025_MM.XLSX","SAP extract รายเดือน แยก Plant (1100/1200/1300)"),
    ("Silver",       "Parquet",       "02_Silver_Cleaned/master_production_2025.parquet","รวม 30 ไฟล์ append, rename garbled columns, cast types"),
    ("Gold (Report)","Excel (.XLSX)", "04_Reports/Production_Cost_2025.xlsx",           "Aggregated ตาม SKU / MatGroup พร้อม Conversion per KG"),
    ("Reference",    "Excel (.XLSX)", "04_Reports/Production_Cost_NRV_Reference_2025.xlsx","เอกสารนี้ — NRV reference สำหรับ Audit"),
]

write_header_row(ws6, 1, flow_rows[0], ORANGE)
for ri, row in enumerate(flow_rows[1:], start=2):
    bg = LIGHT_GRAY if ri % 2 == 0 else "FFFFFF"
    ws6.row_dimensions[ri].height = 24
    write_data_row(ws6, ri, list(row), bg=bg)

ws6.row_dimensions[len(flow_rows) + 2].height = 12

# --- Source file reconciliation ---
r_start = len(flow_rows) + 3
section_label(ws6, r_start, "Source File Reconciliation — ยืนยันความครบถ้วนของข้อมูล", 5)
r_start += 1

write_header_row(ws6, r_start,
    ["Source File", "Plant", "เดือน", "จำนวน Order Rows", "Total Cost (THB)"], GRAY)
r_start += 1

src_summary = df.groupby("Source_File", sort=True).agg(
    Plant    = ("Plant",      "first"),
    Month    = ("Month",      "first"),
    Rows     = ("Material",   "count"),
    TotalCost= ("Total_Cost", "sum"),
).reset_index().sort_values("Source_File")

for ri2, row in enumerate(src_summary.itertuples(index=False), start=r_start):
    bg = LIGHT_BLUE if row.Plant == "1100" else (LIGHT_GREEN if row.Plant == "1200" else LIGHT_RED)
    vals = [row.Source_File, row.Plant, int(row.Month), row.Rows, row.TotalCost]
    write_data_row(ws6, ri2, vals, bg=bg, fmt_map={4: '#,##0', 5: '#,##0'})

# Grand total
ri2 += 1
write_data_row(ws6, ri2,
    ["GRAND TOTAL", "", "", src_summary["Rows"].sum(), src_summary["TotalCost"].sum()],
    bg=YELLOW, fmt_map={4: '#,##0', 5: '#,##0'}, bold=True)

ws6.column_dimensions["A"].width = 28
ws6.column_dimensions["B"].width = 8
ws6.column_dimensions["C"].width = 8
ws6.column_dimensions["D"].width = 18
ws6.column_dimensions["E"].width = 24
ws6.freeze_panes = "A2"

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n  Output: {OUTPUT_FILE}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
