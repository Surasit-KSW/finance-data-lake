"""
NRV RM – Pipe & Pipe GI Analysis  2026 Q1
==========================================
Assumption: RM (HR/GI coil) → FG (Pipe HR / Pipe GI)
NRV(RM) = NRV(FG) weighted by Q1 production volume − Conversion Cost/KG

Inputs
  01_Bronze_Raw/inventory/amc/nrv_sku_202603.xlsx
  04_Reports/AMC_PRD_Q1'2026.XLSX  (Sheet1)

Output
  04_Reports/NRV_RM_Pipe_Analysis_2026_Q1.xlsx
"""

import os, sys, warnings
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── paths ───────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
NRV_FILE = os.path.join(BASE, "01_Bronze_Raw", "inventory", "amc",
                        "nrv_sku_202603.xlsx")
PRD_FILE = os.path.join(BASE, "04_Reports", "AMC_PRD_Q1'2026.XLSX")
OUTPUT   = os.path.join(BASE, "04_Reports", "NRV_RM_Pipe_Analysis_2026_Q1.xlsx")

SELLING_COST = 0.3255   # constant deduction THB/KG

# RM type → list of FG pipe mat_group patterns
RM_TO_FG = {
    "HR":  ["Pipe Rect. HR",  "Pipe Round HR",  "Pipe Square HR"],
    "GI":  ["Pipe Rect. GI",  "Pipe Round GI",  "Pipe Square GI"],
    "GIM": ["Pipe Rect. GIM", "Pipe Round GIM", "Pipe Square GIM"],
    "CR":  ["Pipe Rect. CR",  "Pipe Round CR"],
    "PO":  ["Pipe Round PO",  "Pipe Square PO", "Pipe Rect. PO"],
}

# ── colours ─────────────────────────────────────────────────────────────────
NAVY   = "002060"; WHITE  = "FFFFFF"; LGREEN = "E2EFDA"; LRED   = "FFE0E0"
LBLUE  = "DEEAF1"; LYEL   = "FFF2CC"; LGRAY  = "F2F2F2"; GREEN  = "375623"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def fn(c=None, bold=False, sz=10, italic=False):
    kw = dict(bold=bold, size=sz, italic=italic)
    if c: kw["color"] = c
    return Font(**kw)
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def wh(ws, r, headers, bg=NAVY, txt=WHITE):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = fill(bg); cell.font = fn(txt, bold=True, sz=10)
        cell.alignment = al("center"); cell.border = border_thin()
    ws.row_dimensions[r].height = 18

def wr(ws, r, vals, bg=None, fmts=None):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        if bg: cell.fill = fill(bg)
        cell.font = fn(sz=9); cell.alignment = al(); cell.border = border_thin()
        if fmts and c in fmts:
            cell.number_format = fmts[c]
    ws.row_dimensions[r].height = 16

def aw(ws, extra=2):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + extra, 40)

def title(ws, text, row=1, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(NAVY); c.font = fn(WHITE, bold=True, sz=13)
    c.alignment = al("center"); ws.row_dimensions[row].height = 28

# ════════════════════════════════════════════════════════════════════════════
print("[1/4] Loading data...")

# ── NRV file ────────────────────────────────────────────────────────────────
nrv = pd.read_excel(NRV_FILE)
for col in ["stock_qty_kg","stock_amt","nrv_price","cost_price",
            "avg_price_4m","grp_avg_price","so_price","write_down_amt",
            "val_m1","val_m2","val_m3","val_m4",
            "qty_m1","qty_m2","qty_m3","qty_m4",
            "price_m1","price_m2","price_m3","price_m4"]:
    if col in nrv.columns:
        nrv[col] = pd.to_numeric(nrv[col], errors="coerce")

fg = nrv[nrv["sku_type"] == "FG"].copy()
rm = nrv[nrv["sku_type"] == "RM"].copy()

# Pipe FG only
pipe_fg = fg[fg["mat_group"].str.contains("Pipe", na=False)].copy()

# ── Production Q1 ────────────────────────────────────────────────────────────
prd_raw = pd.read_excel(PRD_FILE, sheet_name="Sheet1", header=0)
pcols = list(prd_raw.columns)

rename_map = {
    pcols[6]:  "Mat_Group_Name",
    pcols[7]:  "Mat_Type_Code",
    pcols[10]: "Mat_Group_1_Name",
    pcols[11]: "Mat_Group_2_Name",
    pcols[12]: "Mat_Group_3_Name",
    pcols[13]: "Mat_Group_4_Name",
}
prd_raw = prd_raw.rename(columns=rename_map)

# Numeric casts
for col in ["Actual GI Amount","Actual GR Amount","Actual GR QTY","Actual GI QTY",
            "Actual ByProduct Scrap Amount","Actual ByProduct Grade B Amount",
            "Actual ByProduct Grade C Amount"]:
    prd_raw[col] = pd.to_numeric(prd_raw[col], errors="coerce").fillna(0)

cost_d = ["Actual D101 Amount","Actual D102 Amount"]
cost_i = [f"Actual I10{n} Amount" for n in range(1, 12)
          if f"Actual I10{n} Amount" in prd_raw.columns and n != 6]
for col in cost_d + cost_i:
    prd_raw[col] = pd.to_numeric(prd_raw[col], errors="coerce").fillna(0)

prd_raw["ByProduct"]  = (prd_raw["Actual ByProduct Scrap Amount"] +
                          prd_raw["Actual ByProduct Grade B Amount"] +
                          prd_raw["Actual ByProduct Grade C Amount"])
prd_raw["DM"]         = prd_raw["Actual GI Amount"] - prd_raw["ByProduct"]
prd_raw["Conversion"] = prd_raw[cost_d].sum(axis=1) + prd_raw[cost_i].sum(axis=1)
prd_raw["Total_Cost"] = prd_raw["DM"] + prd_raw["Conversion"]

# Clean: exclude anomalous rows, completed orders only
prd = prd_raw[
    (prd_raw["Actual GR Amount"] <= 1_000_000_000) &
    (prd_raw["Actual GR QTY"]   > 0)
].copy()

# Pipe production only
pipe_prd = prd[prd["Mat_Group_Name"].str.contains("Pipe", na=False)].copy()

# ── Aggregate production by pipe mat_group ────────────────────────────────
print("[2/4] Calculating production ratios...")

prd_agg = pipe_prd.groupby("Mat_Group_Name").agg(
    GR_QTY       =("Actual GR QTY",    "sum"),
    GI_Amt       =("Actual GI Amount",  "sum"),
    DM_Amt       =("DM",                "sum"),
    Conv_Amt     =("Conversion",        "sum"),
    TotalCost_Amt=("Total_Cost",        "sum"),
    GR_Amt       =("Actual GR Amount",  "sum"),
    Orders       =("Material",          "count"),
).reset_index()

prd_agg["DM_per_KG"]   = np.where(prd_agg["GR_QTY"]>0, prd_agg["DM_Amt"]   / prd_agg["GR_QTY"], 0)
prd_agg["Conv_per_KG"] = np.where(prd_agg["GR_QTY"]>0, prd_agg["Conv_Amt"] / prd_agg["GR_QTY"], 0)
prd_agg["Cost_per_KG"] = np.where(prd_agg["GR_QTY"]>0, prd_agg["TotalCost_Amt"] / prd_agg["GR_QTY"], 0)
prd_agg["GR_per_KG"]   = np.where(prd_agg["GR_QTY"]>0, prd_agg["GR_Amt"]   / prd_agg["GR_QTY"], 0)

# ── FG NRV weighted avg per pipe mat_group (stock-weighted) ──────────────
fg_nrv_agg = pipe_fg.groupby("mat_group").apply(
    lambda d: pd.Series({
        "FG_NRV_WtdAvg": (
            np.average(d["nrv_price"], weights=d["stock_qty_kg"])
            if d["nrv_price"].notna().any() and d["stock_qty_kg"].sum() > 0
            else np.nan
        ),
        "FG_NRV_GrpAvg": d["grp_avg_price"].dropna().mean(),
        "FG_Cost_WtdAvg": (
            np.average(d["cost_price"], weights=d["stock_qty_kg"])
            if d["cost_price"].notna().any() and d["stock_qty_kg"].sum() > 0
            else np.nan
        ),
        "FG_Stock_KG":    d["stock_qty_kg"].sum(),
        "FG_Write_Down":  d["write_down_amt"].sum(),
        "FG_SKUs":        len(d),
        "FG_SKUs_w_NRV":  d["nrv_price"].notna().sum(),
    })
).reset_index()

# ── Build RM-level summary: weight by Q1 production volume ────────────────
records = []
for rm_type, fg_groups in RM_TO_FG.items():
    prod_sub = prd_agg[prd_agg["Mat_Group_Name"].isin(fg_groups)]
    nrv_sub  = fg_nrv_agg[fg_nrv_agg["mat_group"].isin(fg_groups)]

    total_gr = prod_sub["GR_QTY"].sum()
    if total_gr == 0:
        continue

    # Conv_per_KG = production-volume weighted average across all pipe sub-groups
    conv_wtd  = np.average(prod_sub["Conv_per_KG"], weights=prod_sub["GR_QTY"]) if len(prod_sub)>0 else 0
    dm_wtd    = np.average(prod_sub["DM_per_KG"],   weights=prod_sub["GR_QTY"]) if len(prod_sub)>0 else 0
    cost_wtd  = np.average(prod_sub["Cost_per_KG"], weights=prod_sub["GR_QTY"]) if len(prod_sub)>0 else 0

    # FG NRV = stock-weighted NRV across pipe groups (re-weight by FG stock KG)
    nrv_rows = nrv_sub[nrv_sub["FG_NRV_WtdAvg"].notna()]
    fg_nrv_wtd = (
        np.average(nrv_rows["FG_NRV_WtdAvg"], weights=nrv_rows["FG_Stock_KG"])
        if len(nrv_rows)>0 and nrv_rows["FG_Stock_KG"].sum()>0
        else np.nan
    )
    fg_nrv_grp = (
        np.average(nrv_rows["FG_NRV_GrpAvg"], weights=nrv_rows["FG_Stock_KG"])
        if len(nrv_rows)>0 and nrv_rows["FG_Stock_KG"].sum()>0
        else np.nan
    )
    fg_cost_wtd = (
        np.average(nrv_sub["FG_Cost_WtdAvg"], weights=nrv_sub["FG_Stock_KG"])
        if len(nrv_sub)>0 and nrv_sub["FG_Stock_KG"].sum()>0
        else np.nan
    )

    # NRV(RM) derived
    nrv_rm_derived = (fg_nrv_wtd - conv_wtd) if not np.isnan(fg_nrv_wtd) else np.nan

    # NRV(RM) from file
    rm_sub = rm[rm["nrv_source"] == f"RM->FG({rm_type})"].copy()
    if rm_sub.empty:
        rm_sub = rm[rm["nrv_source"].str.contains(f"FG\\({rm_type}\\)", na=False, regex=True)]
    rm_nrv_file = rm_sub["nrv_price"].dropna().mean() if not rm_sub.empty else np.nan
    rm_cost_avg = rm_sub["cost_price"].dropna().mean()   if not rm_sub.empty else np.nan
    rm_stock_kg = rm_sub["stock_qty_kg"].sum()           if not rm_sub.empty else 0

    records.append({
        "RM_Type":         rm_type,
        "FG_Groups":       ", ".join(fg_groups),
        "Q1_GR_QTY_KG":   total_gr,
        "DM_per_KG":       dm_wtd,
        "Conv_per_KG":     conv_wtd,
        "Cost_per_KG":     cost_wtd,
        "FG_NRV_WtdAvg":   fg_nrv_wtd,
        "FG_GrpAvg_Price": fg_nrv_grp,
        "FG_Stock_KG":     nrv_sub["FG_Stock_KG"].sum(),
        "FG_Write_Down":   nrv_sub["FG_Write_Down"].sum(),
        "NRV_RM_Derived":  nrv_rm_derived,
        "NRV_RM_File":     rm_nrv_file,
        "Diff":            (nrv_rm_derived - rm_nrv_file) if not np.isnan(nrv_rm_derived) and not np.isnan(rm_nrv_file) else np.nan,
        "RM_Cost_Avg":     rm_cost_avg,
        "RM_Stock_KG":     rm_stock_kg,
        "RM_SKUs":         len(rm_sub),
    })

rm_summary = pd.DataFrame(records)

# ── Monthly selling price trend (FG pipe by RM type) ─────────────────────
def monthly_price(fg_grp_list):
    sub = pipe_fg[pipe_fg["mat_group"].isin(fg_grp_list)]
    out = {}
    for m in range(1, 5):
        qty = sub[f"qty_m{m}"].fillna(0)
        val = sub[f"val_m{m}"].fillna(0)
        total_q = qty.sum(); total_v = val.sum()
        out[f"M{m}_WtdAvg"] = total_v / total_q if total_q > 0 else np.nan
        out[f"M{m}_QTY_KG"] = total_q
    return out

monthly_rows = []
for rm_type, fg_groups in RM_TO_FG.items():
    if not prd_agg["Mat_Group_Name"].isin(fg_groups).any():
        continue
    row = {"RM_Type": rm_type}
    row.update(monthly_price(fg_groups))
    monthly_rows.append(row)
monthly_df = pd.DataFrame(monthly_rows)

# ── Per-group production detail ───────────────────────────────────────────
grp_detail = prd_agg.merge(
    fg_nrv_agg.rename(columns={"mat_group": "Mat_Group_Name"}),
    on="Mat_Group_Name", how="left"
)
grp_detail["NRV_RM_Derived"] = grp_detail["FG_NRV_WtdAvg"] - grp_detail["Conv_per_KG"]

# Tag RM type
def tag_rm(mg):
    for rm_type, groups in RM_TO_FG.items():
        if mg in groups:
            return rm_type
    return "Other"

grp_detail["RM_Type"] = grp_detail["Mat_Group_Name"].apply(tag_rm)

# ── Individual FG SKU level (pipe, with NRV and implied RM NRV) ────────────
sku_detail = pipe_fg[pipe_fg["nrv_price"].notna()].copy()

# Merge conv from production agg
sku_detail = sku_detail.merge(
    prd_agg[["Mat_Group_Name","Conv_per_KG","DM_per_KG","Cost_per_KG","GR_QTY","Orders"]],
    left_on="mat_group", right_on="Mat_Group_Name", how="left"
)
sku_detail["Implied_RM_NRV"] = sku_detail["nrv_price"] - sku_detail["Conv_per_KG"]
sku_detail = sku_detail.sort_values(["mat_group","sku"])

# ════════════════════════════════════════════════════════════════════════════
print("[3/4] Building Excel...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ───────────────────────────────────────────────────────────────────
# Sheet 1: RM Summary
# ───────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("RM_Summary")
title(ws1, "NRV(RM) Summary — Pipe / Pipe GI  |  2026 Q1", cols=16)

ws1.cell(row=2, column=1, value="Methodology: NRV(RM) = Weighted-Avg NRV(FG Pipe) − Conversion Cost/KG  |  Q1 production volume used as weights").font = fn(NAVY, italic=True)
ws1.row_dimensions[2].height = 16

wh(ws1, 3, [
    "RM Type","FG Pipe Groups",
    "Q1 GR QTY (KG)",
    "DM/KG","Conv/KG","Total Cost/KG",
    "FG NRV WtdAvg","FG GrpAvg Price",
    "NRV(RM) Derived","NRV(RM) in File","Difference",
    "RM Cost Avg","RM Stock KG","RM SKUs",
    "FG Stock KG","FG Write-Down"
], bg=NAVY)

r = 4
for _, row in rm_summary.iterrows():
    diff = row["Diff"]
    bg = LGREEN if (not np.isnan(diff) and abs(diff) < 0.05) else (LRED if not np.isnan(diff) else LYEL)
    wr(ws1, r, [
        row["RM_Type"], row["FG_Groups"],
        row["Q1_GR_QTY_KG"],
        row["DM_per_KG"], row["Conv_per_KG"], row["Cost_per_KG"],
        row["FG_NRV_WtdAvg"], row["FG_GrpAvg_Price"],
        row["NRV_RM_Derived"], row["NRV_RM_File"], row["Diff"],
        row["RM_Cost_Avg"], row["RM_Stock_KG"], row["RM_SKUs"],
        row["FG_Stock_KG"], row["FG_Write_Down"],
    ], bg=bg, fmts={
        3:"#,##0", 4:"#,##0.0000", 5:"#,##0.0000", 6:"#,##0.0000",
        7:"#,##0.0000", 8:"#,##0.0000", 9:"#,##0.0000", 10:"#,##0.0000",
        11:"#,##0.0000", 12:"#,##0.0000", 13:"#,##0", 15:"#,##0", 16:"#,##0",
    })
    r += 1

# Legend
r += 1
ws1.cell(row=r, column=1, value="Colour legend:").font = fn(bold=True)
ws1.cell(row=r+1, column=1).fill = fill(LGREEN)
ws1.cell(row=r+1, column=2, value="Derived NRV matches file (diff < 0.05)")
ws1.cell(row=r+2, column=1).fill = fill(LRED)
ws1.cell(row=r+2, column=2, value="Derived NRV differs from file (investigate)")
ws1.cell(row=r+3, column=1).fill = fill(LYEL)
ws1.cell(row=r+3, column=2, value="No RM NRV in file (RM No FG Data)")
aw(ws1, 3)

# ───────────────────────────────────────────────────────────────────
# Sheet 2: Production Detail by Pipe Group
# ───────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Production_by_Group")
title(ws2, "Q1 Production — Conversion Cost per Pipe Mat Group", cols=13)

wh(ws2, 2, [
    "RM Type","Mat Group",
    "Q1 GR QTY (KG)","Orders",
    "Total GI Amt","Total DM","Total Conv","Total Cost","Total GR Amt",
    "DM/KG","Conv/KG","Cost/KG","GR/KG",
    "FG NRV WtdAvg","FG Cost WtdAvg","FG Stock KG","NRV(RM) Derived",
])

r = 3
for rm_type, fg_groups in RM_TO_FG.items():
    sub = grp_detail[grp_detail["Mat_Group_Name"].isin(fg_groups)].sort_values("GR_QTY", ascending=False)
    if sub.empty:
        continue
    for _, row in sub.iterrows():
        wr(ws2, r, [
            rm_type, row["Mat_Group_Name"],
            row["GR_QTY"], row["Orders"],
            row["GI_Amt"], row["DM_Amt"], row["Conv_Amt"], row["TotalCost_Amt"], row["GR_Amt"],
            row["DM_per_KG"], row["Conv_per_KG"], row["Cost_per_KG"], row["GR_per_KG"],
            row["FG_NRV_WtdAvg"], row["FG_Cost_WtdAvg"], row["FG_Stock_KG"], row["NRV_RM_Derived"],
        ], fmts={
            3:"#,##0", 5:"#,##0", 6:"#,##0", 7:"#,##0", 8:"#,##0", 9:"#,##0",
            10:"#,##0.0000", 11:"#,##0.0000", 12:"#,##0.0000", 13:"#,##0.0000",
            14:"#,##0.0000", 15:"#,##0.0000", 16:"#,##0", 17:"#,##0.0000",
        })
        r += 1
    # sub-total row
    ws2.cell(row=r, column=1, value=f"  Subtotal {rm_type}").font = fn(bold=True)
    ws2.cell(row=r, column=3, value=sub["GR_QTY"].sum()).number_format = "#,##0"
    ws2.cell(row=r, column=5, value=sub["GI_Amt"].sum()).number_format = "#,##0"
    ws2.cell(row=r, column=7, value=sub["Conv_Amt"].sum()).number_format = "#,##0"
    ws2.cell(row=r, column=8, value=sub["TotalCost_Amt"].sum()).number_format = "#,##0"
    for c in range(1, 18):
        ws2.cell(row=r, column=c).fill = fill(LGRAY)
        ws2.cell(row=r, column=c).border = border_thin()
    r += 2

aw(ws2, 3)

# ───────────────────────────────────────────────────────────────────
# Sheet 3: Monthly Selling Price Trend
# ───────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Monthly_Sale_Price")
title(ws3, "Monthly Selling Price — FG Pipe (Vol-Weighted Avg)  2026 Q1", cols=9)

ws3.cell(row=2, column=1, value="Selling price reference used to derive NRV(FG) → NRV(RM)").font = fn(NAVY, italic=True)
ws3.row_dimensions[2].height = 15

wh(ws3, 3, [
    "RM Type",
    "Jan 26 Wtd Avg", "Jan QTY (KG)",
    "Feb 26 Wtd Avg", "Feb QTY (KG)",
    "Mar 26 Wtd Avg", "Mar QTY (KG)",
    "Apr 26 Wtd Avg", "Apr QTY (KG)",
    "4M Wtd Avg",
])

r = 4
for _, row in monthly_df.iterrows():
    m_prices = [row.get(f"M{m}_WtdAvg", np.nan) for m in range(1,5)]
    m_qtys   = [row.get(f"M{m}_QTY_KG",  0)     for m in range(1,5)]
    total_v  = sum(p*q for p,q in zip(m_prices, m_qtys) if not np.isnan(p))
    total_q  = sum(q for p,q in zip(m_prices, m_qtys) if not np.isnan(p))
    avg_4m   = total_v / total_q if total_q > 0 else np.nan

    wr(ws3, r, [
        row["RM_Type"],
        row.get("M1_WtdAvg"), row.get("M1_QTY_KG"),
        row.get("M2_WtdAvg"), row.get("M2_QTY_KG"),
        row.get("M3_WtdAvg"), row.get("M3_QTY_KG"),
        row.get("M4_WtdAvg"), row.get("M4_QTY_KG"),
        avg_4m,
    ], fmts={
        2:"#,##0.0000", 3:"#,##0", 4:"#,##0.0000", 5:"#,##0", 6:"#,##0.0000",
        7:"#,##0", 8:"#,##0.0000", 9:"#,##0", 10:"#,##0.0000",
    })
    r += 1

# NRV calculation walkthrough
r += 1
ws3.cell(row=r, column=1, value="NRV Calculation (IAS 2)").font = fn(NAVY, bold=True, sz=11)
ws3.row_dimensions[r].height = 20; r += 1
steps = [
    ("Step 1", "Reference Price", "= 4-Month Weighted Avg Selling Price (or Group Avg if no sale)"),
    ("Step 2", "NRV(FG)",         f"= Reference Price − {SELLING_COST} THB/KG  (Selling Cost)"),
    ("Step 3", "NRV(RM)",         "= NRV(FG) − Conversion Cost/KG  (from Q1 production)"),
    ("Step 4", "IAS 2 Test",      "= Compare Cost vs NRV → take lower for inventory valuation"),
]
for step, label, desc in steps:
    ws3.cell(row=r, column=1, value=step).font  = fn(WHITE, bold=True); ws3.cell(row=r, column=1).fill = fill(NAVY)
    ws3.cell(row=r, column=2, value=label).font = fn(bold=True)
    ws3.cell(row=r, column=3, value=desc)
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    ws3.row_dimensions[r].height = 16; r += 1

aw(ws3, 3)

# ───────────────────────────────────────────────────────────────────
# Sheet 4: FG SKU Detail (Pipe, with implied RM NRV)
# ───────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("FG_SKU_Detail")
title(ws4, "FG Pipe SKU Detail — NRV & Implied RM NRV", cols=14)

wh(ws4, 2, [
    "SKU","Description","Mat Group",
    "Stock KG","Stock Price","NRV Price","NRV Source",
    "Conv/KG (Q1 Prd)","Implied RM NRV",
    "Cost Price","Price Diff","Write-Down Amt",
    "Trend","Price Chg %",
])

r = 3
for _, row in sku_detail.iterrows():
    bg = LRED if (pd.notna(row.get("write_down_amt")) and row.get("write_down_amt", 0) < -10000) else None
    wr(ws4, r, [
        row["sku"], row["description"], row["mat_group"],
        row["stock_qty_kg"], row["stock_price"], row["nrv_price"], row["nrv_source"],
        row.get("Conv_per_KG"), row.get("Implied_RM_NRV"),
        row["cost_price"], row.get("price_diff"), row.get("write_down_amt"),
        row.get("trend"), row.get("price_chg_pct"),
    ], bg=bg, fmts={
        4:"#,##0", 5:"#,##0.0000", 6:"#,##0.0000",
        8:"#,##0.0000", 9:"#,##0.0000", 10:"#,##0.0000", 11:"#,##0.0000",
        12:"#,##0", 14:"0.00%",
    })
    r += 1

aw(ws4, 3)

# ───────────────────────────────────────────────────────────────────
# Sheet 5: RM SKU Detail
# ───────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("RM_SKU_Detail")
title(ws5, "RM SKU Detail — Coil HR / GI / GIM / CR  (RM in NRV file)", cols=10)

# Merge derived NRV onto RM SKUs
rm_with_derived = rm[rm["nrv_source"].str.contains("RM->FG", na=False)].copy()

# Map rm_type to derived value
derived_map = rm_summary.set_index("RM_Type")["NRV_RM_Derived"].to_dict()
rm_with_derived["RM_Type"] = rm_with_derived["nrv_source"].str.extract(r"RM->FG\((\w+)\)", expand=False)
rm_with_derived["NRV_Derived"] = rm_with_derived["RM_Type"].map(derived_map)
rm_with_derived["Diff_vs_File"] = rm_with_derived["NRV_Derived"] - rm_with_derived["nrv_price"]

wh(ws5, 2, [
    "SKU","Description","Mat Group","NRV Source","RM Type",
    "NRV in File","NRV Derived","Diff","Cost Price",
    "Stock KG","Stock Amt","Write-Down Amt",
])

r = 3
for _, row in rm_with_derived.sort_values(["RM_Type","sku"]).iterrows():
    diff = row.get("Diff_vs_File", np.nan)
    bg = LGREEN if (not pd.isna(diff) and abs(diff) < 0.05) else LRED
    wr(ws5, r, [
        row["sku"], row["description"], row["mat_group"], row["nrv_source"], row.get("RM_Type"),
        row["nrv_price"], row.get("NRV_Derived"), diff, row["cost_price"],
        row["stock_qty_kg"], row["stock_amt"], row.get("write_down_amt"),
    ], bg=bg, fmts={
        6:"#,##0.0000", 7:"#,##0.0000", 8:"#,##0.0000", 9:"#,##0.0000",
        10:"#,##0", 11:"#,##0", 12:"#,##0",
    })
    r += 1

aw(ws5, 3)

# ───────────────────────────────────────────────────────────────────
# Sheet 6: Methodology Note
# ───────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Methodology")
title(ws6, "NRV(RM) Methodology — Pipe / Pipe GI", cols=2)

notes = [
    ("", ""),
    ("ASSUMPTION", "RM (coil) feeds FG pipe through forming/ERW/slitting process"),
    ("", ""),
    ("PIPE LINE MAPPING", ""),
    ("HR coil (Coil HR แคบ/กว้าง)", "→ Pipe Rect. HR / Pipe Round HR / Pipe Square HR"),
    ("GI coil (Coil GI กว้าง)",     "→ Pipe Rect. GI / Pipe Round GI / Pipe Square GI"),
    ("GIM coil (Coil GIM กว้าง)",   "→ Pipe Rect. GIM / Pipe Round GIM / Pipe Square GIM"),
    ("CR coil (Coil CR กว้าง)",     "→ Pipe Rect. CR / Pipe Round CR"),
    ("PO coil (Coil PO)",           "→ Pipe Round PO / Pipe Square PO / Pipe Rect. PO"),
    ("", ""),
    ("CONVERSION COST", ""),
    ("Source",      "SAP CO Production Orders  (AMC_PRD_Q1'2026.XLSX, Sheet1)"),
    ("Formula",     "Conv = D101 + D102 + I101 + I102 + I103 + I104 + I105 + I107 + I108 + I109 + I110 + I111"),
    ("Note",        "I106 excluded (always 0)"),
    ("Conv/KG",     "= Total Conversion ÷ Total GR QTY  (per pipe mat group)"),
    ("Weighting",   "Aggregated by mat_group, then volume-weighted across groups per RM type"),
    ("", ""),
    ("NRV(FG) REFERENCE PRICE", ""),
    ("Priority",    "1) Actual Sale WtdAvg (4M)  →  2) SO Backlog  →  3) MatGroup+Grade Avg"),
    ("NRV(FG)",     f"= Reference Price − {SELLING_COST} THB/KG  (constant selling cost deduction)"),
    ("FG Weighting","Within each RM type, FG groups weighted by inventory stock KG"),
    ("", ""),
    ("NRV(RM) FORMULA", ""),
    ("Derived",     "NRV(RM) = FG NRV Weighted Avg − Conversion Cost/KG  (Q1 actual)"),
    ("IAS 2",       "Inventory write-down = Cost − NRV if Cost > NRV"),
    ("", ""),
    ("FILE COMPARISON", ""),
    ("NRV file",    "NRV_SKU_Analysis_2026_Q1.xlsx  (column: nrv_price, nrv_source = RM->FG(XX))"),
    ("Expectation", "Derived NRV ≈ File NRV if same FG reference price and same Conv/KG assumption"),
    ("Gap reason",  "Differences arise from: (a) file uses specific FG NRV not full group avg, "
                    "(b) Conv/KG may use annual/budget rate not Q1 actual"),
]

r = 2
for label, value in notes:
    if label == "":
        r += 1; continue
    label_cell = ws6.cell(row=r, column=1, value=label)
    value_cell = ws6.cell(row=r, column=2, value=value)
    if label in ("ASSUMPTION","PIPE LINE MAPPING","CONVERSION COST","NRV(FG) REFERENCE PRICE",
                 "NRV(RM) FORMULA","FILE COMPARISON"):
        label_cell.font = fn(NAVY, bold=True, sz=11)
        label_cell.fill = fill(LBLUE)
        value_cell.fill = fill(LBLUE)
    else:
        label_cell.font = fn(bold=True)
    value_cell.alignment = al(wrap=True)
    ws6.row_dimensions[r].height = 20
    r += 1

ws6.column_dimensions["A"].width = 28
ws6.column_dimensions["B"].width = 65

# ════════════════════════════════════════════════════════════════════════════
print("[4/4] Saving...")
wb.save(OUTPUT)
print(f"\n  Output : {OUTPUT}")
print(f"  Sheets : {[s.title for s in wb.worksheets]}")
print()
print("=" * 60)
print("  NRV(RM) — Pipe / Pipe GI Summary")
print("=" * 60)
for _, row in rm_summary.iterrows():
    nrv_d = f"{row['NRV_RM_Derived']:.4f}" if not np.isnan(row['NRV_RM_Derived']) else "N/A"
    nrv_f = f"{row['NRV_RM_File']:.4f}"     if not np.isnan(row['NRV_RM_File'])    else "N/A"
    diff  = f"{row['Diff']:+.4f}"            if not np.isnan(row['Diff'])           else "N/A"
    fg_nrv = f"{row['FG_NRV_WtdAvg']:.4f}" if not np.isnan(row['FG_NRV_WtdAvg']) else "N/A"
    print(f"  RM({row['RM_Type']:4s})  Conv={row['Conv_per_KG']:.4f}  "
          f"FG NRV={fg_nrv}  "
          f"NRV_derived={nrv_d}  NRV_file={nrv_f}  diff={diff}")
print("=" * 60)
