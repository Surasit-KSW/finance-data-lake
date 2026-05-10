"""
nrv_price_methodology_2026_q1.py
==================================
อธิบายที่มาของ NRV Price ใน NRV_SKU_Analysis_2026_Q1_v2.xlsx
ทีละ nrv_source แบบ step-by-step

Output: 04_Reports/NRV_Price_Methodology_2026_Q1.xlsx
  Sheet 1: Summary           — สรุป logic และ priority ทั้งหมด
  Sheet 2: Actual_Sale       — คำนวณ nrv_price จาก Weighted Avg 4 เดือน
  Sheet 3: MatGroup_Avg      — คำนวณ nrv_price จาก grp_avg_price
  Sheet 4: SO_Backlog        — คำนวณ nrv_price จาก SO price
  Sheet 5: No_Data           — SKUs ที่ไม่มีราคาอ้างอิง
  Sheet 6: Selling_Cost_Proof— พิสูจน์ว่า deduction = 0.3255 THB/KG คงที่
"""

import os
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PATH_NRV = os.path.join(BASE_DIR, "01_Bronze_Raw", "Inventory_RollStock", "NRV",
                        "NRV_SKU_Analysis_2026_Q1_v2.xlsx")
OUTPUT   = os.path.join(BASE_DIR, "04_Reports", "NRV_Price_Methodology_2026_Q1.xlsx")
REPORT_DATE = date.today().strftime("%d %B %Y")
SELLING_COST = 0.3255   # THB/KG — deduction คงที่

# ─────────────────────────────────────────────
# Load
# ─────────────────────────────────────────────
print("[1/3] Loading data...")
df = pd.read_excel(PATH_NRV, sheet_name="SKU_Detail")

num_cols = ["price_m1","price_m2","price_m3","price_m4",
            "qty_m1","qty_m2","qty_m3","qty_m4",
            "val_m1","val_m2","val_m3","val_m4",
            "avg_price_4m","so_price","grp_avg_price",
            "nrv_price","cost_price","stock_qty_kg","stock_amt",
            "price_diff","write_down_per_kg","write_down_amt",
            "price_slope","price_chg_pct"]
for c in num_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce")

# Derived
df["wtd_avg_4m"] = (df[["val_m1","val_m2","val_m3","val_m4"]].sum(axis=1) /
                    df[["qty_m1","qty_m2","qty_m3","qty_m4"]].sum(axis=1).replace(0, np.nan))
df["ref_price"]  = np.where(df["nrv_source"]=="Actual Sale M1-M4", df["wtd_avg_4m"],
                   np.where(df["nrv_source"]=="MatGroup+Grade Avg", df["grp_avg_price"],
                   np.where(df["nrv_source"]=="SO Backlog",         df["so_price"],
                   np.nan)))
df["deduction"]  = np.where(df["ref_price"].notna(), df["nrv_price"] - df["ref_price"], np.nan)

fg = df[df["sku_type"] == "FG"].copy()
rm = df[df["sku_type"] == "RM"].copy()

# Source counts
src_counts = df.groupby(["sku_type","nrv_source"]).agg(
    sku_count    = ("sku",          "count"),
    stock_qty    = ("stock_qty_kg", "sum"),
    stock_amt    = ("stock_amt",    "sum"),
    nrv_null     = ("nrv_price",    lambda x: x.isna().sum()),
    total_wd     = ("write_down_amt",lambda x: x.fillna(0).sum()),
).reset_index()

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
NAVY="1F4E79"; GREEN="375623"; RED="7B241C"; ORANGE="C55A11"; PURPLE="4A235A"
LGRAY="F2F2F2"; LBLUE="D9E1F2"; LGREEN="E2EFDA"; LRED="FCE4D6"
LYELLOW="FFF2CC"; LPUR="EAE0F0"; WHITE="FFFFFF"

def _f(h):  return PatternFill("solid", fgColor=h)
def _b(c="CCCCCC"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _fn(color="000000", bold=False, sz=10):
    return Font(color=color, bold=bold, size=sz, name="Calibri")
def _c(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _l(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def wh(ws, r, vals, bg, fg="FFFFFF", col=1):
    for i, v in enumerate(vals, start=col):
        c = ws.cell(row=r, column=i, value=v)
        c.fill=_f(bg); c.font=_fn(fg,bold=True); c.alignment=_c(); c.border=_b("888888")
    ws.row_dimensions[r].height = 22

def wr(ws, r, vals, bg=None, fmts=None, bold=False, col=1):
    for i, v in enumerate(vals, start=col):
        c = ws.cell(row=r, column=i, value=v)
        if bg: c.fill=_f(bg)
        c.font=_fn(bold=bold)
        c.alignment=_l() if isinstance(v,str) else _c()
        c.border=_b()
        if fmts and i in fmts: c.number_format=fmts[i]
    ws.row_dimensions[r].height = 18

def aw(ws, mn=10, mx=46):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[ltr].width = min(max(w+2, mn), mx)

def title_bar(ws, row, text, ncols, bg=NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font=_fn("FFFFFF",bold=True,sz=12); c.fill=_f(bg)
    c.alignment=_c(); c.border=_b(); ws.row_dimensions[row].height=28

def formula_box(ws, row, label, formula, col=1, bg=LYELLOW):
    c1 = ws.cell(row=row, column=col, value=label)
    c1.font=_fn(RED,bold=True,sz=11); c1.fill=_f(bg); c1.alignment=_l(); c1.border=_b()
    c2 = ws.cell(row=row, column=col+1, value=formula)
    c2.font=_fn("000000",bold=True,sz=11); c2.fill=_f(bg); c2.alignment=_l(); c2.border=_b()
    ws.row_dimensions[row].height=24

# ══════════════════════════════════════════════════════════════
# Sheet 1 – Summary (Priority Tree + Overview)
# ══════════════════════════════════════════════════════════════
print("[2/3] Building sheets...")
wb = Workbook()
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False

title_bar(ws1, 1, f"NRV Price — ที่มาและวิธีคำนวณ  |  Q1 2026  |  จัดทำ: {REPORT_DATE}", 8)

# Master formula
ws1.row_dimensions[3].height = 10
formula_box(ws1, 4, "สูตรหลัก:", "NRV Price  =  Reference Price  −  Selling Cost (0.3255 THB/KG)", bg=LYELLOW)
formula_box(ws1, 5, "Selling Cost:", "0.3255 THB/KG  =  ค่าใช้จ่ายในการขายต่อหน่วย (คงที่ทุก SKU)", bg=LYELLOW)

# Priority table
ws1.row_dimensions[7].height = 10
r = 8
wh(ws1, r, ["ลำดับ","nrv_source","Reference Price ที่ใช้","SKU Count","เงื่อนไข","Stock QTY","Write-Down (THB)"], NAVY)
r += 1

priority_data = [
    ("1","Actual Sale M1-M4",  "Weighted Avg ราคาขายจริง M1–M4\n(∑ val / ∑ qty แยก SKU)",
     836, "SKU มีประวัติขายใน 4 เดือน", LGREEN),
    ("2","SO Backlog",          "ราคาใน Sales Order ที่ Confirm แล้ว (so_price)",
     24,  "ไม่มีขายจริงใน 4 เดือน แต่มี SO ค้างอยู่", LGREEN),
    ("3","MatGroup+Grade Avg",  "grp_avg_price = ราคาเฉลี่ยของ\nmat_group + grade เดียวกัน",
     1308,"ไม่มีประวัติขาย ไม่มี SO", LBLUE),
    ("4","RM→FG(HR/GI/GIM/CR)","FG NRV Price − Conversion Cost/KG\n(จากข้อมูลผลิต Q1 2026)",
     342, "RM: ดู NRV ของ FG ที่ผลิตได้ หัก conversion", LPUR),
    ("5","No Data",             "ไม่มี Reference Price",
     1178,"ไม่มีข้อมูลราคาใด ๆ เลย", LRED),
    ("6","RM No FG Data",       "ไม่มี FG ที่ match",
     31,  "RM ประเภทพิเศษ (EG, GA, GL, PCM, PO, SI)", LRED),
]

for pri, src, ref, cnt, cond, bg in priority_data:
    src_row = src_counts[(src_counts["nrv_source"]==src) & (src_counts["sku_type"]=="FG")]
    wd = src_row["total_wd"].sum() if len(src_row) else 0
    qty = src_row["stock_qty"].sum() if len(src_row) else 0
    if src.startswith("RM"):
        src_row2 = src_counts[(src_counts["nrv_source"]==src) & (src_counts["sku_type"]=="RM")]
        wd  += src_row2["total_wd"].sum() if len(src_row2) else 0
        qty += src_row2["stock_qty"].sum() if len(src_row2) else 0
    wr(ws1, r, [pri, src, ref, cnt, cond, qty, wd if wd != 0 else None],
       bg=bg, fmts={4:"#,##0",6:"#,##0",7:"#,##0.00"})
    r += 1

# Total
wr(ws1, r, ["","TOTAL","","","",
            fg["stock_qty_kg"].sum()+rm["stock_qty_kg"].sum(),
            df["write_down_amt"].fillna(0).sum()],
   bg=LYELLOW, fmts={6:"#,##0",7:"#,##0.00"}, bold=True)
r += 2

# Note on grp_avg_price
ws1.cell(row=r,column=1,value="หมายเหตุ grp_avg_price:").font=_fn(NAVY,bold=True)
ws1.cell(row=r,column=2,value=(
    "grp_avg_price = ราคาเฉลี่ยของ SKU ในกลุ่ม mat_group+grade เดียวกัน "
    "คำนวณจาก SKU ที่มีประวัติการขายจริง (Actual Sale) แล้วนำค่าเฉลี่ย (weighted by qty) "
    "ไปใช้กับ SKU ในกลุ่มเดียวกันที่ไม่มีประวัติขาย"
)).font=_fn(sz=10)
ws1.row_dimensions[r].height=36

ws1.column_dimensions["A"].width=10; ws1.column_dimensions["B"].width=26
ws1.column_dimensions["C"].width=46; ws1.column_dimensions["D"].width=12
ws1.column_dimensions["E"].width=38; ws1.column_dimensions["F"].width=16
ws1.column_dimensions["G"].width=18
ws1.freeze_panes="A2"

# ══════════════════════════════════════════════════════════════
# Sheet 2 – Actual Sale M1-M4
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Actual_Sale")
ws2.sheet_view.showGridLines = False
title_bar(ws2, 1, "nrv_source = Actual Sale M1-M4  |  สูตร: NRV = Weighted Avg(M1–M4) − 0.3255", 16)

formula_box(ws2, 2, "สูตร:", "NRV = (val_m1+val_m2+val_m3+val_m4) ÷ (qty_m1+qty_m2+qty_m3+qty_m4) − 0.3255")
ws2.row_dimensions[3].height=6

wh(ws2, 4, [
    "SKU","Description","Mat Group","Grade",
    "ราคา M1","QTY M1","ราคา M2","QTY M2",
    "ราคา M3","QTY M3","ราคา M4","QTY M4",
    "Wtd Avg\n(คำนวณ)","avg_price_4m\n(ไฟล์)","Selling Cost\n(0.3255)","NRV Price",
], GREEN)

sub_s = fg[fg["nrv_source"]=="Actual Sale M1-M4"].sort_values("mat_group")
for ri, row in enumerate(sub_s.itertuples(index=False), start=5):
    has_wd = pd.notna(row.write_down_amt) and row.write_down_amt < 0
    bg = LRED if has_wd else (WHITE if ri%2==0 else LGREEN)
    wr(ws2, ri, [
        row.sku, row.description, row.mat_group, row.grade,
        row.price_m1 if pd.notna(row.price_m1) else None,
        row.qty_m1   if pd.notna(row.qty_m1)   else None,
        row.price_m2 if pd.notna(row.price_m2) else None,
        row.qty_m2   if pd.notna(row.qty_m2)   else None,
        row.price_m3 if pd.notna(row.price_m3) else None,
        row.qty_m3   if pd.notna(row.qty_m3)   else None,
        row.price_m4 if pd.notna(row.price_m4) else None,
        row.qty_m4   if pd.notna(row.qty_m4)   else None,
        row.wtd_avg_4m,
        row.avg_price_4m,
        -SELLING_COST,
        row.nrv_price,
    ], bg=bg, fmts={
        5:"#,##0.0000",6:"#,##0",7:"#,##0.0000",8:"#,##0",
        9:"#,##0.0000",10:"#,##0",11:"#,##0.0000",12:"#,##0",
        13:"#,##0.0000",14:"#,##0.0000",15:"#,##0.0000",16:"#,##0.0000",
    })

aw(ws2)
ws2.freeze_panes="E5"

# ══════════════════════════════════════════════════════════════
# Sheet 3 – MatGroup+Grade Avg
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("MatGroup_Avg")
ws3.sheet_view.showGridLines = False
title_bar(ws3, 1, "nrv_source = MatGroup+Grade Avg  |  สูตร: NRV = grp_avg_price − 0.3255", 10)
formula_box(ws3, 2, "สูตร:","NRV = grp_avg_price (ราคาเฉลี่ยของ mat_group+grade) − 0.3255")
formula_box(ws3, 3, "grp_avg_price:","คำนวณจาก Weighted Avg ของ SKU ในกลุ่มเดียวกันที่มี Actual Sale (แชร์ราคากลางกัน)")
ws3.row_dimensions[4].height=6

# Section A: grp_avg_price reference table
r = 5
ws3.cell(row=r,column=1,value="ตาราง grp_avg_price อ้างอิง (แยก mat_group + grade)").font=_fn(NAVY,bold=True,sz=11)
ws3.row_dimensions[r].height=22
r += 1

wh(ws3, r, ["Mat Group","Grade","grp_avg_price (THB/KG)","SKU ที่ใช้ค่านี้","หมายเหตุ"], ORANGE)
r += 1

grp_ref = fg[fg["nrv_source"]=="MatGroup+Grade Avg"].groupby(["mat_group","grade"]).agg(
    grp_avg_price=("grp_avg_price","first"),
    sku_count    =("sku","count"),
).reset_index().sort_values(["mat_group","grade"])

for _, row in grp_ref.iterrows():
    wr(ws3, r, [
        row["mat_group"], row["grade"],
        row["grp_avg_price"], int(row["sku_count"]),
        f"NRV = {row['grp_avg_price']:.4f} − 0.3255 = {row['grp_avg_price']-SELLING_COST:.4f}",
    ], fmts={3:"#,##0.0000"})
    r += 1

ws3.row_dimensions[r].height=12; r+=1

# Section B: SKU detail
ws3.cell(row=r,column=1,value="รายละเอียดราย SKU").font=_fn(NAVY,bold=True,sz=11)
ws3.row_dimensions[r].height=22; r+=1

wh(ws3, r, ["SKU","Description","Mat Group","Grade",
            "grp_avg_price","Selling Cost (0.3255)","NRV Price",
            "Cost Price","Write-Down/KG"], GREEN)
r += 1

sub3 = fg[fg["nrv_source"]=="MatGroup+Grade Avg"].sort_values(["mat_group","grade"])
for ri2, row in enumerate(sub3.itertuples(index=False), start=r):
    has_wd = pd.notna(row.write_down_per_kg) and row.write_down_per_kg < 0
    bg = LRED if has_wd else (WHITE if ri2%2==0 else LBLUE)
    wr(ws3, ri2, [
        row.sku, row.description, row.mat_group, row.grade,
        row.grp_avg_price, -SELLING_COST, row.nrv_price,
        row.cost_price,
        row.write_down_per_kg if pd.notna(row.write_down_per_kg) else None,
    ], bg=bg, fmts={5:"#,##0.0000",6:"#,##0.0000",7:"#,##0.0000",8:"#,##0.0000",9:"#,##0.0000"})

aw(ws3)
ws3.freeze_panes="A2"

# ══════════════════════════════════════════════════════════════
# Sheet 4 – SO Backlog
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("SO_Backlog")
ws4.sheet_view.showGridLines = False
title_bar(ws4, 1, "nrv_source = SO Backlog  |  สูตร: NRV = so_price − 0.3255", 10)
formula_box(ws4, 2, "สูตร:","NRV = so_price (ราคาใน Sales Order ที่ confirm แล้ว) − 0.3255")
formula_box(ws4, 3, "เงื่อนไข:","ใช้เมื่อ: ไม่มีประวัติขายใน M1–M4 แต่มี SO Backlog ค้างอยู่ (so_qty_kg > 0)")
ws4.row_dimensions[4].height=6

wh(ws4, 5, ["SKU","Description","Mat Group","Grade",
            "so_price","SO QTY (KG)","SO Value",
            "Selling Cost","NRV Price","Cost Price","Write-Down/KG","Write-Down Amt"], RED)

sub4 = fg[fg["nrv_source"]=="SO Backlog"].sort_values("mat_group")
for ri, row in enumerate(sub4.itertuples(index=False), start=6):
    has_wd = pd.notna(row.write_down_per_kg) and row.write_down_per_kg < 0
    bg = LRED if has_wd else (WHITE if ri%2==0 else LGREEN)
    wr(ws4, ri, [
        row.sku, row.description, row.mat_group, row.grade,
        row.so_price,
        row.so_qty_kg   if pd.notna(row.so_qty_kg)  else None,
        row.so_value    if pd.notna(row.so_value)   else None,
        -SELLING_COST,
        row.nrv_price,
        row.cost_price,
        row.write_down_per_kg if pd.notna(row.write_down_per_kg) else None,
        row.write_down_amt    if pd.notna(row.write_down_amt)    else None,
    ], bg=bg, fmts={5:"#,##0.0000",6:"#,##0",7:"#,##0",8:"#,##0.0000",
                    9:"#,##0.0000",10:"#,##0.0000",11:"#,##0.0000",12:"#,##0"})

aw(ws4)
ws4.freeze_panes="A6"

# ══════════════════════════════════════════════════════════════
# Sheet 5 – No Data
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("No_Data")
ws5.sheet_view.showGridLines = False
title_bar(ws5, 1, "nrv_source = No Data  |  nrv_price = ไม่มี (NaN) — ต้องหาราคาอ้างอิงเพิ่มเติม", 8, RED)
ws5.cell(row=2,column=1,value=(
    "SKUs เหล่านี้ไม่มีประวัติขายใน M1–M4, ไม่มี SO Backlog, และไม่มี grp_avg_price ในกลุ่มเดียวกัน\n"
    "→ ควรพิจารณา: ราคาตลาดล่าสุด, ราคา Offer จาก Customer, หรือ cost_price เป็นราคา ceiling"
)).font=_fn(RED, sz=10)
ws5.row_dimensions[2].height=36

wh(ws5, 4, ["SKU","Description","Mat Group","Grade","sku_type",
            "Stock QTY (KG)","Stock Amt (THB)","Cost Price","nrv_price","แนวทาง"], PURPLE)

sub5 = df[df["nrv_source"]=="No Data"].sort_values(["sku_type","mat_group"])
for ri, row in enumerate(sub5.itertuples(index=False), start=5):
    bg = LGRAY if ri%2==0 else WHITE
    wr(ws5, ri, [
        row.sku, row.description, row.mat_group, row.grade, row.sku_type,
        row.stock_qty_kg, row.stock_amt, row.cost_price,
        "N/A",
        "ต้องระบุราคาอ้างอิงเพิ่มเติม",
    ], bg=bg, fmts={6:"#,##0",7:"#,##0",8:"#,##0.0000"})

# Summary by mat_group
ri += 2
ws5.cell(row=ri,column=1,value="สรุป No Data แยก Mat Group").font=_fn(NAVY,bold=True,sz=11)
ws5.row_dimensions[ri].height=22; ri+=1
wh(ws5, ri, ["Mat Group","sku_type","SKU Count","Stock QTY (KG)","Stock Amt (THB)"], PURPLE)
ri+=1
nd_mg = df[df["nrv_source"]=="No Data"].groupby(["mat_group","sku_type"]).agg(
    cnt=("sku","count"),qty=("stock_qty_kg","sum"),amt=("stock_amt","sum")).reset_index()
nd_mg = nd_mg.sort_values(["sku_type","amt"],ascending=[True,False])
for _, row in nd_mg.iterrows():
    wr(ws5, ri, [row["mat_group"],row["sku_type"],row["cnt"],row["qty"],row["amt"]],
       bg=LGRAY, fmts={4:"#,##0",5:"#,##0"}); ri+=1

aw(ws5)
ws5.freeze_panes="A5"

# ══════════════════════════════════════════════════════════════
# Sheet 6 – Selling Cost Proof
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Selling_Cost_Proof")
ws6.sheet_view.showGridLines = False
title_bar(ws6, 1, "พิสูจน์ Selling Cost Deduction = 0.3255 THB/KG (คงที่ทุก SKU)", 8, GREEN)

ws6.cell(row=2,column=1,value=(
    "ตรวจสอบโดยคำนวณ: deduction = NRV Price − Reference Price  สำหรับแต่ละ nrv_source"
)).font=_fn(sz=10); ws6.row_dimensions[2].height=20

r=4
for src_name, ref_col_name in [
    ("Actual Sale M1-M4",  "wtd_avg_4m"),
    ("MatGroup+Grade Avg", "grp_avg_price"),
    ("SO Backlog",         "so_price"),
]:
    sub = fg[fg["nrv_source"]==src_name].copy()
    sub["ref_price"] = sub[ref_col_name]
    sub["deduction"] = sub["nrv_price"] - sub["ref_price"]

    ws6.cell(row=r,column=1,value=f"nrv_source = {src_name}").font=_fn(NAVY,bold=True,sz=11)
    ws6.row_dimensions[r].height=22; r+=1
    wh(ws6, r, ["SKU","Mat Group","Reference Price","NRV Price","Deduction (NRV − Ref)"], GREEN)
    r+=1
    for ri, row_d in enumerate(sub.head(15).itertuples(index=False), start=r):
        wr(ws6, ri, [
            row_d.sku, row_d.mat_group, row_d.ref_price, row_d.nrv_price, row_d.deduction,
        ], bg=LGREEN if abs(row_d.deduction + SELLING_COST) < 0.001 else LRED,
           fmts={3:"#,##0.0000",4:"#,##0.0000",5:"#,##0.0000"})
    r = ri + 2

    # Stats
    stats = sub["deduction"].describe()
    ws6.cell(row=r,column=1,value=f"Stats ({src_name}):").font=_fn(bold=True)
    for si, (stat_name, stat_val) in enumerate(stats.items(), start=r+1):
        ws6.cell(row=si, column=1, value=stat_name)
        ws6.cell(row=si, column=2, value=round(stat_val, 6))
        ws6.row_dimensions[si].height=18
    r = si + 2

aw(ws6)

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
print("[3/3] Saving...")
wb.save(OUTPUT)
print(f"\n  Output : {OUTPUT}")
print(f"  Sheets : {[s.title for s in wb.worksheets]}")
print()
print("=" * 58)
print("  NRV PRICE METHODOLOGY SUMMARY")
print("=" * 58)
print(f"  Formula: NRV = Reference Price - 0.3255 THB/KG")
print(f"  Selling Cost (deduction) : 0.3255 THB/KG (fixed for all SKUs)")
print()
for _, row in src_counts[src_counts["sku_type"]=="FG"].sort_values("sku_count",ascending=False).iterrows():
    print(f"  {row['nrv_source']:<26}: {int(row['sku_count']):>5,} SKUs")
print(f"  {'RM (all sources)':<26}: {len(rm):>5,} SKUs")
print("=" * 58)
