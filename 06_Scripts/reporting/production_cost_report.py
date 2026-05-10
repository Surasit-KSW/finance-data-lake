"""
production_cost_report.py
=========================
Generate Production Cost Excel Report 2025
- DM (Direct Material) = Actual GI - By-Product recovered
- Conversion Cost = Direct (D101+D102) + Indirect (I101-I111)

Output: 04_Reports/Production_Cost_2025.xlsx
  Sheet 1: SKU_Detail
  Sheet 2: MatGroup_Summary
  Sheet 3: Monthly_Trend
"""

import os
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 0. Config
# ─────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
INPUT_FILE  = os.path.join(BASE_DIR, "02_Silver_Cleaned", "master_production_2025.parquet")
OUTPUT_DIR  = os.path.join(BASE_DIR, "04_Reports")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Production_Cost_2025.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# 1. Load Data
# ─────────────────────────────────────────────
print("[1/5] Loading parquet file...")
df = pd.read_parquet(INPUT_FILE, engine="pyarrow")
print(f"      Rows: {len(df):,}  Cols: {len(df.columns)}")

# ─────────────────────────────────────────────
# 2. Rename garbled Thai columns (by position)
# ─────────────────────────────────────────────
cols = list(df.columns)
rename_map = {
    cols[6]:  "Mat_Group_Name",    # e.g. Slit GIM, Pipe Rect. GI
    cols[7]:  "Mat_Type_Code",     # numeric type code
    cols[10]: "Mat_Group_1_Name",  # e.g. Slit Coil, Pipe, Section
    cols[11]: "Mat_Group_2_Name",  # shape group (รูปทรง)
    cols[12]: "Mat_Group_3_Name",  # size (inch)
    cols[13]: "Mat_Group_4_Name",  # thickness
}
df.rename(columns=rename_map, inplace=True)

# Extract Month from Source_File (format: 1100_2025_01.XLSX)
df["Month"] = df["Source_File"].str.extract(r"_(\d{2})\.XLSX$").astype(float).astype("Int64")

# ─────────────────────────────────────────────
# 3. Calculate Cost Components
# ─────────────────────────────────────────────
print("[2/5] Calculating cost components...")

# DM = Actual GI Amount - By-Product recovered (Scrap + Grade B + Grade C)
df["DM_GI_Amount"]     = pd.to_numeric(df["Actual GI Amount"],                errors="coerce").fillna(0)
df["DM_ByProd_Scrap"]  = pd.to_numeric(df["Actual ByProduct Scrap Amount"],   errors="coerce").fillna(0)
df["DM_ByProd_GradeB"] = pd.to_numeric(df["Actual ByProduct Grade B Amount"], errors="coerce").fillna(0)
df["DM_ByProd_GradeC"] = pd.to_numeric(df["Actual ByProduct Grade C Amount"], errors="coerce").fillna(0)
df["DM_ByProd_Total"]  = df["DM_ByProd_Scrap"] + df["DM_ByProd_GradeB"] + df["DM_ByProd_GradeC"]
df["DM_Amount"]        = df["DM_GI_Amount"] - df["DM_ByProd_Total"]

# Conversion = Direct (D101+D102) + Indirect (I101-I111, skip I106 = always 0)
direct_cols = ["Actual D101 Amount", "Actual D102 Amount"]
indirect_cols = [
    "Actual I101 Amount", "Actual I102 Amount", "Actual I103 Amount",
    "Actual I104 Amount", "Actual I105 Amount", "Actual I107 Amount",
    "Actual I108 Amount", "Actual I109 Amount", "Actual I110 Amount",
    "Actual I111 Amount",
]
for c in direct_cols + indirect_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

df["Direct_Amount"]     = df[direct_cols].sum(axis=1)
df["Indirect_Amount"]   = df[indirect_cols].sum(axis=1)
df["Conversion_Amount"] = df["Direct_Amount"] + df["Indirect_Amount"]
df["Total_Cost"]        = df["DM_Amount"] + df["Conversion_Amount"]

# QTY
df["GR_QTY"]        = pd.to_numeric(df["Actual GR QTY"],                errors="coerce").fillna(0)
df["GR_QTY_GradeB"] = pd.to_numeric(df["Actual ByProduct Grade B QTY"], errors="coerce").fillna(0)
df["GR_QTY_GradeC"] = pd.to_numeric(df["Actual ByProduct Grade C QTY"], errors="coerce").fillna(0)
df["GI_QTY"]        = pd.to_numeric(df["Actual GI QTY"],                errors="coerce").fillna(0)

# Total output = Grade A + B + C (ไม่รวม Scrap เพราะเป็น waste)
df["Total_GR_QTY"]  = df["GR_QTY"] + df["GR_QTY_GradeB"] + df["GR_QTY_GradeC"]

# Dimension cleanse
df["Plant"]               = df["Plant"].astype(str).str.strip()
df["Mat_Group_Name"]      = df["Mat_Group_Name"].fillna("Unknown").astype(str)
df["Mat_Group_1_Name"]    = df["Mat_Group_1_Name"].fillna("Unknown").astype(str)
df["Mat_Group_2_Name"]    = df["Mat_Group_2_Name"].fillna("").astype(str)
df["Mat_Group_3_Name"]    = df["Mat_Group_3_Name"].fillna("").astype(str)
df["Mat_Group_4_Name"]    = df["Mat_Group_4_Name"].fillna("").astype(str)
df["Grade"]               = df["Grade"].fillna("").astype(str)
df["Work Center"]         = df["Work Center"].fillna("").astype(str)
df["Work Center Description"] = df["Work Center Description"].fillna("").astype(str)

# ─────────────────────────────────────────────
# 4a. Sheet 1 – SKU Detail
# ─────────────────────────────────────────────
print("[3/5] Building SKU Detail sheet...")

sku_dims = [
    "Plant", "Material", "Description (EN)",
    "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
    "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade",
    "Work Center", "Work Center Description",
]

sku_agg = df.groupby(sku_dims, dropna=False).agg(
    GI_QTY        = ("GI_QTY",           "sum"),
    GR_QTY        = ("GR_QTY",           "sum"),
    GR_QTY_GradeB = ("GR_QTY_GradeB",   "sum"),
    GR_QTY_GradeC = ("GR_QTY_GradeC",   "sum"),
    Total_GR_QTY  = ("Total_GR_QTY",    "sum"),
    DM_GI         = ("DM_GI_Amount",     "sum"),
    DM_ByProd     = ("DM_ByProd_Total",  "sum"),
    DM_Amount     = ("DM_Amount",        "sum"),
    Direct        = ("Direct_Amount",    "sum"),
    Indirect      = ("Indirect_Amount",  "sum"),
    Conversion    = ("Conversion_Amount","sum"),
    Total_Cost    = ("Total_Cost",       "sum"),
).reset_index()

sku_agg["Cost_per_KG"]       = np.where(sku_agg["Total_GR_QTY"] > 0, sku_agg["Total_Cost"] / sku_agg["Total_GR_QTY"], 0)
sku_agg["Conversion_per_KG"] = np.where(sku_agg["Total_GR_QTY"] > 0, sku_agg["Conversion"] / sku_agg["Total_GR_QTY"], 0)
sku_agg["DM_pct"]            = np.where(sku_agg["Total_Cost"] > 0, sku_agg["DM_Amount"]  / sku_agg["Total_Cost"], 0)
sku_agg["Conversion_pct"]    = np.where(sku_agg["Total_Cost"] > 0, sku_agg["Conversion"] / sku_agg["Total_Cost"], 0)
sku_agg = sku_agg.sort_values(["Plant", "Total_Cost"], ascending=[True, False])

sku_agg = sku_agg[[
    "Plant", "Material", "Description (EN)",
    "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
    "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade",
    "Work Center", "Work Center Description",
    "GI_QTY", "GR_QTY", "GR_QTY_GradeB", "GR_QTY_GradeC", "Total_GR_QTY",
    "DM_GI", "DM_ByProd", "DM_Amount",
    "Direct", "Indirect", "Conversion",
    "Total_Cost", "Cost_per_KG", "Conversion_per_KG",
    "DM_pct", "Conversion_pct",
]]

# ─────────────────────────────────────────────
# 4b. Sheet 2 – Mat Group Summary
# ─────────────────────────────────────────────
print("[4/5] Building Mat Group Summary sheet...")

mg_dims = ["Plant", "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name", "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade"]

mg_agg = df.groupby(mg_dims, dropna=False).agg(
    SKU_Count     = ("Material",          "nunique"),
    GR_QTY        = ("GR_QTY",           "sum"),
    GR_QTY_GradeB = ("GR_QTY_GradeB",   "sum"),
    GR_QTY_GradeC = ("GR_QTY_GradeC",   "sum"),
    Total_GR_QTY  = ("Total_GR_QTY",    "sum"),
    DM_GI         = ("DM_GI_Amount",     "sum"),
    DM_ByProd     = ("DM_ByProd_Total",  "sum"),
    DM_Amount     = ("DM_Amount",        "sum"),
    Direct        = ("Direct_Amount",    "sum"),
    Indirect      = ("Indirect_Amount",  "sum"),
    Conversion    = ("Conversion_Amount","sum"),
    Total_Cost    = ("Total_Cost",       "sum"),
).reset_index()

mg_agg["Cost_per_KG"]    = np.where(mg_agg["Total_GR_QTY"] > 0, mg_agg["Total_Cost"] / mg_agg["Total_GR_QTY"], 0)
mg_agg["DM_pct"]         = np.where(mg_agg["Total_Cost"] > 0, mg_agg["DM_Amount"]  / mg_agg["Total_Cost"], 0)
mg_agg["Conversion_pct"] = np.where(mg_agg["Total_Cost"] > 0, mg_agg["Conversion"] / mg_agg["Total_Cost"], 0)
mg_agg = mg_agg.sort_values(["Plant", "Total_Cost"], ascending=[True, False])

mg_agg = mg_agg[[
    "Plant", "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
    "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade", "SKU_Count",
    "GR_QTY", "GR_QTY_GradeB", "GR_QTY_GradeC", "Total_GR_QTY",
    "DM_GI", "DM_ByProd", "DM_Amount",
    "Direct", "Indirect", "Conversion",
    "Total_Cost", "Cost_per_KG",
    "DM_pct", "Conversion_pct",
]]

# ─────────────────────────────────────────────
# 4c. Sheet 3 – Unit Cost Summary (Cost per KG)
# ─────────────────────────────────────────────
# Section A: by Mat Group  (aggregate)
# Section B: by SKU top-N  (detail, sorted by Cost_per_KG desc)

uc_mg = mg_agg[["Plant", "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
                  "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade",
                  "SKU_Count", "GR_QTY", "GR_QTY_GradeB", "GR_QTY_GradeC", "Total_GR_QTY",
                  "DM_Amount", "Conversion", "Total_Cost",
                  "Cost_per_KG", "DM_pct", "Conversion_pct"]].copy()

# DM per KG and Conversion per KG — denominator = Total_GR_QTY (A+B+C)
uc_mg["DM_per_KG"]         = np.where(uc_mg["Total_GR_QTY"] > 0, uc_mg["DM_Amount"]  / uc_mg["Total_GR_QTY"], 0)
uc_mg["Conversion_per_KG"] = np.where(uc_mg["Total_GR_QTY"] > 0, uc_mg["Conversion"] / uc_mg["Total_GR_QTY"], 0)
uc_mg = uc_mg.sort_values(["Plant", "Cost_per_KG"], ascending=[True, False])
uc_mg = uc_mg[[
    "Plant", "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
    "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade", "SKU_Count",
    "GR_QTY", "GR_QTY_GradeB", "GR_QTY_GradeC", "Total_GR_QTY",
    "DM_per_KG", "Conversion_per_KG", "Cost_per_KG",
    "DM_pct", "Conversion_pct",
]]

# SKU-level unit cost (filter Total_GR_QTY > 0)
uc_sku = sku_agg[sku_agg["Total_GR_QTY"] > 0][[
    "Plant", "Material", "Description (EN)",
    "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
    "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade",
    "Work Center",
    "GR_QTY", "GR_QTY_GradeB", "GR_QTY_GradeC", "Total_GR_QTY",
    "DM_Amount", "Conversion", "Total_Cost",
    "Cost_per_KG", "DM_pct", "Conversion_pct",
]].copy()
uc_sku["DM_per_KG"]         = uc_sku["DM_Amount"]  / uc_sku["Total_GR_QTY"]
uc_sku["Conversion_per_KG"] = uc_sku["Conversion"] / uc_sku["Total_GR_QTY"]
uc_sku = uc_sku.sort_values(["Plant", "Cost_per_KG"], ascending=[True, False])
uc_sku = uc_sku[[
    "Plant", "Material", "Description (EN)",
    "Mat_Group_Name", "Mat_Group_1_Name", "Mat_Group_2_Name",
    "Mat_Group_3_Name", "Mat_Group_4_Name", "Grade",
    "Work Center",
    "GR_QTY", "GR_QTY_GradeB", "GR_QTY_GradeC", "Total_GR_QTY",
    "DM_per_KG", "Conversion_per_KG", "Cost_per_KG",
    "DM_pct", "Conversion_pct",
]]

# ─────────────────────────────────────────────
# 4d. Sheet 4 – Monthly Trend (Total Cost by Mat Group)
# ─────────────────────────────────────────────
trend_pivot = df.groupby(["Plant", "Mat_Group_Name", "Month"], dropna=False)["Total_Cost"].sum().reset_index()
trend_pivot = trend_pivot.pivot_table(
    index   = ["Plant", "Mat_Group_Name"],
    columns = "Month",
    values  = "Total_Cost",
    aggfunc = "sum",
    fill_value = 0,
).reset_index()
trend_pivot.columns.name = None

# Ensure month columns are ints
month_cols = [c for c in trend_pivot.columns if isinstance(c, (int, float)) or str(c).isdigit()]
trend_pivot["Grand_Total"] = trend_pivot[month_cols].sum(axis=1)
trend_pivot = trend_pivot.sort_values(["Plant", "Grand_Total"], ascending=[True, False])

# ─────────────────────────────────────────────
# 4e. Sheet 5 – High Conversion Analysis
# ─────────────────────────────────────────────
print("[4e/5] Building High Conversion Analysis sheet...")

# --- Per-cost-code aggregation at SKU level ---
cost_codes = ["D101", "D102", "I101", "I102", "I103", "I104", "I105",
              "I107", "I108", "I109", "I110", "I111"]
code_agg_spec = {f"Actual {c} Amount": "sum" for c in cost_codes}
code_agg_spec["Total_GR_QTY"] = "sum"

code_detail = df.groupby(sku_dims, dropna=False).agg(**{
    **{c: (f"Actual {c} Amount", "sum") for c in cost_codes},
    "Total_GR_QTY_chk": ("Total_GR_QTY", "sum"),
}).reset_index()

# --- Start from sku_agg, filter QTY > 0 ---
hc = sku_agg[sku_agg["Total_GR_QTY"] > 0].copy()

# --- IQR-based outlier threshold per Plant ---
def iqr_upper(s):
    q1, q3 = s.quantile(0.25), s.quantile(0.75)
    return q3 + 1.5 * (q3 - q1)

thresholds = (
    hc.groupby("Plant")["Conversion_per_KG"]
    .apply(iqr_upper)
    .rename("Conv_Threshold")
)
plant_avg_conv = (
    hc.groupby("Plant")["Conversion_per_KG"]
    .mean()
    .rename("Plant_Avg_Conv_KG")
)
low_vol_q25 = (
    hc.groupby("Plant")["Total_GR_QTY"]
    .quantile(0.25)
    .rename("Low_Vol_Q25")
)

hc = hc.join(thresholds,    on="Plant")
hc = hc.join(plant_avg_conv, on="Plant")
hc = hc.join(low_vol_q25,   on="Plant")

# --- Keep only outliers ---
hc = hc[hc["Conversion_per_KG"] > hc["Conv_Threshold"]].copy()

# --- Merge per-code detail ---
hc = hc.merge(code_detail.drop(columns=["Total_GR_QTY_chk"]), on=sku_dims, how="left")

# Per KG for each cost code
for c in cost_codes:
    hc[f"{c}_per_KG"] = np.where(hc["Total_GR_QTY"] > 0, hc[c] / hc["Total_GR_QTY"], 0)
hc.drop(columns=cost_codes, inplace=True)

# Direct / Indirect per KG
hc["Direct_per_KG"]   = np.where(hc["Total_GR_QTY"] > 0, hc["Direct"]   / hc["Total_GR_QTY"], 0)
hc["Indirect_per_KG"] = np.where(hc["Total_GR_QTY"] > 0, hc["Indirect"] / hc["Total_GR_QTY"], 0)

# --- Analysis columns ---
hc["Conv_vs_Avg_pct"] = np.where(
    hc["Plant_Avg_Conv_KG"] > 0,
    (hc["Conversion_per_KG"] - hc["Plant_Avg_Conv_KG"]) / hc["Plant_Avg_Conv_KG"],
    0,
)
hc["Low_Volume_Flag"] = np.where(hc["Total_GR_QTY"] < hc["Low_Vol_Q25"], "YES", "")

# Dominant cost driver (per KG columns)
driver_cols = [f"{c}_per_KG" for c in cost_codes]
hc["Dominant_Driver"] = hc[driver_cols].idxmax(axis=1).str.replace("_per_KG", "")

# Human-readable cause
_D_LABELS = {"D101": "Direct Labor (D101)", "D102": "Direct Machine (D102)"}
_I_LABELS = {
    "I101": "Indirect Labor (I101)",     "I102": "Depreciation (I102)",
    "I103": "Repair/Maint (I103)",       "I104": "Utility (I104)",
    "I105": "Consumable (I105)",         "I107": "Other Indirect (I107)",
    "I108": "Overhead Alloc (I108)",     "I109": "Overhead Alloc (I109)",
    "I110": "Overhead Alloc (I110)",     "I111": "Overhead Alloc (I111)",
}

def explain_cause(row):
    parts = []
    if row["Low_Volume_Flag"] == "YES":
        parts.append("Low Volume — fixed cost dilution")
    driver = row["Dominant_Driver"]
    label  = _D_LABELS.get(driver) or _I_LABELS.get(driver) or driver
    parts.append(f"Dominant: {label}")
    pct = row["Conv_vs_Avg_pct"]
    if pct > 3:
        parts.append(f"Extreme (+{pct:.0%} vs plant avg)")
    elif pct > 1:
        parts.append(f"High (+{pct:.0%} vs plant avg)")
    return " | ".join(parts)

hc["Possible_Cause"] = hc.apply(explain_cause, axis=1)

# Drop helper columns
hc.drop(columns=["Low_Vol_Q25", "Plant_Avg_Conv_KG"], inplace=True)
hc = hc.sort_values(["Plant", "Conversion_per_KG"], ascending=[True, False])

# Final column order
hc_cols = [
    "Plant", "Material", "Description (EN)",
    "Mat_Group_Name", "Mat_Group_1_Name", "Grade", "Work Center",
    "Total_GR_QTY", "Conversion", "Conversion_per_KG",
    "Conv_Threshold", "Conv_vs_Avg_pct",
    "Direct_per_KG", "Indirect_per_KG",
] + driver_cols + [
    "Low_Volume_Flag", "Dominant_Driver", "Possible_Cause",
]
hc = hc[[c for c in hc_cols if c in hc.columns]]

# ─────────────────────────────────────────────
# 5. Export to Excel + Formatting
# ─────────────────────────────────────────────
print("[5/5] Writing Excel file...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    sku_agg.to_excel(    writer, sheet_name="SKU_Detail",          index=False)
    mg_agg.to_excel(     writer, sheet_name="MatGroup_Summary",    index=False)
    trend_pivot.to_excel(writer, sheet_name="Monthly_Trend",       index=False)
    hc.to_excel(         writer, sheet_name="High_Conv_Analysis",  index=False)

    # Unit Cost sheet: Mat Group section + blank row + SKU section
    row_offset = 0
    uc_mg.to_excel(writer, sheet_name="UnitCost_Summary", index=False, startrow=row_offset)
    row_offset += len(uc_mg) + 3  # skip 1 blank row after header+data

    # Section header for SKU block
    ws_uc = writer.sheets["UnitCost_Summary"]
    ws_uc.cell(row=row_offset + 1, column=1).value = "--- Cost per KG by SKU (sorted highest first) ---"

    uc_sku.to_excel(writer, sheet_name="UnitCost_Summary", index=False, startrow=row_offset + 1)

# --- Apply Formatting ---
def style_sheet(ws, df_ref, header_hex="1F4E79"):
    hdr_fill  = PatternFill("solid", fgColor=header_hex)
    hdr_font  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    center_a  = Alignment(horizontal="center", vertical="center")
    left_a    = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="D9D9D9")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    NUM = "#,##0"
    DEC = "#,##0.00"
    PCT = "0.00%"

    # Determine format per column index
    col_fmts = {}
    for i, col in enumerate(df_ref.columns, start=1):
        col_str = str(col).lower()
        if "pct" in col_str or col_str.endswith("%"):
            col_fmts[i] = PCT
        elif df_ref[col].dtype == object:
            col_fmts[i] = None
        elif "per_kg" in col_str or "thickness" in col_str:
            col_fmts[i] = DEC
        else:
            col_fmts[i] = NUM

    max_row = ws.max_row
    max_col = ws.max_column

    # Header
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center_a
        cell.border    = bdr

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font   = body_font
            cell.border = bdr
            fmt = col_fmts.get(cell.column)
            if fmt:
                cell.number_format = fmt
                cell.alignment     = center_a
            else:
                cell.alignment = left_a

    # Auto column width
    for ci in range(1, max_col + 1):
        letter  = get_column_letter(ci)
        max_len = max(
            (len(str(cell.value or "")) for row in ws.iter_rows(min_col=ci, max_col=ci) for cell in row),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)

    ws.freeze_panes = "A2"


def add_grand_total_row(ws, start_col, header_hex="FFD966"):
    tr = ws.max_row + 1
    yellow_fill = PatternFill("solid", fgColor=header_hex)
    bold_font   = Font(bold=True, size=10, name="Calibri")
    thin        = Side(style="thin", color="999999")
    bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci in range(1, ws.max_column + 1):
        cell        = ws.cell(row=tr, column=ci)
        cell.fill   = yellow_fill
        cell.font   = bold_font
        cell.border = bdr
        if ci == 1:
            cell.value     = "GRAND TOTAL"
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif ci >= start_col:
            letter             = get_column_letter(ci)
            cell.value         = f"=SUM({letter}2:{letter}{tr-1})"
            cell.number_format = ws.cell(row=2, column=ci).number_format
            cell.alignment     = Alignment(horizontal="center", vertical="center")


wb = load_workbook(OUTPUT_FILE)

style_sheet(wb["SKU_Detail"],          sku_agg,     header_hex="1F4E79")  # Navy
style_sheet(wb["MatGroup_Summary"],    mg_agg,      header_hex="375623")  # Forest Green
style_sheet(wb["Monthly_Trend"],       trend_pivot, header_hex="7B2D00")  # Dark Orange
style_sheet(wb["High_Conv_Analysis"],  hc,          header_hex="7B241C")  # Dark Red

# Highlight Possible_Cause column in High_Conv_Analysis
ws_hc = wb["High_Conv_Analysis"]
cause_col_idx = list(hc.columns).index("Possible_Cause") + 1
orange_fill = PatternFill("solid", fgColor="FCE4D6")
bold_font_hc = Font(bold=True, size=10, name="Calibri")
for r in range(2, ws_hc.max_row + 1):
    cell = ws_hc.cell(row=r, column=cause_col_idx)
    cell.fill = orange_fill
    cell.font = bold_font_hc

# Conv_vs_Avg_pct column — color scale: green near 0, red if very high
pct_col_idx = list(hc.columns).index("Conv_vs_Avg_pct") + 1
red_fill    = PatternFill("solid", fgColor="FF9999")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
for r in range(2, ws_hc.max_row + 1):
    cell = ws_hc.cell(row=r, column=pct_col_idx)
    v = cell.value or 0
    if isinstance(v, (int, float)):
        if v > 2:
            cell.fill = red_fill
        elif v > 1:
            cell.fill = yellow_fill

# Grand total only on MatGroup sheet (numeric start = col 6 = GR_QTY)
add_grand_total_row(wb["MatGroup_Summary"], start_col=6)

# Style UnitCost_Summary: two sections with separate headers
ws_uc = wb["UnitCost_Summary"]

# Section 1 header: row 1 (Mat Group)
mg_header_row = 1
sku_header_row = len(uc_mg) + 3 + 2  # blank + section label + header

hdr_fill_uc   = PatternFill("solid", fgColor="4A235A")   # Purple
hdr_fill_sku  = PatternFill("solid", fgColor="1A5276")   # Steel Blue
hdr_font_w    = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
body_font_uc  = Font(size=10, name="Calibri")
center_a      = Alignment(horizontal="center", vertical="center")
left_a        = Alignment(horizontal="left",   vertical="center")
thin          = Side(style="thin", color="D9D9D9")
bdr           = Border(left=thin, right=thin, top=thin, bottom=thin)

DEC = "#,##0.00"
PCT = "0.00%"
NUM = "#,##0"

def get_uc_fmt(col_str):
    col_str = col_str.lower()
    if "pct" in col_str:
        return PCT
    if "per_kg" in col_str or "thickness" in col_str:
        return DEC
    if "qty" in col_str or "count" in col_str or "sku" in col_str:
        return NUM
    return None  # text

def style_uc_section(ws, header_row, data_row_start, data_row_end, col_headers, fill):
    # Header row
    for ci, col in enumerate(col_headers, start=1):
        cell = ws.cell(row=header_row, column=ci)
        cell.fill      = fill
        cell.font      = hdr_font_w
        cell.alignment = center_a
        cell.border    = bdr
    # Data rows
    for r in range(data_row_start, data_row_end + 1):
        for ci, col in enumerate(col_headers, start=1):
            cell = ws.cell(row=r, column=ci)
            cell.font   = body_font_uc
            cell.border = bdr
            fmt = get_uc_fmt(col)
            if fmt:
                cell.number_format = fmt
                cell.alignment     = center_a
            else:
                cell.alignment = left_a

style_uc_section(
    ws_uc,
    header_row     = mg_header_row,
    data_row_start = mg_header_row + 1,
    data_row_end   = mg_header_row + len(uc_mg),
    col_headers    = list(uc_mg.columns),
    fill           = hdr_fill_uc,
)

style_uc_section(
    ws_uc,
    header_row     = sku_header_row,
    data_row_start = sku_header_row + 1,
    data_row_end   = sku_header_row + len(uc_sku),
    col_headers    = list(uc_sku.columns),
    fill           = hdr_fill_sku,
)

# Section label styling
label_row = len(uc_mg) + 3 + 1
label_cell = ws_uc.cell(row=label_row, column=1)
label_cell.font  = Font(bold=True, italic=True, size=10, color="4A235A", name="Calibri")
label_cell.value = "--- Cost per KG by SKU (sorted highest first) ---"

# Auto width
max_col_uc = max(len(uc_mg.columns), len(uc_sku.columns))
for ci in range(1, max_col_uc + 1):
    letter  = get_column_letter(ci)
    max_len = max(
        (len(str(ws_uc.cell(row=r, column=ci).value or ""))
         for r in range(1, ws_uc.max_row + 1)),
        default=10,
    )
    ws_uc.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)

ws_uc.freeze_panes = "A2"

wb.save(OUTPUT_FILE)

# ─────────────────────────────────────────────
# Summary
# ─────────────────────────────────────────────
total_dm   = sku_agg["DM_Amount"].sum()
total_conv = sku_agg["Conversion"].sum()
total_all  = sku_agg["Total_Cost"].sum()
total_qty  = sku_agg["Total_GR_QTY"].sum()

print()
print("=" * 56)
print("  Production Cost Summary 2025")
print("=" * 56)
print(f"  GR Output QTY     : {total_qty:>16,.0f}  KG")
print(f"  DM Amount         : {total_dm:>16,.0f}  THB  ({total_dm/total_all*100:.1f}%)")
print(f"  Conversion Amount : {total_conv:>16,.0f}  THB  ({total_conv/total_all*100:.1f}%)")
print(f"  Total Cost        : {total_all:>16,.0f}  THB")
print(f"  Avg Cost/KG       : {total_all/total_qty:>16,.2f}  THB/KG")
print("=" * 56)
print(f"  SKU Count         : {sku_agg['Material'].nunique():,}")
print(f"  Mat Group Count   : {mg_agg['Mat_Group_Name'].nunique():,}")
print(f"  Plants            : {', '.join(sorted(df['Plant'].unique()))}")
print("=" * 56)
print(f"\n  Output: {OUTPUT_FILE}")
