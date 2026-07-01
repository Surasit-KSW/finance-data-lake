"""
build_q1_2026_leadsheet.py
==========================
Builds the AMC Q1'2026 Leadsheet (STAT to client) in Excel format.

Starting point : leadsheet_q2_2026_stat.xlsx
                 (already contains TB Mar 26 + Master TB with Mar 2026 balances)

What this script adds:
  1. Master TB col K  -- Dec 2025 balances (from YE25 leadsheet)
  2. "TB Dec 25" sheet -- SAP-style TB at 31 Dec 2025
  3. "SAM Q1'26" sheet -- empty SAM template for Audit Adjustments
  4. Renames the file correctly to Q1'2026

Output:
  01_Bronze_Raw/templates/leadsheet_q1_2026_stat.xlsx

Usage:
    python build_q1_2026_leadsheet.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import shutil

# --- Paths ------------------------------------------------------------------
BASE_DIR  = Path(__file__).parent.parent.parent   # _Finance_Data_Lake root
TEMPLATES = BASE_DIR / "01_Bronze_Raw" / "templates"
SCRIPTS   = BASE_DIR / "06_Scripts" / "leadsheet"
OUTPUT    = SCRIPTS / "output"

SOURCE_FILE  = TEMPLATES / "leadsheet_q1_2025_stat.xlsx"
OUTPUT_FILE  = TEMPLATES / "leadsheet_q1_2026_stat.xlsx"
DEC25_CSV    = OUTPUT / "tb_dec2025.csv"
MAR26_CSV    = OUTPUT / "tb_mar2026.csv"

# Column indices in Master TB (1-based) — Q1 2025 source file layout:
#   A=AcctCode, B=Name, C=INDEX, D=Caption
#   E=Sep2024(HardClose), F=Mar2025, G=DR, H=CR, I=Adj, J=Dec2024
#   K=empty <-- NEW Dec 2025
#   L=empty <-- NEW Mar 2026
COL_ACCT   = 1   # A - Account Code
COL_NAME   = 2   # B - Account Name
COL_INDEX  = 3   # C - INDEX (schedule reference)
COL_CAP    = 4   # D - Caption
COL_SEP24  = 5   # E - Sep 2024 (Hard Close)
COL_MAR25  = 6   # F - Mar 2025 (existing Q1 2025 balance)
COL_DR     = 7   # G - DR adjustments
COL_CR     = 8   # H - CR adjustments
COL_ADJ    = 9   # I - Adjusted balance
COL_DEC24  = 10  # J - Dec 2024 balance
COL_DEC25  = 11  # K - Dec 2025 balance  <-- NEW
COL_MAR26  = 12  # L - Mar 2026 balance  <-- NEW

MASTER_TB_DATA_START = 4   # first data row in Master TB (rows 1-3 = headers)

DATE_DEC25 = datetime(2025, 12, 31)
DATE_MAR26 = datetime(2026, 3, 31)


# --- Load TB data -----------------------------------------------------------
def load_tb_data() -> tuple[dict, dict]:
    """Return (dec25_map, mar26_map) keyed by account_code string."""
    df_dec = pd.read_csv(DEC25_CSV, dtype={"account_code": str})
    df_mar = pd.read_csv(MAR26_CSV, dtype={"account_code": str})
    dec25_map = dict(zip(df_dec["account_code"], df_dec["balance_dec2025"]))
    mar26_map = dict(zip(df_mar["account_code"], df_mar["balance_mar2026"]))
    return dec25_map, mar26_map


# --- Helpers ----------------------------------------------------------------
def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color="D9E1F2"):
    return PatternFill("solid", fgColor=color)


def _bold(size=10):
    return Font(bold=True, size=size)


# --- Step 1: Copy source file -----------------------------------------------
def copy_source():
    print(f"Copying base file...")
    shutil.copy2(SOURCE_FILE, OUTPUT_FILE)
    print(f"  -> {OUTPUT_FILE.name}")


# --- Step 2: Add new balance columns to Master TB ---------------------------
def _write_balance_column(ws, col_idx: int, label: str, date_val, fill_color: str,
                          balance_map: dict, data_start: int) -> tuple[int, int]:
    """Write a balance column (header + data) to Master TB. Returns (matched, zero)."""
    ws.cell(row=2, column=col_idx).value = label
    ws.cell(row=2, column=col_idx).font = _bold()
    ws.cell(row=2, column=col_idx).fill = _header_fill(fill_color)
    ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=col_idx).value = date_val
    ws.cell(row=3, column=col_idx).number_format = "DD/MM/YYYY"
    ws.cell(row=3, column=col_idx).font = _bold()
    ws.cell(row=3, column=col_idx).fill = _header_fill(fill_color)
    ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal="center")

    matched = zero = 0
    for row_idx in range(data_start, ws.max_row + 1):
        raw_acct = ws.cell(row=row_idx, column=COL_ACCT).value
        if raw_acct is None:
            continue
        try:
            acct = str(int(float(str(raw_acct)))).strip()
        except (ValueError, TypeError):
            continue
        balance = balance_map.get(acct, 0.0)
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = float(balance) if balance else 0.0
        cell.number_format = "#,##0.00;[Red]-#,##0.00"
        cell.alignment = Alignment(horizontal="right")
        if balance:
            matched += 1
        else:
            zero += 1
    return matched, zero


def add_new_columns_to_master_tb(wb: openpyxl.Workbook, dec25_map: dict, mar26_map: dict) -> None:
    """
    Appends two columns to Master TB:
      K = Dec 2025 (YE 2025 balance)
      L = Mar 2026 (Q1 2026 balance)
    """
    print("Adding Dec 2025 and Mar 2026 columns to Master TB...")
    ws = wb["Master TB"]

    m1, z1 = _write_balance_column(ws, COL_DEC25, "YE2025", DATE_DEC25, "FCE4D6",
                                    dec25_map, MASTER_TB_DATA_START)
    print(f"  Col K (Dec 2025): {m1} with balance, {z1} zero/missing")

    m2, z2 = _write_balance_column(ws, COL_MAR26, "Q1FY26", DATE_MAR26, "E2EFDA",
                                    mar26_map, MASTER_TB_DATA_START)
    print(f"  Col L (Mar 2026): {m2} with balance, {z2} zero/missing")

    # Add note row 1 col K-L
    ws.cell(row=1, column=COL_DEC25).value = "NEW - Dec 2025"
    ws.cell(row=1, column=COL_MAR26).value = "NEW - Mar 2026"
    for col in (COL_DEC25, COL_MAR26):
        ws.cell(row=1, column=col).font = Font(bold=True, color="FF0000", size=9)

    # Flag new accounts in Mar 2026 not in Master TB
    master_accts = set()
    for row_idx in range(MASTER_TB_DATA_START, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=COL_ACCT).value
        if raw is not None:
            try:
                master_accts.add(str(int(float(str(raw)))).strip())
            except (ValueError, TypeError):
                pass

    new_accts = set(mar26_map.keys()) - master_accts
    if new_accts:
        print(f"  WARNING: {len(new_accts)} accounts in Mar 2026 not in Master TB:")
        for a in sorted(new_accts):
            print(f"    {a}")


# --- Step 3: Create new TB sheets -------------------------------------------
def _get_master_accts_and_names(ws_master) -> tuple[list, dict]:
    """Extract ordered account list and name map from Master TB."""
    master_accts = []
    acct_names = {}
    for row_idx in range(MASTER_TB_DATA_START, ws_master.max_row + 1):
        raw_acct = ws_master.cell(row=row_idx, column=COL_ACCT).value
        raw_name = ws_master.cell(row=row_idx, column=COL_NAME).value
        if raw_acct is not None:
            try:
                acct = str(int(float(str(raw_acct)))).strip()
                master_accts.append(acct)
                if raw_name:
                    acct_names[acct] = str(raw_name).strip()
            except (ValueError, TypeError):
                pass
    return master_accts, acct_names


def _create_tb_sheet(wb: openpyxl.Workbook, sheet_name: str, col_label: str,
                     balance_map: dict, insert_before: str) -> None:
    """
    Generic function to create a TB sheet:
      Col A: Account Text (code + name)
      Col B: Account Number
      Col C: Balance (from balance_map)
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    insert_idx = wb.sheetnames.index(insert_before)
    ws_new = wb.create_sheet(sheet_name, insert_idx)

    # Header
    headers = ["Text for B/S P&L Item", "Account Number", col_label]
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col_idx)
        cell.value = hdr
        cell.font = _bold()
        cell.fill = _header_fill("D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    ws_master = wb["Master TB"]
    master_accts, acct_names = _get_master_accts_and_names(ws_master)
    master_set = set(master_accts)

    row_out = 2
    written = 0
    for acct in master_accts:
        balance = balance_map.get(acct, 0.0)
        name = acct_names.get(acct, acct)
        ws_new.cell(row=row_out, column=1).value = f"{acct} {name}"
        ws_new.cell(row=row_out, column=2).value = int(acct)
        ws_new.cell(row=row_out, column=3).value = float(balance) if balance else 0.0
        ws_new.cell(row=row_out, column=3).number_format = "#,##0.00;[Red]-#,##0.00"
        row_out += 1
        written += 1

    # Append any extra accounts from balance_map not in Master TB
    for acct in sorted(balance_map.keys()):
        if acct not in master_set and balance_map[acct]:
            name = acct
            ws_new.cell(row=row_out, column=1).value = f"{acct} {name}  [NEW - not in Master TB]"
            ws_new.cell(row=row_out, column=2).value = int(acct)
            ws_new.cell(row=row_out, column=3).value = float(balance_map[acct])
            ws_new.cell(row=row_out, column=3).number_format = "#,##0.00;[Red]-#,##0.00"
            # Highlight new accounts
            ws_new.cell(row=row_out, column=1).fill = PatternFill("solid", fgColor="FFFF00")
            row_out += 1
            written += 1

    ws_new.column_dimensions["A"].width = 55
    ws_new.column_dimensions["B"].width = 14
    ws_new.column_dimensions["C"].width = 24
    return written


def create_tb_dec25(wb: openpyxl.Workbook, dec25_map: dict, mar26_map: dict) -> None:
    print("Creating TB Dec 25 sheet...")
    n = _create_tb_sheet(wb, "TB Dec 25",
                         "Total of Reporting Period Dec25",
                         dec25_map, "Master TB")
    print(f"  TB Dec 25: {n} accounts written")


def create_tb_mar26(wb: openpyxl.Workbook, mar26_map: dict) -> None:
    print("Creating TB Mar 26 sheet...")
    n = _create_tb_sheet(wb, "TB Mar 26",
                         "Total of Reporting Period Mar26",
                         mar26_map, "Master TB")
    print(f"  TB Mar 26: {n} accounts written")


# --- Step 4: Create SAM Q1'26 template sheet --------------------------------
def create_sam_q126(wb: openpyxl.Workbook) -> None:
    """
    Creates 'SAM Q1''26' sheet as an empty template.
    Mirrors the layout of 'SAM' (Q1 2025) with updated header and no data rows.
    """
    print("Creating SAM Q1'26 template sheet...")

    if "SAM Q1'26" in wb.sheetnames:
        del wb["SAM Q1'26"]

    # Insert after SUAM sheet
    insert_after = "SUAM" if "SUAM" in wb.sheetnames else "SAM"
    insert_idx = wb.sheetnames.index(insert_after) + 1
    ws_new = wb.create_sheet("SAM Q1'26", insert_idx)

    # --- Header section (rows 1-8) ---
    headers = [
        (2, "Asia Metal Public Company Limited"),
        (3, "Summary of Corrected Audit Misstatement"),
        (4, "For the period ended 31 March 2026"),
        (5, "Amounts shown in Baht"),
    ]
    for row, text in headers:
        cell = ws_new.cell(row=row, column=1)
        cell.value = text
        cell.font = Font(bold=(row <= 4), size=11 if row == 2 else 10)

    # Row 7-8: Column headers
    col_headers_r7 = {
        2: "W/P Ref",
        3: "Account code",
        4: "Accounts and Description",
        5: "Debit",
        6: "Credit",
        7: "Income Statement",
        10: "Balance Sheet",
    }
    col_headers_r8 = {
        7: "Debit",
        8: "(Credit)",
        9: "Income effect\nDebit/(Credit)",
        10: "Debit",
        11: "(Credit)",
        12: "Balance Sheet effect\nDebit/(Credit)",
    }

    for col, val in col_headers_r7.items():
        cell = ws_new.cell(row=7, column=col)
        cell.value = val
        cell.font = _bold()
        cell.fill = _header_fill("D9E1F2")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for col, val in col_headers_r8.items():
        cell = ws_new.cell(row=8, column=col)
        cell.value = val
        cell.font = _bold()
        cell.fill = _header_fill("D9E1F2")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Instruction note row
    ws_new.cell(row=10, column=1).value = "-- ยังไม่มีรายการปรับปรุง -- กรอก Audit Adjustments เมื่อได้รับจาก Auditor --"
    ws_new.cell(row=10, column=1).font = Font(italic=True, color="808080")

    # Column widths
    col_widths = {1: 5, 2: 10, 3: 14, 4: 45, 5: 18, 6: 18, 7: 18, 8: 18,
                  9: 18, 10: 18, 11: 18, 12: 18}
    for col, width in col_widths.items():
        ws_new.column_dimensions[get_column_letter(col)].width = width

    ws_new.row_dimensions[7].height = 30
    ws_new.row_dimensions[8].height = 40

    print(f"  SAM Q1'26: template created (period: 31 March 2026)")


# --- Step 5: Update file metadata / title -----------------------------------
def update_metadata(wb: openpyxl.Workbook) -> None:
    """Update workbook title and any visible period references."""
    print("Updating workbook metadata...")
    wb.properties.title = "AMC Q1'2026 Leadsheet STAT to client"
    wb.properties.description = "Asia Metal Public Company Limited - Q1 2026 Financial Leadsheet"


# --- Step 6: Print sheet inventory ------------------------------------------
def print_inventory(wb: openpyxl.Workbook) -> None:
    print("\nFinal sheet inventory:")
    for i, name in enumerate(wb.sheetnames, 1):
        marker = " *NEW*" if name in ("TB Dec 25", "SAM Q1'26") else ""
        print(f"  {i:2d}. {name}{marker}")


# --- Main -------------------------------------------------------------------
def main():
    print("=" * 60)
    print("AMC Q1'2026 Leadsheet Builder")
    print("=" * 60)

    # Load TB data
    dec25_map, mar26_map = load_tb_data()
    print(f"Loaded: {len(dec25_map)} Dec 2025 accounts, {len(mar26_map)} Mar 2026 accounts")

    # Step 1: Copy source
    copy_source()

    # Step 2-5: Open and modify
    print("Opening workbook (this may take a moment)...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    add_new_columns_to_master_tb(wb, dec25_map, mar26_map)
    create_tb_dec25(wb, dec25_map, mar26_map)
    create_tb_mar26(wb, mar26_map)
    create_sam_q126(wb)
    update_metadata(wb)
    print_inventory(wb)

    # Save
    print(f"\nSaving output...")
    wb.save(OUTPUT_FILE)
    print(f"  -> {OUTPUT_FILE}")
    print("\nDone. Open in Google Sheets or Excel to verify formulas.")


if __name__ == "__main__":
    main()
