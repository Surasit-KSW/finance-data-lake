"""
extract_tb_data.py
==================
Extracts Trial Balance data from two sources:
  1. Dec 2025 (YE2025) — from AMC Leadsheet YE25.xlsx detail schedules
  2. Mar 2026 (Q1 2026) — from AMC_TB_03.2026_v9.XLSX (SAP export)

Outputs:
  - tb_dec2025.csv  — Account Code, Account Name, Balance (Dec 2025)
  - tb_mar2026.csv  — Account Code, Account Name, Balance, FS Item (Mar 2026)

Usage:
    python extract_tb_data.py
"""

import pandas as pd
import openpyxl
from pathlib import Path

# --- Paths ------------------------------------------------------------------
BASE_DIR  = Path(__file__).parent.parent.parent  # _Finance_Data_Lake root
TEMPLATES = BASE_DIR / "01_Bronze_Raw" / "Templates"
NRV_DIR   = BASE_DIR / "01_Bronze_Raw" / "Inventory_RollStock" / "NRV"
OUTPUT    = BASE_DIR / "06_Scripts" / "leadsheet" / "output"
OUTPUT.mkdir(exist_ok=True)

YE25_FILE   = TEMPLATES / "AMC Leadsheet YE25.xlsx"
TB_MAR26    = NRV_DIR   / "AMC_TB_03.2026_v9.XLSX"


# --- Extract Dec 2025 from YE25 detail schedules ----------------------------
def extract_dec2025() -> pd.DataFrame:
    """
    Each detail schedule (O300, L300, K300, …) contains:
      Col B (index 1): Account Code
      Col C (index 2): Account Name
      Col G (index 6): Dec 31, 2025 balance (Per Book, TB)

    Row structure:
      Row 1-5: headers
      Row 6+:  data
    """
    print("Extracting Dec 2025 balances from YE25 leadsheet…")
    wb = openpyxl.load_workbook(YE25_FILE, read_only=True, data_only=True)

    SKIP_SHEETS = {"F1", "F2", "F3"}
    records = {}

    for sh_name in wb.sheetnames:
        if sh_name in SKIP_SHEETS:
            continue
        ws = wb[sh_name]
        for row in ws.iter_rows(min_row=6, values_only=True):
            acct_raw = row[1] if len(row) > 1 else None
            name_raw = row[2] if len(row) > 2 else None
            bal_raw  = row[6] if len(row) > 6 else None

            if acct_raw is None or not str(acct_raw).strip():
                continue
            try:
                acct = str(int(float(str(acct_raw)))).strip()
            except (ValueError, TypeError):
                continue
            if not isinstance(bal_raw, (int, float)):
                continue

            # First occurrence wins (schedules may share accounts in some edge cases)
            if acct not in records:
                records[acct] = {
                    "account_code": acct,
                    "account_name": str(name_raw).strip() if name_raw else "",
                    "balance_dec2025": bal_raw,
                    "source_sheet": sh_name,
                }

    df = pd.DataFrame(records.values()).sort_values("account_code")
    out = OUTPUT / "tb_dec2025.csv"
    df.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"  -> {len(df):,} accounts saved to {out}")
    return df


# --- Extract Mar 2026 from SAP TB export ------------------------------------
def extract_mar2026() -> pd.DataFrame:
    """
    AMC_TB_03.2026_v9.XLSX columns:
      Col A (0): Financial Statement Item (FS type code)
      Col B (1): Text for B/S P&L Item (Account Name)
      Col C (2): Account Number
      Col D (3): Total of Reporting Period (Balance)

    Row 1: headers; Row 2+: data (skip rows without account number)
    """
    print("Extracting Mar 2026 balances from AMC_TB_03.2026…")
    wb = openpyxl.load_workbook(TB_MAR26, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_raw = row[2]
        if acct_raw is None or not str(acct_raw).strip():
            continue
        if str(acct_raw).strip() == "Account Number":
            continue
        try:
            acct = str(int(float(str(acct_raw)))).strip()
        except (ValueError, TypeError):
            continue

        name_raw = row[1]
        bal_raw  = row[3]
        fs_raw   = row[0]

        # Strip account code prefix from name if present (e.g. "1111010 Petty Cash" → "Petty Cash")
        name = str(name_raw).strip() if name_raw else ""
        if name.startswith(acct):
            name = name[len(acct):].strip()

        records.append({
            "account_code": acct,
            "account_name": name,
            "balance_mar2026": bal_raw if isinstance(bal_raw, (int, float)) else 0.0,
            "fs_item": str(fs_raw) if fs_raw is not None else "",
        })

    df = pd.DataFrame(records).sort_values("account_code")
    out = OUTPUT / "tb_mar2026.csv"
    df.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"  -> {len(df):,} accounts saved to {out}")
    return df


# --- Compare accounts between the two periods -------------------------------
def compare_accounts(df_dec: pd.DataFrame, df_mar: pd.DataFrame) -> None:
    dec_set = set(df_dec["account_code"])
    mar_set = set(df_mar["account_code"])

    new_in_mar  = mar_set - dec_set   # accounts added in Q1 2026
    gone_in_mar = dec_set - mar_set   # accounts in Dec 2025 not in Mar 2026

    print(f"\nAccount comparison:")
    print(f"  Dec 2025: {len(dec_set):,} accounts")
    print(f"  Mar 2026: {len(mar_set):,} accounts")

    if new_in_mar:
        print(f"\n  NEW accounts in Mar 2026 (not in Dec 2025): {len(new_in_mar)}")
        for a in sorted(new_in_mar):
            row = df_mar[df_mar["account_code"] == a].iloc[0]
            print(f"    {a}  {row['account_name'][:50]}")

    if gone_in_mar:
        print(f"\n  Accounts in Dec 2025 NOT in Mar 2026: {len(gone_in_mar)}")
        for a in sorted(list(gone_in_mar))[:20]:
            row = df_dec[df_dec["account_code"] == a].iloc[0]
            print(f"    {a}  {row['account_name'][:50]}")
        if len(gone_in_mar) > 20:
            print(f"    … and {len(gone_in_mar) - 20} more")

    # Save comparison report
    combined = pd.merge(
        df_dec[["account_code", "account_name", "balance_dec2025"]],
        df_mar[["account_code", "balance_mar2026"]],
        on="account_code",
        how="outer",
    ).sort_values("account_code")
    combined["movement"] = combined["balance_mar2026"].fillna(0) - combined["balance_dec2025"].fillna(0)
    combined["status"] = "existing"
    combined.loc[combined["balance_dec2025"].isna(), "status"] = "NEW in Mar26"
    combined.loc[combined["balance_mar2026"].isna(), "status"] = "gone in Mar26"

    out = OUTPUT / "tb_comparison.csv"
    combined.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"\n  Comparison report -> {out}")


# --- Main --------------------------------------------------------------------
if __name__ == "__main__":
    df_dec = extract_dec2025()
    df_mar = extract_mar2026()
    compare_accounts(df_dec, df_mar)
    print("\nDone.")
