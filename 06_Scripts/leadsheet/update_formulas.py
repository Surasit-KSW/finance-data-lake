"""
update_formulas.py
==================
Updates VLOOKUP formula references in the Q1'2026 leadsheet so that
the detail schedules (O300, L300, K300, …) and SAM sheets pull data
from the new Master TB columns instead of the Q1 2025 columns.

Master TB column mapping:
  Old (Q1 2025)          New (Q1 2026)
  ─────────────────────  ────────────────────────────────────
  Col E (5) = Sep 2024   Col K (11) = Dec 2025  (prior YE)
  Col F (6) = Mar 2025   Col L (12) = Mar 2026  (current Q1)
  Col G (7) = DR         Col G (7)  = DR         (unchanged)
  Col H (8) = CR         Col H (8)  = CR         (unchanged)
  Col I (9) = Adj Q125   Col I (9)  = (keep old) (unchanged)
  Col J (10)= Dec 2024   Col J (10) = Dec 2024   (historical)

VLOOKUP patterns updated:
  1. nth_col=6 with range A:J  → nth_col=12 with range A:L  (current period balance)
  2. nth_col=6 with range A:Y  → nth_col=12 with range A:Y  (current period balance)
  3. nth_col=5 with range A:Y  → nth_col=11 with range A:Y  (prior period balance)
  4. nth_col=5 with range A:J  → nth_col=11 with range A:Y  (prior period balance)

Sheets processed: all detail schedules (xxx300, xxx400) + F1, F2, F3

Usage:
    python update_formulas.py
"""

import re
import openpyxl
from pathlib import Path

BASE_DIR    = Path(__file__).parent.parent.parent
TEMPLATES   = BASE_DIR / "01_Bronze_Raw" / "templates"
TARGET_FILE = TEMPLATES / "leadsheet_q1_2026_stat.xlsx"

# Sheets to skip (TB sheets are data-only; SAM sheets use different formulas)
SKIP_SHEETS = {
    "TB 31.12.23", "TB 31Mar24", "TB 30Jun24", "TB30Sep24", "TB Dec24",
    "TB Mar 25", "TB Dec 25", "TB Mar 26",
    "Master TB",
    "SAM Q423", "SAM 24", "SAM", "SUAM", "SAM Q1'26",
}

# --- Regex-based formula transformations ------------------------------------
# Pattern: VLOOKUP(ref,'Master TB'!$A:$J,6,0)  → change col_idx and range
TRANSFORMS = [
    # (pattern, replacement)
    # Current period: col 6 from A:J  →  col 12 from A:L
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A:\$?J\s*,\s*)6(\s*,\s*0\s*\))",
        r"\g<1>12\g<2>"
    ),
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A:\$?J\s*,\s*)6(\s*,\s*FALSE\s*\))",
        r"\g<1>12\g<2>"
    ),
    # And extend the range from A:J to A:L when we change col to 12
    # (done in second pass below)

    # Current period: col 6 from A:Y  →  col 12 from A:Y
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A:\$?Y\s*,\s*)6(\s*,\s*0\s*\))",
        r"\g<1>12\g<2>"
    ),

    # Prior period: col 5 from A:J  →  col 11 from A:Y
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A:\$?J\s*,\s*)5(\s*,\s*0\s*\))",
        r"\g<1>11\g<2>"
    ),
    # Prior period: col 5 from A:Y  →  col 11 from A:Y
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A:\$?Y\s*,\s*)5(\s*,\s*0\s*\))",
        r"\g<1>11\g<2>"
    ),
]

# After changing col index to 12, also widen A:J → A:L if range still A:J
RANGE_EXPAND = [
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A):(\$?J)(\s*,\s*12\s*,)",
        r"\g<1>:L\g<3>"
    ),
    (
        r"(VLOOKUP\([^,]+,'Master TB'!\$?A):(\$?J)(\s*,\s*11\s*,)",
        r"\g<1>:L\g<3>"
    ),
]

# Also update XLOOKUP patterns (Q2 2026 uses XLOOKUP in TB sheets, skip those)
# For detail schedules that use XLOOKUP with the balances
XLOOKUP_TRANSFORMS = [
    # XLOOKUP(val, 'Master TB'!A:A, 'Master TB'!F:F, 0) -> E:E to K:K
    (
        r"(XLOOKUP\([^,]+,'Master TB'!\$?A:\$?A\s*,\s*'Master TB'!)\$?F:\$?F",
        r"\g<1>L:L"
    ),
    (
        r"(XLOOKUP\([^,]+,'Master TB'!\$?A:\$?A\s*,\s*'Master TB'!)\$?E:\$?E",
        r"\g<1>K:K"
    ),
]


def update_cell_formula(formula: str) -> tuple[str, int]:
    """Apply all transformations to a formula string. Returns (new_formula, changes)."""
    changes = 0
    for pattern, replacement in TRANSFORMS:
        new = re.sub(pattern, replacement, formula, flags=re.IGNORECASE)
        if new != formula:
            changes += 1
            formula = new

    for pattern, replacement in RANGE_EXPAND:
        new = re.sub(pattern, replacement, formula, flags=re.IGNORECASE)
        if new != formula:
            changes += 1
            formula = new

    for pattern, replacement in XLOOKUP_TRANSFORMS:
        new = re.sub(pattern, replacement, formula, flags=re.IGNORECASE)
        if new != formula:
            changes += 1
            formula = new

    return formula, changes


def process_workbook():
    print(f"Loading: {TARGET_FILE.name}")
    wb = openpyxl.load_workbook(TARGET_FILE)

    total_sheets_updated = 0
    total_cells_updated = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        sheet_changes = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    new_formula, n = update_cell_formula(cell.value)
                    if n > 0:
                        cell.value = new_formula
                        sheet_changes += n

        if sheet_changes > 0:
            total_sheets_updated += 1
            total_cells_updated += sheet_changes
            print(f"  [{sheet_name}]: {sheet_changes} formula changes")

    print(f"\nTotal: {total_sheets_updated} sheets updated, {total_cells_updated} cell changes")

    print(f"Saving...")
    wb.save(TARGET_FILE)
    print(f"  -> {TARGET_FILE}")


def verify_sample():
    """Spot-check a few formulas after update."""
    print("\nSpot-check after update:")
    wb = openpyxl.load_workbook(TARGET_FILE, data_only=False)

    ws = wb["O300"]
    for row in ws.iter_rows(min_row=6, max_row=8, values_only=False):
        for cell in row[:10]:
            if cell.value and str(cell.value).startswith("="):
                print(f"  O300[{cell.coordinate}]: {cell.value[:80]}")


if __name__ == "__main__":
    process_workbook()
    verify_sample()
    print("\nDone. Open file to verify formulas recalculate correctly.")
