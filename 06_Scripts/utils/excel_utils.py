"""
excel_utils.py — Shared Excel formatting helpers
ใช้ร่วมกันในทุก audit/reporting script เพื่อลด code ซ้ำ

ตัวอย่าง:
    from utils.excel_utils import format_header_row, auto_fit_columns
    ws = wb.active
    format_header_row(ws, row=1)
    auto_fit_columns(ws)
    wb.save("output.xlsx")
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── สีมาตรฐานของโปรเจกต์ ──────────────────────────────────
COLORS = {
    "header_blue":    "1F4E79",   # Dark blue — header row
    "header_green":   "1E5631",   # Dark green — alternative header
    "header_gray":    "404040",   # Dark gray — summary header
    "subheader":      "2E75B6",   # Medium blue — subheader
    "total_row":      "BDD7EE",   # Light blue — total row background
    "highlight":      "FFF2CC",   # Yellow — highlight cell
    "positive":       "E2EFDA",   # Light green — positive variance
    "negative":       "FCE4D6",   # Light orange — negative variance
    "white":          "FFFFFF",
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def format_header_row(ws, row: int = 1, fill_color: str = "header_blue",
                      font_size: int = 11, bold: bool = True):
    """ตกแต่ง header row — พื้นหลังสี, ตัวอักษรขาว, bold, กึ่งกลาง"""
    hex_color = COLORS.get(fill_color, fill_color)
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=bold, size=font_size)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = THIN_BORDER


def format_currency_column(ws, col_letter: str, start_row: int = 2,
                            decimal_places: int = 2):
    """ตั้ง number format เป็น #,##0.00 สำหรับคอลัมน์ตัวเลข"""
    fmt = f'#,##0.{"0" * decimal_places}'
    for row in ws.iter_rows(min_row=start_row, min_col=ws[f"{col_letter}1"].column,
                             max_col=ws[f"{col_letter}1"].column):
        for cell in row:
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")


def auto_fit_columns(ws, min_width: int = 8, max_width: int = 50):
    """ปรับความกว้าง column อัตโนมัติตาม content"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def add_total_row(ws, start_row: int, end_row: int,
                  sum_col_letters: list, label: str = "รวมทั้งหมด",
                  label_col: str = "A", fill_color: str = "total_row"):
    """เพิ่ม total row ด้านล่าง data พร้อม SUM formula และ highlight"""
    total_row = end_row + 1
    hex_color = COLORS.get(fill_color, fill_color)
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    bold_font = Font(bold=True)

    # Label
    label_cell = ws[f"{label_col}{total_row}"]
    label_cell.value = label
    label_cell.font = bold_font
    label_cell.fill = fill
    label_cell.border = THIN_BORDER

    # SUM formulas
    for col_letter in sum_col_letters:
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        cell.number_format = "#,##0.00"
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER

    return total_row


def freeze_header_row(ws, row: int = 1):
    """Freeze แถว header ให้ scroll แล้วยังเห็น header"""
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def set_tab_color(ws, color: str = "1F4E79"):
    """ตั้งสีแถบ sheet tab"""
    ws.sheet_properties.tabColor = COLORS.get(color, color)
