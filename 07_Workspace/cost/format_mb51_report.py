"""
format_mb51_report.py
=====================
ล้าง format เดิมทั้งหมด แล้วใส่ format มาตรฐานรายงานการเงิน

Usage:
    python scripts/format_mb51_report.py [path_to_xlsx]
    (default: ไฟล์ล่าสุดใน 02_Working/mb51_cost_breakdown_*.xlsx)

Prepared by: Claude Code
"""

import sys
import glob
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent.parent.parent  # _Finance_Data_Lake/
WORK    = Path(__file__).resolve().parent.parent / "02_Working"

# ── Colour Palette ────────────────────────────────────────────────────────────
HDR_BG   = "1F3864"   # dark navy
HDR_FG   = "FFFFFF"   # white
SUB_BG   = "2F5496"   # medium blue  (sub-header / section)
SUB_FG   = "FFFFFF"
ROW_ALT  = "EAF0F9"   # very light blue
ROW_ODD  = "FFFFFF"   # white
TOTAL_BG = "D6E4F0"   # light blue (total rows)
BORDER_C = "B0C4DE"   # soft border

# ── Font & Size ───────────────────────────────────────────────────────────────
FONT_FACE = "Calibri"
FONT_SIZE = 10
HDR_SIZE  = 10

# ── Number Formats ────────────────────────────────────────────────────────────
FMT_NUM   = '#,##0.00'
FMT_INT   = '#,##0'
FMT_DATE  = 'DD/MM/YYYY'
FMT_TEXT  = '@'

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=FONT_SIZE, italic=False) -> Font:
    return Font(name=FONT_FACE, bold=bold, color=color, size=size, italic=italic)

def border_thin() -> Border:
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom_only() -> Border:
    s = Side(style="thin", color=BORDER_C)
    return Border(bottom=s)

def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def clear_cell(cell):
    """Strip all existing formatting from one cell."""
    cell.font      = Font(name=FONT_FACE, size=FONT_SIZE)
    cell.fill      = PatternFill(fill_type=None)
    cell.border    = Border()
    cell.alignment = Alignment()
    cell.number_format = "General"


# ── Column Width Auto-fit ─────────────────────────────────────────────────────
def autofit(ws, min_w=8, max_w=45, padding=2):
    """Estimate column widths from cell content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + padding))


# ── Clear entire worksheet ────────────────────────────────────────────────────
def clear_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            clear_cell(cell)
    # Remove row heights
    for rd in ws.row_dimensions.values():
        rd.height = None
    # Remove merge (except keep intentional ones — cleared later per sheet)


# ── Format helpers ────────────────────────────────────────────────────────────
def fmt_header_row(ws, row_num: int, n_cols: int):
    """Apply dark-navy header to a full row."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = font(bold=True, color=HDR_FG, size=HDR_SIZE)
        cell.fill      = fill(HDR_BG)
        cell.alignment = align("center")
        cell.border    = border_thin()
        cell.number_format = FMT_TEXT

def fmt_data_row(ws, row_num: int, n_cols: int, alt: bool, amount_cols: list = None):
    """Apply alternating row colour + number format."""
    bg = ROW_ALT if alt else ROW_ODD
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill      = fill(bg)
        cell.alignment = align("left") if c <= 2 else align("right")
        cell.border    = border_bottom_only()
        # Number format: amount columns
        if amount_cols and c in amount_cols:
            cell.number_format = FMT_NUM
            cell.font = font()
        else:
            cell.font = font()

def fmt_total_row(ws, row_num: int, n_cols: int, amount_cols: list = None):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = font(bold=True)
        cell.fill      = fill(TOTAL_BG)
        cell.alignment = align("right") if c > 2 else align("left")
        cell.border    = border_thin()
        if amount_cols and c in amount_cols:
            cell.number_format = FMT_NUM


# ── Title Banner ──────────────────────────────────────────────────────────────
def add_title(ws, title: str, subtitle: str, n_cols: int):
    """Insert 2 title rows at top of sheet (caller must have shifted data already)."""
    ws.insert_rows(1, 2)

    t = ws.cell(row=1, column=1, value=title)
    t.font      = Font(name=FONT_FACE, bold=True, size=13, color=HDR_FG)
    t.fill      = fill(HDR_BG)
    t.alignment = align("left", "center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 22

    s = ws.cell(row=2, column=1, value=subtitle)
    s.font      = Font(name=FONT_FACE, italic=True, size=9, color="444444")
    s.fill      = fill("F2F7FF")
    s.alignment = align("left", "center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 16


# ═══════════════════════════════════════════════════════════════════════════════
#  Per-sheet formatters
# ═══════════════════════════════════════════════════════════════════════════════

def fmt_summary(ws, sheet_name: str):
    """Format Summary_Monthly and Summary_by_Mvt."""
    clear_sheet(ws)
    max_col = ws.max_column
    max_row = ws.max_row

    # Detect amount columns (cols 3 onward are numeric month/mvt columns)
    amount_cols = list(range(3, max_col + 1))

    # Add title rows
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    add_title(ws,
              f"ASIA METAL PLC — MB51 Cost Breakdown  |  {sheet_name}  |  Q1 2026",
              f"Source: MB51_all plant + MARA  |  Prepared by: Claude Code  |  {ts}",
              max_col)

    # Header row (now row 3 after insert)
    hdr_row = 3
    fmt_header_row(ws, hdr_row, max_col)

    # Data rows
    alt = False
    for r in range(hdr_row + 1, ws.max_row + 1):
        fmt_data_row(ws, r, max_col, alt, amount_cols)
        alt = not alt

    # Freeze pane
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=3)

    # Col widths: plant=10, category=28, rest=16
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    for c in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16


def fmt_detail(ws, sheet_name: str):
    """Format Detail_1300/1100/1200 — row-level data."""
    clear_sheet(ws)
    max_col = ws.max_column
    max_row = ws.max_row

    # Amount column: find 'amount' header
    amount_col_idx = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v and str(v).lower() == "amount":
            amount_col_idx = c
            break
    qty_col_idx = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v and str(v).lower() == "qty":
            qty_col_idx = c
            break

    amount_cols = []
    if amount_col_idx:
        amount_cols.append(amount_col_idx)
    if qty_col_idx:
        amount_cols.append(qty_col_idx)

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    add_title(ws,
              f"ASIA METAL PLC — MB51 Row Detail  |  {sheet_name}  |  Q1 2026",
              f"Source: MB51_all plant + MARA  |  Prepared by: Claude Code  |  {ts}",
              max_col)

    hdr_row = 3
    fmt_header_row(ws, hdr_row, max_col)

    alt = False
    for r in range(hdr_row + 1, ws.max_row + 1):
        fmt_data_row(ws, r, max_col, alt, amount_cols)
        # Left-align text columns
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).alignment = align("left")
        if amount_col_idx:
            ws.cell(row=r, column=amount_col_idx).alignment = align("right")
            ws.cell(row=r, column=amount_col_idx).number_format = FMT_NUM
        alt = not alt

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=3)

    # Column widths
    COL_W = {
        "month": 12, "plant": 8, "material": 18, "description": 40,
        "mat_type": 12, "mat_group": 14, "cost_category": 22,
        "mvt": 8, "amount": 18, "qty": 14, "unit": 8,
        "order": 16, "sloc": 10,
    }
    for c in range(1, max_col + 1):
        hdr_val = str(ws.cell(row=hdr_row, column=c).value or "").lower()
        w = COL_W.get(hdr_val, 14)
        ws.column_dimensions[get_column_letter(c)].width = w


def fmt_ingot(ws):
    """Format Ingot tab — keep important columns visible."""
    clear_sheet(ws)
    max_col = ws.max_column

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    add_title(ws,
              "ASIA METAL PLC — MB51 Ingot  |  Q1 2026",
              f"Source: MB51_ingot  |  Prepared by: Claude Code  |  {ts}",
              max_col)

    hdr_row = 3
    fmt_header_row(ws, hdr_row, max_col)

    alt = False
    for r in range(hdr_row + 1, ws.max_row + 1):
        fmt_data_row(ws, r, max_col, alt)
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).alignment = align("left")
        alt = not alt

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=3)
    autofit(ws, min_w=8, max_w=30)


def fmt_unclassified(ws):
    clear_sheet(ws)
    max_col = ws.max_column

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    add_title(ws,
              "ASIA METAL PLC — Unclassified Materials  |  Q1 2026",
              f"Materials not found in MARA  |  Prepared by: Claude Code  |  {ts}",
              max_col)

    hdr_row = 3
    fmt_header_row(ws, hdr_row, max_col)

    # Highlight unclassified rows in yellow
    WARN_BG = "FFF2CC"
    for r in range(hdr_row + 1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font      = font()
            cell.fill      = fill(WARN_BG)
            cell.alignment = align("left")
            cell.border    = border_bottom_only()

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    autofit(ws, max_w=40)


def fmt_full_detail(ws):
    """Full_Detail — header + freeze only (too many rows for full format)."""
    clear_sheet(ws)
    max_col = ws.max_column

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    add_title(ws,
              "ASIA METAL PLC — Full Detail  |  Q1 2026",
              f"All MB51 rows (201/261/531)  |  Prepared by: Claude Code  |  {ts}",
              max_col)

    hdr_row = 3
    fmt_header_row(ws, hdr_row, max_col)

    # Amount column formatting only (skip row-by-row for speed)
    amount_col_idx = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=hdr_row, column=c).value
        if v and str(v).lower() == "amount":
            amount_col_idx = c
            break
    if amount_col_idx:
        col_letter = get_column_letter(amount_col_idx)
        for r in range(hdr_row + 1, ws.max_row + 1):
            cell = ws.cell(row=r, column=amount_col_idx)
            cell.number_format = FMT_NUM

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=3)
    autofit(ws, max_w=35)


# ═══════════════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    # Find file
    if len(sys.argv) > 1:
        src = Path(sys.argv[1])
    else:
        files = sorted(glob.glob(str(WORK / "mb51_cost_breakdown_*.xlsx")))
        if not files:
            print("No mb51_cost_breakdown_*.xlsx found in 02_Working/")
            return
        src = Path(files[-1])

    print(f"Formatting: {src.name}")
    wb = load_workbook(src)

    tab_map = {
        "Summary_Monthly":  lambda ws: fmt_summary(ws, "Summary_Monthly"),
        "Summary_by_Mvt":   lambda ws: fmt_summary(ws, "Summary_by_Mvt"),
        "Detail_1300":      lambda ws: fmt_detail(ws, "Plant 1300"),
        "Detail_1100":      lambda ws: fmt_detail(ws, "Plant 1100"),
        "Detail_1200":      lambda ws: fmt_detail(ws, "Plant 1200"),
        "Ingot":            fmt_ingot,
        "Unclassified":     fmt_unclassified,
        "Full_Detail":      fmt_full_detail,
    }

    for sh_name in wb.sheetnames:
        fn = tab_map.get(sh_name)
        if fn:
            print(f"  Formatting {sh_name}...")
            fn(wb[sh_name])
        else:
            print(f"  Skipping {sh_name} (no formatter)")

    wb.save(src)
    print(f"\nSaved: {src}")
    print("Prepared by: Claude Code")


if __name__ == "__main__":
    main()
