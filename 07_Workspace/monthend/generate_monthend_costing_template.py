"""
generate_monthend_costing_template.py
======================================
สร้าง Month-End Close Template (Excel) — CO + GL phase — สำหรับ 3 บริษัท:
  AMC (Plants 1100, 1200, 1300)
  GA  (Plants 2100, 2200)
  STC (Plant  3100  — Trading, limited CO steps)

Sheets: Dashboard | CO Close-Master | AMC | GA | STC | GL Close-Master | T-Code Ref

Usage:
  python scripts/generate_monthend_costing_template.py --month 05 --year 2026
  python scripts/generate_monthend_costing_template.py  (default: current month)

Flags:
  --no-upload   สร้างแค่ไฟล์ local ไม่ upload Drive
  --no-fix      ไม่ run fix_master_formulas หลัง upload

Output local: 02_Working/monthend_close_MMYYYY.xlsx
Output Drive: Month-End Close/MM.YYYY/monthend_close_MMYYYY
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

# workspace root สำหรับ import utils
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# 1. STEP DATA  (CO Costing Close — 39 steps across 12 process groups)
# ---------------------------------------------------------------------------
# Dict keys: group, step, task, task_th, tcode, owner, amc, ga, stc, subtasks
# subtasks = list of (description, how_to) — rendered as child rows in CO Master

STEPS = [
  # ── 1. Initial Review ─────────────────────────────────────────────────────
  {"group":"1. Initial Review","step":"010",
   "task":"Cost Centers: Actual/Plan/Variance Analysis",
   "task_th":"วิเคราะห์ต้นทุนจริง/แผน/ผลต่าง CC",
   "tcode":"S_ALR_87013611","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run S_ALR_87013611",      "Controlling Area + Fiscal Year + From/To Period → Execute → drill-down by Cost Center"),
     ("Review large variances",  "Identify CCs with actual vs plan > 5% variance → document root cause before allocation"),
     ("Resolve open items",      "Correct missing accruals or misposted costs via FB50 before starting allocation cycle"),
   ]},
  {"group":"1. Initial Review","step":"020",
   "task":"Adjust GL — General Ledger Adjustments",
   "task_th":"ปรับ GL ก่อนปิด",
   "tcode":"FB50","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Verify period open",      "OB52 → confirm CO/FI posting period is open for current month before any posting"),
     ("Post manual accruals",    "FB50 → document date = last day of month → enter Dr/Cr GL accounts and amounts"),
     ("Review posted documents", "FAGLL03H → filter doc type SA/AA → verify all adjustments correct before proceeding"),
   ]},

  # ── 2. Template Allocation ────────────────────────────────────────────────
  {"group":"2. Template Allocation","step":"030",
   "task":"Template Allocation — Combined Orders",
   "task_th":"Template Allocation (Combined)",
   "tcode":"CPTD","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Review template definition","CPTD → Controlling Area + Fiscal Year + Period → verify sender/receiver assignments"),
     ("Execute test mode",       "Run CPTD simulation → review proposed allocation quantities → confirm no errors"),
     ("Execute actual run",      "Production mode → verify allocation documents created → check results via KOB1/KOB3"),
   ]},
  {"group":"2. Template Allocation","step":"040",
   "task":"Template Allocation — per Plant",
   "task_th":"Template Allocation (แยก Plant)",
   "tcode":"CPTD","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run per plant separately", "Execute CPTD for each plant (AMC: 1100/1200/1300; GA: 2100/2200) individually"),
     ("Compare vs plan",         "Verify allocated quantities match planned activity consumption per plant"),
     ("Reconcile deviations",    "Investigate allocation variance > 10% between plants before proceeding"),
   ]},

  # ── 3. Allocation Cycle ───────────────────────────────────────────────────
  {"group":"3. Allocation Cycle","step":"050",
   "task":"Execute Allocation Cycle (Assessment)",
   "task_th":"รัน Allocation Cycle Assessment (KSV5)",
   "tcode":"KSV5","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run KSV5 test mode",      "KSV5 → Controlling area + fiscal year + period → test → review receiver amounts"),
     ("Verify cycle definition", "Check sender cost centers and receiver tracing factors are correct for the period"),
     ("Execute actual run",      "Production mode → post assessment documents → verify via KSB1 on receiver CCs"),
   ]},
  {"group":"3. Allocation Cycle","step":"060",
   "task":"Execute Allocation Cycle (Distribution)",
   "task_th":"รัน Allocation Cycle Distribution (KSU5)",
   "tcode":"KSU5","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run KSU5 test mode",      "KSU5 → same parameters as KSV5 → test distribution results before actual"),
     ("Review distribution basis","Verify tracing factors (headcount/machine hours) are updated for the period"),
     ("Execute actual run",      "Post distribution cycle → verify GL documents created → check receiver CCs in FAGLL03H"),
   ]},
  {"group":"3. Allocation Cycle","step":"070",
   "task":"Cost Centers: Actual/Plan/Variance (Post-Alloc)",
   "task_th":"ตรวจสอบ CC หลัง Allocation",
   "tcode":"S_ALR_87013628","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run post-allocation report","S_ALR_87013628 → compare actual vs plan after all allocations complete"),
     ("Verify receiving CCs",    "Check receiver cost centers reflect correct allocated amounts"),
     ("Sign off allocation",     "Document allocation result is acceptable before proceeding to revaluation"),
   ]},

  # ── 4. Revaluation @ Actual Price ─────────────────────────────────────────
  {"group":"4. Revaluation @ Actual","step":"080",
   "task":"Actual Cost Splitting",
   "task_th":"แยกต้นทุน Activity จริง",
   "tcode":"KSS2","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KSS2",            "KSS2 → Controlling area + fiscal year + period + business process → run splitting"),
     ("Verify split results",    "Check activity type costs split correctly across cost centers per splitting rules"),
     ("Investigate exceptions",  "Review CCs flagged with split errors or zero activity quantities → resolve before KSII"),
   ]},
  {"group":"4. Revaluation @ Actual","step":"090",
   "task":"Actual Cost Center Splitting Report",
   "task_th":"รายงาน CC Splitting จริง",
   "tcode":"S_ALR_87013628","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run report after KSS2",   "S_ALR_87013628 → verify CC balances updated after cost splitting step"),
     ("Check activity quantities","Confirm activity quantities match production confirmations (CO11N / MIGO)"),
     ("Document exceptions",     "Flag CCs with unexpected splits for management review before price calculation"),
   ]},
  {"group":"4. Revaluation @ Actual","step":"100",
   "task":"Actual Activity Price Calculation",
   "task_th":"คำนวณราคา Activity จริง",
   "tcode":"KSII","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KSII",            "KSII → Controlling area + fiscal year + period → calculate actual prices (actual cost / actual qty)"),
     ("Review calculated prices","Compare actual vs standard activity prices → document variances > 10%"),
     ("Approve before revaluation","Sign off actual prices before CON2 to avoid re-running the entire revaluation"),
   ]},
  {"group":"4. Revaluation @ Actual","step":"110",
   "task":"Display Activity Price Report",
   "task_th":"แสดงรายงานราคา Activity",
   "tcode":"KSBT","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run KSBT",                "KSBT → display activity type prices for all cost centers in controlling area"),
     ("Compare actual vs standard","Document significant price changes vs prior month for management commentary"),
     ("Archive report",          "Export to Excel → save in month-end folder as audit trail supporting document"),
   ]},
  {"group":"4. Revaluation @ Actual","step":"120",
   "task":"Activity Revaluation — Combined Orders",
   "task_th":"Revaluation ที่ราคาจริง (Combined)",
   "tcode":"CON2","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute CON2 test mode",  "CON2 → Combined Orders → test mode → review proposed revaluation amounts per order"),
     ("Execute actual run",      "Production mode → post revaluation documents for Combined Orders"),
     ("Verify revaluation docs", "KOB3 or KKBC_ORD → confirm amounts are reasonable vs prior revaluation"),
   ]},
  {"group":"4. Revaluation @ Actual","step":"130",
   "task":"Activity Revaluation — per Plant",
   "task_th":"Revaluation ที่ราคาจริง (แยก Plant)",
   "tcode":"CON2","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute CON2 per plant",  "Run CON2 for each plant separately (AMC: 1100/1200/1300; GA: 2100/2200)"),
     ("Verify all plants done",  "Confirm no plant left unprocessed before moving to settlement step"),
     ("Review revaluation impact","Total revaluation amount should be consistent with activity price changes from KSBT"),
   ]},
  {"group":"4. Revaluation @ Actual","step":"140",
   "task":"Activity Revaluation Report",
   "task_th":"รายงาน Revaluation จริง",
   "tcode":"S_ALR_87013643","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run S_ALR_87013643",      "Display actual activity costs after revaluation → compare before and after values"),
     ("Review material cost impact","Verify that cost changes due to revaluation are reasonable and explained"),
     ("Document results",        "Prepare revaluation summary for management review and audit documentation"),
   ]},

  # ── 5. Work in Process ────────────────────────────────────────────────────
  {"group":"5. Work in Process","step":"150",
   "task":"Change Cutoff Period",
   "task_th":"กำหนด Cutoff Period (1 ครั้งต่อ Plant)",
   "tcode":"KKA0","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KKA0",            "KKA0 → Select plant → set WIP cutoff period to current period (YYYY/MM) → save"),
     ("Run for all plants",      "Repeat for each plant (AMC: 1100/1200/1300; GA: 2100/2200) — must complete BEFORE KKAOH"),
     ("Verify cutoff period",    "Incorrect fiscal year/period causes wrong WIP → double-check before running KKAOH"),
   ]},
  {"group":"5. Work in Process","step":"160",
   "task":"Calculate Work in Process",
   "task_th":"คำนวณ WIP",
   "tcode":"KKAOH","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KKAOH test mode", "KKAOH → Plant + order selection → test run → review WIP amounts by order"),
     ("Execute actual run",      "Production mode → post WIP calculation → verify FI documents for WIP GL accounts"),
     ("Reconcile WIP with GL",   "FAGLL03H WIP account (e.g. 131xxxx) total must match KKAOH report total"),
   ]},
  {"group":"5. Work in Process","step":"170",
   "task":"Display Work in Process Report",
   "task_th":"แสดงรายงาน WIP",
   "tcode":"KKAQ","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run KKAQ",                "KKAQ → Plant + period → display WIP balances per production order"),
     ("Flag large WIP orders",   "Identify orders with WIP > threshold → review with operations for possible early settlement"),
     ("Archive WIP report",      "Export KKAQ as month-end WIP supporting document"),
   ]},

  # ── 6. Variance (Combined) ────────────────────────────────────────────────
  {"group":"6. Variance (Combined)","step":"180",
   "task":"Variance Calculation — Combined Orders",
   "task_th":"คำนวณ Variance (Combined Orders)",
   "tcode":"KKS1","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KKS1 test mode",  "KKS1 → Plant + Combined order type + period → test → review variance by order"),
     ("Review variance categories","Check Input Price / Quantity / Resource-Usage variance split for Combined Orders"),
     ("Execute actual run",      "Post variance calculation documents → verify before proceeding to CO88H settlement"),
   ]},

  # ── 7. Settlement (Combined) ──────────────────────────────────────────────
  {"group":"7. Settlement (Combined)","step":"190",
   "task":"Production Cost Settlement — Combined",
   "task_th":"ปิด Order ต้นทุน (Combined)",
   "tcode":"CO88H","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute CO88H test mode", "CO88H → Combined Orders → test mode → check settlement amounts to PA/material"),
     ("Execute actual settlement","Production mode → post settlement → verify FI documents GL 5391xxx / cost object"),
     ("Verify zero balance",     "KOC4 or KOB1 → confirm all Combined Orders settled → order balances = zero"),
   ]},

  # ── 8. Variance (Regular) ─────────────────────────────────────────────────
  {"group":"8. Variance (Regular)","step":"200",
   "task":"Variance Calculation — Regular Orders",
   "task_th":"คำนวณ Variance (Regular Orders)",
   "tcode":"KKS1","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KKS1 for Regular","KKS1 → Plant + regular production orders + period → test mode first"),
     ("Review variance by order","Check top-variance orders → confirm Input/Output price vs quantity drivers"),
     ("Execute actual run",      "Post variance documents for regular orders → verify before CO88H settlement"),
   ]},

  # ── 9. Settlement (Regular) ───────────────────────────────────────────────
  {"group":"9. Settlement (Regular)","step":"210",
   "task":"Production Cost Settlement — Regular",
   "task_th":"ปิด Order ต้นทุน (Regular)",
   "tcode":"CO88H","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute CO88H for Regular","CO88H → Regular production orders → test mode → review settlement amounts"),
     ("Execute actual run",      "Post settlement → verify FI documents → GL 5391xxx credit / cost receiver debit"),
     ("Review KOC4 list",        "KOC4 → confirm all regular orders settled → document any orders with settlement errors"),
   ]},
  {"group":"9. Settlement (Regular)","step":"220",
   "task":"Display GL Summary Report",
   "task_th":"ตรวจสอบ GL Summary หลัง Settlement",
   "tcode":"F.01","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run F.01",                "F.01 → Company code + fiscal year + period → financial statements balance check"),
     ("Review inventory accounts","FG / WIP / RM account balances should reflect settlement impact → investigate surprises"),
     ("Save GL snapshot",        "Export F.01 report → save as pre-ML closing GL reference for reconciliation"),
   ]},
  {"group":"9. Settlement (Regular)","step":"230",
   "task":"Adjust GL — PC Variance (Dr/Cr 5391020 / 5119010)",
   "task_th":"ปรับ GL ผลต่าง PC",
   "tcode":"F-02 / F-50 / KOC4","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Review variance via KOC4","KOC4 → production cost variance summary → identify net Dr/Cr to 5391020 / 5119010"),
     ("Post via F-02 or F-50",   "Doc type SA → Dr 5391020 (PC variance) / Cr 5119010 (offset account) → save"),
     ("Verify balance",          "FAGLL03H 5391020 + 5119010 → net effect correct and accounts properly balanced"),
   ]},

  # ── 10. ML Closing ────────────────────────────────────────────────────────
  {"group":"10. ML Closing","step":"240",
   "task":"ML Closing — Actual Costing Run",
   "task_th":"รัน ML Closing (CKMLCP)",
   "tcode":"CKMLCP","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Step 1 — Selection",      "CKMLCP → Company code + plant + period → execute Selection → verify all materials included"),
     ("Step 2 — Qty Determination","Run Quantity Determination → verify consumption quantities match MB51/MB52 data"),
     ("Step 3 — Single-level",   "Run Single-level Price Determination → calculate actual cost per material"),
     ("Step 4 — Multi-level",    "Run Multilevel Price Determination → propagate actual costs across all BOM levels"),
     ("Step 5 — Revaluation",    "Run Revaluation/FI Transfer → verify FI documents posted → check inventory revaluation"),
   ]},
  {"group":"10. ML Closing","step":"250",
   "task":"Display GL Summary Report (Post-ML)",
   "task_th":"ตรวจสอบ GL หลัง ML",
   "tcode":"F.01","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run F.01 after CKMLCP",   "F.01 → verify inventory accounts updated with ML actual costing values"),
     ("Compare with pre-ML GL",  "Reconcile inventory balance change vs CKMLCP revaluation amount → must match"),
     ("Review COGS impact",      "Check COGS accounts (5xxx) for ML-driven revaluation postings"),
   ]},
  {"group":"10. ML Closing","step":"260",
   "task":"Value Flow Monitor Report",
   "task_th":"ตรวจสอบ Value Flow",
   "tcode":"CKMVFM","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run CKMVFM",              "CKMVFM → Company code + plant + period → display material ledger value flows"),
     ("Check errors/warnings",   "Investigate materials with red/yellow status → resolve before period lock"),
     ("Confirm all processed",   "No materials in 'Not yet processed' or error status before closing"),
   ]},
  {"group":"10. ML Closing","step":"270",
   "task":"Adjust GL — ML Variance (Dr/Cr 5391010 / 5119010)",
   "task_th":"ปรับ GL ผลต่าง ML",
   "tcode":"F-02 / F-50 / CKMPDB","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Run CKMPDB",              "CKMPDB → display ML price difference account balances → identify 5391010/5119010 amounts"),
     ("Post via F-02 or F-50",   "Doc type SA → Dr/Cr 5391010 (ML variance) / 5119010 (offset) → save document"),
     ("Verify GL balance",       "FAGLL03H 5391010 + 5119010 → confirm correct → carry-forward to next period check"),
   ]},
  {"group":"10. ML Closing","step":"280",
   "task":"Periodic Valuation (PA Transfer)",
   "task_th":"ส่งมูลค่า ML ไป PA",
   "tcode":"KE27","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Execute KE27",            "KE27 → Controlling area + fiscal year + period → transfer ML actual costs to PA"),
     ("Review PA transfer docs", "Verify cost components (material/labor/overhead) transferred to PA line items correctly"),
     ("Verify in KE24",          "KE24 → confirm PA received updated actual cost values from ML closing"),
   ]},

  # ── 11. Reporting ─────────────────────────────────────────────────────────
  {"group":"11. Reporting","step":"290",
   "task":"Profitability Analysis — Detail Line Item",
   "task_th":"รายงาน PA รายละเอียด",
   "tcode":"KE24","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run KE24",                "KE24 → actual version + fiscal year + period → drill down by product/customer/plant"),
     ("Review revenue vs COGS",  "Check margin by product group → flag unusual profitability vs prior month"),
     ("Export PA detail",        "Download to Excel → save as PA detail report for management pack"),
   ]},
  {"group":"11. Reporting","step":"300",
   "task":"Materials List: Prices and Inventory",
   "task_th":"รายการวัตถุดิบ ราคา + Inventory",
   "tcode":"S_P99_41000062","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run S_P99_41000062",      "Plant + material group + period → execute → display material prices and inventory values"),
     ("Compare standard vs actual","Note materials with actual price > 10% deviation from standard → explain movement"),
     ("Archive report",          "Save as month-end material price supporting document"),
   ]},
  {"group":"11. Reporting","step":"310",
   "task":"Material Price Analysis Report",
   "task_th":"วิเคราะห์ราคาวัตถุดิบ (ML)",
   "tcode":"CKM3N","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run CKM3N",               "CKM3N → Material + plant + period → display ML actual cost component breakdown"),
     ("Verify cost components",  "Check material/labor/overhead cost split matches expectations vs standard"),
     ("Document price movements","Flag materials with > 5% price change vs prior period for management commentary"),
   ]},
  {"group":"11. Reporting","step":"320",
   "task":"Stock Card Report (Standard)",
   "task_th":"Stock Card รายงานมาตรฐาน",
   "tcode":"S_P00_07000139","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run S_P00_07000139",      "Plant + material + period → display opening/closing stock and all movements"),
     ("Verify opening balance",  "Confirm opening balance = prior month closing balance → flag discrepancies"),
     ("Reconcile with warehouse","Cross-check quantities with warehouse physical count → investigate differences > tolerance"),
   ]},
  {"group":"11. Reporting","step":"330",
   "task":"Stock Card Report (Custom)",
   "task_th":"Stock Card รายงาน Custom",
   "tcode":"ZCOR001","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run ZCOR001",             "ZCOR001 → Plant + period → execute custom stock card report with extended fields"),
     ("Cross-reference standard","Verify ZCOR001 totals match S_P00_07000139 → flag discrepancies"),
     ("Archive final report",    "Save as final stock card supporting document for audit"),
   ]},

  # ── 12. Close & PA ────────────────────────────────────────────────────────
  {"group":"12. Close & PA","step":"340",
   "task":"Delete Production Order — Combined",
   "task_th":"ลบ Production Order (Combined)",
   "tcode":"CO78","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Identify eligible orders", "CO78 → filter Combined Orders with status TECO + DLT eligible + zero balance"),
     ("Execute deletion",         "Run CO78 deletion → verify orders removed → confirm in COOIS order list"),
     ("Archive order list",       "Document deleted Combined Orders list for audit trail"),
   ]},
  {"group":"12. Close & PA","step":"350",
   "task":"Close Production Order",
   "task_th":"ปิด Production Order",
   "tcode":"COHV","owner":"Controlling","amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Select orders for CLSD",  "COHV → mass processing → filter by status TECO + zero settlement balance"),
     ("Execute CLSD status",     "Run status change to CLSD (Closed) → cannot post after CLSD status set"),
     ("Verify closure count",    "Check number of orders closed matches expected count → document for period-end report"),
   ]},
  {"group":"12. Close & PA","step":"360",
   "task":"Profitability Analysis — Detail (Final)",
   "task_th":"รายงาน PA รายละเอียด (Final)",
   "tcode":"KE24","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run KE24 final",          "Same parameters as step 290 → run after ALL closings complete for final PA data"),
     ("Compare with preliminary","Compare final PA vs step 290 preliminary → explain any differences"),
     ("Sign off PA data",        "Document PA is final and ready for management reporting"),
   ]},
  {"group":"12. Close & PA","step":"370",
   "task":"Profitability Analysis — Summary Report",
   "task_th":"รายงาน PA สรุป",
   "tcode":"KE30","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Execute KE30",            "KE30 → Select PA report form + actual version + period → run P&L summary by segment"),
     ("Review profitability",    "Check product/customer/channel profitability → identify top margin contributors and drains"),
     ("Export for management",   "Format KE30 output → include in monthly management reporting package"),
   ]},
  {"group":"12. Close & PA","step":"380",
   "task":"Transfer SD Billing to PA",
   "task_th":"โอน SD Billing ไป PA",
   "tcode":"KE4S","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Execute KE4S",            "KE4S → Company code + billing date range → transfer all SD billing docs to PA"),
     ("Verify revenue in PA",    "KE24 → confirm revenue line items updated → compare with FI revenue GL total"),
     ("Reconcile PA vs FI",      "KE24 revenue total vs FAGLL03H GL revenue accounts → tolerance <= 100 THB"),
   ]},
  {"group":"12. Close & PA","step":"390",
   "task":"Lock Period CO",
   "task_th":"ล็อค Period CO",
   "tcode":"OKP1","owner":"Controlling","amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Execute OKP1",            "OKP1 → Controlling area + company code → set current period status to CLOSED"),
     ("Verify period lock",      "Test posting → should return error 'Posting period not open' → period confirmed locked"),
     ("Notify stakeholders",     "Inform AP/GL teams that CO period is closed → no further CO postings for the month"),
   ]},
]

# ---------------------------------------------------------------------------
# 1b. GL STEP DATA  (company-level — AMC / GA / STC)
# ---------------------------------------------------------------------------
# Dict keys: group, step, task, tcode, owner, target, amc, ga, stc, subtasks
# subtasks = list of (description, how_to) — shown as child rows below main step

GL_STEPS = [
  # ── A. GR/IR & AP Clearing ────────────────────────────────────────────────
  {"group":"A. GR/IR & AP Clearing","step":"A010",
   "task":"GR/IR Reconciliation — Verify & Clear",
   "tcode":"F.19 / F.13","owner":"AP/Controlling","target":"Day 1-3",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run F.13 — Test Mode first",        "F.13 → Enter Company Code + Fiscal Year → Select clearing type (GR/IR) → Run test → Review proposed items ก่อน execute จริง"),
     ("Execute F.13 — Final Run",           "หลัง test pass → run production mode → verify clearing document generated → ตรวจสอบใน FBL1H/FBL5H"),
     ("Run F.19 — GR/IR Regrouping",        "วิเคราะห์ GR/IR account balance → post adjustment entries (goods received not invoiced / invoiced not received) → verify GL postings"),
     ("Verify outstanding items",           "FAGLL03H GR/IR account → document remaining open items > 30 days → escalate to AP team for resolution"),
   ]},
  {"group":"A. GR/IR & AP Clearing","step":"A020",
   "task":"Invoice Verification — Inter-company (STC billing / AMC billing)",
   "tcode":"MIRO / FB60","owner":"AP","target":"Day 5-15",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Match STC invoice to AMC billing",   "ตรวจสอบ invoice จาก STC vs billing document ใน AMC → ยืนยันจำนวนและ GL account ตรงกัน"),
     ("Post unmatched invoices via MIRO",   "MIRO → enter PO number → verify quantities and prices → simulate → post"),
     ("Reverse incorrect invoices (FB60)",  "FB60 / MR8M สำหรับ reverse invoice ที่ post ผิด → re-post with correct data"),
     ("Confirm inter-company balance nets to zero", "ตรวจสอบว่า AR ที่ entity หนึ่ง = AP ที่อีก entity → balance ต้อง net zero"),
   ]},
  {"group":"A. GR/IR & AP Clearing","step":"A030",
   "task":"Clear Open Items — Vendor / Customer",
   "tcode":"F-44 / F-32","owner":"AP","target":"Day 5",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("F-44 Vendor manual clearing",        "F-44 → enter vendor code + company code → select items to clear (invoices vs payments) → simulate → post clearing document"),
     ("F-32 Customer manual clearing",       "F-32 → enter customer code → select items (down payments vs invoices) → simulate → post"),
     ("Verify cleared items in FBL1H/FBL5H","FBL1H (vendor) / FBL5H (customer) → filter cleared items → confirm balance accurate"),
   ]},

  # ── B. Recurring SAP Documents ────────────────────────────────────────────
  {"group":"B. Recurring SAP","step":"B010",
   "task":"Execute Recurring Entries (Prepaid / Rent / ROU)",
   "tcode":"F.14 / FBD2","owner":"Controlling","target":"Day 3-5",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Review F.15 list before execution",  "F.15 → verify ทุก recurring doc active + amounts correct + no deletion indicator ก่อน run"),
     ("Check SM35 for pending sessions",    "SM35 → ตรวจสอบ batch input session ค้างอยู่ก่อน → execute หรือ delete session เก่าที่ค้าง"),
     ("Execute F.14 per company code",      "F.14 → Company Code → Recurring doc numbers → Settlement period → enter session name → execute"),
     ("Process batch session & verify",     "SM35 → execute session → verify posting documents created → ตรวจสอบ GL amounts ตรงกับ schedule"),
   ]},
  {"group":"B. Recurring SAP","step":"B020",
   "task":"Verify Recurring Entry List",
   "tcode":"F.15","owner":"Controlling","target":"Day 5",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Display all active recurring docs",  "F.15 → company code → review list: amounts, posting dates, GL accounts ครบถ้วน"),
     ("Verify amounts match contracts",     "ตรวจสอบจำนวนแต่ละ recurring doc กับ contract/agreement (rent, SAP fee, insurance)"),
     ("Flag obsolete documents",            "ระบุ doc ที่ควร set deletion indicator (expired contracts, changed amounts) → FBD2 → update"),
   ]},
  {"group":"B. Recurring SAP","step":"B030",
   "task":"IT Support / SAP Maintenance Accrual",
   "tcode":"FB01 / FBD2","owner":"Controlling","target":"Day 5",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Calculate monthly IT accrual",       "รับ invoice/contract จาก IT → คำนวณ monthly allocation (annual fee / 12)"),
     ("Post accrual via FBS1/FB01",         "FB01 → Dr IT Expense / Cr Accrued Liabilities → posting date = last day of month"),
     ("Set auto-reversal for next month",   "ตั้ง reversal date = วันที่ 1 ของเดือนถัดไป → ป้องกัน double posting เมื่อ invoice จริงมาถึง"),
   ]},

  # ── C. Accrued Expenses ───────────────────────────────────────────────────
  {"group":"C. Accrued Expenses","step":"C010",
   "task":"Salary / Compensation Accrual (A1, A2 split)",
   "tcode":"FBS1 / FBD2","owner":"HR/Acctg","target":"Day 5",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Receive payroll summary from HR",    "รับ payroll summary แยก A1/A2 → ตรวจสอบจำนวนกับ approved headcount + salary grade"),
     ("Post FBS1 accrual per entity",       "FBS1 → Dr Salary Expense (A1/A2) / Cr Salary Payable → posting date = last day → set auto-reversal Day 1 next month"),
     ("Reconcile to payroll report",        "เปรียบเทียบ GL posting กับ official payroll report → variance ต้องไม่เกิน ±1,000 THB"),
   ]},
  {"group":"C. Accrued Expenses","step":"C020",
   "task":"Bonus Provision (MD Approval required)",
   "tcode":"FB01","owner":"Acctg/MD","target":"Day 5",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Obtain MD approval + calculation",   "รับ bonus calculation จาก HR + MD approval email ก่อน post ทุกครั้ง → attach เป็น supporting doc"),
     ("Post provision via FB01",            "FB01 → Dr Bonus Expense / Cr Bonus Payable → post with month-end date"),
     ("Document calculation basis",         "บันทึก basis: % of salary / performance score / headcount → เก็บไว้สำหรับ audit"),
   ]},
  {"group":"C. Accrued Expenses","step":"C030",
   "task":"EBO — Employee Benefit Obligation",
   "tcode":"FBS1 / FB01","owner":"HR/Acctg","target":"Day 5",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Receive EBO schedule from HR/Actuary","รับ EBO liability schedule (actuarial basis) → verify monthly accrual amount vs prior month"),
     ("Post FBS1 monthly accrual",          "FBS1 → Dr EBO Expense / Cr EBO Liability → set reversal date"),
     ("Reconcile GL to EBO schedule",       "FAGLL03H EBO GL account → verify balance matches cumulative schedule → document variance if any"),
   ]},
  {"group":"C. Accrued Expenses","step":"C040",
   "task":"Audit Fee Accrual",
   "tcode":"FBS1 / FBD2","owner":"Acctg","target":"Day 5",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Calculate monthly audit fee",        "Annual fee per engagement letter / 12 months → verify against FY budget (AMC: 280K, GA: 280K, STC: 100K annual approx.)"),
     ("Post FBS1 accrual",                  "FBS1 → Dr Audit Fee Expense / Cr Accrued Audit Fee → posting date = last day of month → set auto-reversal"),
     ("Match to engagement letter",         "ตรวจสอบ cumulative accrual vs engagement letter amount → ถ้า actual invoice มาถึง → reverse accrual แล้ว post actual"),
   ]},
  {"group":"C. Accrued Expenses","step":"C050",
   "task":"Professional Services Accrual (Lawyer / AIS / TOT)",
   "tcode":"FB01","owner":"Acctg","target":"Day 5",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Identify services rendered not invoiced","รวบรวม services ที่ใช้แล้วในเดือนนี้แต่ยังไม่มี invoice (lawyer, consultants, telecom)"),
     ("Post accrual per service type",      "FB01 → Dr Professional Fee Expense (แยก type) / Cr Accrued Liabilities → post month-end"),
     ("Reverse when actual invoice received","เมื่อ invoice จริงมาถึง → reverse accrual → post actual invoice via MIRO/FB60"),
   ]},
  {"group":"C. Accrued Expenses","step":"C060",
   "task":"Insurance Premium Accrual (AIA)",
   "tcode":"FBS1 / FBD2","owner":"Acctg","target":"Day 5",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Review AIA policy premium schedule", "ตรวจสอบ premium schedule ตาม policy type (group life, health, property) → คำนวณ monthly allocation"),
     ("Post recurring accrual via FBS1",    "FBS1 → Dr Insurance Expense / Cr Insurance Payable → ใช้ FBD2 recurring doc ถ้ามีการตั้ง"),
     ("Verify against policy documents",    "FAGLL03H insurance GL → reconcile cumulative accrual vs policy premium schedule → ตรวจสอบ coverage period"),
   ]},
  {"group":"C. Accrued Expenses","step":"C070",
   "task":"Other Salary / K.staff Accrual",
   "tcode":"FB01","owner":"HR/Acctg","target":"Day 5",
   "amc":True,"ga":False,"stc":False,
   "subtasks":[
     ("Receive K.staff salary data",        "รับข้อมูล K.staff / contractor salary จาก HR → verify จำนวน headcount และ rate"),
     ("Post FBS1 accrual",                  "FBS1 → Dr Salary Expense (Other) / Cr Salary Payable → month-end date → auto-reversal"),
   ]},

  # ── D. Lease / ROU ────────────────────────────────────────────────────────
  {"group":"D. Lease / ROU","step":"D010",
   "task":"ROU Asset Reconciliation (Prepaid vs Accrual)",
   "tcode":"FAGLL03H","owner":"Acctg","target":"Day 1-3",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Export GL 1159010 / 1151030 balances","FAGLL03H → GL accounts 1159010 (prepaid rent) + 1151030 (prepaid other) → export line items"),
     ("Compare to prepaid schedule",        "ตรวจสอบ GL balance vs prepaid schedule: beginning + payments - amortization = ending → must match"),
     ("Verify monthly ROU amortization",    "ยืนยันว่า FBD2 recurring doc ตัดค่า ROU amortization ถูกต้อง (monthly = total ROU / lease term months)"),
     ("Document IFRS16 current/non-current","แยก lease liability: portion due within 12 months (current) vs beyond 12 months (non-current) → verify BS presentation"),
   ]},
  {"group":"D. Lease / ROU","step":"D020",
   "task":"Prepaid Lease Allocation (Recurring FBD2)",
   "tcode":"FBD2 / F.14","owner":"Acctg","target":"Day 3",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Verify FBD2 recurring documents active","F.15 → ตรวจสอบ prepaid allocation recurring docs ครบ (rent A1/A2/GI, SAP fee) → amounts correct"),
     ("Execute F.14 allocation run",        "F.14 → company code + recurring doc numbers → settlement period → execute batch session"),
     ("Verify posting amounts vs schedule", "ตรวจสอบ GL postings ตรงกับ prepaid schedule: Dr Rent Expense / Cr Prepaid → ต้องตรงทุก entity"),
   ]},
  {"group":"D. Lease / ROU","step":"D030",
   "task":"Vehicle Lease Accrual (BENZ / Majesty / BYD / Zeekr)",
   "tcode":"FB01","owner":"Acctg","target":"Day 5",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Review vehicle lease contracts",     "ตรวจสอบ lease payment schedule ทุกคัน → คำนวณ monthly payment (principal + interest component IFRS16)"),
     ("Post lease liability entry",         "FB01 → Dr Right-of-Use Asset Expense / Cr Lease Liability → post with month-end date"),
     ("Reconcile to lease amortization table","ตรวจสอบ outstanding lease liability balance vs amortization schedule → verify interest expense reasonable"),
   ]},

  # ── E. Fixed Assets ───────────────────────────────────────────────────────
  {"group":"E. Fixed Assets","step":"E010",
   "task":"Add / Retire Asset",
   "tcode":"AS01 / AS02 / AS03 / ABAON","owner":"Acctg/FA","target":"Day 5",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Create new asset master (AS01)",     "AS01 → company code → asset class → enter description, useful life, depreciation key → save → record asset number"),
     ("Post asset acquisition",             "F-90 / MIRO → link to asset master → post acquisition → verify GL: Dr Asset account / Cr Vendor/AP"),
     ("Retire/dispose asset (ABAON/AS02)",  "ABAON → enter asset number → retirement date → sale revenue → post → verify gain/loss GL posting"),
     ("Run depreciation after retirement",  "AFAB → execute for retirement period → ensure partial month depreciation calculated and posted correctly"),
   ]},
  {"group":"E. Fixed Assets","step":"E020",
   "task":"Post Depreciation (AFAB) / Adjustment",
   "tcode":"AFAB / ABAON / AB08","owner":"Acctg/FA","target":"Day 7",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run AFAB — Test Mode",               "AFAB → company code + fiscal year + posting period → select 'Planned' → Test Run → review amounts and errors list"),
     ("Fix errors before final run",        "ตรวจสอบ error log → fix asset master data หรือ GL account issues → re-run test until error-free"),
     ("Execute AFAB — Production Run",      "AFAB → uncheck test run → execute → verify depreciation posting document created"),
     ("Verify depreciation in GL",          "FAGLL03H depreciation expense account → verify amount reasonable (cost x rate / 12) → match YAARP001 report"),
   ]},
  {"group":"E. Fixed Assets","step":"E030",
   "task":"Depreciation Verification Report",
   "tcode":"YAARP001 / S_ALR_87011963","owner":"Acctg/FA","target":"Day 7",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run YAARP001 asset balance report",  "YAARP001 → company code + period → export: gross value, accumulated depreciation, NBV per asset"),
     ("Run S_ALR_87011963 comparison",      "S_ALR_87011963 → compare current period vs prior period → verify additions/retirements/depreciation movements"),
     ("Reconcile FA register to GL",        "FA register total gross = GL asset account balance → FA accum. depreciation = GL contra-account → must match"),
     ("Investigate variances",              "ถ้า variance → check unposted asset transactions → AFAB error log → manual adjustments pending"),
   ]},

  # ── F. Tax & Compliance ───────────────────────────────────────────────────
  {"group":"F. Tax & Compliance","step":"F010",
   "task":"WHT Report — Form PND3/PND53 (Day 7)",
   "tcode":"ZAPC001","owner":"Tax/Acctg","target":"Day 7",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run ZAPC001 WHT report",             "ZAPC001 → company code + posting period → generate PND3 (individuals) and PND53 (companies) report"),
     ("Verify WHT codes on all invoices",   "ตรวจสอบว่า invoice ทุกใบที่เข้าข่ายมี WHT code ถูกต้อง → rate: 3% (service) / 5% (rent) / 1% (goods)"),
     ("Reconcile WHT GL to report",         "FAGLL03H WHT payable GL accounts → verify balance matches ZAPC001 report total → document difference if any"),
     ("Prepare PND3/PND53 for filing",      "Export report → review → submit via e-withholding tax system (ภาษีหัก ณ ที่จ่าย) → internal deadline: Day 7, RD deadline: Day 15"),
   ]},
  {"group":"F. Tax & Compliance","step":"F020",
   "task":"VAT Reconciliation — Input / Output Tax",
   "tcode":"FAGLL03H","owner":"Tax/Acctg","target":"Day 7",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Export input VAT GL (ภาษีซื้อ)",     "FAGLL03H → input VAT GL accounts → verify all AP invoices have correct VAT code (V1=7% standard, V0=0% exempt)"),
     ("Export output VAT GL (ภาษีขาย)",     "FAGLL03H → output VAT GL accounts → match to SD billing documents → verify VAT rate applied correctly"),
     ("Clear matched VAT items",            "Run automatic clearing for matched VAT pairs → convert remaining to down payments if needed"),
     ("Reconcile to VAT register for PP.30","ยืนยัน VAT GL balance matches VAT return (แบบ ภ.พ.30) → filing deadline Day 15 (paper) / Day 23 (online RD e-filing)"),
   ]},
  {"group":"F. Tax & Compliance","step":"F030",
   "task":"Tax Filing — Form PND1/PND3/PND53 Submission (Day 23)",
   "tcode":"ZAPC001 / RD e-filing","owner":"Tax/Acctg","target":"Day 23",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Finalize PND3/PND53 amounts",        "รวบรวม WHT data ทุก vendor → ตรวจสอบ ZAPC001 final → reconcile to GL one last time before filing"),
     ("Submit via RD e-filing system",      "Login RD e-filing → upload / enter PND3 (บุคคลธรรมดา) + PND53 (นิติบุคคล) → submit before Day 23"),
     ("Confirm filing receipt",             "Download e-filing receipt / confirmation → attach to month-end working paper as evidence"),
     ("Post WHT payment if applicable",     "FB05 → post WHT payment to RD → Dr WHT Payable / Cr Bank → verify GL cleared"),
   ]},
  {"group":"F. Tax & Compliance","step":"F040",
   "task":"Tax Invoice Report (AP side)",
   "tcode":"S_ALR_87009895","owner":"Tax/Acctg","target":"Day 7",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run S_ALR_87009895 invoice report",  "S_ALR_87009895 → company code + period → generate AP invoice list with VAT amounts"),
     ("Verify all invoices have tax code",   "ตรวจสอบ invoice ทุกใบมี tax code → flag invoices without code → post missing tax adjustments"),
     ("Reconcile to VAT return",            "Total input VAT from report must match PP.30 input VAT section → document any differences"),
   ]},

  # ── G. BS Reconciliation ──────────────────────────────────────────────────
  {"group":"G. BS Reconciliation","step":"G010",
   "task":"Prepaid Expenses Recon (GL 1151030)",
   "tcode":"FAGLL03H","owner":"Acctg","target":"Day 9",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Export GL 1151030 line items",       "FAGLL03H → GL 1151030 (prepaid other) → period + company code → export all postings"),
     ("Match to prepaid amortization schedule","Beg. balance + new prepayments - monthly amortization = End balance → must match GL exactly"),
     ("Verify recurring amortization posted","ตรวจสอบ FBD2 recurring doc ตัดค่าถูกต้องในเดือนนี้ → ถ้าขาด → post manual adjustment"),
   ]},
  {"group":"G. BS Reconciliation","step":"G020",
   "task":"Premium / Security Deposits Recon",
   "tcode":"FAGLL03H","owner":"Acctg","target":"Day 9",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Export deposit GL accounts",         "FAGLL03H → security deposit + premium GL accounts → list all active deposits"),
     ("Match to deposit register",          "ตรวจสอบ deposit ทุกรายการมีสัญญา/ใบเสร็จสนับสนุน → ระบุ deposits ที่ครบกำหนดคืนแล้ว"),
     ("Flag aged deposits > 12 months",     "Deposit ที่ค้างเกิน 12 เดือน → verify ยังมีสัญญา active อยู่ → ถ้าไม่มี → follow up คืนเงิน"),
   ]},
  {"group":"G. BS Reconciliation","step":"G030",
   "task":"Input Tax (Not Yet Due) Recon",
   "tcode":"FAGLL03H","owner":"Tax/Acctg","target":"Day 9",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Export deferred input VAT GL",       "FAGLL03H → input tax not yet due GL → verify items with future VAT claim dates"),
     ("Clear items where VAT is now due",   "ระบุ items ที่ครบกำหนดแล้วในเดือนนี้ → transfer to claimable VAT account → clear deferred balance"),
     ("Reconcile to pending invoice list",  "Balance ใน deferred VAT must match open invoices awaiting VAT claim → document aged items"),
   ]},
  {"group":"G. BS Reconciliation","step":"G040",
   "task":"Accrued Expenses Recon",
   "tcode":"FAGLL03H","owner":"Acctg","target":"Day 9",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Export all accrued expense GL accounts","FAGLL03H → all accrued liability accounts → export current month postings"),
     ("Match each accrual to support doc",  "ทุก accrual ต้องมี: calculation sheet, approval email, reversal document → verify all present"),
     ("Verify prior month reversals posted","ตรวจสอบว่า reversal จากเดือนก่อนถูก post ในวันที่ 1 → ถ้าขาด → post manual reversal ทันที"),
     ("Investigate unusual movements",     "Flag items ที่ increase > 20% vs prior month → ขอ explanation จาก owner → document in working paper"),
   ]},
  {"group":"G. BS Reconciliation","step":"G050",
   "task":"Leasing / ROU Recon",
   "tcode":"FAGLL03H","owner":"Acctg","target":"Day 9",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Export lease liability GL",          "FAGLL03H → lease liability GL accounts (current + non-current) → verify monthly payment and interest postings"),
     ("Reconcile to amortization schedule", "Beg. balance - principal payment + new leases = End balance → match to lease amortization table"),
     ("Verify IFRS16 current/non-current split","Portion due within 12 months = current → beyond 12 months = non-current → verify BS presentation correct"),
   ]},
  {"group":"G. BS Reconciliation","step":"G060",
   "task":"EBO Reconciliation",
   "tcode":"FAGLL03H","owner":"HR/Acctg","target":"Day 9",
   "amc":True,"ga":True,"stc":False,
   "subtasks":[
     ("Export EBO liability GL",            "FAGLL03H → EBO / employee benefit liability GL → current month postings"),
     ("Match to HR/actuary schedule",       "GL balance must match EBO schedule from HR/actuary → variance > 5% requires explanation"),
     ("Document assumption changes",        "บันทึก changes: discount rate, mortality table, salary increase % → อธิบาย impact ต่อ EBO amount"),
   ]},

  # ── H. Financial Statement ────────────────────────────────────────────────
  {"group":"H. Financial Statement","step":"H010",
   "task":"Trial Balance / GL Summary Report (F.01)",
   "tcode":"F.01","owner":"Acctg","target":"Day 10",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Run F.01 Trial Balance",             "F.01 → company code + financial statement version + period → generate trial balance → verify Dr = Cr"),
     ("Check unassigned GL accounts",       "Accounts in 'Not Assigned' section → assign to FSV node in OB58 → re-run until no unassigned items"),
     ("Export to Excel for review",         "Export F.01 → Excel → compare to prior month + budget → flag variances > 10% for investigation"),
     ("Confirm Balance Sheet balances",     "Assets = Liabilities + Equity → verify net income flows to retained earnings correctly"),
   ]},
  {"group":"H. Financial Statement","step":"H020",
   "task":"GL Line Item Detail Review",
   "tcode":"FAGLL03H / FBL5N / FBL1H","owner":"Acctg","target":"Day 10",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("FAGLL03H key GL accounts review",    "FAGLL03H → review significant accounts: cash, AR, AP, accruals → look for large/unusual items"),
     ("AR subledger reconciliation",        "FBL5H → total customer balances must equal GL AR control account → investigate variance immediately"),
     ("AP subledger reconciliation",        "FBL1H → total vendor balances must equal GL AP control account → verify no unmatched items"),
     ("Investigate large manual entries",   "Filter document type: SA (manual GL) → review all manual entries > 100K → verify support docs attached"),
   ]},
  {"group":"H. Financial Statement","step":"H030",
   "task":"Print Journal Vouchers (JV / PV / RV)",
   "tcode":"ZGLF001 / ZAPF001 / ZARF001","owner":"Acctg","target":"Day 10",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Print JV — Journal Vouchers (ZGLF001)","ZGLF001 → company code + period → print all manual GL postings → attach support docs → get approver signature"),
     ("Print PV — Payment Vouchers (ZAPF001)","ZAPF001 → company code + period → print payment vouchers → verify payee, amount, bank details"),
     ("Print RV — Receipt Vouchers (ZARF001)","ZARF001 → print receipt vouchers → verify all customer receipts documented"),
     ("Archive physical/digital copies",    "Organize vouchers by company + month → scan and upload to shared drive → maintain 7-year retention policy"),
   ]},

  # ── I. Audit Support ──────────────────────────────────────────────────────
  {"group":"I. Audit Support","step":"I010",
   "task":"Prepare Audit Support List (PBC — Provided By Client)",
   "tcode":"— (Manual / F.01 / FAGLL03H)","owner":"Acctg/Audit","target":"Day 15",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Prepare Financial Statements",       "F.01 → export BS + P&L → format with prior period comparison → include GL account codes and descriptions"),
     ("Prepare GL detail schedules",        "FAGLL03H → significant accounts → export line item detail → reconcile total to F.01 balance"),
     ("Compile accrual support",            "รวบรวม: accrual calculations, approval emails, reversal documents → organize by account → label clearly"),
     ("Prepare subledger reconciliations",  "AR (FBL5H), AP (FBL1H), FA register (YAARP001) → reconcile each to GL control account → package as reconciliation schedules"),
     ("Prepare intercompany reconciliation","ตรวจสอบ AMC-GA-STC intercompany balances → AR ที่ entity หนึ่ง = AP ที่อีก entity → net zero → document timing differences"),
   ]},
  {"group":"I. Audit Support","step":"I020",
   "task":"Leadsheet Review & Sign-off",
   "tcode":"— (Manual)","owner":"Controller/Mgr","target":"Day 15",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Controller reviews all schedules",   "Controller signs off on: FS, reconciliations, accrual support, subledger recons → document review date"),
     ("Note open items for auditors",       "List items still pending / estimates → document assumptions → communicate to audit team proactively"),
     ("Respond to auditor queries",         "Track auditor PBC request list → assign owner → set deadline → follow up until all items closed"),
   ]},
  {"group":"I. Audit Support","step":"I030",
   "task":"WHT Annual TDS Form 50 (End of Year Only)",
   "tcode":"ZAPC001","owner":"Tax/Acctg","target":"Day 31/5",
   "amc":False,"ga":False,"stc":False,
   "subtasks":[
     ("Generate annual WHT summary",        "ZAPC001 → full year → generate PND50 for all employees → verify against payroll tax records"),
     ("Issue WHT certificates to employees","ออก ใบรับรองการหักภาษี ณ ที่จ่าย (TDS certificate) ให้พนักงานทุกคน → deadline: 31 March"),
     ("Submit PND50 to Revenue Department", "Submit PND50 via RD e-filing → attach annual summary → keep filing receipt"),
   ]},

  # ── J. Period Close FI ────────────────────────────────────────────────────
  {"group":"J. Period Close FI","step":"J010",
   "task":"Lock Period FI (OB52) — Close AR / AP / GL / FA",
   "tcode":"OB52","owner":"Controlling","target":"End",
   "amc":True,"ga":True,"stc":True,
   "subtasks":[
     ("Close AR/AP/MM (D/K/M account types)","OB52 → posting period variant → update 'From Per.2' for Account Types D (Customer), K (Vendor), M (Material) → save → verify postings fail for prior period"),
     ("Close GL/FA (S/A account types)",    "OB52 → update 'From Per.2' for Account Types S (G/L) and A (Assets) → do this AFTER all GL and FA postings complete"),
     ("Verify period locked",               "Attempt test posting to closed period → must get error 'Posting period X/YYYY is not open' → screenshot as evidence"),
     ("Lock CO period (OKP1)",              "OKP1 → company code + fiscal year → lock current CO period → open next period → verify CO postings fail for locked period"),
   ]},
]

GL_GROUP_COLOURS = {
    "A. GR/IR & AP Clearing":  "FF154360",
    "B. Recurring SAP":        "FF145A32",
    "C. Accrued Expenses":     "FF4A235A",
    "D. Lease / ROU":          "FF7B2C2C",
    "E. Fixed Assets":         "FF6E2F1A",
    "F. Tax & Compliance":     "FF7E5109",
    "G. BS Reconciliation":    "FF1B4F72",
    "H. Financial Statement":  "FF212F3D",
    "I. Audit Support":        "FF424949",
    "J. Period Close FI":      "FF1F4E79",
}

# column indices for GL Master (1-based)
# A=1 No., B=2 Group, C=3 Step, D=4 Task, E=5 How-to, F=6 T-Code,
# G=7 Owner, H=8 Target Day, I=9 AMC, J=10 GA, K=11 STC, L=12 Remark
GL_STATUS_COLS = {"AMC": 9, "GA": 10, "STC": 11}

# sub-task row fills
SUBTASK_FILL   = PatternFill("solid", fgColor="FFF0F7FF")   # very light blue
SUBTASK_FONT   = Font(name="Calibri", size=8, italic=False, color="FF2C3E50")

# ---------------------------------------------------------------------------
# 2. COLOUR PALETTE
# ---------------------------------------------------------------------------
# Process group header colours (Excel ARGB hex)
GROUP_COLOURS = {
    "1. Initial Review":       "FF1F4E79",  # dark blue
    "2. Template Allocation":  "FF375623",  # dark green
    "3. Allocation Cycle":     "FF375623",
    "4. Revaluation @ Actual": "FF7B2C2C",  # dark red
    "5. Work in Process":      "FF4A235A",  # dark purple
    "6. Variance (Combined)":  "FF7E5109",  # dark amber
    "7. Settlement (Combined)":"FF7E5109",
    "8. Variance (Regular)":   "FF7E5109",
    "9. Settlement (Regular)": "FF7E5109",
    "10. ML Closing":          "FF1B4F72",  # teal
    "11. Reporting":           "FF424949",  # dark gray
    "12. Close & PA":          "FF1F4E79",
}
GROUP_TEXT_COLOUR = "FFFFFFFF"  # white

STATUS_FILLS = {
    "Pass":        PatternFill("solid", fgColor="FF92D050"),  # green
    "In Progress": PatternFill("solid", fgColor="FFFFC000"),  # amber
    "Pending":     PatternFill("solid", fgColor="FFFFFFFF"),  # white
    "N/A":         PatternFill("solid", fgColor="FFD3D3D3"),  # light gray
}

HDR_FILL  = PatternFill("solid", fgColor="FF2E4057")   # sheet header
HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
DATA_FONT = Font(name="Calibri", size=9)
GRP_FONT  = Font(name="Calibri", bold=True, color=GROUP_TEXT_COLOUR, size=9)

THIN  = Side(style="thin",   color="FFB0B0B0")
MED   = Side(style="medium", color="FF808080")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

STATUS_OPTIONS = '"Pending,In Progress,Pass,N/A"'

# Plants per company
PLANTS = {
    "AMC": ["1100", "1200", "1300"],
    "GA":  ["2100", "2200"],
    "STC": ["3100"],
}
ALL_PLANTS = [f"AMC-{p}" for p in PLANTS["AMC"]] + \
             [f"GA-{p}"  for p in PLANTS["GA"]]  + \
             [f"STC-{p}" for p in PLANTS["STC"]]

# ---------------------------------------------------------------------------
# 3. HELPERS
# ---------------------------------------------------------------------------

def _cell(ws, row, col, value=None, font=None, fill=None, alignment=None,
          border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:            c.font            = font
    if fill:            c.fill            = fill
    if alignment:       c.alignment       = alignment
    if border:          c.border          = border
    if number_format:   c.number_format   = number_format
    return c


def _auto_col_width(ws, extra=2):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                # take the longest line in a multiline value
                max_line = max((len(ln) for ln in val.split("\n")), default=0)
                if max_line > max_len:
                    max_len = max_line
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)


def _add_status_dv(ws, col_letter, start_row, end_row):
    dv = DataValidation(
        type="list", formula1=STATUS_OPTIONS,
        allow_blank=True, showDropDown=False,
        sqref=f"{col_letter}{start_row}:{col_letter}{end_row}"
    )
    dv.error   = "Invalid status"
    dv.prompt  = "Select: Pending / In Progress / Pass / N/A"
    ws.add_data_validation(dv)


def _freeze(ws, cell="A3"):
    ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# 4. SHEET BUILDERS
# ---------------------------------------------------------------------------

def build_master_sheet(wb, period_label: str):
    """
    CO Close — Master: 15-column layout with sub-task rows.

    Columns:
      A=No.  B=Process Group  C=Step  D=Task(EN)  E=How-to/Detail
      F=Task(TH)  G=T-Code  H=Owner
      I=AMC-1100  J=AMC-1200  K=AMC-1300  L=GA-2100  M=GA-2200  N=STC-3100
      O=Remark
    """
    ws = wb.create_sheet("CO Close \u2014 Master")
    ws.sheet_view.showGridLines = False

    NUM_COLS = 15  # A-O

    # ── Title row 1 ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    _cell(ws, 1, 1,
          value=f"Month-End Costing Close Checklist  |  {period_label}",
          font=Font(name="Calibri", bold=True, size=13, color="FFFFFFFF"),
          fill=PatternFill("solid", fgColor="FF1F3864"),
          alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    # ── Column headers row 2 ───────────────────────────────────────────────
    HDR_LABELS = ["No.", "Process Group", "Step", "Task (EN)", "How-to / Detail",
                  "Task (TH)", "T-Code", "Owner",
                  "AMC-1100", "AMC-1200", "AMC-1300",
                  "GA-2100",  "GA-2200",
                  "STC-3100", "Remark"]
    for ci, hdr in enumerate(HDR_LABELS, start=1):
        _cell(ws, 2, ci, value=hdr,
              font=HDR_FONT, fill=HDR_FILL,
              alignment=Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
              border=THIN_BORDER)
    ws.row_dimensions[2].height = 32

    # ── Column widths ──────────────────────────────────────────────────────
    #       A    B     C    D     E     F     G     H    I    J    K    L    M    N    O
    col_widths = [5, 22, 6, 40, 52, 32, 20, 12, 11, 11, 11, 11, 11, 11, 26]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows ──────────────────────────────────────────────────────────
    # Plant status cols: I=9..N=14
    plant_cols = {
        "AMC-1100":  9, "AMC-1200": 10, "AMC-1300": 11,
        "GA-2100":  12, "GA-2200":  13,
        "STC-3100": 14,
    }
    REMARK_COL = 15

    current_group    = None
    row_num          = 3
    step_count       = 0
    status_start_row = row_num

    for s in STEPS:
        grp      = s["group"]
        step     = s["step"]
        task_en  = s["task"]
        task_th  = s["task_th"]
        tcode    = s["tcode"]
        owner    = s["owner"]
        use_map  = {
            "AMC-1100": s["amc"], "AMC-1200": s["amc"], "AMC-1300": s["amc"],
            "GA-2100":  s["ga"],  "GA-2200":  s["ga"],
            "STC-3100": s["stc"],
        }
        subtasks = s.get("subtasks", [])

        # ── Group header row ──────────────────────────────────────────────
        if grp != current_group:
            current_group = grp
            grp_fill = PatternFill("solid", fgColor=GROUP_COLOURS.get(grp, "FF333333"))
            ws.merge_cells(f"A{row_num}:{get_column_letter(NUM_COLS)}{row_num}")
            _cell(ws, row_num, 1, value=f"  {grp}",
                  font=GRP_FONT, fill=grp_fill,
                  alignment=Alignment(vertical="center"),
                  border=THIN_BORDER)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # ── Main step row ─────────────────────────────────────────────────
        step_count += 1
        row_fill = (PatternFill("solid", fgColor="FFF7F9FC") if step_count % 2 == 0
                    else PatternFill("solid", fgColor="FFFFFFFF"))

        #  A            B     C     D        E(blank)  F        G      H
        data = [step_count, grp, step, task_en, "",   task_th, tcode, owner]
        for ci, val in enumerate(data, start=1):
            _cell(ws, row_num, ci, value=val,
                  font=DATA_FONT, fill=row_fill,
                  alignment=Alignment(vertical="center",
                                      wrap_text=(ci in (4, 5, 6))),
                  border=THIN_BORDER)

        # Status cols I-N
        for pname, pcol in plant_cols.items():
            init_val = "Pending" if use_map[pname] else "N/A"
            pfill    = row_fill if use_map[pname] else STATUS_FILLS["N/A"]
            _cell(ws, row_num, pcol, value=init_val,
                  font=Font(name="Calibri", size=9, bold=True),
                  fill=pfill,
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=THIN_BORDER)

        # Remark col O
        _cell(ws, row_num, REMARK_COL, font=DATA_FONT, fill=row_fill, border=THIN_BORDER)

        ws.row_dimensions[row_num].height = 26
        row_num += 1

        # ── Subtask rows ──────────────────────────────────────────────────
        for sub_desc, sub_howto in subtasks:
            for ci in range(1, NUM_COLS + 1):
                _cell(ws, row_num, ci,
                      font=SUBTASK_FONT, fill=SUBTASK_FILL,
                      border=THIN_BORDER,
                      alignment=Alignment(vertical="center", wrap_text=True))

            # D = sub-task description (indented, italic blue)
            ws.cell(row=row_num, column=4).value = f"  \u2192 {sub_desc}"
            ws.cell(row=row_num, column=4).font  = Font(
                name="Calibri", size=8, italic=True, color="FF1A5276")
            ws.cell(row=row_num, column=4).alignment = Alignment(
                vertical="center", wrap_text=True, indent=1)

            # E = how-to detail
            ws.cell(row=row_num, column=5).value = sub_howto
            ws.cell(row=row_num, column=5).font  = Font(
                name="Calibri", size=8, color="FF2C3E50")
            ws.cell(row=row_num, column=5).alignment = Alignment(
                vertical="top", wrap_text=True)

            ws.row_dimensions[row_num].height = 40
            row_num += 1

    status_end_row = row_num - 1

    # ── Data validation + conditional formatting for plant status cols ─────
    for pcol in plant_cols.values():   # 9-14
        cl = get_column_letter(pcol)
        _add_status_dv(ws, cl, status_start_row, status_end_row)
        rng = f"{cl}{status_start_row}:{cl}{status_end_row}"
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"Pass"'],
                       fill=STATUS_FILLS["Pass"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"In Progress"'],
                       fill=STATUS_FILLS["In Progress"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"N/A"'],
                       fill=STATUS_FILLS["N/A"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"Pending"'],
                       fill=STATUS_FILLS["Pending"]))

    # ── Progress header row 2 — matching GL Master style ──────────────────
    for pname, pcol in plant_cols.items():
        cl = get_column_letter(pcol)
        s  = status_start_row
        e  = status_end_row
        formula = (
            f'=IFERROR('
            f'COUNTIF({cl}{s}:{cl}{e},"Pass")'
            f'/(COUNTIF({cl}{s}:{cl}{e},"Pending")'
            f'+COUNTIF({cl}{s}:{cl}{e},"In Progress")'
            f'+COUNTIF({cl}{s}:{cl}{e},"Pass")),0)'
        )
        c = ws.cell(row=2, column=pcol, value=formula)
        c.font          = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        c.fill          = PatternFill("solid", fgColor="FF1F3864")
        c.number_format = "0%"
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = THIN_BORDER

    _freeze(ws, "D3")
    ws.print_title_rows = "1:2"
    return ws, status_start_row, status_end_row


def build_company_sheet(wb, company: str, plants: list,
                        period_label: str,
                        master_sheet_name: str = "CO Close \u2014 Master"):
    """
    One sheet per company — step rows + subtask child rows, plant status columns.

    Columns:
      A=#  B=Group  C=Step  D=Task(EN)  E=How-to/Detail  F=T-Code  G=Owner
      H+=plant status (one col per plant)
    """
    ws = wb.create_sheet(company)
    ws.sheet_view.showGridLines = False

    n_plants   = len(plants)
    PLANT_START = 8                  # first plant status col (H)
    last_col    = 7 + n_plants       # A-G fixed + plant cols

    # ── Title row 1 ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    _cell(ws, 1, 1,
          value=f"{company} \u2014 Costing Close  |  {period_label}",
          font=Font(name="Calibri", bold=True, size=12, color="FFFFFFFF"),
          fill=PatternFill("solid", fgColor="FF1F3864"),
          alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 26

    # ── Progress row 2 — filled after data rows ────────────────────────────
    ws.merge_cells("A2:G2")
    _cell(ws, 2, 1, value="Progress",
          font=HDR_FONT, fill=HDR_FILL,
          alignment=Alignment(horizontal="center", vertical="center"),
          border=THIN_BORDER)
    for pi, plant in enumerate(plants):
        pc = PLANT_START + pi
        _cell(ws, 2, pc, value=f"{company}-{plant}",
              font=HDR_FONT, fill=HDR_FILL,
              alignment=Alignment(horizontal="center"),
              border=THIN_BORDER)
    ws.row_dimensions[2].height = 20

    # ── Column headers row 3 ───────────────────────────────────────────────
    hdrs = ["#", "Group", "Step", "Task (EN)", "How-to / Detail",
            "T-Code", "Owner"] + [f"{company}-{p}" for p in plants]
    for ci, h in enumerate(hdrs, start=1):
        _cell(ws, 3, ci, value=h,
              font=HDR_FONT, fill=HDR_FILL,
              alignment=Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
              border=THIN_BORDER)
    ws.row_dimensions[3].height = 26

    # ── Column widths ──────────────────────────────────────────────────────
    #       A    B     C    D     E     F     G
    widths = [4, 22, 6, 40, 52, 18, 12] + [11] * n_plants
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Helpers ────────────────────────────────────────────────────────────
    def applies(s):
        if company == "AMC": return s["amc"]
        if company == "GA":  return s["ga"]
        if company == "STC": return s["stc"]
        return True

    current_group = None
    row_num       = 4
    step_count    = 0
    data_start    = row_num

    # ── Data rows ──────────────────────────────────────────────────────────
    for s in STEPS:
        grp      = s["group"]
        step     = s["step"]
        task_en  = s["task"]
        tcode    = s["tcode"]
        owner    = s["owner"]
        subtasks = s.get("subtasks", [])

        # Group header
        if grp != current_group:
            current_group = grp
            grp_fill = PatternFill("solid", fgColor=GROUP_COLOURS.get(grp, "FF333333"))
            ws.merge_cells(f"A{row_num}:{get_column_letter(last_col)}{row_num}")
            _cell(ws, row_num, 1, value=f"  {grp}",
                  font=GRP_FONT, fill=grp_fill,
                  alignment=Alignment(vertical="center"),
                  border=THIN_BORDER)
            ws.row_dimensions[row_num].height = 16
            row_num += 1

        # Main step row
        step_count += 1
        row_fill = (PatternFill("solid", fgColor="FFF7F9FC") if step_count % 2 == 0
                    else PatternFill("solid", fgColor="FFFFFFFF"))

        #  A            B     C     D         E(blank)  F      G
        data = [step_count, grp, step, task_en, "",    tcode, owner]
        for ci, val in enumerate(data, start=1):
            _cell(ws, row_num, ci, value=val,
                  font=DATA_FONT, fill=row_fill,
                  alignment=Alignment(vertical="center",
                                      wrap_text=(ci in (4, 5))),
                  border=THIN_BORDER)

        is_applicable = applies(s)
        for pi in range(n_plants):
            pc = PLANT_START + pi
            init_val = "Pending" if is_applicable else "N/A"
            pf       = row_fill if is_applicable else STATUS_FILLS["N/A"]
            _cell(ws, row_num, pc, value=init_val,
                  font=Font(name="Calibri", size=9, bold=True),
                  fill=pf,
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=THIN_BORDER)

        ws.row_dimensions[row_num].height = 26
        row_num += 1

        # Subtask rows
        for sub_desc, sub_howto in subtasks:
            for ci in range(1, last_col + 1):
                _cell(ws, row_num, ci,
                      font=SUBTASK_FONT, fill=SUBTASK_FILL,
                      border=THIN_BORDER,
                      alignment=Alignment(vertical="center", wrap_text=True))

            # D = description
            ws.cell(row=row_num, column=4).value = f"  \u2192 {sub_desc}"
            ws.cell(row=row_num, column=4).font  = Font(
                name="Calibri", size=8, italic=True, color="FF1A5276")
            ws.cell(row=row_num, column=4).alignment = Alignment(
                vertical="center", wrap_text=True, indent=1)

            # E = how-to detail
            ws.cell(row=row_num, column=5).value = sub_howto
            ws.cell(row=row_num, column=5).font  = Font(
                name="Calibri", size=8, color="FF2C3E50")
            ws.cell(row=row_num, column=5).alignment = Alignment(
                vertical="top", wrap_text=True)

            ws.row_dimensions[row_num].height = 40
            row_num += 1

    data_end = row_num - 1

    # ── Data validation + conditional formatting ───────────────────────────
    for pi in range(n_plants):
        cl  = get_column_letter(PLANT_START + pi)
        _add_status_dv(ws, cl, data_start, data_end)
        rng = f"{cl}{data_start}:{cl}{data_end}"
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"Pass"'],
                       fill=STATUS_FILLS["Pass"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"In Progress"'],
                       fill=STATUS_FILLS["In Progress"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"N/A"'],
                       fill=STATUS_FILLS["N/A"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"Pending"'],
                       fill=STATUS_FILLS["Pending"]))

    # ── Progress formula row 2 — exclude subtask blank rows ───────────────
    for pi, plant in enumerate(plants):
        pc = PLANT_START + pi
        cl = get_column_letter(pc)
        s, e = data_start, data_end
        formula = (
            f'=IFERROR('
            f'COUNTIF({cl}{s}:{cl}{e},"Pass")'
            f'/(COUNTIF({cl}{s}:{cl}{e},"Pending")'
            f'+COUNTIF({cl}{s}:{cl}{e},"In Progress")'
            f'+COUNTIF({cl}{s}:{cl}{e},"Pass")),0)'
        )
        c = ws.cell(row=2, column=pc, value=formula)
        c.font          = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        c.fill          = PatternFill("solid", fgColor="FF1F3864")
        c.number_format = "0%"
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = THIN_BORDER

    _freeze(ws, "D4")
    return ws


def build_gl_master_sheet(wb, period_label: str):
    """
    GL Close — Master: 12-column layout with sub-task rows.

    Columns:
      A=No.  B=Group  C=Step  D=Task(EN)  E=How-to/Detail  F=T-Code
      G=Owner  H=Target Day  I=AMC  J=GA  K=STC  L=Remark
    """
    ws = wb.create_sheet("GL Close \u2014 Master")
    ws.sheet_view.showGridLines = False

    NUM_COLS = 12  # A-L

    # ── Title row 1 ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    _cell(ws, 1, 1,
          value=f"Month-End GL Close Checklist  |  {period_label}",
          font=Font(name="Calibri", bold=True, size=13, color="FFFFFFFF"),
          fill=PatternFill("solid", fgColor="FF154360"),
          alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    # ── Column headers row 2 ───────────────────────────────────────────────
    HDR_LABELS = ["No.", "Group", "Step", "Task (EN)", "How-to / Detail",
                  "T-Code", "Owner", "Target Day", "AMC", "GA", "STC", "Remark"]
    for ci, hdr in enumerate(HDR_LABELS, start=1):
        _cell(ws, 2, ci, value=hdr,
              font=HDR_FONT,
              fill=PatternFill("solid", fgColor="FF1F3864"),
              alignment=Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
              border=THIN_BORDER)
    ws.row_dimensions[2].height = 32

    # ── Column widths ──────────────────────────────────────────────────────
    #       A    B     C    D     E     F     G    H    I    J    K    L
    col_widths = [5, 22, 7, 42, 52, 22, 14, 11, 11, 11, 11, 28]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows ──────────────────────────────────────────────────────────
    # GL_STATUS_COLS = {"AMC": 9, "GA": 10, "STC": 11}
    company_cols = GL_STATUS_COLS   # {AMC:9, GA:10, STC:11}
    REMARK_COL   = 12

    current_group = None
    row_num       = 3
    step_count    = 0
    status_start_row = row_num

    for s in GL_STEPS:
        grp        = s["group"]
        step       = s["step"]
        task_en    = s["task"]
        tcode      = s["tcode"]
        owner      = s["owner"]
        target_day = s["target"]
        use_map    = {"AMC": s["amc"], "GA": s["ga"], "STC": s["stc"]}
        subtasks   = s.get("subtasks", [])

        # ── Group header row ──────────────────────────────────────────────
        if grp != current_group:
            current_group = grp
            grp_fill = PatternFill("solid",
                                   fgColor=GL_GROUP_COLOURS.get(grp, "FF2C3E50"))
            ws.merge_cells(f"A{row_num}:{get_column_letter(NUM_COLS)}{row_num}")
            _cell(ws, row_num, 1, value=f"  {grp}",
                  font=GRP_FONT, fill=grp_fill,
                  alignment=Alignment(vertical="center"),
                  border=THIN_BORDER)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # ── Main step row ─────────────────────────────────────────────────
        step_count += 1
        row_fill = (PatternFill("solid", fgColor="FFF7F9FC") if step_count % 2 == 0
                    else PatternFill("solid", fgColor="FFFFFFFF"))

        #  A   B      C     D        E(blank)  F      G      H
        data = [step_count, grp, step, task_en, "",   tcode, owner, target_day]
        for ci, val in enumerate(data, start=1):
            _cell(ws, row_num, ci, value=val,
                  font=DATA_FONT, fill=row_fill,
                  alignment=Alignment(vertical="center",
                                      wrap_text=(ci in (4, 5))),
                  border=THIN_BORDER)

        # Status cols I / J / K
        for cname, ccol in company_cols.items():
            init_val = "Pending" if use_map[cname] else "N/A"
            pfill    = row_fill if use_map[cname] else STATUS_FILLS["N/A"]
            _cell(ws, row_num, ccol, value=init_val,
                  font=Font(name="Calibri", size=9, bold=True),
                  fill=pfill,
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=THIN_BORDER)

        # Remark col L
        _cell(ws, row_num, REMARK_COL, font=DATA_FONT, fill=row_fill, border=THIN_BORDER)

        ws.row_dimensions[row_num].height = 26
        row_num += 1

        # ── Subtask rows ──────────────────────────────────────────────────
        for sub_desc, sub_howto in subtasks:
            # A: blank (no number), B-C: blank, D: indented description, E: how-to
            # F-H: blank, I-K: blank (no status), L: blank
            for ci in range(1, NUM_COLS + 1):
                _cell(ws, row_num, ci,
                      font=SUBTASK_FONT, fill=SUBTASK_FILL,
                      border=THIN_BORDER,
                      alignment=Alignment(vertical="center", wrap_text=True))

            # D = sub-task description
            ws.cell(row=row_num, column=4).value = f"  \u2192 {sub_desc}"
            ws.cell(row=row_num, column=4).font  = Font(
                name="Calibri", size=8, italic=True, color="FF1A5276")
            ws.cell(row=row_num, column=4).alignment = Alignment(
                vertical="center", wrap_text=True, indent=1)

            # E = how-to detail
            ws.cell(row=row_num, column=5).value = sub_howto
            ws.cell(row=row_num, column=5).font  = Font(
                name="Calibri", size=8, color="FF2C3E50")
            ws.cell(row=row_num, column=5).alignment = Alignment(
                vertical="top", wrap_text=True)

            ws.row_dimensions[row_num].height = 42
            row_num += 1

    status_end_row = row_num - 1

    # ── Data validation + conditional formatting for status cols ──────────
    for ccol in company_cols.values():   # 9, 10, 11
        cl = get_column_letter(ccol)
        _add_status_dv(ws, cl, status_start_row, status_end_row)
        rng = f"{cl}{status_start_row}:{cl}{status_end_row}"
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"Pass"'],
                       fill=STATUS_FILLS["Pass"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"In Progress"'],
                       fill=STATUS_FILLS["In Progress"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"N/A"'],
                       fill=STATUS_FILLS["N/A"]))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="equal", formula=['"Pending"'],
                       fill=STATUS_FILLS["Pending"]))

    # ── Progress header in row 2 — match CO Master style ──────────────────
    # Use COUNTIF("Pass") / (COUNTIF("Pending")+COUNTIF("In Progress")+COUNTIF("Pass"))
    # so that subtask rows (blank) are excluded from denominator
    for cname, ccol in company_cols.items():
        cl = get_column_letter(ccol)
        s  = status_start_row
        e  = status_end_row
        formula = (
            f'=IFERROR('
            f'COUNTIF({cl}{s}:{cl}{e},"Pass")'
            f'/(COUNTIF({cl}{s}:{cl}{e},"Pending")'
            f'+COUNTIF({cl}{s}:{cl}{e},"In Progress")'
            f'+COUNTIF({cl}{s}:{cl}{e},"Pass")),0)'
        )
        c = ws.cell(row=2, column=ccol, value=formula)
        c.font          = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        c.fill          = PatternFill("solid", fgColor="FF1F3864")
        c.number_format = "0%"
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = THIN_BORDER

    _freeze(ws, "D3")
    ws.print_title_rows = "1:2"
    return ws, status_start_row, status_end_row


def build_dashboard(wb, period_label: str, master_ws_name: str = "CO Close — Master"):
    ws = wb.create_sheet("Dashboard", 0)   # insert first
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    # Title
    ws.merge_cells("B1:M1")
    _cell(ws, 1, 2,
          value=f"Month-End Close Checklist  —  {period_label}",
          font=Font(name="Calibri", bold=True, size=16, color="FFFFFFFF"),
          fill=PatternFill("solid", fgColor="FF1F3864"),
          alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 36

    # Sub-title
    ws.merge_cells("B2:M2")
    _cell(ws, 2, 2, value="CO Costing Close + GL Close  (AMC / GA / STC)",
          font=Font(name="Calibri", italic=True, size=10, color="FF4A4A4A"),
          alignment=Alignment(horizontal="center"))
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6  # spacer

    # ── Section header helper ───────────────────────────────────────────────
    def _section_hdr(row, label, colour):
        ws.merge_cells(f"B{row}:M{row}")
        _cell(ws, row, 2, value=f"  {label}",
              font=Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
              fill=PatternFill("solid", fgColor=colour),
              alignment=Alignment(vertical="center"))
        ws.row_dimensions[row].height = 20

    # ── Batch helper: one row per plant/company ─────────────────────────────
    def _status_row(row, label, pct_formula, badge_formula,
                    label_col=2, pct_col=7, badge_col=10,
                    label_span="B:F", pct_span="G:I", badge_span="J:M",
                    row_fill="FFF0F4F8"):
        # label
        ws.merge_cells(f"{label_span.split(':')[0]}{row}:{label_span.split(':')[1]}{row}")
        _cell(ws, row, label_col, value=f"  {label}",
              font=Font(name="Calibri", size=10),
              fill=PatternFill("solid", fgColor=row_fill),
              alignment=Alignment(vertical="center"),
              border=THIN_BORDER)
        # pct
        c = ws.cell(row=row, column=pct_col, value=pct_formula)
        c.font          = Font(name="Calibri", bold=True, size=10)
        c.number_format = "0%"
        c.alignment     = Alignment(horizontal="center", vertical="center")
        c.border        = THIN_BORDER
        pct_end = pct_span.split(":")[1]
        ws.merge_cells(f"{get_column_letter(pct_col)}{row}:{pct_end}{row}")
        # badge
        c2 = ws.cell(row=row, column=badge_col, value=badge_formula)
        c2.font       = Font(name="Calibri", bold=True, size=9)
        c2.alignment  = Alignment(horizontal="center", vertical="center")
        c2.border     = THIN_BORDER
        badge_end = badge_span.split(":")[1]
        ws.merge_cells(f"{get_column_letter(badge_col)}{row}:{badge_end}{row}")
        # CF on badge
        brange = f"{get_column_letter(badge_col)}{row}:{badge_end}{row}"
        ws.conditional_formatting.add(brange,
            CellIsRule(operator="equal", formula=['"COMPLETE"'],
                       fill=STATUS_FILLS["Pass"]))
        ws.conditional_formatting.add(brange,
            CellIsRule(operator="equal", formula=['"IN PROGRESS"'],
                       fill=STATUS_FILLS["In Progress"]))
        ws.row_dimensions[row].height = 22

    comp_colour = {"AMC": "FF2E4057", "GA": "FF375623", "STC": "FF7B2C2C"}

    # ── CO COSTING section ──────────────────────────────────────────────────
    cur = 4
    _section_hdr(cur, "CO COSTING CLOSE  (per Plant)", "FF1F4E79")
    cur += 1

    co_companies = {
        "AMC": ["1100", "1200", "1300"],
        "GA":  ["2100", "2200"],
        "STC": ["3100"],
    }
    for company, plants in co_companies.items():
        # Company sub-header
        ws.merge_cells(f"B{cur}:M{cur}")
        _cell(ws, cur, 2, value=f"  {company}",
              font=Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
              fill=PatternFill("solid", fgColor=comp_colour[company]),
              alignment=Alignment(vertical="center"),
              border=THIN_BORDER)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for plant in plants:
            cl_plant = get_column_letter(8 + plants.index(plant))  # H+ after How-to col added
            pct_f  = (f"=IFERROR("
                      f"COUNTIF('{company}'!{cl_plant}4:{cl_plant}500,\"Pass\")"
                      f"/(COUNTIF('{company}'!{cl_plant}4:{cl_plant}500,\"Pending\")"
                      f"+COUNTIF('{company}'!{cl_plant}4:{cl_plant}500,\"In Progress\")"
                      f"+COUNTIF('{company}'!{cl_plant}4:{cl_plant}500,\"Pass\")),0)")
            badge_f = f'=IF(G{cur}=1,"COMPLETE",IF(G{cur}>0,"IN PROGRESS","PENDING"))'
            _status_row(cur, f"Plant {plant}", pct_f, badge_f)
            cur += 1

    ws.row_dimensions[cur].height = 6  # spacer
    cur += 1

    # ── GL CLOSE section ────────────────────────────────────────────────────
    _section_hdr(cur, "GL CLOSE  (per Company)", "FF154360")
    cur += 1

    gl_col_map = {"AMC": 9, "GA": 10, "STC": 11}  # in GL Close-Master sheet (I/J/K)
    for company in ("AMC", "GA", "STC"):
        ccol = gl_col_map[company]
        cl_co = get_column_letter(ccol)
        pct_f  = (f"=IFERROR("
                  f"COUNTIF('GL Close \u2014 Master'!{cl_co}3:{cl_co}500,\"Pass\")"
                  f"/(COUNTIF('GL Close \u2014 Master'!{cl_co}3:{cl_co}500,\"Pending\")"
                  f"+COUNTIF('GL Close \u2014 Master'!{cl_co}3:{cl_co}500,\"In Progress\")"
                  f"+COUNTIF('GL Close \u2014 Master'!{cl_co}3:{cl_co}500,\"Pass\")),0)")
        badge_f = f'=IF(G{cur}=1,"COMPLETE",IF(G{cur}>0,"IN PROGRESS","PENDING"))'
        ws.merge_cells(f"B{cur}:M{cur}")
        _cell(ws, cur, 2, value=f"  {company}",
              font=Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
              fill=PatternFill("solid", fgColor=comp_colour[company]),
              alignment=Alignment(vertical="center"),
              border=THIN_BORDER)
        ws.row_dimensions[cur].height = 18
        cur += 1
        _status_row(cur, f"{company} GL Close", pct_f, badge_f)
        cur += 1

    ws.row_dimensions[cur].height = 6
    cur += 1

    # ── Instructions ────────────────────────────────────────────────────────
    ws.merge_cells(f"B{cur}:M{cur}")
    _cell(ws, cur, 2, value="HOW TO USE",
          font=Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
          fill=PatternFill("solid", fgColor="FF4A4A4A"),
          alignment=Alignment(horizontal="center"))
    instructions = [
        "1. CO Phase: ไปที่ sheet 'CO Close -- Master' หรือ sheet บริษัท (AMC / GA / STC) -- แก้ Status per plant",
        "2. GL Phase: ไปที่ sheet 'GL Close -- Master' -- แก้ Status per company (AMC / GA / STC)",
        "3. N/A = ขั้นตอนที่ไม่ใช้ -- ไม่ต้องแก้ไข",
        "4. Dashboard คำนวณ % completion อัตโนมัติจากทุก sheet",
        "5. ดู T-Code reference ใน sheet 'T-Code Ref'",
    ]
    cur += 1
    for text in instructions:
        ws.merge_cells(f"B{cur}:M{cur}")
        _cell(ws, cur, 2, value=text,
              font=Font(name="Calibri", size=9),
              fill=PatternFill("solid", fgColor="FFFFFDE7"),
              alignment=Alignment(vertical="center"))
        ws.row_dimensions[cur].height = 18
        cur += 1

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 4
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 4
    ws.column_dimensions["I"].width = 4
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 4
    ws.column_dimensions["L"].width = 4
    ws.column_dimensions["M"].width = 4

    return ws


def build_tcode_sheet(wb):
    ws = wb.create_sheet("T-Code Ref")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    _cell(ws, 1, 1, value="SAP Transaction Code Reference — Costing Close",
          font=Font(name="Calibri", bold=True, size=12, color="FFFFFFFF"),
          fill=PatternFill("solid", fgColor="FF1F3864"),
          alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 26

    hdrs = ["#", "T-Code", "Description (EN)", "Use Case / Phase", "Notes"]
    for ci, h in enumerate(hdrs, start=1):
        _cell(ws, 2, ci, value=h,
              font=HDR_FONT, fill=HDR_FILL,
              alignment=Alignment(horizontal="center", vertical="center"),
              border=THIN_BORDER)
    ws.row_dimensions[2].height = 20

    TCODES = [
        # CO Costing Phase
        ("S_ALR_87013611", "CC Actual/Plan/Variance",          "Initial Review",              ""),
        ("S_ALR_87013628", "CC Actual/Plan Variance (post)",   "Allocation Cycle / Review",   ""),
        ("S_ALR_87013643", "Activity Revaluation Report",      "Revaluation @ Actual",        ""),
        ("CPTD",           "Template Allocation",              "Template Allocation",         "1 รัน ต่อ Plant"),
        ("KSV5",           "Assessment Cycle",                 "Allocation Cycle",            ""),
        ("KSU5",           "Distribution Cycle",               "Allocation Cycle",            ""),
        ("KSS2",           "Actual Cost Splitting",            "Revaluation @ Actual",        ""),
        ("KSII",           "Actual Activity Price Calc",       "Revaluation @ Actual",        ""),
        ("KSBT",           "Activity Price Report",            "Revaluation @ Actual",        ""),
        ("CON2",           "Activity Revaluation @ Actual",    "Revaluation @ Actual",        "Combined + per Plant"),
        ("KKA0",           "Change Cutoff Period",             "WIP",                         "1 ครั้ง ต่อ Plant"),
        ("KKAOH",          "Calculate WIP",                    "WIP",                         ""),
        ("KKAQ",           "Display WIP",                      "WIP",                         ""),
        ("KKS1",           "Variance Calculation",             "Variance Calc",               "Combined + Regular"),
        ("CO88H",          "Production Cost Settlement",       "Actual Settlement",           "Combined + Regular"),
        ("KOC4",           "Order Selection",                  "Settlement — Regular",        "Used with F-02 JE"),
        ("CKMLCP",         "ML Closing / Actual Costing Run",  "ML Closing",                  "AMC + GA only"),
        ("CKMVFM",         "Value Flow Monitor",               "ML Closing",                  ""),
        ("CKMPDB",         "Price Difference Balance",         "ML Closing — Adj GL",         ""),
        ("KE27",           "Periodic Valuation (PA)",          "ML Closing → PA",             ""),
        ("KE24",           "PA Report — Line Item",            "Reporting / Final",           ""),
        ("KE30",           "PA Summary Report",                "Final PA",                    ""),
        ("KE4S",           "Transfer SD Billing to PA",        "Final PA",                    ""),
        ("CO78",           "Delete Combined Production Order", "Close & PA",                  ""),
        ("COHV",           "Close Production Order",           "Close & PA",                  ""),
        ("OKP1",           "Lock Period CO",                   "Period Lock",                 "ทุก company"),
        ("S_P99_41000062", "Materials List: Prices & Inv.",    "Reporting",                   ""),
        ("CKM3N",          "Material Price Analysis",          "Reporting",                   "ML materials"),
        ("S_P00_07000139", "Stock Card (Standard)",            "Reporting",                   ""),
        ("ZCOR001",        "Stock Card (Custom)",              "Reporting",                   "Custom report"),
        # GL Phase (reserved)
        ("FB50",           "GL Journal Entry",                 "GL Phase — Adjust",           ""),
        ("F-02",           "Post General Document",            "GL Phase — Adjust",           ""),
        ("F-50",           "Enter G/L Account Posting",        "GL Phase — Adjust",           ""),
        ("F.01",           "GL Summary / Trial Balance",       "GL Phase — Review",           ""),
        ("FBD1",           "Create Recurring Entry",           "GL Phase — Recurring",        ""),
        ("FBD2",           "Execute Recurring Entry",          "GL Phase — Recurring",        ""),
        ("F.13",           "GR/IR Clearing",                   "GL Phase — Clearing",         ""),
        ("F.14",           "Run Recurring Entry",              "GL Phase — Recurring",        ""),
        ("F.15",           "Accrual List",                     "GL Phase — Accrual",          ""),
        ("F.19",           "GR/IR Clearing 2",                 "GL Phase — Clearing",         ""),
        ("FAGLL03H",       "GL Line Item Display",             "GL Phase — Review",           ""),
        ("ZAPC001",        "WHT Report (Form 3/53)",           "GL Phase — Tax",              ""),
    ]

    # Section headers
    CO_COUNT = 30   # first 30 rows are CO phase
    for i, (tc, desc, phase, note) in enumerate(TCODES, start=1):
        r = i + 2
        section_fill = PatternFill("solid", fgColor="FFEAF4FB") if i <= CO_COUNT \
                       else PatternFill("solid", fgColor="FFFFF9E7")
        row_font = Font(name="Calibri", size=9)
        vals = [i, tc, desc, phase, note]
        for ci, v in enumerate(vals, start=1):
            _cell(ws, r, ci, value=v,
                  font=row_font if ci > 1 else Font(name="Calibri", size=9, bold=True),
                  fill=section_fill,
                  alignment=Alignment(vertical="center", wrap_text=(ci == 3)),
                  border=THIN_BORDER)
        ws.row_dimensions[r].height = 18

    # section label
    sect_r = CO_COUNT + 3
    ws.merge_cells(f"A{sect_r}:E{sect_r}")
    _cell(ws, sect_r, 1, value="  ─── GL Phase T-Codes (Reserved — Phase 2) ───",
          font=Font(name="Calibri", bold=True, italic=True, size=9, color="FF7E5109"),
          fill=PatternFill("solid", fgColor="FFFFF3CD"),
          alignment=Alignment(vertical="center"))

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 28

    _freeze(ws, "A3")
    return ws


# ---------------------------------------------------------------------------
# 5. MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate Month-End Costing Close Template (.xlsx)")
    now = datetime.now()
    parser.add_argument("--month", default=f"{now.month:02d}",
                        help="Month number, e.g. 02  (default: current)")
    parser.add_argument("--year",  default=str(now.year),
                        help="Year 4-digit, e.g. 2026  (default: current)")
    parser.add_argument("--no-upload", action="store_true",
                        help="Skip Google Drive upload (local file only)")
    parser.add_argument("--no-fix",    action="store_true",
                        help="Skip fix_master_formulas after upload")
    args = parser.parse_args()

    month  = args.month.zfill(2)
    year   = args.year
    period = f"{month}/{year}"

    # Output path
    script_dir   = Path(__file__).parent
    workspace    = script_dir.parent
    out_dir      = Path(__file__).resolve().parent.parent / "02_Working"
    out_dir.mkdir(exist_ok=True)
    out_path     = out_dir / f"monthend_close_{month}{year}.xlsx"

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print(f"Building template for period: {period}")

    # 1. Dashboard (inserted at pos 0 inside builder)
    build_dashboard(wb, period)
    print("  [OK] Dashboard")

    # 2. CO Master sheet
    build_master_sheet(wb, period)
    print("  [OK] CO Close -- Master")

    # 3. Company CO sheets (per plant)
    for company, plants in PLANTS.items():
        build_company_sheet(wb, company, plants, period)
        print(f"  [OK] CO -- {company}")

    # 4. GL Master sheet
    build_gl_master_sheet(wb, period)
    print("  [OK] GL Close -- Master")

    # 5. T-Code reference
    build_tcode_sheet(wb)
    print("  [OK] T-Code Ref")

    wb.save(out_path)
    print(f"Saved -> {out_path}")

    # ── Step 6: Upload to Google Drive ────────────────────────────────────
    if args.no_upload:
        print("\n[--no-upload] Skipping Drive upload.")
        return out_path

    print("\n[Drive] Starting upload and folder organization...")
    try:
        from utils.auth import get_credentials
        from utils.drive_organizer import DriveOrganizer
        import gspread

        creds    = get_credentials()
        org      = DriveOrganizer(creds)
        sheet_id = org.upload_and_organize(out_path, month=month, year=year)

        if sheet_id is None:
            print("[Drive] Upload skipped by user.")
            return out_path

        print(f"\n[Drive] Sheet ID: {sheet_id}")

        # ── Step 7: Fix Master formulas ───────────────────────────────────
        if args.no_fix:
            print("[--no-fix] Skipping formula fix.")
            return sheet_id

        print("\n[Formula Fix] Updating CO Close Master to pull from AMC/GA/STC...")
        _fix_master_formulas(sheet_id, creds)
        print("[Formula Fix] Done.")

    except Exception as e:
        print(f"[Drive ERROR] {e}")
        print("Local file still saved at:", out_path)

    return out_path


def _fix_master_formulas(spreadsheet_id: str, creds) -> None:
    """
    Update CO Close — Master status columns (H-M) ให้ใช้ INDEX/MATCH
    ดึงจาก AMC / GA / STC sheets อัตโนมัติ
    """
    import gspread

    MASTER_SHEET = "CO Close \u2014 Master"
    # Master: plant status at I-N (9-14)
    # Company sheets: plant col starts at H (8) after How-to col E added
    FORMULA_MAP  = {
        9:  ("AMC", "H"),   # AMC-1100  Master col I → AMC sheet col H
        10: ("AMC", "I"),   # AMC-1200  Master col J → AMC sheet col I
        11: ("AMC", "J"),   # AMC-1300  Master col K → AMC sheet col J
        12: ("GA",  "H"),   # GA-2100   Master col L → GA  sheet col H
        13: ("GA",  "I"),   # GA-2200   Master col M → GA  sheet col I
        14: ("STC", "H"),   # STC-3100  Master col N → STC sheet col H
    }

    def col_letter(n):
        result = ""
        while n > 0:
            n, rem = divmod(n - 1, 26)
            result = chr(65 + rem) + result
        return result

    def make_formula(row, src_sheet, src_col):
        return (
            f"=IFERROR("
            f"INDEX('{src_sheet}'!${src_col}:${src_col},"
            f"MATCH($C{row},'{src_sheet}'!$C:$C,0)),"
            f"\"N/A\")"
        )

    client     = gspread.authorize(creds)
    ss         = client.open_by_key(spreadsheet_id)
    ws         = ss.worksheet(MASTER_SHEET)
    all_values = ws.get_all_values()

    updates = []
    for i, row_vals in enumerate(all_values):
        sheet_row = i + 1
        col_a     = row_vals[0].strip() if row_vals else ""
        if not col_a.isdigit():
            continue
        for master_col, (src_sheet, src_col) in FORMULA_MAP.items():
            cell_addr = f"{col_letter(master_col)}{sheet_row}"
            formula   = make_formula(sheet_row, src_sheet, src_col)
            updates.append({"range": cell_addr, "values": [[formula]]})

    for start in range(0, len(updates), 500):
        chunk = updates[start: start + 500]
        ws.batch_update(
            [{"range": u["range"], "values": u["values"]} for u in chunk],
            value_input_option="USER_ENTERED"
        )
    print(f"  Updated {len(updates)} formula cells in {MASTER_SHEET}")


if __name__ == "__main__":
    main()
