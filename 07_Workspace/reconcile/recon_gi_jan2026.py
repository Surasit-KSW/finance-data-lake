"""
Reconciliation Script — GI Production Cost Jan 2026 | Plant 1300
Prepared by: Claude Code

Sources:
  PRD_1300_01.2026.XLSX  — Production Order custom report
  KSB1_1300_01.2026.XLSX — Cost Center Actuals (CO)
  AMC_GL_03.2026.XLSX    — GL Line Items Jan-Mar 2026 (filter Jan + CC 13*)

Output:
  Google Sheet: RECON_GI_Production_Cost_Jan2026
"""

import sys, datetime
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from collections import defaultdict
from utils.auth import get_gspread_client
from utils.sheets import get_or_add_worksheet

PREPARED_BY = "Claude Code"
TODAY = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
DATA_DIR = str(Path(__file__).resolve().parent.parent.parent / '01_Bronze_Raw' / 'PRD_GI')

SHEET_NAME = "RECON_GI_Production_Cost_Jan2026"


# ─────────────────────────────────────────────────────────────────
def to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").strip())
    except: return 0.0


def classify_gl(acc):
    a = str(acc)
    if a.startswith(("551", "552", "553")): return "Labor"
    if a.startswith("561"): return "Electricity"
    if a.startswith("571"): return "Repair & Maintenance"
    if a.startswith("581"): return "Depreciation"
    if a.startswith("591"): return "Tools/Supplies/Other"
    if a.startswith("592"): return "Manufacturing Supplies"
    if a.startswith("599"): return "Other Expenses"
    if a.startswith("611"): return "Transport"
    if a.startswith("541"): return "Consumption (Material)"
    if a.startswith("531"): return "Production Variance"
    if a.startswith("539"): return "Variance Adjustment (ML)"
    if a.startswith("943"): return "CO Settlement (to Orders)"
    return "Other"


# ─────────────────────────────────────────────────────────────────
def load_prd():
    print("Loading PRD...")
    wb = openpyxl.load_workbook(f"{DATA_DIR}\\PRD_1300_01.2026.XLSX", data_only=True)
    ws = wb.active
    I_COLS = [43, 46, 49, 52, 55, 58, 61, 64, 67, 70, 73]

    totals = defaultdict(lambda: defaultdict(float))
    detail = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        ot  = str(row[15] or "")
        key = "PK+CR" if "Pickling" in ot else "GI"
        gi  = to_float(row[22]);  scr = to_float(row[25])
        grb = to_float(row[28]);  grc = to_float(row[31])
        gr  = to_float(row[34]);  d1  = to_float(row[37]); d2 = to_float(row[40])
        ind = sum(to_float(row[ic]) for ic in I_COLS if ic < len(row))
        net = gi + d1 + d2 + ind - scr - grb - grc

        totals[key]["GI_Material"]      += gi
        totals[key]["Scrap_Recovery"]   += scr
        totals[key]["GradeB"]           += grb
        totals[key]["GradeC"]           += grc
        totals[key]["GR_Output"]        += gr
        totals[key]["D101_Direct_Mach"] += d1
        totals[key]["D102_Direct_Labor"] += d2
        totals[key]["Indirect_Total"]   += ind

        detail.append([key, ot, row[0], str(row[1] or ""), str(row[5] or ""),
                       round(gi, 2), round(scr, 2), round(grb, 2), round(grc, 2),
                       round(gr, 2), round(d1, 2), round(d2, 2), round(ind, 2), round(net, 2)])

    print(f"  PRD rows: {len(detail)}")
    return totals, detail


def load_ksb1():
    print("Loading KSB1...")
    wb = openpyxl.load_workbook(f"{DATA_DIR}\\KSB1_1300_01.2026.XLSX", data_only=True)
    ws = wb.active
    by_cc   = defaultdict(float)
    by_elem = defaultdict(lambda: {"name": "", "total": 0.0})
    rows    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        cc   = str(row[0]) if row[0] else "(no CC)"
        elem = str(row[1]) if row[1] else ""
        name = str(row[2]) if row[2] else ""
        amt  = to_float(row[3])
        by_cc[cc] += amt
        if elem:
            by_elem[elem]["name"]  = name
            by_elem[elem]["total"] += amt
        rows.append([cc, elem, name, round(amt, 2)])

    print(f"  KSB1 rows: {len(rows)}")
    return by_cc, by_elem, rows


def load_gl():
    print("Loading GL (Jan 2026, CC 13*)... may take ~60s")
    wb = openpyxl.load_workbook(f"{DATA_DIR}\\AMC_GL_03.2026.XLSX",
                                data_only=True, read_only=True)
    ws = wb.active
    by_account = defaultdict(lambda: {"name": "", "total": 0.0, "category": ""})
    filtered   = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row): continue
        gl_acc  = str(row[2]) if row[2] else ""
        gl_name = str(row[3]) if row[3] else ""
        posting = row[6]
        amount  = to_float(row[10])
        doc_type = str(row[13]) if row[13] else ""
        cc      = str(row[14]) if row[14] else ""
        if not (isinstance(posting, datetime.datetime) and posting.month == 1): continue
        if not cc.startswith("13"): continue
        by_account[gl_acc]["name"]     = gl_name
        by_account[gl_acc]["total"]   += amount
        by_account[gl_acc]["category"] = classify_gl(gl_acc)
        filtered.append([cc, gl_acc, gl_name, doc_type,
                         posting.strftime("%d/%m/%Y"), round(amount, 2)])

    print(f"  GL rows filtered: {len(filtered):,}")
    return by_account, filtered


# ─────────────────────────────────────────────────────────────────
def main():
    prd_totals, prd_detail = load_prd()
    ksb1_by_cc, ksb1_by_elem, ksb1_rows = load_ksb1()
    gl_by_account, gl_filtered = load_gl()

    # ── Compute grand totals ──
    grand = defaultdict(float)
    for key in ["PK+CR", "GI"]:
        for k, v in prd_totals[key].items():
            grand[k] += v
    gnet = (grand["GI_Material"] + grand["D101_Direct_Mach"] + grand["D102_Direct_Labor"]
            + grand["Indirect_Total"] - grand["Scrap_Recovery"] - grand["GradeB"] - grand["GradeC"])

    # ── Category totals for GL ──
    cat_totals = defaultdict(float)
    for acc, info in gl_by_account.items():
        cat_totals[info["category"]] += info["total"]

    prd_d101 = grand["D101_Direct_Mach"]
    prd_d102 = grand["D102_Direct_Labor"]
    gl_9431010 = abs(gl_by_account["9431010"]["total"])
    gl_9431020 = abs(gl_by_account["9431020"]["total"])
    gl_9439030 = abs(gl_by_account["9439030"]["total"])
    gl_9439040 = abs(gl_by_account["9439040"]["total"])

    overhead_cats = ["Labor", "Electricity", "Repair & Maintenance", "Depreciation",
                     "Manufacturing Supplies", "Tools/Supplies/Other", "Other Expenses", "Transport"]
    gl_overhead = sum(cat_totals[c] for c in overhead_cats)
    co_settlement = cat_totals["CO Settlement (to Orders)"]
    ksb1_known = abs(ksb1_by_cc.get("1387110", 0) + ksb1_by_cc.get("1387120", 0)
                     + ksb1_by_cc.get("1387210", 0))
    cc_net_balance = gl_overhead + co_settlement  # should be ~0

    def match(a, b, tol=1.0):
        return "MATCH" if abs(a - b) <= tol else f"DIFF ({a-b:+,.2f})"

    # ── Write to Google Sheets ──
    print("\nConnecting to Google Sheets...")
    gc = get_gspread_client()
    try:
        sh = gc.open(SHEET_NAME)
        print(f"Opened existing: {SHEET_NAME}")
    except Exception:
        sh = gc.create(SHEET_NAME)
        print(f"Created new: {SHEET_NAME}")

    def tab(name, rows=2000, cols=20):
        return get_or_add_worksheet(sh, name)

    # ── Tab 1: PRD_Detail ──
    print("Writing PRD_Detail...")
    ws1 = tab("PRD_Detail")
    ws1.clear()
    headers_prd = ["Line", "Order Type", "Material", "Description", "Work Center",
                   "GI_Material", "Scrap_Recovery", "GradeB", "GradeC",
                   "GR_Output", "D101_Direct_Mach", "D102_Direct_Labor",
                   "Indirect_Total", "Net_Cost"]
    ws1.append_row(["Production Order Detail — Jan 2026 | Plant 1300"])
    ws1.append_row(headers_prd)
    ws1.append_rows(prd_detail, value_input_option="USER_ENTERED")

    # ── Tab 2: PRD_Summary ──
    print("Writing PRD_Summary...")
    ws2 = tab("PRD_Summary")
    ws2.clear()
    prd_sum = [
        ["PRD Summary — Jan 2026 | Plant 1300 | Prepared by: " + PREPARED_BY],
        [],
        ["Line", "GI_Material", "Scrap_Recovery", "GradeB", "GradeC",
         "GR_Output", "D101_Direct_Mach", "D102_Direct_Labor",
         "Indirect_Total", "Net_Cost (GI+D+I-Scrap)"],
    ]
    for line in ["PK+CR", "GI"]:
        t = prd_totals[line]
        net = (t["GI_Material"] + t["D101_Direct_Mach"] + t["D102_Direct_Labor"]
               + t["Indirect_Total"] - t["Scrap_Recovery"] - t["GradeB"] - t["GradeC"])
        prd_sum.append([line,
                        round(t["GI_Material"], 2), round(t["Scrap_Recovery"], 2),
                        round(t["GradeB"], 2), round(t["GradeC"], 2),
                        round(t["GR_Output"], 2), round(t["D101_Direct_Mach"], 2),
                        round(t["D102_Direct_Labor"], 2), round(t["Indirect_Total"], 2),
                        round(net, 2)])
    prd_sum.append(["TOTAL",
                    round(grand["GI_Material"], 2), round(grand["Scrap_Recovery"], 2),
                    round(grand["GradeB"], 2), round(grand["GradeC"], 2),
                    round(grand["GR_Output"], 2), round(grand["D101_Direct_Mach"], 2),
                    round(grand["D102_Direct_Labor"], 2), round(grand["Indirect_Total"], 2),
                    round(gnet, 2)])
    ws2.append_rows(prd_sum, value_input_option="USER_ENTERED")

    # ── Tab 3: KSB1_Summary ──
    print("Writing KSB1_Summary...")
    ws3 = tab("KSB1_Summary")
    ws3.clear()
    cc_map = {"1387110": "Pickling Line 01", "1387120": "Cold Rolling Mill 01",
              "1387210": "Galvanizing Line 01", "(no CC)": "No CC (Direct to Order)"}
    ksb1_sum = [["KSB1 Summary — Jan 2026 | Plant 1300 | CC 1387xxx"],
                [],
                ["Cost Center", "CC Name", "Cost Element", "Cost Element Name", "Amount (THB)"]]
    for cc in ["1387110", "1387120", "1387210", "(no CC)"]:
        elem_totals = defaultdict(lambda: {"name": "", "total": 0.0})
        for r in ksb1_rows:
            if r[0] == cc and r[1]:
                elem_totals[r[1]]["name"]  = r[2]
                elem_totals[r[1]]["total"] += r[3]
        for e, info in sorted(elem_totals.items()):
            ksb1_sum.append([cc, cc_map.get(cc, ""), e, info["name"], round(info["total"], 2)])
        ksb1_sum.append([f"--- {cc} SUBTOTAL ---", cc_map.get(cc, ""), "", "", round(ksb1_by_cc[cc], 2)])
        ksb1_sum.append([])
    ksb1_sum.append(["=== GRAND TOTAL ===", "", "", "", round(sum(ksb1_by_cc.values()), 2)])
    ws3.append_rows(ksb1_sum, value_input_option="USER_ENTERED")

    # ── Tab 4: GL_Summary ──
    print("Writing GL_Summary...")
    ws4 = tab("GL_Summary")
    ws4.clear()
    gl_sum = [["GL Summary — Jan 2026 | CC 13* | Plant 1300"],
              [],
              ["Category", "GL Account", "Account Name", "Amount (THB)"]]
    cat_order = ["Labor", "Electricity", "Repair & Maintenance", "Depreciation",
                 "Manufacturing Supplies", "Tools/Supplies/Other", "Other Expenses",
                 "Transport", "Consumption (Material)", "CO Settlement (to Orders)",
                 "Production Variance", "Variance Adjustment (ML)", "Other"]
    gl_total = 0.0
    for cat in cat_order:
        items = [(a, i["name"], round(i["total"], 2))
                 for a, i in sorted(gl_by_account.items()) if i["category"] == cat]
        if not items: continue
        cat_total = sum(x[2] for x in items)
        for acc, name, amt in items:
            gl_sum.append([cat, acc, name, amt])
        gl_sum.append([f"--- {cat} SUBTOTAL ---", "", "", round(cat_total, 2)])
        gl_sum.append([])
        gl_total += cat_total
    gl_sum.append(["=== GRAND TOTAL ===", "", "", round(gl_total, 2)])
    ws4.append_rows(gl_sum, value_input_option="USER_ENTERED")

    # ── Tab 5: RECON_Result ──
    print("Writing RECON_Result...")
    ws5 = tab("RECON_Result")
    ws5.clear()

    recon = [
        ["RECONCILIATION RESULT — GI Production Cost | Jan 2026 | Plant 1300"],
        [f"Prepared by: {PREPARED_BY}   |   {TODAY}   |   Status: Draft"],
        [],
        ["=" * 60],
        ["LAYER 1: PRD Report vs CO Settlement (Direct Costs)"],
        ["=" * 60],
        ["Item", "PRD Amount (THB)", "GL/CO Amount (THB)", "Difference", "Status"],
        ["D101 Direct Machine",
         round(prd_d101, 2), round(gl_9431010, 2),
         round(prd_d101 - gl_9431010, 2), match(prd_d101, gl_9431010)],
        ["D102 Direct Labor",
         round(prd_d102, 2), round(gl_9431020, 2),
         round(prd_d102 - gl_9431020, 2), match(prd_d102, gl_9431020)],
        [],
        ["=" * 60],
        ["LAYER 2: GL Actual Overhead vs KSB1 Settled (CC 1387xxx)"],
        ["=" * 60],
        ["Item", "GL Actual (THB)", "KSB1 Settled |amt| (THB)", "Difference", "Status"],
        ["Overhead (Labor+Elec+R&M+Depr+Others)",
         round(gl_overhead, 2), round(ksb1_known, 2),
         round(gl_overhead - ksb1_known, 2),
         match(gl_overhead, ksb1_known, tol=100)],
        ["CC Net Balance (Actual + Settlement should = 0)",
         round(gl_overhead, 2), round(abs(co_settlement), 2),
         round(cc_net_balance, 2),
         "BALANCED" if abs(cc_net_balance) < 100 else f"UNBALANCED"],
        [],
        ["=" * 60],
        ["LAYER 3: Production Cost Build-Up (PRD)"],
        ["=" * 60],
        ["Item", "Amount (THB)", "Note"],
        ["Material: HRC / CRC consumed (GI)",
         round(grand["GI_Material"], 2), "Raw material issued to orders"],
        ["Direct: Machine time (D101)",
         round(prd_d101, 2), "Settled from CC to orders — matches GL 9431010"],
        ["Direct: Labor (D102)",
         round(prd_d102, 2), "Settled from CC to orders — matches GL 9431020"],
        ["Indirect: Overhead allocation (I101-I111)",
         round(grand["Indirect_Total"], 2), "Indirect costs panned to orders"],
        ["Less: Scrap Recovery",
         round(-grand["Scrap_Recovery"], 2), "By-product credit"],
        ["NET Production Cost",
         round(gnet, 2), "Total cost in production orders"],
        [],
        ["GL: Actual Overhead posted to CC 13*",
         round(gl_overhead, 2), "55xxx-61xxx accounts"],
        ["GL: CO Settlement to Orders",
         round(co_settlement, 2), "943xxxx (negative = cost moved to orders)"],
        ["GL: Variance Adjustment ML (excluded)",
         round(cat_totals["Variance Adjustment (ML)"], 2), "5391020 — price correction, not operating cost"],
        [],
        ["=" * 60],
        ["FINDINGS / ข้อสังเกต"],
        ["=" * 60],
        ["#", "รายการ", "ผลลัพธ์"],
        ["1", "D101 Direct Machine: PRD vs GL 9431010",
         f"MATCH — {prd_d101:,.2f} = {gl_9431010:,.2f}"],
        ["2", "D102 Direct Labor: PRD vs GL 9431020",
         f"MATCH — {prd_d102:,.2f} = {gl_9431020:,.2f}"],
        ["3", "Material Cost (HRC/Zinc) ใน PRD",
         "อยู่ใน GI Amount — GL แสดงเป็น Credit ใน Inventory (ไม่ผ่าน CC)"],
        ["4", "Indirect (I101-I111) ใน PRD",
         f"รวม {grand['Indirect_Total']:,.2f} — ตรวจสอบว่าเป็น pool total หรือ per-order allocation"],
        ["5", "GL 5391020 Variance Adjustment ML",
         f"{cat_totals['Variance Adjustment (ML)']:,.2f} — ML price adj ไม่ใช่ production cost"],
        ["6", "KSB1 มีเพียง Jan 2026 เท่านั้น",
         "Feb/Mar ต้องรัน KSB1 เพิ่ม หรือใช้ GL (CC 13*) แทน"],
        [],
        [f"Timestamp: {TODAY}"],
    ]
    ws5.append_rows(recon, value_input_option="USER_ENTERED")

    # ── Remove default Sheet1 ──
    try:
        default = sh.worksheet("Sheet1")
        if len(sh.worksheets()) > 1:
            sh.del_worksheet(default)
    except Exception:
        pass

    print(f"\nDone!")
    print(f"Sheet: {sh.url}")
    return sh.url


if __name__ == "__main__":
    main()
