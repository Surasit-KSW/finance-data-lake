"""
recon_plant1300_cost.py
=======================
Plant 1300 Production Cost Reconciliation
Reconcile path: MB51 → PRD → KSB1 → GL → TB (P&L)

Tabs produced (10):
  1. Overview      — executive cost summary + match matrix (meeting-ready)
  2. Cost_BuildUp  — layer-by-layer cost: HRC → PK+CR → Zinc → GI (THB + THB/MT)
  3. GL_vs_TB      — GL actuals vs Trial Balance P&L alignment
  4. GL_vs_KSB1    — GL overhead vs KSB1 CC totals + CO settlement ≈ 0
  5. KSB1_vs_PRD   — KSB1 settlement vs PRD D101/D102 (tol +/-1 THB)
  6. MB51_vs_PRD   — MB51 mvt 261 issues vs PRD GI material amount
  7. GL_Detail     — raw GL line items (CC 13*, month-filtered)
  8. KSB1_Detail   — raw KSB1 rows by CC + cost element
  9. PRD_Detail    — raw PRD production orders (PK+CR + GI)
 10. MB51_Detail   — raw MB51 Plant 1300 movements (261 + 531)

Usage:
  python scripts/recon_plant1300_cost.py --month 3          # March 2026
  python scripts/recon_plant1300_cost.py --month 3 --dry-run  # preview only
  python scripts/recon_plant1300_cost.py --month 4          # April (PRD only — KSB1/GL not yet)

Sources (01_Raw/PRD_GI/):
  PRD_1300_MM.2026.XLSX         — Production Order cost per month
  KSB1_1300_MM.2026.XLSX        — CC Actuals CO per month
  AMC_GL_03.2026.XLSX           — GL line items Jan–Mar combined (filter by month)
  MB51_all plant_03.2026.XLSX   — Material movements all plants (filter Plant=1300)
  AMC_TB_MM.2026.XLSX           — Trial Balance per month

Stop conditions:
  D101/D102 PRD vs GL diff > +/-1,000 THB → WARNING + DIFF status
  MB51 vs PRD material diff > +/-50,000 THB → WARNING (investigate separately)
  GL vs KSB1 diff > +/-5% → FLAG for review

Prepared by: Claude Code
"""

import sys
import datetime
import argparse
import os
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from utils.auth import get_gspread_client
from utils.sheets import get_or_add_worksheet

# ─── Constants ────────────────────────────────────────────────────────────────

PREPARED_BY = "Claude Code"
TODAY = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
YEAR = 2026

DATA_DIR = str(Path(__file__).resolve().parent.parent.parent / '01_Bronze_Raw' / 'PRD_GI')

MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

CC_MAP = {
    "1387110": "Pickling Line 01 (PK)",
    "1387120": "Cold Rolling Mill 01 (CR)",
    "1387210": "Galvanizing Line 01 (GI)",
    "(no CC)": "No CC (Direct to Order)",
}

EXCLUDE_GL = {"5391020", "5211010"}   # ML price adj + COGM journal — not production cost

OVERHEAD_CATS = [
    "Labor", "Electricity", "Repair & Maintenance", "Depreciation",
    "Manufacturing Supplies", "Tools/Supplies/Other", "Other Expenses", "Transport",
]

# PRD column indices (0-based, from recon_gi_plant1300.py)
PRD_ORDER_TYPE = 15
PRD_GI_AMT   = 22   # Raw material consumed (THB)
PRD_SCRAP    = 25   # Scrap / by-product recovery
PRD_GRADEB   = 28   # Grade B recovery
PRD_GRADEC   = 31   # Grade C recovery
PRD_GR_QTY   = 32   # GR output qty (kg)
PRD_GR_AMT   = 34   # GR output value (THB)
PRD_D101     = 37   # Direct machine cost
PRD_D102     = 40   # Direct labor cost
PRD_I_COLS   = [43, 46, 49, 52, 55, 58, 61, 64, 67, 70, 73]  # I101–I111


# ─── Helpers ──────────────────────────────────────────────────────────────────

def to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").strip())
    except: return 0.0


def thb_mt(thb, qty_kg):
    """Return THB per MT; 0.0 if qty is zero."""
    return round(thb / (qty_kg / 1000), 2) if qty_kg > 0 else 0.0


def classify_gl(acc):
    a = str(acc)
    if a.startswith(("551", "552", "553")): return "Labor"
    if a.startswith("561"):  return "Electricity"
    if a.startswith("571"):  return "Repair & Maintenance"
    if a.startswith("581"):  return "Depreciation"
    if a.startswith("591"):  return "Tools/Supplies/Other"
    if a.startswith("592"):  return "Manufacturing Supplies"
    if a.startswith("599"):  return "Other Expenses"
    if a.startswith("611"):  return "Transport"
    if a.startswith("541"):  return "Consumption (Material)"
    if a.startswith("531"):  return "Production Variance"
    if a.startswith("539"):  return "Variance Adjustment (ML)"
    if a.startswith("943"):  return "CO Settlement (to Orders)"
    return "Other"


def match_status(a, b, tol=1.0):
    diff = round(a - b, 2)
    if abs(diff) <= tol:
        return f"MATCH (diff={diff:+,.2f})"
    return f"DIFF ({diff:+,.2f})"


def sep(label=""):
    return [f"{'─'*10} {label} {'─'*(max(0,50-len(label)))}"] if label else ["─" * 62]


# ─── Data Loaders ─────────────────────────────────────────────────────────────

def load_prd(month: int):
    """Load PRD_1300_MM.2026.XLSX — return (totals_by_line, gr_qty, detail_rows)."""
    mm   = f"{month:02d}"
    path = os.path.join(DATA_DIR, f"PRD_1300_{mm}.2026.XLSX")
    print(f"Loading PRD: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    totals = {
        "PK+CR": defaultdict(float),
        "GI":    defaultdict(float),
    }
    detail = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row[:10]):
            continue
        ot  = str(row[PRD_ORDER_TYPE] or "")
        key = "PK+CR" if "Pickling" in ot else "GI"

        gi   = to_float(row[PRD_GI_AMT])
        scr  = to_float(row[PRD_SCRAP])
        grb  = to_float(row[PRD_GRADEB])
        grc  = to_float(row[PRD_GRADEC])
        grq  = to_float(row[PRD_GR_QTY])   # kg
        gra  = to_float(row[PRD_GR_AMT])
        d1   = to_float(row[PRD_D101])
        d2   = to_float(row[PRD_D102])
        ind  = sum(to_float(row[ic]) for ic in PRD_I_COLS if ic < len(row))

        totals[key]["GI_Material"]       += gi
        totals[key]["Scrap_Recovery"]    += scr
        totals[key]["GradeB"]            += grb
        totals[key]["GradeC"]            += grc
        totals[key]["GR_QTY_kg"]         += grq
        totals[key]["GR_Output_THB"]     += gra
        totals[key]["D101_Direct_Mach"]  += d1
        totals[key]["D102_Direct_Labor"] += d2
        totals[key]["Indirect_Total"]    += ind

        detail.append([
            key, ot,
            str(row[0] or ""), str(row[1] or ""), str(row[5] or ""),
            round(gi, 2), round(scr, 2), round(grb, 2), round(grc, 2),
            round(grq, 2), round(gra, 2), round(d1, 2), round(d2, 2), round(ind, 2),
        ])

    print(f"  PRD rows: {len(detail)}")
    return totals, detail


def load_ksb1(month: int):
    """Load KSB1_1300_MM.2026.XLSX — return (by_cc, by_elem, rows)."""
    mm   = f"{month:02d}"
    path = os.path.join(DATA_DIR, f"KSB1_1300_{mm}.2026.XLSX")
    print(f"Loading KSB1: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    by_cc   = defaultdict(float)
    by_elem = defaultdict(lambda: {"name": "", "total": 0.0})
    rows    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row[:4]):
            continue
        cc   = str(row[0]).strip() if row[0] else ""
        elem = str(row[1]).strip() if row[1] else ""
        # Skip SAP-generated subtotal rows (empty CC AND empty cost element)
        if not cc and not elem:
            continue
        cc   = cc if cc else "(no CC)"
        name = str(row[2]).strip() if row[2] else ""
        amt  = to_float(row[3])

        by_cc[cc] += amt
        if elem:
            by_elem[elem]["name"]   = name
            by_elem[elem]["total"] += amt
        rows.append([cc, elem, name, round(amt, 2)])

    print(f"  KSB1 rows: {len(rows)}")
    return by_cc, by_elem, rows


def load_gl(month: int):
    """Load AMC_GL_03.2026.XLSX — filter CC 13*, posting month, exclude EXCLUDE_GL."""
    path = os.path.join(DATA_DIR, "AMC_GL_03.2026.XLSX")
    print(f"Loading GL (month={month}, CC 13*)... may take ~60s")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    by_account = defaultdict(lambda: {"name": "", "total": 0.0, "category": "",
                                      "excluded": False})
    cat_totals = defaultdict(float)
    filtered   = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue
        gl_acc   = str(row[2]).strip() if row[2] else ""
        gl_name  = str(row[3]).strip() if row[3] else ""
        posting  = row[6]
        amount   = to_float(row[10])
        doc_type = str(row[13]).strip() if row[13] else ""
        cc       = str(row[14]).strip() if row[14] else ""

        if not (isinstance(posting, datetime.datetime) and posting.month == month):
            continue
        if not cc.startswith("13"):
            continue

        category = classify_gl(gl_acc)
        excluded = gl_acc in EXCLUDE_GL

        by_account[gl_acc]["name"]     = gl_name
        by_account[gl_acc]["total"]   += amount
        by_account[gl_acc]["category"] = category
        by_account[gl_acc]["excluded"] = excluded

        if not excluded:
            cat_totals[category] += amount

        filtered.append([
            cc, gl_acc, gl_name, doc_type,
            posting.strftime("%d/%m/%Y"), round(amount, 2),
            "EXCLUDED" if excluded else "",
        ])

    print(f"  GL rows filtered: {len(filtered):,}")
    wb.close()
    return by_account, cat_totals, filtered


def load_mb51(month: int):
    """Load MB51_all plant_03.2026.XLSX — filter Plant=1300, mvt 261/531, month."""
    path = os.path.join(DATA_DIR, "MB51_all plant_03.2026.XLSX")
    print(f"Loading MB51: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    # Detect column indices from header row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip().lower() if c else "" for c in row]
        break

    def find_col(candidates):
        for c in candidates:
            for i, h in enumerate(headers):
                if c in h:
                    return i
        return None

    idx_plant   = find_col(["plant"])
    idx_mvt     = find_col(["movement type", "mvt type", "move type", "mvt", "movement"])
    idx_date    = find_col(["posting date", "pstng date", "posting", "doc date"])
    idx_amount  = find_col(["amt.in loc", "amount in lc", "amount in local", "loc.cur", "val in lc", "amount"])
    idx_mat     = find_col(["material"])
    idx_matdesc = find_col(["material description", "material desc", "description"])

    print(f"  MB51 columns — plant:{idx_plant} mvt:{idx_mvt} date:{idx_date} "
          f"amount:{idx_amount} mat:{idx_mat}")

    issues   = 0.0   # mvt 261: material issued to production
    recovery = 0.0   # mvt 531: by-product received back
    rows     = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue

        plant = str(row[idx_plant]).strip() if idx_plant is not None and row[idx_plant] else ""
        mvt   = str(row[idx_mvt]).strip() if idx_mvt is not None and row[idx_mvt] else ""
        mvt   = mvt.lstrip("0")  # "0261" → "261"
        date  = row[idx_date] if idx_date is not None else None
        amt   = to_float(row[idx_amount]) if idx_amount is not None else 0.0
        mat   = str(row[idx_mat] or "") if idx_mat is not None else ""
        desc  = str(row[idx_matdesc] or "") if idx_matdesc is not None else ""

        if plant != "1300":
            continue
        if not (isinstance(date, datetime.datetime) and date.month == month):
            continue
        if mvt not in {"261", "531"}:
            continue

        if mvt == "261":
            issues += amt
        else:
            recovery += amt

        rows.append([
            plant, mvt,
            date.strftime("%d/%m/%Y") if isinstance(date, datetime.datetime) else str(date),
            mat, desc, round(amt, 2),
        ])

    wb.close()
    print(f"  MB51 Plant 1300 rows: {len(rows)} (mvt 261+531)")
    return {"issues": round(issues, 2), "recovery": round(recovery, 2), "rows": rows}


def load_tb(month: int):
    """Load AMC_TB_MM.2026.XLSX — extract production cost accounts (55x–61x, 59x)."""
    mm   = f"{month:02d}"
    # Try individual month file first, fallback to combined
    candidates = [
        os.path.join(DATA_DIR, f"AMC_TB_{mm}.2026.XLSX"),
        os.path.join(DATA_DIR, "AMC_TB_01-03.2026.XLSX"),
    ]
    path = None
    for c in candidates:
        if os.path.exists(c):
            path = c
            break

    if path is None:
        print(f"  TB file not found (tried {candidates}) — skipping GL_vs_TB tab")
        return None

    print(f"Loading TB: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Detect columns
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip().lower() if c else "" for c in row]
        break

    def find_col(candidates_list):
        for c in candidates_list:
            for i, h in enumerate(headers):
                if c in h:
                    return i
        return None

    idx_acc  = find_col(["gl account", "account", "g/l account", "gl acct"])
    idx_name = find_col(["account name", "account desc", "text for b/s", "description", "name", "text"])
    idx_amt  = find_col(["total of reporting", "balance", "amount", "total", "net"])
    idx_mon  = find_col(["posting period", "posting per"])

    print(f"  TB columns — acc:{idx_acc} name:{idx_name} amt:{idx_amt} period:{idx_mon}")

    accounts = {}
    rows     = []

    TB_PREFIXES = ("55", "56", "57", "58", "59", "61")  # production cost P&L accounts

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue

        acc  = str(row[idx_acc]).strip() if idx_acc is not None and row[idx_acc] else ""
        name = str(row[idx_name]).strip() if idx_name is not None and row[idx_name] else ""
        amt  = to_float(row[idx_amt]) if idx_amt is not None else 0.0
        per  = row[idx_mon] if idx_mon is not None else None

        # Filter by period if column exists
        if per is not None:
            if isinstance(per, (int, float)) and int(per) != month:
                continue
            if isinstance(per, datetime.datetime) and per.month != month:
                continue

        if not acc.startswith(TB_PREFIXES):
            continue

        if acc not in accounts:
            accounts[acc] = {"name": name, "amount": 0.0}
        accounts[acc]["amount"] += amt
        rows.append([acc, name, round(amt, 2)])

    print(f"  TB production accounts: {len(accounts)}")
    return {"accounts": accounts, "rows": rows}


# ─── Tab Builders ─────────────────────────────────────────────────────────────

def build_overview(month, mn, prd_totals, gl_cat, ksb1_by_cc, mb51, tb_data, warnings):
    """Tab 1: Overview — executive summary for meeting."""
    pk   = prd_totals["PK+CR"]
    gi   = prd_totals["GI"]
    crc_qty = pk["GR_QTY_kg"]
    gi_qty  = gi["GR_QTY_kg"]

    # Cost components (PK+CR)
    pk_hrc   = pk["GI_Material"]
    pk_scrap = pk["Scrap_Recovery"] + pk["GradeB"] + pk["GradeC"]
    pk_d101  = pk["D101_Direct_Mach"]
    pk_d102  = pk["D102_Direct_Labor"]
    pk_ind   = pk["Indirect_Total"]
    pk_total = pk_hrc + pk_d101 + pk_d102 + pk_ind - pk_scrap

    # Cost components (GI)
    gi_zinc  = gi["GI_Material"]
    gi_scrap = gi["Scrap_Recovery"] + gi["GradeB"] + gi["GradeC"]
    gi_d101  = gi["D101_Direct_Mach"]
    gi_d102  = gi["D102_Direct_Labor"]
    gi_ind   = gi["Indirect_Total"]
    gi_total = gi_zinc + gi_d101 + gi_d102 + gi_ind - gi_scrap

    # Reconciliation checks
    prd_d101_all = pk["D101_Direct_Mach"] + gi["D101_Direct_Mach"]
    prd_d102_all = pk["D102_Direct_Labor"] + gi["D102_Direct_Labor"]
    gl_overhead   = sum(gl_cat.get(c, 0) for c in OVERHEAD_CATS)
    co_settlement = gl_cat.get("CO Settlement (to Orders)", 0.0)
    cc_net = gl_overhead + co_settlement

    # KSB1 = settlement entries only; 9439030 = CRC→GI intra-CC material transfer (not pure overhead)
    ksb1_no_cc  = ksb1_by_cc.get("(no CC)", 0.0)
    ksb1_d101   = abs(ksb1_by_cc.get("1387110", 0) + ksb1_by_cc.get("1387120", 0))
    # Use total CC settlement as reference (for overview only)
    ksb1_total  = sum(v for k, v in ksb1_by_cc.items() if k != "(no CC)")

    # MB51 mvt 261 is negative in SAP (goods issue = credit to inventory) — use abs
    mb51_issues_abs = abs(mb51["issues"]) if mb51 else 0.0
    prd_material  = pk_hrc + gi_zinc
    mb51_diff     = mb51_issues_abs - prd_material if mb51 else None

    rows = [
        [f"PLANT 1300 — Cost Structure & Reconciliation | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}   |   {TODAY}   |   Status: Draft"],
        [""],
    ]

    if warnings:
        rows.append(["[!]  WARNINGS"])
        for w in warnings:
            rows.append(["", w])
        rows.append([""])

    rows += [
        sep("PRODUCTION OUTPUT"),
        ["", "CRC Output (GR)", f"{crc_qty:,.0f} kg", f"{crc_qty/1000:,.1f} MT"],
        ["", "GI Output  (GR)", f"{gi_qty:,.0f} kg",  f"{gi_qty/1000:,.1f} MT"],
        [""],
        sep("COST SUMMARY — PK+CR (THB)"),
        ["", "Items", "THB", "THB/MT"],
        ["", "Raw Material (HRC)", f"{pk_hrc:,.2f}",
         f"{thb_mt(pk_hrc, crc_qty):,.2f}" if crc_qty else "n/a"],
        ["", "Direct Machine (D101)", f"{pk_d101:,.2f}",
         f"{thb_mt(pk_d101, crc_qty):,.2f}" if crc_qty else "n/a"],
        ["", "Direct Labor  (D102)", f"{pk_d102:,.2f}",
         f"{thb_mt(pk_d102, crc_qty):,.2f}" if crc_qty else "n/a"],
        ["", "Indirect Pool (I101–I111)", f"{pk_ind:,.2f}",
         f"{thb_mt(pk_ind, crc_qty):,.2f}" if crc_qty else "n/a"],
        ["", "Less: Scrap Recovery", f"{-pk_scrap:,.2f}", ""],
        ["", "═" * 30, "═" * 14, "═" * 12],
        ["", "CRC TOTAL", f"{pk_total:,.2f}",
         f"{thb_mt(pk_total, crc_qty):,.2f}" if crc_qty else "n/a"],
        [""],
        sep("COST SUMMARY — GI LINE (THB)"),
        ["", "Items", "THB", "THB/MT"],
        ["", "Zinc / ZAM Material", f"{gi_zinc:,.2f}",
         f"{thb_mt(gi_zinc, gi_qty):,.2f}" if gi_qty else "n/a"],
        ["", "Direct Machine (D101)", f"{gi_d101:,.2f}",
         f"{thb_mt(gi_d101, gi_qty):,.2f}" if gi_qty else "n/a"],
        ["", "Direct Labor  (D102)", f"{gi_d102:,.2f}",
         f"{thb_mt(gi_d102, gi_qty):,.2f}" if gi_qty else "n/a"],
        ["", "Indirect Pool (I101–I111)", f"{gi_ind:,.2f}",
         f"{thb_mt(gi_ind, gi_qty):,.2f}" if gi_qty else "n/a"],
        ["", "Less: Scrap/Dross Recovery", f"{-gi_scrap:,.2f}", ""],
        ["", "═" * 30, "═" * 14, "═" * 12],
        ["", "GI TOTAL", f"{gi_total:,.2f}",
         f"{thb_mt(gi_total, gi_qty):,.2f}" if gi_qty else "n/a"],
        [""],
        sep("RECONCILIATION STATUS"),
        ["", "Layer", "Check", "Value", "Status"],
    ]

    # Layer checks
    def chk(label, val_a, val_b, tol=1.0, note=""):
        diff = round(val_a - val_b, 2)
        st   = "MATCH" if abs(diff) <= tol else f"DIFF  {diff:+,.2f}"
        return ["", label, f"{val_a:,.2f}  vs  {val_b:,.2f}", "", st, note]

    rows.append(["", "GL CC net balance", "GL overhead (CC13*) + CO settlement ≈ 0",
                 f"net={cc_net:,.2f}",
                 "BALANCED" if abs(cc_net) < 1000 else f"UNBALANCED {cc_net:+,.2f}",
                 "see GL_vs_KSB1"])
    rows.append(["", "KSB1 settlement total", "Total settled from CC 1387xxx to orders",
                 f"{abs(ksb1_total):,.2f} THB",
                 "INFO (incl. CRC->GI material transfer)",
                 "see GL_vs_KSB1"])

    if mb51:
        rows.append(["", "MB51 vs PRD", "MB51 mvt261 vs PRD material",
                     f"|MB51|={mb51_issues_abs:,.2f}  PRD={prd_material:,.2f}",
                     "MATCH" if abs(mb51_diff) < 50000 else f"DIFF  {mb51_diff:+,.2f}",
                     "see MB51_vs_PRD"])
    else:
        rows.append(["", "MB51 vs PRD", "MB51 data not available", "", "SKIP", ""])

    if tb_data:
        tb_total = sum(v["amount"] for v in tb_data["accounts"].values())
        rows.append(["", "GL vs TB", "GL overhead vs TB P&L accounts",
                     f"GL={gl_overhead:,.2f}  TB={tb_total:,.2f}",
                     "MATCH" if abs(gl_overhead - tb_total) < 5000 else f"DIFF  {gl_overhead - tb_total:+,.2f}",
                     "see GL_vs_TB"])
    else:
        rows.append(["", "GL vs TB", "TB data not available for this month", "", "SKIP", ""])

    rows += [
        [""],
        sep("CKMLCP / MATERIAL LEDGER — CONTEXT (Jan-Mar 2026)"),
        ["", "[!] Jan 2026: CKMLCP closed late — Jan amounts rolled into Feb closing"],
        ["", "[!] Jan-Mar 2026: ML closed WITHOUT completing 100% of pending orders"],
        ["", "[!] From Apr 2026: pending orders will be completed + ML closed 100%"],
        ["", "=> Diffs shown (GL<>TB, MB51<>PRD, variance) are EXPECTED for Jan-Mar"],
        ["", "=> These numbers are SNAPSHOTS / directional — not final until Apr closing"],
        ["", "=> Some variance sitting in CCs due to un-allocated amounts from incomplete ML"],
        ["", "=> GL 5391020 (ML price adj) excluded from all recon — not production cost"],
        [""],
        sep("NOTES"),
        ["", "• GL (CC 13*) is partial — some overhead posts without CC in FI doc; use TB for total"],
        ["", "• GL excludes: 5391020 (ML price adj) + 5211010 (COGM journal)"],
        ["", "• KSB1 = settlement entries only (943x/9439x credit from CC to orders)"],
        ["", "• KSB1 9439030 Indirect Labour includes CRC->GI intra-CC material transfer"],
        ["", f"• CRC output qty ({crc_qty/1000:,.1f} MT) used as basis for PK+CR THB/MT"],
        ["", f"• GI output qty ({gi_qty/1000:,.1f} MT) used as basis for GI THB/MT"],
        ["", f"Timestamp: {TODAY}"],
    ]

    return rows


def build_cost_buildup(month, mn, prd_totals):
    """Tab 2: Layer-by-layer cost build-up with THB/MT."""
    pk  = prd_totals["PK+CR"]
    gi  = prd_totals["GI"]
    crc = pk["GR_QTY_kg"]
    giq = gi["GR_QTY_kg"]

    rows = [
        [f"COST BUILD-UP — Plant 1300 | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}  |  {TODAY}"],
        [""],
        ["Layer", "Items", "Source", "THB", "THB/MT (basis)", "Note"],
        [""],
        sep("LAYER 0 — RAW MATERIAL (HRC into PK+CR)"),
        ["L0", "HRC Material", "PRD col 22 (PK+CR orders)",
         round(pk["GI_Material"], 2), thb_mt(pk["GI_Material"], crc), f"CRC output {crc/1000:,.1f} MT"],
        [""],
        sep("LAYER 1 — PK+CR CONVERSION"),
        ["L1", "Direct Machine (D101)", "PRD col 37",
         round(pk["D101_Direct_Mach"], 2), thb_mt(pk["D101_Direct_Mach"], crc), ""],
        ["L1", "Direct Labor  (D102)", "PRD col 40",
         round(pk["D102_Direct_Labor"], 2), thb_mt(pk["D102_Direct_Labor"], crc), ""],
        ["L1", "Indirect Pool (I101–I111)", "PRD cols 43–73",
         round(pk["Indirect_Total"], 2), thb_mt(pk["Indirect_Total"], crc), ""],
        ["L1", "Less: Scrap Recovery (PK+CR)", "PRD col 25+28+31",
         round(-pk["Scrap_Recovery"] - pk["GradeB"] - pk["GradeC"], 2),
         thb_mt(-(pk["Scrap_Recovery"] + pk["GradeB"] + pk["GradeC"]), crc), "credit"],
        [""],
        ["L1", "PK+CR Conversion Subtotal", "(D101+D102+Ind–Scrap)",
         round(pk["D101_Direct_Mach"] + pk["D102_Direct_Labor"] + pk["Indirect_Total"]
               - pk["Scrap_Recovery"] - pk["GradeB"] - pk["GradeC"], 2),
         thb_mt(pk["D101_Direct_Mach"] + pk["D102_Direct_Labor"] + pk["Indirect_Total"]
                - pk["Scrap_Recovery"] - pk["GradeB"] - pk["GradeC"], crc), ""],
        [""],
        ["L0+L1", "══ CRC Cost Total ══", "HRC + PK+CR conversion",
         round(pk["GI_Material"] + pk["D101_Direct_Mach"] + pk["D102_Direct_Labor"]
               + pk["Indirect_Total"] - pk["Scrap_Recovery"] - pk["GradeB"] - pk["GradeC"], 2),
         thb_mt(pk["GI_Material"] + pk["D101_Direct_Mach"] + pk["D102_Direct_Labor"]
                + pk["Indirect_Total"] - pk["Scrap_Recovery"] - pk["GradeB"] - pk["GradeC"], crc),
         f"CRC {crc/1000:,.1f} MT"],
        [""],
        sep("LAYER 2 — ZINC / ZAM MATERIAL (into GI Line)"),
        ["L2", "Zinc + ZAM Material", "PRD col 22 (GI orders)",
         round(gi["GI_Material"], 2), thb_mt(gi["GI_Material"], giq), f"GI output {giq/1000:,.1f} MT"],
        [""],
        sep("LAYER 3 — GI LINE CONVERSION"),
        ["L3", "Direct Machine (D101)", "PRD col 37",
         round(gi["D101_Direct_Mach"], 2), thb_mt(gi["D101_Direct_Mach"], giq), ""],
        ["L3", "Direct Labor  (D102)", "PRD col 40",
         round(gi["D102_Direct_Labor"], 2), thb_mt(gi["D102_Direct_Labor"], giq), ""],
        ["L3", "Indirect Pool (I101–I111)", "PRD cols 43–73",
         round(gi["Indirect_Total"], 2), thb_mt(gi["Indirect_Total"], giq), ""],
        ["L3", "Less: Scrap/Dross Recovery (GI)", "PRD col 25+28+31",
         round(-gi["Scrap_Recovery"] - gi["GradeB"] - gi["GradeC"], 2),
         thb_mt(-(gi["Scrap_Recovery"] + gi["GradeB"] + gi["GradeC"]), giq), "credit"],
        [""],
        ["L3", "GI Conversion Subtotal", "(D101+D102+Ind–Scrap)",
         round(gi["D101_Direct_Mach"] + gi["D102_Direct_Labor"] + gi["Indirect_Total"]
               - gi["Scrap_Recovery"] - gi["GradeB"] - gi["GradeC"], 2),
         thb_mt(gi["D101_Direct_Mach"] + gi["D102_Direct_Labor"] + gi["Indirect_Total"]
                - gi["Scrap_Recovery"] - gi["GradeB"] - gi["GradeC"], giq), ""],
        [""],
    ]

    pk_total = (pk["GI_Material"] + pk["D101_Direct_Mach"] + pk["D102_Direct_Labor"]
                + pk["Indirect_Total"] - pk["Scrap_Recovery"] - pk["GradeB"] - pk["GradeC"])
    gi_total = (gi["GI_Material"] + gi["D101_Direct_Mach"] + gi["D102_Direct_Labor"]
                + gi["Indirect_Total"] - gi["Scrap_Recovery"] - gi["GradeB"] - gi["GradeC"])
    grand_total = pk_total + gi_total

    rows += [
        sep("GRAND TOTAL — GI COIL OUTPUT"),
        ["TOTAL", "══ CRC + Zinc/ZAM + GI Conversion ══", "",
         round(grand_total, 2), thb_mt(grand_total, giq),
         f"GI output {giq/1000:,.1f} MT"],
        ["", "(CRC Cost portion)", "",
         round(pk_total, 2), thb_mt(pk_total, giq), ""],
        ["", "(GI Cost portion)",  "",
         round(gi_total, 2), thb_mt(gi_total, giq), ""],
        [""],
        sep("GR OUTPUT VALUE (Standard / Actual)"),
        ["", "PK+CR GR Output Value", "PRD col 34",
         round(pk["GR_Output_THB"], 2), thb_mt(pk["GR_Output_THB"], crc), ""],
        ["", "GI GR Output Value", "PRD col 34",
         round(gi["GR_Output_THB"], 2), thb_mt(gi["GR_Output_THB"], giq), ""],
        [""],
        [f"Source: PRD_1300_{month:02d}.2026.XLSX  |  Timestamp: {TODAY}"],
    ]
    return rows


def build_gl_vs_tb(month, mn, gl_cat, gl_by_acc, tb_data):
    """Tab 3: GL actuals vs Trial Balance — P&L alignment."""
    if tb_data is None:
        return [
            [f"GL vs TB — Plant 1300 | {mn} {YEAR}"],
            [""],
            ["TB file not found — tab skipped"],
            [f"  Expected: AMC_TB_{month:02d}.2026.XLSX or AMC_TB_01-03.2026.XLSX"],
        ]

    rows = [
        [f"GL vs TB — P&L Alignment | Plant 1300 | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}  |  {TODAY}"],
        [""],
        ["Note: Compare GL posted amounts (CC 13*) vs Trial Balance production cost accounts."],
        ["       Difference = timing, accruals, or accounts not in TB extract."],
        [""],
        sep("GL SUMMARY (by category, CC 13*, excl. EXCLUDE_GL)"),
        ["Category", "GL Amount (THB)"],
    ]
    gl_overhead = 0.0
    for cat in OVERHEAD_CATS:
        amt = gl_cat.get(cat, 0.0)
        rows.append([cat, round(amt, 2)])
        gl_overhead += amt
    rows += [
        ["── GL Overhead Total ──", round(gl_overhead, 2)],
        ["CO Settlement (to Orders)", round(gl_cat.get("CO Settlement (to Orders)", 0.0), 2)],
        [""],
        sep("TB PRODUCTION ACCOUNTS (55x–61x)"),
        ["GL Account", "Account Name", "TB Amount (THB)"],
    ]
    tb_total = 0.0
    for acc in sorted(tb_data["accounts"]):
        info = tb_data["accounts"][acc]
        rows.append([acc, info["name"], round(info["amount"], 2)])
        tb_total += info["amount"]
    rows += [
        ["── TB Total ──", "", round(tb_total, 2)],
        [""],
        sep("COMPARISON"),
        ["", "GL Overhead Total", round(gl_overhead, 2)],
        ["", "TB Total", round(tb_total, 2)],
        ["", "Difference (GL – TB)", round(gl_overhead - tb_total, 2)],
        ["", "Status",
         "MATCH" if abs(gl_overhead - tb_total) < 5000 else f"DIFF  {gl_overhead - tb_total:+,.2f}"],
        [""],
        [f"Source: AMC_GL_03.2026.XLSX + TB file  |  Timestamp: {TODAY}"],
    ]
    return rows


def build_gl_vs_ksb1(month, mn, gl_cat, gl_by_acc, ksb1_by_cc, ksb1_by_elem):
    """Tab 4: GL overhead (CC 13*) + KSB1 settlement breakdown.

    Note: KSB1_1300_MM.XLSX contains SETTLEMENT entries only (943x/9439xxx).
    These are credits FROM the CC TO production orders.
    Overhead debits (55x–61x) are NOT in this KSB1 extract — see TB for full overhead picture.
    The primary check here is: KSB1 9439030 Indirect Labour includes CRC→GI material
    transfer, which explains why KSB1 total >> GL overhead.
    """
    gl_overhead   = sum(gl_cat.get(c, 0) for c in OVERHEAD_CATS)
    co_settlement = gl_cat.get("CO Settlement (to Orders)", 0.0)
    cc_net        = gl_overhead + co_settlement

    ksb1_all = sum(v for v in ksb1_by_cc.values())
    ksb1_no_cc = ksb1_by_cc.get("(no CC)", 0.0)
    ksb1_known = ksb1_all - ksb1_no_cc

    # KSB1 pure overhead settlement (9431+9439 excl. 9439030 material transfer)
    ksb1_d101    = ksb1_by_elem.get("9431010", {}).get("total", 0.0)
    ksb1_d102    = ksb1_by_elem.get("9431020", {}).get("total", 0.0)
    ksb1_indirect = sum(
        v["total"] for k, v in ksb1_by_elem.items()
        if k.startswith("9439") and k != "9439030"   # 9439030 = CRC→GI material transfer
    )
    ksb1_crc_transfer = ksb1_by_elem.get("9439030", {}).get("total", 0.0)
    ksb1_overhead_only = ksb1_d101 + ksb1_d102 + ksb1_indirect

    rows = [
        [f"GL vs KSB1 — Settlement Analysis | Plant 1300 | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}  |  {TODAY}"],
        [""],
        ["NOTE: KSB1_1300_MM.XLSX = SETTLEMENT entries only (credit side: 943xxx/9439xxx)."],
        ["      Overhead debits (payroll, elec, R&M) are NOT in this KSB1 extract."],
        ["      Use GL_vs_TB tab for overhead total comparison."],
        ["      9439030 Indirect Labour includes CRC cost transferred to GI line (intra-CC)."],
        [""],
        sep("GL OVERHEAD posted to CC 13* (FI view, subset of TB)"),
        ["Category", "GL Amount (THB)", "Note"],
    ]
    for cat in OVERHEAD_CATS:
        amt = gl_cat.get(cat, 0.0)
        rows.append([cat, round(amt, 2), "CC 13* rows only"])
    rows += [
        ["── GL Overhead Total (CC 13*) ──", round(gl_overhead, 2),
         "Partial — some overhead posts without CC in GL line"],
        ["CO Settlement (943xxx in GL)", round(co_settlement, 2),
         "credit — offsets overhead"],
        ["── GL CC Net (overhead + settlement) ──", round(cc_net, 2),
         "BALANCED" if abs(cc_net) < 1000 else f"DIFF {cc_net:+,.2f}"],
        [""],
        sep("KSB1 SETTLEMENT BREAKDOWN (credit from CC to orders)"),
        ["Cost Element", "Name", "Amount (THB)", "Note"],
    ]
    for elem in sorted(ksb1_by_elem.keys()):
        info = ksb1_by_elem[elem]
        note = ("CRC→GI material transfer (intra-CC, NOT pure overhead)"
                if elem == "9439030" else "")
        rows.append([elem, info["name"], round(info["total"], 2), note])
    rows += [
        ["── D101+D102 settlement ──", "",
         round(ksb1_d101 + ksb1_d102, 2), "verify vs PRD: see KSB1_vs_PRD tab"],
        ["── Indirect overhead settlement ──", "",
         round(ksb1_indirect, 2), "excl. 9439030"],
        ["── 9439030 CRC→GI transfer ──", "",
         round(ksb1_crc_transfer, 2), "intra-CC material cost"],
        ["── KSB1 Grand Total ──", "", round(ksb1_all, 2), ""],
        [""],
        sep("KSB1 BY COST CENTER"),
        ["CC Code", "CC Name", "Settlement Total (THB)"],
    ]
    for cc in sorted(ksb1_by_cc.keys()):
        rows.append([cc, CC_MAP.get(cc, ""), round(ksb1_by_cc[cc], 2)])
    rows += [
        [""],
        sep("GL EXCLUDED ACCOUNTS"),
        ["GL Account", "Name", "Amount (THB)", "Reason"],
    ]
    for acc, info in sorted(gl_by_acc.items()):
        if info.get("excluded"):
            rows.append([acc, info["name"], round(info["total"], 2),
                         "ML Price Adj" if acc == "5391020" else "COGM Journal"])
    rows += [
        [""],
        [f"Source: AMC_GL_03.2026.XLSX + KSB1_1300_{month:02d}.2026.XLSX  |  Timestamp: {TODAY}"],
    ]
    return rows


def build_ksb1_vs_prd(month, mn, prd_totals, ksb1_by_elem, warnings):
    """Tab 5: KSB1 settlement elements vs PRD D101/D102."""
    pk = prd_totals["PK+CR"]
    gi = prd_totals["GI"]

    prd_d101 = pk["D101_Direct_Mach"]  + gi["D101_Direct_Mach"]
    prd_d102 = pk["D102_Direct_Labor"] + gi["D102_Direct_Labor"]

    # KSB1 cost elements for D101/D102
    ksb1_d101 = abs(ksb1_by_elem.get("9431010", {}).get("total", 0.0))
    ksb1_d102 = abs(ksb1_by_elem.get("9431020", {}).get("total", 0.0))

    d101_diff = round(prd_d101 - ksb1_d101, 2)
    d102_diff = round(prd_d102 - ksb1_d102, 2)

    if abs(d101_diff) > 1000:
        warnings.append(f"D101 PRD vs KSB1: diff={d101_diff:+,.2f} THB exceeds +/-1,000 tolerance")
    if abs(d102_diff) > 1000:
        warnings.append(f"D102 PRD vs KSB1: diff={d102_diff:+,.2f} THB exceeds +/-1,000 tolerance")

    rows = [
        [f"KSB1 vs PRD — Settlement Reconciliation | Plant 1300 | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}  |  {TODAY}"],
        [""],
        ["Logic: CO settlement posts to Orders as D101 (machine) + D102 (labor)."],
        ["       KSB1 CE 9431010 = D101 credit; KSB1 CE 9431020 = D102 credit."],
        ["       Tolerance: +/-1 THB."],
        [""],
        sep("D101 — DIRECT MACHINE COST"),
        ["Source", "Amount (THB)", ""],
        ["PRD D101 (PK+CR)", round(pk["D101_Direct_Mach"], 2), ""],
        ["PRD D101 (GI)", round(gi["D101_Direct_Mach"], 2), ""],
        ["PRD D101 TOTAL", round(prd_d101, 2), "sum of all orders"],
        ["KSB1 CE 9431010 (credit, abs)", round(ksb1_d101, 2),
         "CO settlement credit in CC"],
        ["Difference (PRD – KSB1)", d101_diff,
         match_status(prd_d101, ksb1_d101)],
        [""],
        sep("D102 — DIRECT LABOR COST"),
        ["Source", "Amount (THB)", ""],
        ["PRD D102 (PK+CR)", round(pk["D102_Direct_Labor"], 2), ""],
        ["PRD D102 (GI)", round(gi["D102_Direct_Labor"], 2), ""],
        ["PRD D102 TOTAL", round(prd_d102, 2), "sum of all orders"],
        ["KSB1 CE 9431020 (credit, abs)", round(ksb1_d102, 2),
         "CO settlement credit in CC"],
        ["Difference (PRD – KSB1)", d102_diff,
         match_status(prd_d102, ksb1_d102)],
        [""],
        sep("KSB1 ALL COST ELEMENTS"),
        ["Cost Element", "Name", "Amount (THB)"],
    ]
    for elem in sorted(ksb1_by_elem.keys()):
        info = ksb1_by_elem[elem]
        rows.append([elem, info["name"], round(info["total"], 2)])
    rows += [
        [""],
        [f"Source: KSB1_1300_{month:02d}.2026.XLSX + PRD_1300_{month:02d}.2026.XLSX  |  {TODAY}"],
    ]
    return rows


def build_mb51_vs_prd(month, mn, mb51, prd_totals, warnings):
    """Tab 6: MB51 material movements vs PRD."""
    if mb51 is None:
        return [
            [f"MB51 vs PRD — Material Reconciliation | Plant 1300 | {mn} {YEAR}"],
            [""],
            ["MB51 file not found — tab skipped"],
        ]

    pk = prd_totals["PK+CR"]
    gi = prd_totals["GI"]
    prd_material = pk["GI_Material"] + gi["GI_Material"]
    prd_scrap    = (pk["Scrap_Recovery"] + pk["GradeB"] + pk["GradeC"]
                    + gi["Scrap_Recovery"] + gi["GradeB"] + gi["GradeC"])

    # MB51 mvt 261: goods issue = credit to inventory = negative amount in SAP
    mb51_issues_abs = abs(mb51["issues"])
    mat_diff   = round(mb51_issues_abs - prd_material, 2)
    scrap_diff = round(mb51["recovery"] - prd_scrap, 2)   # mvt531 is positive

    if abs(mat_diff) > 50000:
        warnings.append(f"MB51 vs PRD material: diff={mat_diff:+,.2f} THB -- investigate separately")

    rows = [
        [f"MB51 vs PRD — Material Movement Reconciliation | Plant 1300 | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}  |  {TODAY}"],
        [""],
        ["Logic: MB51 mvt 261 = material issued to production orders (negative in SAP = goods issue)."],
        ["       PRD GI_Material = same cost from production order perspective (positive)."],
        ["       MB51 mvt 531 = by-products received back (positive)."],
        ["       Tolerance: +/-50,000 THB (timing diff + valuation diff acceptable)."],
        [""],
        sep("MATERIAL ISSUED (mvt 261 vs PRD GI_Material)"),
        ["Source", "Amount (THB)", ""],
        ["MB51 mvt 261 raw sum (negative in SAP)", round(mb51["issues"], 2),
         "Plant 1300, posting month filtered"],
        ["MB51 mvt 261 abs (for comparison)", round(mb51_issues_abs, 2),
         "abs value — goods issue = credit to inventory"],
        ["PRD GI_Material (PK+CR)", round(pk["GI_Material"], 2), ""],
        ["PRD GI_Material (GI)", round(gi["GI_Material"], 2), ""],
        ["PRD GI_Material TOTAL", round(prd_material, 2), "sum of all orders"],
        ["Difference (|MB51 261| - PRD)", mat_diff,
         "MATCH" if abs(mat_diff) < 50000 else f"DIFF -- investigate  {mat_diff:+,.2f}"],
        [""],
        sep("BY-PRODUCTS RECEIVED (mvt 531 vs PRD Scrap Recovery)"),
        ["Source", "Amount (THB)", ""],
        ["MB51 mvt 531 (by-products received, positive)", round(mb51["recovery"], 2), ""],
        ["PRD Scrap+GradeB+GradeC (PK+CR)", round(pk["Scrap_Recovery"]+pk["GradeB"]+pk["GradeC"], 2), ""],
        ["PRD Scrap+GradeB+GradeC (GI)", round(gi["Scrap_Recovery"]+gi["GradeB"]+gi["GradeC"], 2), ""],
        ["PRD Total Recovery", round(prd_scrap, 2), ""],
        ["Difference (MB51 531 - PRD Scrap)", scrap_diff,
         "MATCH" if abs(scrap_diff) < 50000 else f"DIFF  {scrap_diff:+,.2f}"],
        [""],
        sep("NOTE"),
        ["", "Differences may arise from: valuation date difference, standard vs actual price,"],
        ["", "partial settlement timing, or materials flowing to multiple plants."],
        [""],
        [f"Source: MB51_all plant_03.2026.XLSX (Plant=1300, month={month})  |  {TODAY}"],
    ]
    return rows


# ─── Sources Tab ──────────────────────────────────────────────────────────────

def build_sources(month, mn):
    """Tab 7: Traceability — where every number comes from."""
    mm = f"{month:02d}"
    rows = [
        [f"SOURCES & CALCULATION METHODOLOGY | Plant 1300 | {mn} {YEAR}"],
        [f"Prepared by: {PREPARED_BY}  |  {TODAY}"],
        [""],
        ["This tab documents the source file, column, and filter criteria for every metric",
         "in this reconciliation. Use it to trace any number back to raw data."],
        [""],
        sep("PRD — PRODUCTION ORDER DATA"),
        ["Source file", f"01_Raw/PRD_GI/PRD_1300_{mm}.2026.XLSX"],
        ["Filter", "All rows (min_row=2, skip blank)"],
        ["Order type split", '"Pickling" in Order Type → PK+CR   |   else → GI'],
        [""],
        ["Metric", "Tab(s) used", "Column index (0-based)", "Calculation"],
        ["CRC Output GR qty (kg)", "Overview, Cost_BuildUp", "col 32 — GR Output Qty",
         "SUM where order type = PK+CR"],
        ["GI Output GR qty (kg)",  "Overview, Cost_BuildUp", "col 32 — GR Output Qty",
         "SUM where order type = GI"],
        ["GI GR Output THB",       "Cost_BuildUp",           "col 34 — GR Output THB",
         "SUM per order group"],
        ["HRC Material (L0)",      "Cost_BuildUp",           "col 22 — GI_Material",
         "SUM where PK+CR"],
        ["Zinc / ZAM Material (L2)","Cost_BuildUp",          "col 22 — GI_Material",
         "SUM where GI"],
        ["Scrap Recovery",         "Cost_BuildUp",           "col 25+28+31 — Scrap+GradeB+GradeC",
         "SUM per order group"],
        ["D101 Direct Machine",    "Cost_BuildUp, KSB1_vs_PRD", "col 37 — D101",
         "SUM ALL orders (PK+CR + GI)"],
        ["D102 Direct Labor",      "Cost_BuildUp, KSB1_vs_PRD", "col 40 — D102",
         "SUM ALL orders"],
        ["Indirect I101–I111",     "Cost_BuildUp",           "cols 43,46,49,52,55,58,61,64,67,70,73",
         "SUM ALL orders"],
        [""],
        sep("KSB1 — COST CENTER ACTUALS (SETTLEMENT ENTRIES ONLY)"),
        ["Source file", f"01_Raw/PRD_GI/KSB1_1300_{mm}.2026.XLSX"],
        ["Filter", "Skip rows where BOTH CC and Cost Element are empty (SAP subtotal rows)"],
        ["Content", "This KSB1 extract = settlement entries only (943x/9439x credit from CC to orders)"],
        [""],
        ["Metric", "Tab(s) used", "Cost Element", "Calculation"],
        ["KSB1 D101 settlement",   "KSB1_vs_PRD",  "9431010 — Direct Machine",
         "SUM(abs) — credit amount"],
        ["KSB1 D102 settlement",   "KSB1_vs_PRD",  "9431020 — Direct Labour",
         "SUM(abs) — credit amount"],
        ["KSB1 9439030",           "GL_vs_KSB1",   "9439030 — Indirect Labour",
         "SUM(abs) — includes CRC→GI intra-CC material transfer (not pure overhead)"],
        ["KSB1 9439020",           "GL_vs_KSB1",   "9439020 — Indirect Electricity",
         "SUM(abs)"],
        ["KSB1 9439040",           "GL_vs_KSB1",   "9439040 — Indirect R&M",
         "SUM(abs)"],
        ["KSB1 Total Settlement",  "GL_vs_KSB1",   "All elements",
         "SUM(abs) all CCs excl. subtotal rows"],
        [""],
        sep("GL — FINANCIAL ACCOUNTING LINE ITEMS"),
        ["Source file", "01_Raw/PRD_GI/AMC_GL_03.2026.XLSX  (Jan–Mar combined)"],
        ["Filter applied", f"(1) Posting date month = {month}  "
         "(2) Cost Center starts with '13'  "
         "(3) GL 5391020 + 5211010 EXCLUDED"],
        [""],
        ["Metric", "Tab(s) used", "GL Account Prefix", "Category"],
        ["Labor",                "GL_vs_KSB1, GL_vs_TB", "551xxx / 552xxx / 553xxx", "Included in overhead"],
        ["Electricity",         "GL_vs_KSB1, GL_vs_TB", "561xxx",                  "Included in overhead"],
        ["Repair & Maintenance","GL_vs_KSB1, GL_vs_TB", "571xxx",                  "Included in overhead"],
        ["Depreciation",        "GL_vs_KSB1, GL_vs_TB", "581xxx",                  "Included in overhead"],
        ["Tools/Supplies/Other","GL_vs_KSB1, GL_vs_TB", "591xxx",                  "Included in overhead"],
        ["Mfg Supplies",        "GL_vs_KSB1, GL_vs_TB", "592xxx",                  "Included in overhead"],
        ["Other Expenses",      "GL_vs_KSB1, GL_vs_TB", "599xxx",                  "Included in overhead"],
        ["Transport",           "GL_vs_KSB1, GL_vs_TB", "611xxx",                  "Included in overhead"],
        ["CO Settlement",       "GL_vs_KSB1",            "943xxx",                  "Credit — offsets overhead"],
        ["ML Price Adj",        "GL_vs_KSB1 (excl.)",    "5391020",
         "EXCLUDED — ML revaluation, not production cost. Will be allocated to orders when CKMLCP completes"],
        ["COGM Journal",        "GL_vs_KSB1 (excl.)",    "5211010",                 "EXCLUDED — inventory transfer entry"],
        [""],
        ["Note: GL overhead (CC 13*) is a PARTIAL view.",
         "Some overhead (payroll, depreciation allocation) posts to FI without explicit CC in the document.",
         "Use TB for the complete overhead picture."],
        [""],
        sep("TB — TRIAL BALANCE"),
        ["Source file", f"01_Raw/PRD_GI/AMC_TB_{mm}.2026.XLSX  (or AMC_TB_01-03.2026.XLSX)"],
        ["Filter applied", f"GL Account starts with 55/56/57/58/59/61  |  Posting period = {month}"],
        ["Content", "All postings to production cost P&L accounts regardless of CC"],
        ["Used in", "GL_vs_TB — compare GL overhead (CC 13* partial) vs TB (full overhead total)"],
        [""],
        sep("MB51 — MATERIAL MOVEMENTS"),
        ["Source file", "01_Raw/PRD_GI/MB51_all plant_03.2026.XLSX  (Jan–Mar combined, all plants)"],
        ["Filter applied",
         f"Plant = '1300'  |  Movement Type in {{261, 531}}  |  Posting Date month = {month}"],
        [""],
        ["Metric", "Tab(s) used", "Movement Type", "Sign in SAP", "Calculation"],
        ["Material issued to orders", "MB51_vs_PRD", "261 — Goods Issue for Order",
         "NEGATIVE (credit to inventory)", "SUM then abs()"],
        ["By-products received",      "MB51_vs_PRD", "531 — Receipt of By-Products",
         "POSITIVE (debit to inventory)", "SUM directly"],
        [""],
        sep("CKMLCP / MATERIAL LEDGER CONTEXT"),
        ["Applies to", "Jan–Mar 2026 data in this file"],
        ["Jan 2026", "CKMLCP closed late — Jan amounts included in Feb closing (Jan shows near-zero)"],
        ["Jan–Mar 2026", "ML closed WITHOUT completing 100% pending orders"],
        ["GL 5391020 balance", "Represents ML price adj NOT YET allocated to orders"],
        ["From Apr 2026", "Pending orders will be completed + ML closed 100% — numbers become final"],
        ["Implication",
         "Diffs in GL<>TB, MB51<>PRD, and GL CC net balance reflect INCOMPLETE ML, not errors"],
        [""],
        [f"Script: recon_plant1300_cost.py --month {month}  |  Timestamp: {TODAY}"],
    ]
    return rows


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Plant 1300 Production Cost Reconciliation — 10-tab Google Sheet"
    )
    parser.add_argument("--month", type=int, default=3,
                        help="Month number 1–4 (default: 3=March)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview tab row counts without writing to Google Sheets")
    args  = parser.parse_args()
    month = args.month
    mn    = MONTH_NAMES.get(month, str(month))

    print(f"\n{'='*65}")
    print(f"  Plant 1300 Cost Reconciliation | {mn} {YEAR}")
    print(f"{'='*65}")

    # ── Load Data ──────────────────────────────────────────────────────────────
    prd_totals, prd_detail = load_prd(month)

    # KSB1 / GL / MB51 / TB available for months 1–3
    ksb1_available = month in (1, 2, 3)
    gl_available   = month in (1, 2, 3)
    mb51_available = month in (1, 2, 3)

    if ksb1_available:
        ksb1_by_cc, ksb1_by_elem, ksb1_rows = load_ksb1(month)
    else:
        print(f"  KSB1 not available for month {month} — skipping KSB1 layers")
        ksb1_by_cc = defaultdict(float)
        ksb1_by_elem = {}
        ksb1_rows = []

    if gl_available:
        gl_by_acc, gl_cat, gl_filtered = load_gl(month)
    else:
        print(f"  GL not available for month {month} — skipping GL layers")
        gl_by_acc = {}
        gl_cat    = defaultdict(float)
        gl_filtered = []

    mb51 = load_mb51(month) if mb51_available else None
    tb   = load_tb(month)

    # ── Build Tabs ─────────────────────────────────────────────────────────────
    warnings_list = []

    # Pre-compute D101/D102 warnings before building overview
    if ksb1_available:
        pk = prd_totals["PK+CR"]
        gi = prd_totals["GI"]
        prd_d101 = pk["D101_Direct_Mach"]  + gi["D101_Direct_Mach"]
        prd_d102 = pk["D102_Direct_Labor"] + gi["D102_Direct_Labor"]
        ksb1_d101 = abs(ksb1_by_elem.get("9431010", {}).get("total", 0.0))
        ksb1_d102 = abs(ksb1_by_elem.get("9431020", {}).get("total", 0.0))
        if abs(prd_d101 - ksb1_d101) > 1000:
            warnings_list.append(
                f"D101 PRD vs KSB1: diff={prd_d101-ksb1_d101:+,.2f} THB (tol +/-1,000)")
        if abs(prd_d102 - ksb1_d102) > 1000:
            warnings_list.append(
                f"D102 PRD vs KSB1: diff={prd_d102-ksb1_d102:+,.2f} THB (tol +/-1,000)")

    # MB51 warning is generated inside build_mb51_vs_prd — no pre-compute needed here

    print("\nBuilding tab content...")

    tab_data = {
        "Overview":    build_overview(month, mn, prd_totals, gl_cat,
                                      ksb1_by_cc, mb51, tb, warnings_list),
        "Cost_BuildUp": build_cost_buildup(month, mn, prd_totals),
        "GL_vs_TB":    build_gl_vs_tb(month, mn, gl_cat, gl_by_acc, tb),
        "GL_vs_KSB1":  (build_gl_vs_ksb1(month, mn, gl_cat, gl_by_acc, ksb1_by_cc, ksb1_by_elem)
                        if gl_available else [[f"GL not available for month {month}"]]),
        "KSB1_vs_PRD": (build_ksb1_vs_prd(month, mn, prd_totals, ksb1_by_elem, warnings_list)
                        if ksb1_available else [[f"KSB1 not available for month {month}"]]),
        "MB51_vs_PRD": build_mb51_vs_prd(month, mn, mb51, prd_totals, warnings_list),
        "Sources":     build_sources(month, mn),
    }

    # Detail tabs — raw data
    prd_hdr = [f"PRD Detail — {mn} {YEAR} | Plant 1300"]
    prd_col = ["Line", "Order Type", "Material", "Description", "WC",
               "GI_Material", "Scrap", "GradeB", "GradeC",
               "GR_QTY_kg", "GR_Output_THB", "D101", "D102", "Indirect"]

    ksb1_hdr = [f"KSB1 Detail — {mn} {YEAR} | Plant 1300 CC 1387xxx"]
    ksb1_col = ["CC", "Cost Element", "Name", "Amount (THB)"]

    gl_hdr = [f"GL Detail — {mn} {YEAR} | CC 13* (excl. {EXCLUDE_GL})"]
    gl_col = ["CC", "GL Account", "Account Name", "Doc Type", "Posting Date",
              "Amount (THB)", "Status"]

    mb51_hdr = [f"MB51 Detail — {mn} {YEAR} | Plant 1300 mvt 261+531"]
    mb51_col = ["Plant", "Mvt", "Posting Date", "Material", "Description", "Amount (THB)"]

    detail_tabs = {
        "GL_Detail":   [[gl_hdr[0]], gl_col] + (gl_filtered if gl_available else [["n/a"]]),
        "KSB1_Detail": [[ksb1_hdr[0]], ksb1_col] + (ksb1_rows if ksb1_available else [["n/a"]]),
        "PRD_Detail":  [[prd_hdr[0]], prd_col] + prd_detail,
        "MB51_Detail": [[mb51_hdr[0]], mb51_col] + (mb51["rows"] if mb51 else [["n/a"]]),
    }

    # Summary
    if warnings_list:
        print(f"\n  [!]  WARNINGS ({len(warnings_list)}):")
        for w in warnings_list:
            print(f"     {w}")

    all_tabs = list(tab_data.items()) + list(detail_tabs.items())
    for tname, rows in all_tabs:
        print(f"  Tab {tname:20s}: {len(rows):>4} rows")

    if args.dry_run:
        print(f"\nDry run complete — nothing written.")
        return

    # ── Write to Google Sheets ─────────────────────────────────────────────────
    sheet_name = f"Plant_1300_Cost_Recon_{mn}_{YEAR}"
    print(f"\nConnecting to Google Sheets: {sheet_name}")
    gc = get_gspread_client()
    try:
        sh = gc.open(sheet_name)
        print(f"  Opened existing sheet")
    except Exception:
        sh = gc.create(sheet_name)
        print(f"  Created new sheet: {sh.url}")

    for tname, rows in all_tabs:
        print(f"  Writing {tname}...")
        ws = get_or_add_worksheet(sh, tname)
        ws.clear()
        if rows:
            ws.append_rows(rows, value_input_option="USER_ENTERED")

    # Remove default Sheet1 if still present
    try:
        default = sh.worksheet("Sheet1")
        if len(sh.worksheets()) > 1:
            sh.del_worksheet(default)
    except Exception:
        pass

    print(f"\nDone!  Sheet URL: {sh.url}")
    print(f"Sheet name: {sheet_name}")
    if warnings_list:
        print(f"\n[!]  {len(warnings_list)} warning(s) — see Overview tab for details.")
    return sh.url


if __name__ == "__main__":
    main()
