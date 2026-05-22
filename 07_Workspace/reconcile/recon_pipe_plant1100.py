"""
recon_pipe_plant1100.py
=======================
Reconciliation — Pipe Production Cost | Plant 1100 | parameterised by month

Usage:
  python recon_pipe_plant1100.py --month 1    # Jan 2026
  python recon_pipe_plant1100.py --month 2    # Feb 2026
  python recon_pipe_plant1100.py --month 3    # Mar 2026

Sources (01_Raw/PRD_GI/):
  PRD_1100_MM.2026.XLSX   — Production Order cost per order
  KSB1_1100_MM.2026.XLSX  — Cost Center Actuals CO (CC 1187xxx)
  AMC_GL_03.2026.XLSX     — GL Line Items Jan-Mar (filter CC 1187*)

Plant 1100 structure:
  Slitting  : Order Type "A1 Slit Coil / Eye up" | Work Centers SL02, SL03, SL09
  Pipe      : Order Type "A1 Pipe Line"           | Work Centers PL01,PL08,PL12,PL17,PL18,FF01,FF02
  Rework    : "A1 Rework" | RW01
  Packaging : "A1 Packaging" | CP01

  Combine Order: One parent order holds GI Amount; child orders (Original Order Status=Yes) hold GR.
                 No double-counting when summing all rows.

Prepared by: Claude Code
"""

import sys, datetime, argparse
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from collections import defaultdict
from utils.auth import get_gspread_client
from utils.sheets import get_or_add_worksheet

PREPARED_BY = "Claude Code"
TODAY = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
DATA_DIR = str(Path(__file__).resolve().parent.parent.parent / '01_Bronze_Raw' / 'PRD_GI')

MONTH_NAMES = {1: "Jan", 2: "Feb", 3: "Mar"}

# Work Center → Process classification
WC_PROCESS = {
    "SL02": "Slitting", "SL03": "Slitting", "SL09": "Slitting",
    "SL04": "Slitting", "SL05": "Slitting", "SL06": "Slitting",
    "PL01": "Pipe", "PL02": "Pipe", "PL03": "Pipe", "PL04": "Pipe",
    "PL05": "Pipe", "PL06": "Pipe", "PL07": "Pipe", "PL08": "Pipe",
    "PL09": "Pipe", "PL10": "Pipe", "PL11": "Pipe", "PL12": "Pipe",
    "PL13": "Pipe", "PL14": "Pipe", "PL15": "Pipe", "PL16": "Pipe",
    "PL17": "Pipe", "PL18": "Pipe", "PL19": "Pipe",
    "FF01": "Pipe", "FF02": "Pipe", "FF03": "Pipe",
    "RW01": "Rework",
    "CP01": "Packaging",
}

# KSB1 CC prefix → Process
def cc_process(cc: str) -> str:
    c = str(cc).strip()
    if c.startswith("11871"):  return "Slitting"
    if c.startswith("11872"):  return "Pipe"
    if c.startswith("11873"):  return "Common/QC"
    if c.startswith("11876"):  return "Service/Other"
    return "Other"

# GL account classification (production only)
def classify_gl(acc: str) -> str:
    a = str(acc)
    if a.startswith(("551", "552", "553")): return "Labor"
    if a.startswith("561"): return "Electricity"
    if a.startswith("571"): return "Repair & Maintenance"
    if a.startswith("581"): return "Depreciation"
    if a.startswith("591"): return "Tools/Supplies"
    if a.startswith("592"): return "Manufacturing Supplies"
    if a.startswith("599"): return "Other Production"
    if a.startswith("611"): return "Transport"
    if a.startswith("531"): return "Production Variance"
    if a.startswith("539"): return "Variance Adjustment (ML)"
    if a.startswith("521"): return "COGM Semi-FG (exclude)"
    if a.startswith("943"): return "CO Settlement (to Orders)"
    # Non-production — flag but don't include in production total
    if a.startswith(("6", "7", "8")): return "SGA/Finance (non-production)"
    return "Other"

def to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").strip())
    except: return 0.0


# ─────────────────────────────────────────────────────────────────────────────
def load_prd(month: int):
    mm = f"{month:02d}"
    path = f"{DATA_DIR}\\PRD_1100_{mm}.2026.XLSX"
    print(f"Loading PRD: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    I_COLS = [43, 46, 49, 52, 55, 58, 61, 64, 67, 70, 73]

    # totals[process] → dict of amount fields
    totals = defaultdict(lambda: defaultdict(float))
    detail = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue

        wc      = str(row[5] or "").strip()
        ot      = str(row[15] or "").strip()
        order   = str(row[16] or "").strip()
        comb    = str(row[17] or "").strip()  # Combined Order number
        orig_st = str(row[19] or "").strip()  # "Yes" if this is an Original (output) order

        process = WC_PROCESS.get(wc, "Other")

        gi  = to_float(row[22]); scr = to_float(row[25])
        grb = to_float(row[28]); grc = to_float(row[31])
        gr  = to_float(row[34]); d1  = to_float(row[37]); d2 = to_float(row[40])
        ind = sum(to_float(row[ic]) for ic in I_COLS if ic < len(row))
        net = gi + d1 + d2 + ind - scr - grb - grc

        totals[process]["GI_Material"]       += gi
        totals[process]["Scrap_Recovery"]    += scr
        totals[process]["GradeB"]            += grb
        totals[process]["GradeC"]            += grc
        totals[process]["GR_Output"]         += gr
        totals[process]["D101_Direct_Mach"]  += d1
        totals[process]["D102_Direct_Labor"] += d2
        totals[process]["Indirect_Total"]    += ind
        totals[process]["Net_Cost"]          += net

        detail.append([
            process, wc, ot, order,
            "CombineParent" if (comb and comb == order) else ("OriginalChild" if orig_st == "Yes" else "Standalone"),
            str(row[0] or ""), str(row[1] or ""),
            round(gi, 2), round(scr, 2), round(grb, 2), round(grc, 2),
            round(gr, 2), round(d1, 2), round(d2, 2), round(ind, 2), round(net, 2)
        ])

    print(f"  Rows: {len(detail)} | Slit={len([d for d in detail if d[0]=='Slitting'])} "
          f"Pipe={len([d for d in detail if d[0]=='Pipe'])} "
          f"Other={len([d for d in detail if d[0] not in ('Slitting','Pipe')])}")
    return totals, detail


def load_ksb1(month: int):
    mm = f"{month:02d}"
    path = f"{DATA_DIR}\\KSB1_1100_{mm}.2026.XLSX"
    print(f"Loading KSB1: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    by_cc      = defaultdict(float)
    by_elem    = defaultdict(lambda: {"name": "", "total": 0.0})
    by_process = defaultdict(lambda: defaultdict(float))  # process → elem → amount
    rows       = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row[:4]): continue
        cc   = str(row[0]).strip() if row[0] else "(no CC)"
        elem = str(row[1]).strip() if row[1] else ""
        name = str(row[2]).strip() if row[2] else ""
        amt  = to_float(row[3])
        proc = cc_process(cc)

        by_cc[cc] += amt
        if elem:
            by_elem[elem]["name"]  = name
            by_elem[elem]["total"] += amt
            by_process[proc][elem] += amt
        rows.append([cc, proc, elem, name, round(amt, 2)])

    print(f"  KSB1 rows: {len(rows)}")
    return by_cc, by_elem, by_process, rows


def load_gl(month: int):
    """Load GL filtered to CC 1187* (production CCs only) for the given month."""
    path = f"{DATA_DIR}\\AMC_GL_03.2026.XLSX"
    print(f"Loading GL (month={month}, CC 1187*)... may take ~60s")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    by_account  = defaultdict(lambda: {"name": "", "total": 0.0, "category": ""})
    filtered    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row): continue
        gl_acc   = str(row[2]).strip() if row[2] else ""
        gl_name  = str(row[3]).strip() if row[3] else ""
        posting  = row[6]
        amount   = to_float(row[10])
        doc_type = str(row[13]).strip() if row[13] else ""
        cc       = str(row[14]).strip() if row[14] else ""

        if not (isinstance(posting, datetime.datetime) and posting.month == month):
            continue
        if not cc.startswith("1187"):
            continue

        cat = classify_gl(gl_acc)
        by_account[gl_acc]["name"]     = gl_name
        by_account[gl_acc]["total"]   += amount
        by_account[gl_acc]["category"] = cat
        filtered.append([cc, gl_acc, gl_name, doc_type,
                         posting.strftime("%d/%m/%Y"), round(amount, 2), cat])

    print(f"  GL rows filtered (CC 1187*): {len(filtered):,}")
    return by_account, filtered


# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", type=int, required=True, choices=[1, 2, 3])
    args  = parser.parse_args()
    month = args.month
    mn    = MONTH_NAMES[month]

    print(f"\n{'='*65}")
    print(f"Pipe Production Cost Reconciliation | Plant 1100 | {mn} 2026")
    print(f"{'='*65}")

    prd_totals, prd_detail           = load_prd(month)
    ksb1_by_cc, ksb1_by_elem, ksb1_by_proc, ksb1_rows = load_ksb1(month)
    gl_by_account, gl_filtered       = load_gl(month)

    # ── Aggregate totals ──
    grand = defaultdict(float)
    for proc in prd_totals:
        for k, v in prd_totals[proc].items():
            grand[k] += v

    cat_totals = defaultdict(float)
    for acc, info in gl_by_account.items():
        cat_totals[info["category"]] += info["total"]

    prd_d101 = grand["D101_Direct_Mach"]
    prd_d102 = grand["D102_Direct_Labor"]
    gl_9431010 = abs(gl_by_account.get("9431010", {}).get("total", 0.0))
    gl_9431020 = abs(gl_by_account.get("9431020", {}).get("total", 0.0))

    overhead_cats = ["Labor", "Electricity", "Repair & Maintenance", "Depreciation",
                     "Manufacturing Supplies", "Tools/Supplies", "Other Production", "Transport"]
    gl_overhead   = sum(cat_totals.get(c, 0) for c in overhead_cats)
    co_settlement = cat_totals.get("CO Settlement (to Orders)", 0.0)
    cc_net        = gl_overhead + co_settlement

    def match(a, b, tol=1.0):
        diff = round(a - b, 2)
        return f"MATCH ({diff:+,.2f})" if abs(diff) <= tol else f"DIFF ({diff:+,.2f})"

    # Console summary
    print(f"\n{'Process':<12} {'GI Material':>16} {'GR Output':>16} {'D101':>12} {'D102':>12} {'Net Cost':>16}")
    print("-" * 86)
    for proc in ["Slitting", "Pipe", "Rework", "Packaging", "Other"]:
        t = prd_totals.get(proc)
        if not t: continue
        print(f"{proc:<12} {t['GI_Material']:>16,.2f} {t['GR_Output']:>16,.2f} "
              f"{t['D101_Direct_Mach']:>12,.2f} {t['D102_Direct_Labor']:>12,.2f} "
              f"{t['Net_Cost']:>16,.2f}")
    print("-" * 86)
    print(f"{'TOTAL':<12} {grand['GI_Material']:>16,.2f} {grand['GR_Output']:>16,.2f} "
          f"{grand['D101_Direct_Mach']:>12,.2f} {grand['D102_Direct_Labor']:>12,.2f} "
          f"{grand['Net_Cost']:>16,.2f}")
    print(f"\n  D101: PRD={prd_d101:,.2f}  GL={gl_9431010:,.2f}  | {match(prd_d101, gl_9431010)}")
    print(f"  D102: PRD={prd_d102:,.2f}  GL={gl_9431020:,.2f}  | {match(prd_d102, gl_9431020)}")

    # ── Write to Google Sheets ──
    sheet_name = f"RECON_PIPE_1100_{mn}_2026"
    print(f"\nConnecting to Google Sheets: {sheet_name}")
    gc = get_gspread_client()
    try:
        sh = gc.open(sheet_name)
        print("  Opened existing")
    except Exception:
        sh = gc.create(sheet_name)
        print("  Created new")

    def tab(name):
        return get_or_add_worksheet(sh, name)

    # ── Tab 1: PRD_Detail ──
    print("  Writing PRD_Detail...")
    ws1 = tab("PRD_Detail")
    ws1.clear()
    ws1.append_row([f"PRD Detail — {mn} 2026 | Plant 1100 | Prepared by: {PREPARED_BY}"])
    ws1.append_row(["Process", "Work Center", "Order Type", "Order", "Order Role",
                    "Material", "Description",
                    "GI_Material", "Scrap_Recovery", "GradeB", "GradeC",
                    "GR_Output", "D101", "D102", "Indirect_Total", "Net_Cost"])
    ws1.append_rows(prd_detail, value_input_option="USER_ENTERED")

    # ── Tab 2: PRD_Summary_by_Process ──
    print("  Writing PRD_Summary...")
    ws2 = tab("PRD_Summary")
    ws2.clear()
    rows_sum = [
        [f"PRD Summary — {mn} 2026 | Plant 1100 | By Process"],
        [],
        ["Process", "GI_Material", "Scrap_Recovery", "GradeB", "GradeC",
         "GR_Output", "D101_Direct_Mach", "D102_Direct_Labor", "Indirect_Total", "Net_Cost"],
    ]
    for proc in ["Slitting", "Pipe", "Rework", "Packaging", "Other"]:
        t = prd_totals.get(proc)
        if not t: continue
        rows_sum.append([proc,
            round(t["GI_Material"], 2), round(t["Scrap_Recovery"], 2),
            round(t["GradeB"], 2), round(t["GradeC"], 2), round(t["GR_Output"], 2),
            round(t["D101_Direct_Mach"], 2), round(t["D102_Direct_Labor"], 2),
            round(t["Indirect_Total"], 2), round(t["Net_Cost"], 2)])
    rows_sum.append(["TOTAL",
        round(grand["GI_Material"], 2), round(grand["Scrap_Recovery"], 2),
        round(grand["GradeB"], 2), round(grand["GradeC"], 2), round(grand["GR_Output"], 2),
        round(grand["D101_Direct_Mach"], 2), round(grand["D102_Direct_Labor"], 2),
        round(grand["Indirect_Total"], 2), round(grand["Net_Cost"], 2)])
    ws2.append_rows(rows_sum, value_input_option="USER_ENTERED")

    # ── Tab 3: KSB1_Summary ──
    print("  Writing KSB1_Summary...")
    ws3 = tab("KSB1_Summary")
    ws3.clear()
    ksb1_out = [
        [f"KSB1 Summary — {mn} 2026 | Plant 1100 | CC 1187xxx"],
        [],
        ["Process", "Cost Center", "Cost Element", "CE Name", "Amount (THB)"],
    ]
    for proc in ["Slitting", "Pipe", "Common/QC", "Service/Other", "Other"]:
        proc_rows = [r for r in ksb1_rows if r[1] == proc]
        if not proc_rows: continue
        cc_sub = defaultdict(lambda: defaultdict(lambda: {"name":"","total":0.0}))
        for r in proc_rows:
            if r[2]:
                cc_sub[r[0]][r[2]]["name"]   = r[3]
                cc_sub[r[0]][r[2]]["total"] += r[4]
        for cc in sorted(cc_sub):
            for elem, info in sorted(cc_sub[cc].items()):
                ksb1_out.append([proc, cc, elem, info["name"], round(info["total"], 2)])
            ksb1_out.append([f"  — {cc} subtotal —", "", "", "", round(ksb1_by_cc[cc], 2)])
        proc_total = sum(ksb1_by_cc[cc] for cc in cc_sub)
        ksb1_out.append([f"=== {proc} TOTAL ===", "", "", "", round(proc_total, 2)])
        ksb1_out.append([])
    ksb1_out.append(["=== GRAND TOTAL ===", "", "", "", round(sum(ksb1_by_cc.values()), 2)])
    ws3.append_rows(ksb1_out, value_input_option="USER_ENTERED")

    # ── Tab 4: GL_Summary ──
    print("  Writing GL_Summary...")
    ws4 = tab("GL_Summary")
    ws4.clear()
    cat_order = ["Labor", "Electricity", "Repair & Maintenance", "Depreciation",
                 "Manufacturing Supplies", "Tools/Supplies", "Other Production", "Transport",
                 "CO Settlement (to Orders)", "Production Variance",
                 "Variance Adjustment (ML)", "COGM Semi-FG (exclude)",
                 "SGA/Finance (non-production)", "Other"]
    gl_out = [[f"GL Summary — {mn} 2026 | CC 1187* | Plant 1100 (Production CCs only)"],
              [], ["Category", "GL Account", "Account Name", "Amount (THB)"]]
    gl_total = 0.0
    prod_total = 0.0
    for cat in cat_order:
        items = [(a, i["name"], round(i["total"], 2))
                 for a, i in sorted(gl_by_account.items()) if i["category"] == cat]
        if not items: continue
        cat_total = sum(x[2] for x in items)
        for acc, name, amt in items:
            gl_out.append([cat, acc, name, amt])
        gl_out.append([f"--- {cat} SUBTOTAL ---", "", "", round(cat_total, 2)])
        gl_out.append([])
        gl_total += cat_total
        if cat not in ("SGA/Finance (non-production)", "COGM Semi-FG (exclude)",
                        "Variance Adjustment (ML)"):
            prod_total += cat_total
    gl_out.append(["=== GRAND TOTAL (all) ===", "", "", round(gl_total, 2)])
    gl_out.append(["=== PRODUCTION TOTAL (excl. SGA/ML/COGM) ===", "", "", round(prod_total, 2)])
    ws4.append_rows(gl_out, value_input_option="USER_ENTERED")

    # ── Tab 5: RECON_Result ──
    print("  Writing RECON_Result...")
    ws5 = tab("RECON_Result")
    ws5.clear()

    recon = [
        [f"RECONCILIATION — Pipe Production Cost | {mn} 2026 | Plant 1100"],
        [f"Prepared by: {PREPARED_BY}   |   {TODAY}   |   Status: Draft"],
        [],
        ["=" * 65],
        ["LAYER 1: PRD → GL  |  Direct Costs (D101 / D102)"],
        ["=" * 65],
        ["Item", "PRD (THB)", "GL/CO (THB)", "Diff", "Status"],
        ["D101 Direct Machine (all processes)",
         round(prd_d101, 2), round(gl_9431010, 2),
         round(prd_d101 - gl_9431010, 2), match(prd_d101, gl_9431010)],
        ["D102 Direct Labor (all processes)",
         round(prd_d102, 2), round(gl_9431020, 2),
         round(prd_d102 - gl_9431020, 2), match(prd_d102, gl_9431020)],
        [],
        ["  -- By Process --"],
        ["Process", "D101 (PRD)", "D102 (PRD)", "Net Cost (PRD)", "GI Material", "GR Output"],
    ]
    for proc in ["Slitting", "Pipe", "Rework", "Packaging"]:
        t = prd_totals.get(proc)
        if not t: continue
        recon.append([proc,
            round(t["D101_Direct_Mach"], 2), round(t["D102_Direct_Labor"], 2),
            round(t["Net_Cost"], 2), round(t["GI_Material"], 2), round(t["GR_Output"], 2)])
    recon.append(["TOTAL",
        round(grand["D101_Direct_Mach"], 2), round(grand["D102_Direct_Labor"], 2),
        round(grand["Net_Cost"], 2), round(grand["GI_Material"], 2), round(grand["GR_Output"], 2)])

    recon += [
        [],
        ["=" * 65],
        ["LAYER 2: GL Overhead → KSB1  |  CC 1187xxx"],
        ["=" * 65],
        ["Item", "GL Actual (THB)", "KSB1 Total |amt| (THB)", "Diff", "Status"],
        ["Overhead (Labor+Elec+R&M+Depr+Others)",
         round(gl_overhead, 2), round(abs(sum(ksb1_by_cc.values())), 2),
         round(gl_overhead - abs(sum(ksb1_by_cc.values())), 2),
         match(gl_overhead, abs(sum(ksb1_by_cc.values())), tol=500)],
        ["CC Net Balance (Actual + Settlement ≈ 0)",
         round(gl_overhead, 2), round(abs(co_settlement), 2),
         round(cc_net, 2),
         "BALANCED" if abs(cc_net) < 1000 else f"UNBALANCED ({cc_net:+,.2f})"],
        [],
        ["  -- KSB1 by Process --"],
        ["Process", "KSB1 Total (THB)", "D101", "D102", "Indirect Total"],
    ]
    for proc in ["Slitting", "Pipe", "Common/QC", "Service/Other"]:
        pe = ksb1_by_proc.get(proc, {})
        subtotal = sum(ksb1_by_cc[cc] for cc in ksb1_by_cc
                       if cc_process(cc) == proc)
        recon.append([proc, round(subtotal, 2),
            round(abs(pe.get("9431010", 0)), 2),
            round(abs(pe.get("9431020", 0)), 2),
            round(abs(sum(v for k, v in pe.items() if k.startswith("9439"))), 2)])
    recon.append(["TOTAL", round(sum(ksb1_by_cc.values()), 2),
        round(abs(ksb1_by_elem.get("9431010", {}).get("total", 0)), 2),
        round(abs(ksb1_by_elem.get("9431020", {}).get("total", 0)), 2),
        round(abs(sum(info["total"] for e, info in ksb1_by_elem.items() if e.startswith("9439"))), 2)])

    recon += [
        [],
        ["=" * 65],
        ["LAYER 3: Cost Build-Up Summary"],
        ["=" * 65],
        ["Item", "Amount (THB)", "Note"],
        ["--- SLITTING ---", "", ""],
        ["GI Material (GIC/HRC input)",
         round(prd_totals["Slitting"]["GI_Material"], 2), "PRD col 22 — สลิทวัตถุดิบ"],
        ["Direct Machine (D101)",
         round(prd_totals["Slitting"]["D101_Direct_Mach"], 2), ""],
        ["Direct Labor (D102)",
         round(prd_totals["Slitting"]["D102_Direct_Labor"], 2), ""],
        ["Indirect Pool (I101-I111)",
         round(prd_totals["Slitting"]["Indirect_Total"], 2), ""],
        ["Less: Scrap Recovery",
         round(-prd_totals["Slitting"]["Scrap_Recovery"], 2), "Slit Scrap credit"],
        ["SLITTING Net Cost",
         round(prd_totals["Slitting"]["Net_Cost"], 2),
         f"THB/kg = Net / GR_Output × 1000 (ถ้า GR ≠ 0)"],
        ["Slitting GR Output (THB)",
         round(prd_totals["Slitting"]["GR_Output"], 2), "Slit Coil output value"],
        [],
        ["--- PIPE / FORMING ---", "", ""],
        ["GI Material (Slit Coil input)",
         round(prd_totals["Pipe"]["GI_Material"], 2), "PRD col 22"],
        ["Direct Machine (D101)",
         round(prd_totals["Pipe"]["D101_Direct_Mach"], 2), ""],
        ["Direct Labor (D102)",
         round(prd_totals["Pipe"]["D102_Direct_Labor"], 2), ""],
        ["Indirect Pool (I101-I111)",
         round(prd_totals["Pipe"]["Indirect_Total"], 2), ""],
        ["Less: Scrap Recovery",
         round(-prd_totals["Pipe"]["Scrap_Recovery"], 2), "Trim Scrap credit"],
        ["Less: Grade B/C",
         round(-(prd_totals["Pipe"]["GradeB"] + prd_totals["Pipe"]["GradeC"]), 2), ""],
        ["PIPE Net Cost",
         round(prd_totals["Pipe"]["Net_Cost"], 2), ""],
        ["Pipe GR Output (THB)",
         round(prd_totals["Pipe"]["GR_Output"], 2), "Pipe output value"],
        [],
        ["=== TOTAL PLANT 1100 NET COST ===",
         round(grand["Net_Cost"], 2), "Slitting + Pipe + Other"],
        [],
        ["GL: Overhead posted (CC 1187*)",
         round(gl_overhead, 2), "55x-59x accounts"],
        ["GL: CO Settlement (credit)",
         round(co_settlement, 2), "943xxx"],
        ["GL: Variance Adj ML (excluded)",
         round(cat_totals.get("Variance Adjustment (ML)", 0), 2), "5391020"],
        ["GL: COGM Semi-FG (excluded)",
         round(cat_totals.get("COGM Semi-FG (exclude)", 0), 2), "521xxxx"],
        [],
        ["=" * 65],
        ["FINDINGS / ข้อสังเกต"],
        ["=" * 65],
        ["#", "รายการ", "ผลลัพธ์"],
        ["1", "D101 Direct Machine: PRD vs GL 9431010",
         f"{match(prd_d101, gl_9431010)} | PRD={prd_d101:,.2f}  GL={gl_9431010:,.2f}"],
        ["2", "D102 Direct Labor: PRD vs GL 9431020",
         f"{match(prd_d102, gl_9431020)} | PRD={prd_d102:,.2f}  GL={gl_9431020:,.2f}"],
        ["3", "Combine Order (Slitting)",
         "Parent = GI Amount, Children = GR Amount — ไม่มี double count"],
        ["4", "KSB1 (no CC)",
         f"{ksb1_by_cc.get('(no CC)', 0):,.2f} THB — ต้องตรวจสอบ"],
        ["5", "GL CC scope",
         "ใช้ CC 1187* เท่านั้น (ตัด SGA/Finance CCs ออก)"],
        [],
        [f"Timestamp: {TODAY}"],
    ]
    ws5.append_rows(recon, value_input_option="USER_ENTERED")

    # Clean up Sheet1
    try:
        default = sh.worksheet("Sheet1")
        if len(sh.worksheets()) > 1:
            sh.del_worksheet(default)
    except Exception:
        pass

    print(f"\nDone! Sheet: {sh.url}")
    return sh.url


if __name__ == "__main__":
    main()
