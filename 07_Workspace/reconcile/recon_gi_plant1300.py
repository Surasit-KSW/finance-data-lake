"""
recon_gi_plant1300.py
=====================
Reconciliation — GI Production Cost | Plant 1300 | parameterised by month

Usage:
  python recon_gi_plant1300.py --month 2    # Feb 2026
  python recon_gi_plant1300.py --month 3    # Mar 2026
  python recon_gi_plant1300.py --month 1    # Jan 2026 (re-run)

Sources (01_Raw/PRD_GI/):
  PRD_1300_MM.2026.XLSX   — Production Order cost (per month)
  KSB1_1300_MM.2026.XLSX  — Cost Center Actuals CO (per month)
  AMC_GL_03.2026.XLSX     — GL Line Items Jan-Mar (filter by posting month)

Layer: PRD → KSB1 → GL
Output: Google Sheet RECON_GI_1300_MM_2026

Prepared by: Claude Code
"""

import sys, datetime, argparse
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from collections import defaultdict
from utils.auth import get_gspread_client
from utils.sheets import get_or_add_worksheet

PREPARED_BY = "Claude Code"
TODAY = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

DATA_DIR = str(Path(__file__).resolve().parent.parent.parent / '01_Bronze_Raw' / 'PRD_GI')

MONTH_NAMES = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
               7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

CC_MAP = {
    "1387110": "Pickling Line 01",
    "1387120": "Cold Rolling Mill 01",
    "1387210": "Galvanizing Line 01",
    "(no CC)": "No CC (Direct to Order)",
}


# ─────────────────────────────────────────────────────────────────────────────
def to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").strip())
    except: return 0.0


def classify_gl(acc):
    a = str(acc)
    if a.startswith(("551", "552", "553")): return "Labor"
    if a.startswith("561"): return "Electricity"
    if a.startswith("571"): return "Repair & Maintenance"
    if a.startswith("581"): return "Depreciation"
    if a.startswith("591"): return "Tools/Supplies/Other"
    if a.startswith("592"): return "Manufacturing Supplies"
    if a.startswith("599"): return "Other Expenses"
    if a.startswith("611"): return "Transport"
    if a.startswith("541"): return "Consumption (Material)"
    if a.startswith("531"): return "Production Variance"
    if a.startswith("539"): return "Variance Adjustment (ML)"
    if a.startswith("943"): return "CO Settlement (to Orders)"
    return "Other"


# ─────────────────────────────────────────────────────────────────────────────
def load_prd(month: int):
    mm = f"{month:02d}"
    path = f"{DATA_DIR}\\PRD_1300_{mm}.2026.XLSX"
    print(f"Loading PRD: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    I_COLS = [43, 46, 49, 52, 55, 58, 61, 64, 67, 70, 73]

    totals = defaultdict(lambda: defaultdict(float))
    detail = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row[:10]):
            continue
        ot  = str(row[15] or "")
        key = "PK+CR" if "Pickling" in ot else "GI"
        gi  = to_float(row[22]);  scr = to_float(row[25])
        grb = to_float(row[28]);  grc = to_float(row[31])
        gr  = to_float(row[34]);  d1  = to_float(row[37]); d2 = to_float(row[40])
        ind = sum(to_float(row[ic]) for ic in I_COLS if ic < len(row))
        net = gi + d1 + d2 + ind - scr - grb - grc

        totals[key]["GI_Material"]       += gi
        totals[key]["Scrap_Recovery"]    += scr
        totals[key]["GradeB"]            += grb
        totals[key]["GradeC"]            += grc
        totals[key]["GR_Output"]         += gr
        totals[key]["D101_Direct_Mach"]  += d1
        totals[key]["D102_Direct_Labor"] += d2
        totals[key]["Indirect_Total"]    += ind

        detail.append([key, ot,
                        str(row[0] or ""), str(row[1] or ""), str(row[5] or ""),
                        round(gi, 2), round(scr, 2), round(grb, 2), round(grc, 2),
                        round(gr, 2), round(d1, 2), round(d2, 2),
                        round(ind, 2), round(net, 2)])

    print(f"  PRD rows: {len(detail)}")
    return totals, detail


def load_ksb1(month: int):
    mm = f"{month:02d}"
    path = f"{DATA_DIR}\\KSB1_1300_{mm}.2026.XLSX"
    print(f"Loading KSB1: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    by_cc   = defaultdict(float)
    by_elem = defaultdict(lambda: {"name": "", "total": 0.0})
    rows    = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row[:4]):
            continue
        cc   = str(row[0]).strip() if row[0] else "(no CC)"
        elem = str(row[1]).strip() if row[1] else ""
        name = str(row[2]).strip() if row[2] else ""
        amt  = to_float(row[3])
        by_cc[cc] += amt
        if elem:
            by_elem[elem]["name"]  = name
            by_elem[elem]["total"] += amt
        rows.append([cc, elem, name, round(amt, 2)])

    print(f"  KSB1 rows: {len(rows)}")
    return by_cc, by_elem, rows


def load_gl(month: int):
    path = f"{DATA_DIR}\\AMC_GL_03.2026.XLSX"
    print(f"Loading GL (month={month}, CC 13*)... may take ~60s")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    by_account = defaultdict(lambda: {"name": "", "total": 0.0, "category": ""})
    filtered   = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row): continue
        gl_acc   = str(row[2]).strip() if row[2] else ""
        gl_name  = str(row[3]).strip() if row[3] else ""
        posting  = row[6]
        amount   = to_float(row[10])
        doc_type = str(row[13]).strip() if row[13] else ""
        cc       = str(row[14]).strip() if row[14] else ""

        if not (isinstance(posting, datetime.datetime) and posting.month == month):
            continue
        if not cc.startswith("13"):
            continue

        by_account[gl_acc]["name"]     = gl_name
        by_account[gl_acc]["total"]   += amount
        by_account[gl_acc]["category"] = classify_gl(gl_acc)
        filtered.append([cc, gl_acc, gl_name, doc_type,
                         posting.strftime("%d/%m/%Y"), round(amount, 2)])

    print(f"  GL rows filtered: {len(filtered):,}")
    return by_account, filtered


# ─────────────────────────────────────────────────────────────────────────────
def build_recon_summary(month: int, prd_totals, grand, gnet,
                        gl_by_account, cat_totals, ksb1_by_cc):
    mn = MONTH_NAMES[month]

    prd_d101   = grand["D101_Direct_Mach"]
    prd_d102   = grand["D102_Direct_Labor"]
    gl_9431010 = abs(gl_by_account.get("9431010", {}).get("total", 0.0))
    gl_9431020 = abs(gl_by_account.get("9431020", {}).get("total", 0.0))

    overhead_cats = ["Labor", "Electricity", "Repair & Maintenance", "Depreciation",
                     "Manufacturing Supplies", "Tools/Supplies/Other",
                     "Other Expenses", "Transport"]
    gl_overhead    = sum(cat_totals.get(c, 0) for c in overhead_cats)
    co_settlement  = cat_totals.get("CO Settlement (to Orders)", 0.0)
    cc_net_balance = gl_overhead + co_settlement

    # Key CC actuals from KSB1
    ksb1_pk  = ksb1_by_cc.get("1387110", 0) + sum(
        v for k, v in ksb1_by_cc.items() if k.startswith("138711") and k != "1387110")
    ksb1_cr  = ksb1_by_cc.get("1387120", 0) + sum(
        v for k, v in ksb1_by_cc.items() if k.startswith("138712") and k != "1387120")
    ksb1_gi  = ksb1_by_cc.get("1387210", 0) + sum(
        v for k, v in ksb1_by_cc.items() if k.startswith("138721"))
    ksb1_known = abs(ksb1_pk + ksb1_cr + ksb1_gi)

    def match(a, b, tol=1.0):
        diff = round(a - b, 2)
        return f"MATCH (diff={diff:+,.2f})" if abs(diff) <= tol else f"DIFF ({diff:+,.2f})"

    rows = [
        [f"RECONCILIATION — GI Production Cost | {mn} 2026 | Plant 1300"],
        [f"Prepared by: {PREPARED_BY}   |   {TODAY}   |   Status: Draft"],
        [],
        ["=" * 65],
        [f"LAYER 1: PRD → GL  |  Direct Costs (D101 / D102)"],
        ["=" * 65],
        ["Item", "PRD (THB)", "GL/CO (THB)", "Diff", "Status"],
        ["D101 Direct Machine",
         round(prd_d101, 2), round(gl_9431010, 2),
         round(prd_d101 - gl_9431010, 2), match(prd_d101, gl_9431010)],
        ["D102 Direct Labor",
         round(prd_d102, 2), round(gl_9431020, 2),
         round(prd_d102 - gl_9431020, 2), match(prd_d102, gl_9431020)],
        [],
        ["=" * 65],
        [f"LAYER 2: GL Overhead → KSB1  |  CC 1387xxx Actual vs Settled"],
        ["=" * 65],
        ["Item", "GL Actual (THB)", "KSB1 Main CC |amt| (THB)", "Diff", "Status"],
        ["Overhead (55x-61x excl. excl.)",
         round(gl_overhead, 2), round(ksb1_known, 2),
         round(gl_overhead - ksb1_known, 2),
         match(gl_overhead, ksb1_known, tol=100)],
        ["CC Net (Actual + CO Settlement  ≈ 0)",
         round(gl_overhead, 2), round(abs(co_settlement), 2),
         round(cc_net_balance, 2),
         "BALANCED" if abs(cc_net_balance) < 1000 else f"UNBALANCED ({cc_net_balance:+,.2f})"],
        [],
        ["=" * 65],
        [f"LAYER 3: Cost Build-Up (PRD)"],
        ["=" * 65],
        ["Item", "Amount (THB)", "Note"],
        ["Material: HRC/CRC GI Amount",
         round(grand["GI_Material"], 2), "Raw material consumed — from PRD col 22"],
        ["Direct Machine (D101)",
         round(prd_d101, 2), "CC→Order settlement | verified vs GL 9431010"],
        ["Direct Labor (D102)",
         round(prd_d102, 2), "CC→Order settlement | verified vs GL 9431020"],
        ["Indirect Pool (I101-I111)",
         round(grand["Indirect_Total"], 2), "Overhead allocation to orders"],
        ["Less: Scrap Recovery",
         round(-grand["Scrap_Recovery"], 2), "By-product credit (Mvt 531 receipts)"],
        ["Less: Grade B Recovery",
         round(-grand["GradeB"], 2), ""],
        ["Less: Grade C Recovery",
         round(-grand["GradeC"], 2), ""],
        ["NET Production Cost (Orders)",
         round(gnet, 2), "Total cost settled to production orders"],
        [],
        ["GR Output Value", round(grand["GR_Output"], 2), "Finished goods received"],
        [],
        ["GL Overhead posted (CC 13*)",
         round(gl_overhead, 2), "55x-61x accounts"],
        ["GL CO Settlement (credit)",
         round(co_settlement, 2), "943xxx — cost moved to orders"],
        ["GL Variance Adj ML (excluded)",
         round(cat_totals.get("Variance Adjustment (ML)", 0), 2),
         "5391020 — ML price correction, not production cost"],
        [],
        ["=" * 65],
        ["FINDINGS / ข้อสังเกต"],
        ["=" * 65],
        ["#", "รายการ", "ผลลัพธ์"],
        ["1", f"D101 Direct Machine: PRD vs GL 9431010",
         f"{match(prd_d101, gl_9431010)} | PRD={prd_d101:,.2f} GL={gl_9431010:,.2f}"],
        ["2", f"D102 Direct Labor: PRD vs GL 9431020",
         f"{match(prd_d102, gl_9431020)} | PRD={prd_d102:,.2f} GL={gl_9431020:,.2f}"],
        ["3", "Material Cost (HRC/CRC) ใน PRD",
         "GI Amount — GL = Inventory Credit (ไม่ผ่าน CC)"],
        ["4", "Indirect I101-I111",
         f"รวม {grand['Indirect_Total']:,.2f} THB"],
        ["5", "GL 5391020 ML Adj",
         f"{cat_totals.get('Variance Adjustment (ML)', 0):,.2f} THB — excluded"],
        ["6", "KSB1 (no CC) entries",
         f"{ksb1_by_cc.get('(no CC)', 0):,.2f} THB — ยังไม่ classify"],
        [],
        [f"Timestamp: {TODAY}"],
    ]
    return rows


# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Reconcile GI Plant 1300 by month")
    parser.add_argument("--month", type=int, required=True, choices=[1, 2, 3],
                        help="Month number (1=Jan, 2=Feb, 3=Mar)")
    args = parser.parse_args()
    month = args.month
    mn    = MONTH_NAMES[month]

    print(f"\n{'='*60}")
    print(f"GI Production Cost Reconciliation | Plant 1300 | {mn} 2026")
    print(f"{'='*60}")

    prd_totals, prd_detail = load_prd(month)
    ksb1_by_cc, ksb1_by_elem, ksb1_rows = load_ksb1(month)
    gl_by_account, gl_filtered = load_gl(month)

    # Grand totals
    grand = defaultdict(float)
    for key in ["PK+CR", "GI"]:
        for k, v in prd_totals[key].items():
            grand[k] += v
    gnet = (grand["GI_Material"] + grand["D101_Direct_Mach"] + grand["D102_Direct_Labor"]
            + grand["Indirect_Total"] - grand["Scrap_Recovery"]
            - grand["GradeB"] - grand["GradeC"])

    cat_totals = defaultdict(float)
    for acc, info in gl_by_account.items():
        cat_totals[info["category"]] += info["total"]

    # Quick console summary
    print(f"\n  PRD Net Cost:       {gnet:>18,.2f} THB")
    print(f"  D101 (PRD):        {grand['D101_Direct_Mach']:>18,.2f}")
    print(f"  D101 (GL 9431010): {abs(gl_by_account.get('9431010',{}).get('total',0)):>18,.2f}")
    print(f"  D102 (PRD):        {grand['D102_Direct_Labor']:>18,.2f}")
    print(f"  D102 (GL 9431020): {abs(gl_by_account.get('9431020',{}).get('total',0)):>18,.2f}")

    # Write to Google Sheets
    sheet_name = f"RECON_GI_1300_{mn}_2026"
    print(f"\nConnecting to Google Sheets: {sheet_name}")
    gc = get_gspread_client()
    try:
        sh = gc.open(sheet_name)
        print(f"  Opened existing sheet")
    except Exception:
        sh = gc.create(sheet_name)
        print(f"  Created new sheet")

    def tab(name):
        return get_or_add_worksheet(sh, name)

    # Tab 1: PRD Detail
    print("  Writing PRD_Detail...")
    ws1 = tab("PRD_Detail")
    ws1.clear()
    headers = [f"PRD Detail — {mn} 2026 | Plant 1300"]
    ws1.append_row(headers)
    ws1.append_row(["Line", "Order Type", "Material", "Description", "Work Center",
                    "GI_Material", "Scrap_Recovery", "GradeB", "GradeC",
                    "GR_Output", "D101_Direct_Mach", "D102_Direct_Labor",
                    "Indirect_Total", "Net_Cost"])
    ws1.append_rows(prd_detail, value_input_option="USER_ENTERED")

    # Tab 2: PRD Summary
    print("  Writing PRD_Summary...")
    ws2 = tab("PRD_Summary")
    ws2.clear()
    prd_sum = [[f"PRD Summary — {mn} 2026 | Plant 1300 | Prepared by: {PREPARED_BY}"],
               [],
               ["Line", "GI_Material", "Scrap_Recovery", "GradeB", "GradeC",
                "GR_Output", "D101_Direct_Mach", "D102_Direct_Labor",
                "Indirect_Total", "Net_Cost"]]
    for line in ["PK+CR", "GI"]:
        t = prd_totals[line]
        net = (t["GI_Material"] + t["D101_Direct_Mach"] + t["D102_Direct_Labor"]
               + t["Indirect_Total"] - t["Scrap_Recovery"] - t["GradeB"] - t["GradeC"])
        prd_sum.append([line,
                        round(t["GI_Material"], 2), round(t["Scrap_Recovery"], 2),
                        round(t["GradeB"], 2), round(t["GradeC"], 2),
                        round(t["GR_Output"], 2), round(t["D101_Direct_Mach"], 2),
                        round(t["D102_Direct_Labor"], 2), round(t["Indirect_Total"], 2),
                        round(net, 2)])
    prd_sum.append(["TOTAL",
                    round(grand["GI_Material"], 2), round(grand["Scrap_Recovery"], 2),
                    round(grand["GradeB"], 2), round(grand["GradeC"], 2),
                    round(grand["GR_Output"], 2), round(grand["D101_Direct_Mach"], 2),
                    round(grand["D102_Direct_Labor"], 2), round(grand["Indirect_Total"], 2),
                    round(gnet, 2)])
    ws2.append_rows(prd_sum, value_input_option="USER_ENTERED")

    # Tab 3: KSB1 Summary
    print("  Writing KSB1_Summary...")
    ws3 = tab("KSB1_Summary")
    ws3.clear()
    ksb1_sum = [[f"KSB1 Summary — {mn} 2026 | Plant 1300 | CC 1387xxx"], [],
                ["Cost Center", "CC Name", "Cost Element", "CE Name", "Amount (THB)"]]
    for cc in sorted(set(r[0] for r in ksb1_rows)):
        elem_totals = defaultdict(lambda: {"name": "", "total": 0.0})
        for r in ksb1_rows:
            if r[0] == cc and r[1]:
                elem_totals[r[1]]["name"]   = r[2]
                elem_totals[r[1]]["total"] += r[3]
        for e, info in sorted(elem_totals.items()):
            ksb1_sum.append([cc, CC_MAP.get(cc, ""), e, info["name"], round(info["total"], 2)])
        ksb1_sum.append([f"--- {cc} SUBTOTAL ---", CC_MAP.get(cc, ""), "", "",
                          round(ksb1_by_cc[cc], 2)])
        ksb1_sum.append([])
    ksb1_sum.append(["=== GRAND TOTAL ===", "", "", "", round(sum(ksb1_by_cc.values()), 2)])
    ws3.append_rows(ksb1_sum, value_input_option="USER_ENTERED")

    # Tab 4: GL Summary
    print("  Writing GL_Summary...")
    ws4 = tab("GL_Summary")
    ws4.clear()
    cat_order = ["Labor", "Electricity", "Repair & Maintenance", "Depreciation",
                 "Manufacturing Supplies", "Tools/Supplies/Other", "Other Expenses",
                 "Transport", "Consumption (Material)", "CO Settlement (to Orders)",
                 "Production Variance", "Variance Adjustment (ML)", "Other"]
    gl_sum = [[f"GL Summary — {mn} 2026 | CC 13* | Plant 1300"], [],
              ["Category", "GL Account", "Account Name", "Amount (THB)"]]
    gl_total = 0.0
    for cat in cat_order:
        items = [(a, i["name"], round(i["total"], 2))
                 for a, i in sorted(gl_by_account.items()) if i["category"] == cat]
        if not items: continue
        cat_total = sum(x[2] for x in items)
        for acc, name, amt in items:
            gl_sum.append([cat, acc, name, amt])
        gl_sum.append([f"--- {cat} SUBTOTAL ---", "", "", round(cat_total, 2)])
        gl_sum.append([])
        gl_total += cat_total
    gl_sum.append(["=== GRAND TOTAL ===", "", "", round(gl_total, 2)])
    ws4.append_rows(gl_sum, value_input_option="USER_ENTERED")

    # Tab 5: RECON Result
    print("  Writing RECON_Result...")
    ws5 = tab("RECON_Result")
    ws5.clear()
    recon_rows = build_recon_summary(month, prd_totals, grand, gnet,
                                     gl_by_account, cat_totals, ksb1_by_cc)
    ws5.append_rows(recon_rows, value_input_option="USER_ENTERED")

    # Clean up default Sheet1
    try:
        default = sh.worksheet("Sheet1")
        if len(sh.worksheets()) > 1:
            sh.del_worksheet(default)
    except Exception:
        pass

    print(f"\nDone! Sheet: {sh.url}")
    return sh.url


if __name__ == "__main__":
    main()
