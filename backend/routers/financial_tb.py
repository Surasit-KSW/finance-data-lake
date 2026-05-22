"""
routers/financial_tb.py
========================
Financial Trial Balance & Leadsheet endpoints — /api/v1/financial/...

Serves:
  - Trial Balance data (from Bronze raw Excel files via openpyxl)
  - Master TB account index mapping
  - Trigger leadsheet build script

Consumers: leadsheet builder, main-dashboard, fin-dashboard
"""
import subprocess
import sys
from pathlib import Path
from fastapi import APIRouter, HTTPException, BackgroundTasks, Header
import openpyxl

from backend.core.config import settings
from backend.services.db_service import query_df

router = APIRouter(prefix="/api/v1/financial", tags=["Financial v1"])

BRONZE = settings.BRONZE
TEMPLATES = BRONZE / "Templates"
NRV_DIR   = BRONZE / "Inventory_RollStock" / "NRV"


def _latest(directory: Path, pattern: str) -> Path | None:
    """Return the last (alphabetically) file matching pattern, or None."""
    matches = sorted(directory.glob(pattern))
    return matches[-1] if matches else None


# TB files keyed by period — versioned filenames resolved via glob at startup
TB_FILES: dict[str, Path | None] = {
    "2026-03-31": _latest(NRV_DIR, "AMC_TB_03.2026_v*.XLSX"),
    "2025-12-31": TEMPLATES / "AMC Leadsheet YE25.xlsx",
}

LEADSHEET_SCRIPT = settings.PROJECT_ROOT / "06_Scripts" / "leadsheet" / "build_q1_2026_leadsheet.py"


# ---------------------------------------------------------------------------
def _read_tb_from_sap_export(file: Path) -> list[dict]:
    """Read a SAP-style TB export (4-column format)."""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_raw = row[2]
        if acct_raw is None or not str(acct_raw).strip():
            continue
        if str(acct_raw).strip() == "Account Number":
            continue
        try:
            acct = str(int(float(str(acct_raw)))).strip()
        except (ValueError, TypeError):
            continue
        name = str(row[1]).strip() if row[1] else ""
        if name.startswith(acct):
            name = name[len(acct):].strip()
        records.append({
            "account_code": acct,
            "account_name": name,
            "balance":      float(row[3]) if isinstance(row[3], (int, float)) else 0.0,
            "fs_item":      str(row[0]) if row[0] else "",
        })
    wb.close()
    return records


def _read_tb_from_ye25(file: Path) -> list[dict]:
    """Extract Dec 2025 balances from YE25 leadsheet detail schedules (col G = Dec31 balance)."""
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    SKIP = {"F1", "F2", "F3"}
    records = {}
    for sh in wb.sheetnames:
        if sh in SKIP:
            continue
        ws = wb[sh]
        for row in ws.iter_rows(min_row=6, values_only=True):
            acct_raw = row[1] if len(row) > 1 else None
            if acct_raw is None:
                continue
            try:
                acct = str(int(float(str(acct_raw)))).strip()
            except (ValueError, TypeError):
                continue
            bal = row[6] if len(row) > 6 else None
            if isinstance(bal, (int, float)) and acct not in records:
                name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                records[acct] = {
                    "account_code": acct,
                    "account_name": name,
                    "balance":      float(bal),
                    "fs_item":      "",
                }
    wb.close()
    return list(records.values())


# ---------------------------------------------------------------------------
@router.get("/tb/{period}")
def get_trial_balance(period: str):
    """
    Trial Balance ณ วันที่ระบุ
    period format: YYYY-MM-DD (e.g. 2026-03-31, 2025-12-31)

    Source priority:
      1. PostgreSQL tb_data table (when DATABASE_URL is set — Vercel)
      2. Local Excel files (local dev with DuckDB)
    """
    # ── Cloud path: read from PostgreSQL tb_data ─────────────────────────────
    if settings.use_postgres:
        df = query_df(
            "SELECT account_code, account_name, balance, fs_item "
            "FROM tb_data WHERE period = ? ORDER BY account_code",
            [period],
        )
        if df.empty:
            raise HTTPException(
                status_code=404,
                detail=f"No TB data for period '{period}' in database. Run migrate_to_neon.py first.",
            )
        data = df.to_dict(orient="records")
        total = float(df["balance"].sum()) if not df.empty else 0.0
        return {
            "status":    "ok",
            "period":    period,
            "source":    "postgresql",
            "count":     len(data),
            "net_total": round(total, 2),
            "data":      data,
        }

    # ── Local path: read from Excel files (original behavior) ────────────────
    tb_file = TB_FILES.get(period)
    if not tb_file:
        available = list(TB_FILES.keys())
        raise HTTPException(
            status_code=404,
            detail=f"No TB data for period '{period}'. Available: {available}",
        )
    if not tb_file.exists():
        raise HTTPException(
            status_code=503,
            detail=f"TB file not found on disk: {tb_file.name}",
        )

    if period == "2025-12-31":
        data = _read_tb_from_ye25(tb_file)
    else:
        data = _read_tb_from_sap_export(tb_file)

    total = sum(r["balance"] for r in data)
    return {
        "status":    "ok",
        "period":    period,
        "source":    "excel",
        "count":     len(data),
        "net_total": round(total, 2),
        "data":      data,
    }


@router.get("/tb")
def list_available_periods():
    """รายการงวดที่มีข้อมูล Trial Balance"""
    if settings.use_postgres:
        df = query_df(
            "SELECT period, COUNT(*) AS account_count FROM tb_data GROUP BY period ORDER BY period"
        )
        result = [
            {"period": row["period"], "available": True, "account_count": row["account_count"], "source": "postgresql"}
            for _, row in df.iterrows()
        ]
    else:
        result = [
            {"period": p, "available": path.exists(), "file": path.name, "source": "excel"}
            for p, path in TB_FILES.items()
        ]
    return {"status": "ok", "periods": result}


@router.get("/master-tb")
def get_master_tb():
    """
    Master TB account index mapping
    Returns: account_code, account_name, index (schedule ref), caption
    Reads from Q1 2026 leadsheet Master TB sheet.
    """
    q126 = TEMPLATES / "AMC_Q12026_Leadsheet STAT to client_.xlsx"
    if not q126.exists():
        # fallback to Q1 2025
        q126 = TEMPLATES / "AMC_Q12025_Leadsheet STAT to client.xlsx"
    if not q126.exists():
        raise HTTPException(status_code=503, detail="Master TB file not found")

    wb = openpyxl.load_workbook(q126, read_only=True, data_only=True)
    ws = wb["Master TB"]
    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        acct_raw = row[0]
        if acct_raw is None:
            continue
        try:
            acct = str(int(float(str(acct_raw)))).strip()
        except (ValueError, TypeError):
            continue
        records.append({
            "account_code": acct,
            "account_name": str(row[1]).strip() if row[1] else "",
            "index":        str(row[2]).strip() if row[2] else "",
            "caption":      str(row[3]).strip() if row[3] else "",
            "balance_dec25": float(row[10]) if isinstance(row[10], (int, float)) else 0.0,
            "balance_mar26": float(row[11]) if isinstance(row[11], (int, float)) else 0.0,
        })
    wb.close()

    return {
        "status": "ok",
        "count":  len(records),
        "data":   records,
    }


@router.post("/leadsheet/build")
def trigger_leadsheet_build(
    background_tasks: BackgroundTasks,
    x_api_key: str | None = Header(None),
):
    """
    Trigger the Q1'2026 leadsheet builder script in the background.
    Returns immediately; check /api/v1/health for status.
    Requires X-Api-Key header when LEADSHEET_API_KEY is set in .env.
    """
    if settings.LEADSHEET_API_KEY and x_api_key != settings.LEADSHEET_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid or missing X-Api-Key")

    if not LEADSHEET_SCRIPT.exists():
        raise HTTPException(
            status_code=503,
            detail=f"Leadsheet script not found: {LEADSHEET_SCRIPT}",
        )

    def _run():
        subprocess.run(
            [sys.executable, str(LEADSHEET_SCRIPT)],
            cwd=str(settings.PROJECT_ROOT),
            capture_output=True,
        )

    background_tasks.add_task(_run)
    return {
        "status":  "accepted",
        "message": "Leadsheet build triggered in background",
        "script":  LEADSHEET_SCRIPT.name,
    }
